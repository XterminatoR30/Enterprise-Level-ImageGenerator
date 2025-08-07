import os
import requests
import json
import logging
import time
import zipfile
import io
import base64
import gradio as gr
import tempfile
import shutil
from PIL import Image
import replicate

# Enable AVIF support
try:
    import pillow_avif_plugin
    print("AVIF support enabled via pillow-avif-plugin")
except ImportError:
    try:
        import pillow_heif
        pillow_heif.register_heif_opener()
        print("AVIF support enabled via pillow-heif")
    except ImportError:
        print("No AVIF support available - install pillow-avif-plugin or pillow-heif")

from fastapi import FastAPI, HTTPException, Query, Body, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from typing import Optional, List, Dict, Tuple, Union
import uvicorn
from dotenv import load_dotenv
import asyncio
import torch
from torchvision import transforms
from transformers import AutoModelForImageSegmentation, pipeline
from transformers import AutoModelForImageSegmentation
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
from rembg import remove
from openai import OpenAI
import numpy as np
import cv2
from pathlib import Path
from PIL import ImageDraw, ImageFont, ImageFilter
from datetime import datetime
import pytz
import configparser
import boto3
import re
import traceback
import random
from io import BytesIO
import sys
import uuid
import pandas as pd
import pickle

from gdrive_debug_utils import (
    check_google_drive_dependencies_enhanced,
    create_google_drive_service_enhanced,
    get_or_create_folder_enhanced,
    upload_to_google_drive_enhanced,
    upload_multiple_files_to_google_drive_enhanced
)

from google_auth_oauthlib.flow import Flow
from googleapiclient.discovery import build
import mimetypes # Added for GDrive
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload # Added for GDrive
from google.oauth2.credentials import Credentials # Added for GDrive
from google_auth_oauthlib.flow import InstalledAppFlow, Flow # Added for GDrive
from google.auth.transport.requests import Request # Added for GDrive
from tqdm import tqdm

def Create_Service(client_secret_file, api_name, api_version, *scopes):
    print(client_secret_file, api_name, api_version, scopes, sep='-')
    CLIENT_SECRET_FILE = client_secret_file
    API_SERVICE_NAME = api_name
    API_VERSION = api_version
    SCOPES = [scope for scope in scopes[0]]
    print(SCOPES)

    cred = None

    pickle_file = f'token_{API_SERVICE_NAME}_{API_VERSION}.pickle'
    # print(pickle_file)

    if os.path.exists(pickle_file):
        with open(pickle_file, 'rb') as token:
            cred = pickle.load(token)

    if not cred or not cred.valid:
        if cred and cred.expired and cred.refresh_token:
            cred.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRET_FILE, SCOPES)
            cred = flow.run_local_server()

        with open(pickle_file, 'wb') as token:
            pickle.dump(cred, token)

    try:
        service = build(API_SERVICE_NAME, API_VERSION, credentials=cred)
        print(API_SERVICE_NAME, 'service created successfully')
        return service
    except Exception as e:
        print('Unable to connect.')
        print(e)
        return None

def convert_to_RFC_datetime(year=1900, month=1, day=1, hour=0, minute=0):
    dt = datetime(year, month, day, hour, minute, 0).isoformat() + 'Z'
    return dt

def get_gmt7_timestamp(include_seconds=False):
    """Get current timestamp in GMT+7 timezone"""
    gmt7 = pytz.timezone('Asia/Bangkok')
    now = datetime.now(gmt7)
    if include_seconds:
        return now.strftime('%Y-%m-%d %H:%M:%S')
    else:
        return now.strftime('%Y-%m-%d %H:%M')

def get_gmt7_filename_timestamp():
    """Get current timestamp for filenames in GMT+7 timezone without seconds"""
    gmt7 = pytz.timezone('Asia/Bangkok')
    now = datetime.now(gmt7)
    return now.strftime('%Y%m%d_%H%M')

# Import the process_generated_images function

# Process generated_images function defined inline below
# (replaces import from separate file)
PHOTOROOM_API_KEY = os.getenv("PHOTOROOM_API_KEY", "")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
openai_client = OpenAI(api_key=OPENAI_API_KEY)


REPLICATE_API_KEY = os.getenv("REPLICATE_API_TOKEN", "")
os.environ["REPLICATE_API_TOKEN"] = REPLICATE_API_KEY

# ======= IMAGEN-4 FUNCTIONS =======

def generate_single_image_imagen4(primary_image_bytes, reference_images_bytes_list, prompt, aspect_ratio="1:1", num_images=1, safety_filter_level="block_only_high", model="google/imagen-4"):
    """
    Generate images using Imagen-4 or Imagen-4-Ultra from Replicate with one primary image and reference images.
    
    Args:
        primary_image_bytes: Bytes of the primary image
        reference_images_bytes_list: List of reference image bytes (not used in basic Imagen-4)
        prompt: Text prompt for generation
        aspect_ratio: Image aspect ratio ("1:1", "16:9", "9:16", "3:4", "4:3")
        num_images: Number of images to generate (1-4 for Imagen-4, 1 for Imagen-4-Ultra)
        safety_filter_level: Safety filter level ("block_low_and_above", "block_medium_and_above", "block_only_high")
        model: Model to use ("google/imagen-4" or "google/imagen-4-ultra")
    
    Returns:
        List of base64 encoded images or None if failed
    """
    try:
        # Limit num_images based on model
        if model == "google/imagen-4-ultra":
            max_images = 1
        else:
            max_images = 4
        
        logger.info(f"Imagen-4 Generation - Model: {model}, Parameters: aspect_ratio={aspect_ratio}, num_images={num_images}, safety_filter_level={safety_filter_level}")
        
        # Enhance prompt for better style transfer
        enhanced_prompt = prompt
        if primary_image_bytes is not None:
            # Add style guidance to prompt for better adherence to reference image
            style_keywords = [
                "in the exact same style as the reference image",
                "matching the visual style and aesthetic",
                "maintaining the same art style and color palette",
                "with identical artistic technique"
            ]
            if not any(keyword in prompt.lower() for keyword in ["style", "aesthetic", "technique"]):
                enhanced_prompt = f"{prompt}, in the exact same style as the reference image"
                logger.info(f"🎨 Enhanced prompt for style transfer: {enhanced_prompt[:100]}...")
        
        # Prepare input for Imagen-4
        input_data = {
            "prompt": enhanced_prompt,
            "aspect_ratio": aspect_ratio,
            "safety_filter_level": safety_filter_level,
            "num_outputs": max(1, min(max_images, num_images))  # Ensure num_images is within model limits
        }
        
        # If we have a reference image, use it as style reference
        if primary_image_bytes is not None:
            try:
                # Convert bytes to base64 data URI for Replicate API
                style_reference_base64 = base64.b64encode(primary_image_bytes).decode('utf-8')
                data_uri = f"data:image/png;base64,{style_reference_base64}"
                input_data["style_reference_images"] = [data_uri]
                logger.info(f"✅ Imagen-4: Using style reference image with enhanced prompt")
            except Exception as ref_error:
                logger.warning(f"❌ Imagen-4: Could not use reference image: {ref_error}")
                logger.info("ℹ️ Imagen-4: Proceeding without reference image")
        
        # Generate image using Replicate API
        logger.info(f"🎨 Imagen-4: Generating image with model {model} and enhanced prompt: {enhanced_prompt[:100]}...")
        
        output = replicate.run(
            model,
            input=input_data
        )
        
        # Handle different types of responses from Replicate
        result_images = []
        if output:
            logger.info(f"📋 Imagen-4: Received output type: {type(output)}")
            
            # Handle list of URLs or other data (most common case)
            if isinstance(output, list):
                for item in output:
                    try:
                        if isinstance(item, str) and (item.startswith('http') or item.startswith('data:')):
                            # Download image from URL and convert to base64
                            import requests
                            response = requests.get(item)
                            response.raise_for_status()
                            img_b64 = base64.b64encode(response.content).decode('utf-8')
                            result_images.append(img_b64)
                            logger.info(f"✅ Imagen-4: Downloaded and converted image to base64")
                        elif isinstance(item, list):
                            # Handle pixel data arrays - this might be a numpy array or raw pixel data
                            logger.warning(f"❌ Imagen-4: Received pixel data array instead of image URL. Array length: {len(item)}")
                            logger.warning("This suggests the API response format has changed or there's an issue with the model")
                            continue
                        else:
                            logger.warning(f"❌ Imagen-4: Unexpected item format: {type(item)}, content: {str(item)[:100]}...")
                    except Exception as item_error:
                        logger.error(f"❌ Imagen-4: Error processing item: {item_error}")
                        
            # Handle single URL
            elif isinstance(output, str) and (output.startswith('http') or output.startswith('data:')):
                try:
                    import requests
                    response = requests.get(output)
                    response.raise_for_status()
                    img_b64 = base64.b64encode(response.content).decode('utf-8')
                    result_images.append(img_b64)
                    logger.info(f"✅ Imagen-4: Downloaded and converted single image to base64")
                except Exception as url_error:
                    logger.error(f"❌ Imagen-4: Error downloading from URL: {url_error}")
                    
            # Handle file-like object
            elif hasattr(output, 'read'):
                try:
                    image_data = output.read()
                    img_b64 = base64.b64encode(image_data).decode('utf-8')
                    result_images.append(img_b64)
                    logger.info(f"✅ Imagen-4: Read file object and converted to base64")
                except Exception as file_error:
                    logger.error(f"❌ Imagen-4: Error reading file object: {file_error}")
                    
            else:
                logger.warning(f"❌ Imagen-4: Unhandled output format: {type(output)}")
                logger.info(f"❌ Imagen-4: Output content: {str(output)[:200]}...")
        
        if result_images:
            logger.info(f"✅ Imagen-4: Successfully generated {len(result_images)} image(s)")
            return result_images
        else:
            logger.error(f"❌ Imagen-4: No valid images generated")
            return None
        
    except Exception as e:
        logger.error(f"❌ Error generating image with Imagen-4: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None

def safe_file_handler_imagen4(file_data):
    """
    Handle file data safely to avoid Windows path length issues for Imagen-4 processing.
    """
    import shutil  # Import shutil locally to ensure it's available
    
    if file_data is None:
        return None
    
    # If it's already bytes, return as-is
    if isinstance(file_data, bytes):
        return file_data
    
    # If it's a file path, read the content
    if isinstance(file_data, str):
        try:
            # For Windows, try to use short path if possible
            if os.name == 'nt' and len(file_data) > 200:
                # Try to copy to a shorter temp path first
                temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.tmp')
                temp_file.close()
                try:
                    shutil.copy2(file_data, temp_file.name)
                    with open(temp_file.name, 'rb') as f:
                        data = f.read()
                    os.unlink(temp_file.name)  # Clean up
                    return data
                except:
                    os.unlink(temp_file.name)  # Clean up even if failed
                    pass  # Fall back to original method
            
            with open(file_data, 'rb') as f:
                return f.read()
        except Exception as e:
            logger.error(f"Error reading file: {e}")
            return None
    
    # If it's a file-like object with a read method
    if hasattr(file_data, 'read'):
        try:
            if hasattr(file_data, 'seek'):
                file_data.seek(0)  # Reset to beginning
            return file_data.read()
        except Exception as e:
            logger.error(f"Error reading file object: {e}")
            return None
    
    return None

LEONARDO_API_KEY = os.getenv("LEONARDO_API_KEY")
if not LEONARDO_API_KEY:
    LEONARDO_API_KEY = ""
    logging.warning("Using fallback Leonardo API key. Please set LEONARDO_API_KEY environment variable for security.")

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Add a console handler to ensure logs are displayed
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)
logger.propagate = False  # Prevent duplicate logs

# Initialize FastAPI app
app = FastAPI(
    title=" Image Generator API",
    description="A wrapper for Leonardo AI API to generate images",
    version="1.0.0"
)

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Leonardo API base URL
LEONARDO_API_BASE_URL = ""

# Directory to save generated images
IMAGES_DIR = "generated_output"
os.makedirs(IMAGES_DIR, exist_ok=True)

# Theme and category mappings
THEME_CATEGORIES = {
    "Pets": [
        "Dogs", "Cats", "Alpaca", "Rabbit", "Polar Bear", "Panda", 
        "Hamster", "Tiger", "Turtle", "Hippo", "Ikan Cupang", 
        "Sugar Glider", "Capybara"
    ],
    "Sports": [
        "Football", "Basketball", "Tennis", "Running", "Paddel", 
        "Gym", "Pilates", "Yoga", "Cycling", "Hiking"
    ],
    "Hobbies": [
        "Car", "Motorbike", "Game", "Travelling", "Movies", 
        "Music", "Foodies", "Martial Arts"
    ],
    "Zodiac": [
        "Aries", "Taurus", "Gemini", "Cancer", "Leo", "Virgo", 
        "Libra", "Scorpio", "Sagittarius", "Capricorn", "Aquarius", "Pisces"
    ]
}

# Add theme and category mappings for filename convention after THEME_CATEGORIES definition
THEME_MAPPING = {
    "Pets": "01",
    "Sports": "02",
    "Hobbies": "03",
    "Zodiac": "04"
}

CATEGORY_MAPPING = {
    # Pets categories
    "Dogs": "001",
    "Cats": "002",
    "Alpaca": "003",
    "Rabbit": "004",
    "Polar Bear": "005",
    "Panda": "006",
    "Hamster": "007",
    "Tiger": "008",
    "Turtle": "009",
    "Hippo": "010",
    "Ikan Cupang": "011",
    "Sugar Glider": "012",
    "Capybara": "013",
    
    # Sports categories
    "Football": "001",
    "Basketball": "002",
    "Tennis": "003",
    "Running": "004",
    "Paddel": "005",
    "Gym": "006",
    "Pilates": "007",
    "Yoga": "008",
    "Cycling": "009",
    "Hiking": "010",
    
    # Hobbies categories
    "Car": "001",
    "Motorbike": "002",
    "Game": "003",
    "Travelling": "004",
    "Movies": "005",
    "Music": "006",
    "Foodies": "007",
    "Martial Arts": "008",
    
    # Zodiac categories
    "Aries": "001",
    "Taurus": "002",
    "Gemini": "003",
    "Cancer": "004",
    "Leo": "005",
    "Virgo": "006",
    "Libra": "007",
    "Scorpio": "008",
    "Sagittarius": "009",
    "Capricorn": "010",
    "Aquarius": "011",
    "Pisces": "012"
}

# Model name mapping
MODEL_NAMES = {
    "Flux Dev": "b2614463-296c-462a-9586-aafdb8f00e36",
    "Flux Schnell": "1dd50843-d653-4516-a8e3-f0238ee453ff",
    "AlbedoBase XL": "2067ae52-33fd-4a82-bb92-c2c55e7d2786",
    "Phoenix 1.0": "de7d3faf-762f-48e0-b3b7-9d0ac3a3fcf3",
}

# Image processing modes and preprocessor IDs
IMAGE_PROCESS_MODES = {
    "None": None,
    "Style Reference": 67,   # Style Reference preprocessor ID
    "Character Reference": 133,  # Character Reference preprocessor ID
    "Content Reference": 100  # Content Reference preprocessor ID
}

# Strength types
STRENGTH_TYPES = ["Low", "Mid", "High"]

# Preset styles mapping
PRESET_STYLES = {
    "Creative": "6fedbf1f-4a17-45ec-84fb-92fe524a29ef",
    "3D Render": "debdf72a-91a4-467b-bf61-cc02bdeb69c6",
}

# Store for generated images
generated_images = {}

# Initialize OpenAI client for Qwen
client = OpenAI(
    api_key=os.getenv("DASHSCOPE_API_KEY", ''),
    base_url="https://dashscope-intl.aliyuncs.com/compatible-mode/v1"
)

# Add Qwen helper functions
def encode_image(image_path):
    try:
        with open(image_path, "rb") as f:
            image_bytes = f.read()
        return base64.b64encode(image_bytes).decode('utf-8')
    except Exception as e:
        logger.error(f"Error in encode_image: {str(e)}")
        raise

def encode_image_to_base64_file(image_path, output_dir=None, include_metadata=True):
    """
    Encode image to base64 and save as .txt file
    
    Args:
        image_path (str): Path to the image file
        output_dir (str, optional): Directory to save base64 file. If None, saves in same directory as image
        include_metadata (bool): Whether to include image metadata in the base64 file
        
    Returns:
        str: Path to the created base64 .txt file, or None if failed
    """
    try:
        # Determine output directory
        if output_dir is None:
            output_dir = os.path.dirname(image_path)
        
        # Create base64 filename
        base_filename = os.path.splitext(os.path.basename(image_path))[0]
        base64_filename = f"{base_filename}_base64.txt"
        base64_path = os.path.join(output_dir, base64_filename)
        
        # Encode image to base64
        with open(image_path, "rb") as image_file:
            base64_content = base64.b64encode(image_file.read()).decode('utf-8')
        
        # Write only the base64 content without metadata headers
        with open(base64_path, 'w', encoding='utf-8') as f:
            f.write(base64_content)
        
        logger.info(f"Successfully encoded image to base64 file: {base64_path}")
        return base64_path
        
    except Exception as e:
        logger.error(f"Error encoding image to base64 file: {str(e)}")
        return None

def batch_encode_images_to_base64(image_paths, output_dir=None, include_metadata=True):
    """
    Encode multiple images to base64 files
    
    Args:
        image_paths (list): List of image file paths
        output_dir (str, optional): Directory to save base64 files
        include_metadata (bool): Whether to include metadata in base64 files
        
    Returns:
        list: List of created base64 file paths
    """
    base64_files = []
    
    for image_path in image_paths:
        # Handle different types of image paths/objects
        actual_path = None
        
        # Check if it's a string path
        if isinstance(image_path, str):
            actual_path = image_path
        # Check if it's a PIL Image object
        elif hasattr(image_path, 'save'):
            # It's a PIL Image, save it to temporary file first
            import tempfile
            from PIL import Image
            
            temp_dir = tempfile.mkdtemp()
            temp_filename = f"temp_image_{len(os.listdir(temp_dir))}.png"
            actual_path = os.path.join(temp_dir, temp_filename)
            
            try:
                image_path.save(actual_path)
                logger.info(f"Saved PIL Image to temporary file for base64: {actual_path}")
            except Exception as save_error:
                logger.error(f"Error saving PIL Image to file for base64: {str(save_error)}")
                continue
        # Check if it's a tuple/dict format
        elif isinstance(image_path, (tuple, list)) and len(image_path) > 0:
            actual_path = image_path[0] if isinstance(image_path[0], str) else None
        elif isinstance(image_path, dict) and 'path' in image_path:
            actual_path = image_path['path']
        else:
            logger.warning(f"Unsupported image path format for base64 encoding: {type(image_path)}")
            continue
        
        if actual_path and os.path.exists(actual_path):
            base64_file = encode_image_to_base64_file(actual_path, output_dir, include_metadata)
            if base64_file:
                base64_files.append(base64_file)
        else:
            logger.warning(f"Image file not found for base64 encoding: {actual_path} (original: {type(image_path)})")
    
    logger.info(f"Successfully encoded {len(base64_files)} images to base64 files")
    return base64_files

def inference_with_api(image_path, prompt, sys_prompt="You are a helpful assistant.", model_id="", min_pixels=512*28*28, max_pixels=2048*28*28):
    """Use Qwen API to generate a description from an image or text-only prompt"""
    try:
        # Add default instruction to the prompt
        default_instruction = "NO CROPPED IMAGE ARE ALLOWED OF ANY FORM. Whether it's cropped hair, face, hands, legs, limbs etc. "
        prompt = default_instruction + str(prompt)
        # Create a fresh message for each request
        messages = [
            {
                "role": "system",
                "content": [{"type": "text", "text": sys_prompt}]
            }
        ]
        
        # If image_path is provided, add image content
        if image_path:
            # Encode the image to base64
            base64_image = encode_image(image_path)
            
            # Add user message with both image and text
            messages.append({
                "role": "user",
                "content": [
                    {
                        "type": "image_url",
                        "min_pixels": min_pixels,
                        "max_pixels": max_pixels,
                        "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"},
                    },
                    {"type": "text", "text": prompt},
                ],
            })
            
            # Log the request to help with debugging
            logger.info(f"Sending image+text request to Qwen API for image: {image_path}")
        else:
            # Text-only request (no image)
            messages.append({
                "role": "user",
                "content": [{"type": "text", "text": prompt}],
            })
            
            # Log the request to help with debugging
            logger.info(f"Sending text-only request to Qwen API")
        
        retries = 3
        for attempt in range(retries):
            try:
                # Create a new completion request
                completion = client.chat.completions.create(
                    model=model_id,
                    messages=messages,
                    timeout=15  # timeout in seconds
                )
                
                # Extract and return the response
                response_text = completion.choices[0].message.content
                logger.info(f"Received response from Qwen API: {response_text[:100]}...")
                # Clean asterisks from the response
                if response_text:
                    # Apply basic asterisk cleaning
                    cleaned_response = response_text.strip()
                    # Remove leading and trailing asterisks
                    while cleaned_response.startswith('*') or cleaned_response.startswith('**'):
                        if cleaned_response.startswith('**'):
                            cleaned_response = cleaned_response[2:]
                        elif cleaned_response.startswith('*'):
                            cleaned_response = cleaned_response[1:]
                        cleaned_response = cleaned_response.strip()
                    
                    while cleaned_response.endswith('*') or cleaned_response.endswith('**'):
                        if cleaned_response.endswith('**'):
                            cleaned_response = cleaned_response[:-2]
                        elif cleaned_response.endswith('*'):
                            cleaned_response = cleaned_response[:-1]
                        cleaned_response = cleaned_response.strip()
                    
                    # Remove any remaining double asterisks within the text (markdown bold formatting)
                    cleaned_response = cleaned_response.replace('**', '')
                    
                    if cleaned_response != response_text:
                        logger.info(f"Cleaned asterisks from API response")
                    
                    return cleaned_response
                return response_text
            except Exception as inner_e:
                # If the error message contains "Connection error", retry
                if "Connection error" in str(inner_e):
                    print(f"Connection error on attempt {attempt+1}: {inner_e}. Retrying in 2 seconds...")
                    time.sleep(2)
                else:
                    raise
        raise Exception("Failed to complete API call after multiple retries due to connection errors.")
    except Exception as e:
        logger.error(f"Error in inference_with_api: {str(e)}")
        return "3D Cartoon, Plain White Background, Full Body Shot, detailed character with vibrant colors and distinctive features"

def detect_human_in_image(image_path):
    """Detect if an image contains human subjects using Qwen vision model"""
    try:
        image = Image.open(image_path).convert("RGB")
        image = image.resize((512, 512), Image.LANCZOS)
        
        logger.info(f"Detecting human subjects in image: {image_path}")
        detection_prompt = (
            "Look at this image carefully and focus ONLY on the visual content, completely ignoring any text, writing, or signs. "
            "Does this image contain any human beings, people, or human characters? "
            "Answer with only 'YES' if there are any humans, human faces, human bodies, or human-like characters visible. "
            "Answer with only 'NO' if there are no humans at all (only animals, objects, landscapes, etc.). "
            "Be very specific - even cartoon humans, anime characters, or stylized human figures count as humans. "
            "IGNORE any text content - focus only on the actual subjects in the image."
        )
        
        # Save resized image to a temporary file
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as temp_file:
            image.save(temp_file.name, format='PNG')
            temp_image_path = temp_file.name
        
        # Get detection result from API
        detection_result = inference_with_api(temp_image_path, detection_prompt)
        logger.info(f"Human detection result: {detection_result}")
        
        # Clean up temporary file
        try:
            os.unlink(temp_image_path)
        except Exception as e:
            logger.warning(f"Could not delete temporary file {temp_image_path}: {str(e)}")
        
        # Check if the result indicates humans are present
        contains_human = "YES" in detection_result.upper()
        logger.info(f"Contains human: {contains_human}")
        return contains_human
    
    except Exception as e:
        logger.error(f"Error detecting humans in image: {str(e)}")
        # Default to False (no humans) if detection fails
        return False

def detect_animal_breed(image_path):
    """Detect specific cat or dog breeds from the reference image using Qwen vision model"""
    try:
        image = Image.open(image_path).convert("RGB")
        image = image.resize((512, 512), Image.LANCZOS)
        
        logger.info(f"Detecting animal breed in image: {image_path}")
        
        # First detect if it's a cat or dog
        animal_detection_prompt = (
            "Look at this image carefully and focus ONLY on the visual content, completely ignoring any text, writing, or signs. "
            "What is the main subject? "
            "Answer with only 'CAT' if the main subject is a cat, "
            "Answer with only 'DOG' if the main subject is a dog, "
            "Answer with only 'NEITHER' if the main subject is neither a cat nor a dog. "
            "IGNORE any text content - focus only on the actual animals/subjects in the image."
        )
        
        # Save resized image to a temporary file
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as temp_file:
            image.save(temp_file.name, format='PNG')
            temp_image_path = temp_file.name
        
        # Get animal type detection result from API
        animal_type_result = inference_with_api(temp_image_path, animal_detection_prompt)
        logger.info(f"Animal type detection result: {animal_type_result}")
        
        breed = "Unknown"
        
        if "CAT" in animal_type_result.upper():
            # Detect cat breed
            cat_breed_prompt = (
                "Look at this cat image carefully and focus ONLY on the cat's physical features, completely ignoring any text, writing, or signs. "
                "What specific breed is this cat based on its visual characteristics? "
                "Choose ONLY from these options: Siamese, Anggora, British Shorthair, Munchkin, Persian, Tabby. "
                "Answer with only the breed name that best matches this cat's physical appearance. "
                "If you're unsure, choose the closest match from the list. "
                "IGNORE any text content - focus only on the cat's physical features like fur pattern, body shape, and facial structure."
            )
            breed_result = inference_with_api(temp_image_path, cat_breed_prompt)
            logger.info(f"Cat breed detection result: {breed_result}")
            
            # Extract breed from response
            cat_breeds = ["Siamese", "Anggora", "British Shorthair", "Munchkin", "Persian", "Tabby"]
            for cat_breed in cat_breeds:
                if cat_breed.upper() in breed_result.upper():
                    breed = f"Cat - {cat_breed}"
                    break
            
        elif "DOG" in animal_type_result.upper():
            # Detect dog breed
            dog_breed_prompt = (
                "Look at this dog image carefully and focus ONLY on the dog's physical features, completely ignoring any text, writing, or signs. "
                "What specific breed is this dog based on its visual characteristics? "
                "Choose ONLY from these options: Pomeranian, Labrador, French Bulldog, Husky, Golden Retriever. "
                "Answer with only the breed name that best matches this dog's physical appearance. "
                "If you're unsure, choose the closest match from the list. "
                "IGNORE any text content - focus only on the dog's physical features like size, fur type, body shape, and facial structure."
            )
            breed_result = inference_with_api(temp_image_path, dog_breed_prompt)
            logger.info(f"Dog breed detection result: {breed_result}")
            
            # Extract breed from response
            dog_breeds = ["Pomeranian", "Labrador", "French Bulldog", "Husky", "Golden Retriever"]
            for dog_breed in dog_breeds:
                if dog_breed.upper() in breed_result.upper():
                    breed = f"Dog - {dog_breed}"
                    break
        
        # Clean up temporary file
        try:
            os.unlink(temp_image_path)
        except Exception as e:
            logger.warning(f"Could not delete temporary file {temp_image_path}: {str(e)}")
        
        logger.info(f"Detected breed: {breed}")
        return breed
    
    except Exception as e:
        logger.error(f"Error detecting animal breed: {str(e)}")
        return "Unknown"

def generate_prompt_from_image(image_path):
    """Generate a detailed prompt from an image using Qwen"""
    try:
        image = Image.open(image_path).convert("RGB")
        image = image.resize((512, 512), Image.LANCZOS)
        
        logger.info(f"Generating prompt for image: {image_path}")
        
        # First detect if the image contains humans
        contains_human = detect_human_in_image(image_path)
        
        # Detect animal breed for non-human subjects
        detected_breed = "Unknown"
        breed_info = ""
        if not contains_human:
            detected_breed = detect_animal_breed(image_path)
            logger.info(f"🐾 BREED DETECTION IN PROMPT GENERATION: {detected_breed}")
            
            # Create breed-specific information to inject into the prompt
            if detected_breed != "Unknown" and detected_breed != "Neither":
                if "Cat -" in detected_breed:
                    breed_name = detected_breed.replace("Cat - ", "")
                    breed_info = f" This is specifically a {breed_name} cat breed. "
                    logger.info(f"🐱 Injecting cat breed info: {breed_info.strip()}")
                elif "Dog -" in detected_breed:
                    breed_name = detected_breed.replace("Dog - ", "")
                    breed_info = f" This is specifically a {breed_name} dog breed. "
                    logger.info(f"🐶 Injecting dog breed info: {breed_info.strip()}")
        
        # Detect if it's an animal or object for non-human subjects
        is_animal = False
        if not contains_human and detected_breed != "Unknown" and detected_breed != "Neither":
            is_animal = True
        
        # Adjust prompt based on subject type
        if contains_human:
            # For human subjects, don't use "3D Cartoon" at the beginning
            prompt_request = (
                "Describe this image focusing ONLY on the visual and physical characteristics of the subject, "
                "always add 'Plain White Background', ', Full Body Shot' at the very beginning of the prompt. "
                "Ensure proportionality of head and body to 1:1, never make any character with a head that is too big or too small. "
                "Focus on colors, textures, facial expressions, clothing, poses, and distinctive physical features. One paragraph maximum. No Hallucinations and No more than one character. "
                "CRITICAL TEXT EXCLUSION RULES: "
                "- COMPLETELY IGNORE any text, writing, signs, labels, letters, numbers, or written words visible in the image "
                "- DO NOT mention any readable text content, brand names, or textual elements "
                "- DO NOT describe what any signs, labels, or text say "
                "- Focus ONLY on visual elements like colors, shapes, clothing, expressions, poses, and physical appearance "
                "OBJECT RULES: "
                "1. Any objects in the scene must be appropriately sized and NEVER larger than the main subject. "
                "2. Limit the number of objects to 1-2 maximum. "
                "3. The main subject must always be the focal point and dominant element in the image. "
                "4. Objects should be proportionate and realistic in size compared to the character. "
                "5. Do not add extra objects that weren't in the original image. "
                "6. If there are multiple objects in the original image, only include the 1-2 most important ones. "
            )
            logger.info("Human detected - using text-insensitive prompt without '3D Cartoon'")
        elif is_animal:
            # For animal subjects, use "3D Cartoon" and "Full Body Shot" with breed info
            prompt_request = (
                "Describe this image focusing ONLY on the visual and physical characteristics of the subject, "
                "always add '3D Cartoon', ', Plain White Background', ', Full Body Shot' at the very beginning of the prompt. "
                "Ensure proportionality of head and body to 1:1, never make any character with a head that is too big or too small. "
                "Focus on colors, textures, facial expressions, poses, and distinctive physical features. One paragraph maximum. No Hallucinations and No more than one character. "
                f"{breed_info}"  # Inject breed information here
                "CRITICAL TEXT EXCLUSION RULES: "
                "- COMPLETELY IGNORE any text, writing, signs, labels, letters, numbers, or written words visible in the image "
                "- DO NOT mention any readable text content, brand names, or textual elements "
                "- DO NOT describe what any signs, labels, or text say "
                "- Focus ONLY on visual elements like colors, shapes, fur patterns, expressions, poses, and physical appearance "
                "OBJECT RULES: "
                "1. Any objects in the scene must be appropriately sized and NEVER larger than the main subject. "
                "2. Limit the number of objects to 1-2 maximum. "
                "3. The main subject must always be the focal point and dominant element in the image. "
                "4. Objects should be proportionate and realistic in size compared to the character. "
                "5. Do not add extra objects that weren't in the original image. "
                "6. If there are multiple objects in the original image, only include the 1-2 most important ones. "
            )
            if breed_info:
                logger.info(f"Animal detected - using text-insensitive prompt with '3D Cartoon', 'Full Body Shot' and breed info: {breed_info}")
            else:
                logger.info("Animal detected - using text-insensitive prompt with '3D Cartoon' and 'Full Body Shot'")
        else:
            # For non-living objects, use "Full Frame Shot" instead of "Full Body Shot" and no body part references
            prompt_request = (
                "Describe this image focusing ONLY on the visual and physical characteristics of the subject, "
                "always add 'Plain White Background', ', Full Frame Shot' at the very beginning of the prompt. "
                "Focus on colors, textures, materials, shapes, and distinctive physical properties. One paragraph maximum. No Hallucinations and focus on one main subject. "
                "DO NOT include any body part references like eyes, mouth, ears, face, hands, feet, or facial expressions. "
                "DO NOT use gender-specific language or character references. "
                "CRITICAL TEXT EXCLUSION RULES: "
                "- COMPLETELY IGNORE any text, writing, signs, labels, letters, numbers, or written words visible in the image "
                "- DO NOT mention any readable text content, brand names, or textual elements "
                "- DO NOT describe what any signs, labels, or text say "
                "- Focus ONLY on visual elements like colors, shapes, materials, textures, and physical properties "
                "OBJECT RULES: "
                "1. Any additional objects in the scene must be appropriately sized and NEVER larger than the main subject. "
                "2. Limit the number of objects to 1-2 maximum. "
                "3. The main subject must always be the focal point and dominant element in the image. "
                "4. Objects should be proportionate and realistic in size. "
                "5. Do not add extra objects that weren't in the original image. "
                "6. If there are multiple objects in the original image, only include the 1-2 most important ones. "
            )
            logger.info("Non-living object detected - using text-insensitive prompt with 'Full Frame Shot' and no body part references")
        
        # Save resized image to a temporary file
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as temp_file:
            image.save(temp_file.name, format='PNG')
            temp_image_path = temp_file.name
        
        # Get prompt from API - ensure we're passing the correct prompt_request
        generated_prompt = inference_with_api(temp_image_path, prompt_request)
        logger.info(f"Generated prompt: {generated_prompt}")
        
        # ALWAYS ensure gender is explicitly mentioned in the generated prompt for humans only
        if generated_prompt and contains_human:
            logger.info("🚀 Ensuring gender is explicitly mentioned in generated prompt")
            generated_prompt = ensure_gender_in_prompt(generated_prompt.strip(), contains_human=True)
            logger.info(f"✅ Gender-enhanced generated prompt: {generated_prompt[:100]}...")
        elif generated_prompt and not contains_human:
            logger.info("🐾 Non-human subject detected - skipping gender enhancement")
            generated_prompt = generated_prompt.strip()
        
        # Clean up temporary file
        try:
            os.unlink(temp_image_path)
        except Exception as e:
            logger.warning(f"Could not delete temporary file {temp_image_path}: {str(e)}")
        
        return generated_prompt.strip() if generated_prompt else "Plain White Background, Full Body Shot, detailed character with vibrant colors and distinctive features"
    
    except Exception as e:
        logger.error(f"Error generating prompt from image: {str(e)}")
        # Return a more generic default prompt that doesn't specifically mention 3D Cartoon for safety
        return "Plain White Background, Full Body Shot, detailed character with vibrant colors and distinctive features"

# Global variables for BiRefNet_HR model
_birefnet_hr_model = None
_birefnet_hr_transform = None

def get_birefnet_hr_model():
    global _birefnet_hr_model, _birefnet_hr_transform
    if _birefnet_hr_model is None:
        device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
        _birefnet_hr_model = AutoModelForImageSegmentation.from_pretrained(
            'ZhengPeng7/BiRefNet_HR',
            trust_remote_code=True,
            torch_dtype=torch.float32
        ).to(device)
        if not hasattr(_birefnet_hr_model.config, "get_text_config"):
            _birefnet_hr_model.config.get_text_config = lambda: None
        _birefnet_hr_model.eval()
        _birefnet_hr_transform = transforms.Compose([
            transforms.Resize((2048, 2048)),
            transforms.ToTensor(),
            transforms.Normalize([0.485, 0.456, 0.406], [0.229, 0.224, 0.225])
        ])
    return _birefnet_hr_model, _birefnet_hr_transform

def remove_background_birefnet_hr(input_path):
    try:
        model, transform_img = get_birefnet_hr_model()
        device = next(model.parameters()).device
        
        # Load the image and ensure it's in RGB mode for processing
        img = Image.open(input_path).convert("RGB")
        
        # Transform and predict
        t_in = transform_img(img).unsqueeze(0).to(device)
        with torch.no_grad():
            preds = model(t_in)[-1].sigmoid()
            mask = preds[0].squeeze().cpu()
        
        # Convert the mask to a PIL image and resize to match the original image
        mask_pil = transforms.ToPILImage()(mask).resize(img.size, Image.LANCZOS)
        
        # Enhanced mask processing for smoother results
        from PIL import ImageEnhance, ImageFilter
        
        # Apply subtle contrast enhancement instead of aggressive
        mask_pil = ImageEnhance.Contrast(mask_pil).enhance(1.2)
        
        # Apply subtle sharpening to preserve edge details
        mask_pil = mask_pil.filter(ImageFilter.UnsharpMask(radius=1, percent=50, threshold=3))
        
        # Create a new transparent image
        out = img.copy()
        
        # Apply the alpha mask
        out.putalpha(mask_pil)
        
        # Convert to RGBA for alpha processing
        out = out.convert('RGBA')
        
        # Apply improved alpha edge processing with reduced blur for sharper results
        out = improve_alpha_edges(out, threshold=180, edge_feather=4, use_gaussian_blur=True, feather_intensity=0.6)
        
        # Save to a temporary file to ensure alpha channel is preserved
        with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp:
            out.save(tmp.name, 'PNG')
            # Load it back to ensure proper alpha channel data
            result = Image.open(tmp.name).convert('RGBA')
            
            # Clean up the temporary file
            try:
                os.unlink(tmp.name)
            except:
                pass
                
        return result
    except Exception as e:
        logger.error(f"remove_background_birefnet_hr: {e}")
        logger.error(traceback.format_exc())
        return None

def improve_alpha_edges(image, threshold=180, edge_feather=4, use_gaussian_blur=True, feather_intensity=0.6):
    """
    Improve alpha channel edges by removing greyish-white artifacts and creating smoother transparency.
    Enhanced version with better smoothing for card applications.
    
    Args:
        image: PIL Image in RGBA mode
        threshold: Pixel brightness threshold for edge detection (0-255) - optimized for better artifact removal
        edge_feather: Number of pixels to feather the edges (2-6 recommended for smoother results)
        use_gaussian_blur: Whether to apply subtle Gaussian blur for smoother edges
        feather_intensity: Intensity of feathering effect (0.0-1.0) - reduced for less blur
    
    Returns:
        PIL Image with improved alpha channel
    """
    try:
        import numpy as np
        from PIL import Image, ImageFilter
        
        if image.mode != 'RGBA':
            image = image.convert('RGBA')
        
        # Convert to numpy array for processing
        img_array = np.array(image)
        
        # Separate RGB and Alpha channels
        rgb = img_array[:, :, :3]
        alpha = img_array[:, :, 3]
        
        # Create a copy for processing
        new_alpha = alpha.copy()
        
        # Step 1: Enhanced artifact removal with better thresholding
        # Find pixels that are mostly transparent but not fully transparent
        semi_transparent = (alpha > 0) & (alpha < 180)
        
        # For semi-transparent pixels, check if they're likely edge artifacts
        # by looking at their RGB values
        if np.any(semi_transparent):
            # Calculate brightness of semi-transparent pixels
            brightness = np.mean(rgb[semi_transparent], axis=1)
            
            # Enhanced artifact detection: also check for uniform color (greyish artifacts)
            semi_transparent_coords = np.where(semi_transparent)
            for i in range(len(semi_transparent_coords[0])):
                row, col = semi_transparent_coords[0][i], semi_transparent_coords[1][i]
                pixel_brightness = np.mean(rgb[row, col])
                
                # Check if pixel is too bright OR if it's a uniform greyish color
                is_bright_artifact = pixel_brightness > threshold
                is_uniform_grey = (np.std(rgb[row, col]) < 15) and (pixel_brightness > 160)
                
                if is_bright_artifact or is_uniform_grey:
                    new_alpha[row, col] = 0
        
        # Step 2: Improve edge quality by applying alpha threshold
        # Make pixels with very low alpha fully transparent
        new_alpha[new_alpha < 25] = 0
        
        # Step 3: Create cleaner edges by processing border pixels
        # Find edge pixels (pixels with alpha > 0 that are adjacent to transparent pixels)
        from scipy import ndimage
        
        # Create binary mask of non-transparent pixels
        solid_mask = new_alpha > 0
        
        # Find edge pixels using morphological operations
        eroded = ndimage.binary_erosion(solid_mask, structure=np.ones((3, 3)))
        edge_pixels = solid_mask & ~eroded
        
        # For edge pixels, apply stronger alpha if they're not artifacts
        edge_coords = np.where(edge_pixels)
        for i in range(len(edge_coords[0])):
            row, col = edge_coords[0][i], edge_coords[1][i]
            
            # Check if this edge pixel is likely an artifact
            pixel_brightness = np.mean(rgb[row, col])
            
            if pixel_brightness > threshold:
                # Likely an artifact, reduce alpha significantly
                new_alpha[row, col] = min(new_alpha[row, col], 40)
            else:
                # Likely a real edge, ensure it's solid
                if new_alpha[row, col] > 80:
                    new_alpha[row, col] = 255
        
        # Step 4: Enhanced edge feathering with better smoothing
        if edge_feather > 0:
            # Create a distance transform for smooth falloff
            distance_transform = ndimage.distance_transform_edt(~solid_mask)
            
            # Apply feathering to pixels within the feather range
            feather_mask = (distance_transform > 0) & (distance_transform <= edge_feather)
            
            if np.any(feather_mask):
                # Calculate feathering factor based on distance with improved curve
                feather_coords = np.where(feather_mask)
                for i in range(len(feather_coords[0])):
                    row, col = feather_coords[0][i], feather_coords[1][i]
                    distance = distance_transform[row, col]
                    
                    # Only apply feathering if the pixel isn't already processed as an artifact
                    if new_alpha[row, col] > 0:
                        # Create smoother falloff using cosine curve for more natural transition
                        normalized_distance = distance / edge_feather
                        falloff_factor = max(0, 0.5 * (1 + np.cos(np.pi * normalized_distance)))
                        falloff_factor = falloff_factor * feather_intensity + (1 - feather_intensity)
                        new_alpha[row, col] = int(new_alpha[row, col] * falloff_factor)
        
        # Step 5: Final cleanup - remove isolated pixels
        # Remove small isolated transparent regions that might be artifacts
        labeled_array, num_features = ndimage.label(new_alpha > 0)
        
        if num_features > 1:
            # Find the largest connected component (main subject)
            component_sizes = ndimage.sum(new_alpha > 0, labeled_array, range(1, num_features + 1))
            largest_component = np.argmax(component_sizes) + 1
            
            # Remove small components (likely artifacts)
            min_size = max(50, np.max(component_sizes) * 0.005)  # Reduced to 0.5% of largest component
            
            for i in range(1, num_features + 1):
                if i != largest_component and component_sizes[i - 1] < min_size:
                    new_alpha[labeled_array == i] = 0
        
        # Reconstruct the image
        improved_array = img_array.copy()
        improved_array[:, :, 3] = new_alpha
        
        # Convert back to PIL Image
        improved_image = Image.fromarray(improved_array, 'RGBA')
        
        # Step 6: Apply Gaussian blur to alpha channel for antialiasing (if enabled)
        if use_gaussian_blur:
            alpha = improved_image.split()[3]
            # Apply very subtle Gaussian blur for smoother edges without losing details
            alpha = alpha.filter(ImageFilter.GaussianBlur(radius=0.8))
            improved_image.putalpha(alpha)
            logger.info("🎨 APPLIED: Minimal Gaussian blur for smoother edges")
        
        # Step 7: Optional super-sampling for ultra-smooth edges (advanced technique)
        # This creates a higher resolution alpha, applies blur, then downsamples back
        if use_gaussian_blur and edge_feather >= 3:  # Reduced threshold for better quality
            try:
                alpha = improved_image.split()[3]
                # Resize alpha to 2x resolution
                large_alpha = alpha.resize((alpha.width*2, alpha.height*2), Image.LANCZOS)
                # Apply subtle blur at higher resolution for less blurry results
                large_alpha = large_alpha.filter(ImageFilter.GaussianBlur(radius=1.2))
                # Downsample back with antialiasing
                small_alpha = large_alpha.resize(alpha.size, Image.LANCZOS)
                improved_image.putalpha(small_alpha)
                logger.info("🎨 APPLIED: Enhanced super-sampling technique for smoother edges")
            except Exception as e:
                logger.warning(f"⚠️ Super-sampling failed, continuing with standard processing: {e}")
        
        logger.info("✅ ALPHA IMPROVEMENT: Successfully improved alpha channel edges with enhanced smoothing")
        return improved_image
        
    except Exception as e:
        logger.error(f"❌ ALPHA IMPROVEMENT ERROR: {str(e)}")
        logger.error(f"❌ ALPHA IMPROVEMENT TRACEBACK: {traceback.format_exc()}")
        # Return original image if processing fails
        return image

@app.get("/")
def read_root():
    return {"message": " Image Generator API", "status": "running", "api_key_set": bool(LEONARDO_API_KEY)}

@app.get("/generate-image")
async def generate_image(
    prompt: str = Query(..., description="The prompt to generate the image"),
    model_id: Optional[str] = Query("6bef9f1b-29cb-40c7-b9df-32b51c1f67d3", description="Leonardo AI model ID"),
    width: int = Query(512, description="Image width"),
    height: int = Query(512, description="Image width"),
    photo_real: bool = Query(False, description="Use PhotoReal feature"),
    image_prompt_id: Optional[str] = Query(None, description="Image ID to use as reference"),
    init_image_id: Optional[str] = Query(None, description="Init image ID for image-to-image generation"),
    init_strength: Optional[float] = Query(0.5, description="Init strength for image-to-image (0.1-0.9)")
):
    """
    Generate an image using the Leonardo AI API
    """
    
    if not LEONARDO_API_KEY:
        raise HTTPException(status_code=500, detail="API key not configured")
    
    headers = {
        "Authorization": f"Bearer {LEONARDO_API_KEY}",
        "Content-Type": "application/json"
    }
    
    # Prepare the payload based on the PhotoReal setting
    payload = {
        "prompt": prompt,
        "width": width,
        "height": height
    }
    
    # Add model ID
    payload["modelId"] = model_id
    
    # Add PhotoReal settings if enabled
    if photo_real:
        payload["photoReal"] = True
        payload["photoRealVersion"] = "v2"
        payload["alchemy"] = True
        payload["presetStyle"] = "CINEMATIC"
        
        # PhotoReal v2 requires compatible model (Leonardo Kino XL)
        if model_id == "6bef9f1b-29cb-40c7-b9df-32b51c1f67d3":
            payload["modelId"] = "aa77f04e-3eec-4034-9c07-d0f619684628"  # Leonardo Kino XL
    
    # Add reference image if provided
    if image_prompt_id:
        payload["imagePrompts"] = [{"id": image_prompt_id, "weight": 0.5}]
    
    # Add init image for image-to-image generation if provided
    if init_image_id:
        payload["init_image_id"] = init_image_id
        payload["init_strength"] = max(0.1, min(0.9, init_strength))  # Clamp between 0.1 and 0.9
    
    try:
        logger.info(f"Sending request to Leonardo API: {payload}")
        response = requests.post(
            f"{LEONARDO_API_BASE_URL}/generations",
            headers=headers,
            json=payload
        )
        
        response.raise_for_status()
        generation_data = response.json()
        logger.info(f"Generation initiated with ID: {generation_data['sdGenerationJob']['generationId']}")
        return generation_data
    except requests.exceptions.RequestException as e:
        logger.error(f"Error calling Leonardo API: {str(e)}")
        if hasattr(e, 'response') and e.response:
            error_detail = e.response.text
            status_code = e.response.status_code
        else:
            error_detail = str(e)
            status_code = 500
        
        raise HTTPException(status_code=status_code, detail=error_detail)

@app.get("/generation/{generation_id}")
async def get_generation(generation_id: str, wait: bool = Query(False, description="Wait for generation to complete")):
    """
    Get the status and result of a generation by ID
    """
    if not LEONARDO_API_KEY:
        raise HTTPException(status_code=500, detail="API key not configured")
    
    headers = {
        "Authorization": f"Bearer {LEONARDO_API_KEY}",
        "Content-Type": "application/json"
    }
    
    try:
        # If wait parameter is True, wait for generation to complete
        if wait:
            logger.info(f"Waiting for generation {generation_id} to complete...")
            time.sleep(20)  # Wait for 20 seconds as in the example
        
        response = requests.get(
            f"{LEONARDO_API_BASE_URL}/generations/{generation_id}",
            headers=headers
        )
        
        response.raise_for_status()
        generation_data = response.json()
        
        # If generation is complete and has generated images, save them
        if 'generations_by_pk' in generation_data and generation_data['generations_by_pk']['status'] == 'COMPLETE':
            if 'generated_images' in generation_data['generations_by_pk']:
                # Save references to generated images
                images = []
                for img in generation_data['generations_by_pk']['generated_images']:
                    img_url = img.get('url')
                    if img_url:
                        img_id = img.get('id', 'unknown')
                        images.append({
                            'id': img_id,
                            'url': img_url
                        })
                        
                if images:
                    generated_images[generation_id] = images
                    
        return generation_data
    except requests.exceptions.RequestException as e:
        logger.error(f"Error calling Leonardo API: {str(e)}")
        if hasattr(e, 'response') and e.response:
            error_detail = e.response.text
            status_code = e.response.status_code
        else:
            error_detail = str(e)
            status_code = 500
        
        raise HTTPException(status_code=status_code, detail=error_detail)

@app.post("/upload-reference-image")
async def upload_reference_image(file: UploadFile = File(...)):
    """
    Upload a reference image to Leonardo AI using the two-step process
    """
    if not LEONARDO_API_KEY:
        raise HTTPException(status_code=500, detail="API key not configured")
    
    if not file.filename.lower().endswith(('.jpg', '.jpeg', '.png', '.avif')):
        raise HTTPException(status_code=400, detail="Only JPG, JPEG, PNG, and AVIF files are allowed")
    
    try:
        # Get file extension
        extension = file.filename.split('.')[-1].lower()
        
        # Handle AVIF conversion
        if extension == 'avif':
            # Create a temporary file for the AVIF
            with tempfile.NamedTemporaryFile(delete=False, suffix='.avif') as avif_temp_file:
                shutil.copyfileobj(file.file, avif_temp_file)
                avif_temp_path = avif_temp_file.name
            
            # Convert AVIF to PNG
            png_temp_path = avif_temp_path.rsplit('.', 1)[0] + '.png'
            converted_path = convert_avif(avif_temp_path, png_temp_path, 'PNG')
            
            if converted_path == avif_temp_path:
                # Conversion failed, clean up and raise error
                try:
                    os.unlink(avif_temp_path)
                except:
                    pass
                raise HTTPException(status_code=400, detail="Failed to convert AVIF file")
            
            # Clean up original AVIF file
            try:
                os.unlink(avif_temp_path)
            except:
                pass
            
            # Update extension and file path for further processing
            extension = 'png'
            temp_file_path = converted_path
            
            # Create a new file-like object for the converted PNG
            with open(converted_path, 'rb') as converted_file:
                file_content = converted_file.read()
        else:
            # For non-AVIF files, create temporary file as before
            with tempfile.NamedTemporaryFile(delete=False, suffix=f'.{extension}') as temp_file:
                shutil.copyfileobj(file.file, temp_file)
                temp_file_path = temp_file.name
            
            with open(temp_file_path, 'rb') as temp_file_obj:
                file_content = temp_file_obj.read()
        
        # Step 1: Get a presigned URL for uploading
        headers = {
            "Authorization": f"Bearer {LEONARDO_API_KEY}",
            "Content-Type": "application/json",
            "accept": "application/json"
        }
        
        payload = {"extension": extension}
        
        presigned_response = requests.post(
            f"{LEONARDO_API_BASE_URL}/init-image",
            json=payload,
            headers=headers
        )
        
        presigned_response.raise_for_status()
        presigned_data = presigned_response.json()
        
        logger.info(f"Presigned URL response: {presigned_data}")
        
        # Extract upload information
        upload_url = presigned_data['uploadInitImage']['url']
        upload_fields = json.loads(presigned_data['uploadInitImage']['fields'])
        image_id = presigned_data['uploadInitImage']['id']
        
        # Step 2: Upload the image to the presigned URL
        # Upload to the presigned URL using the file content
        files = {'file': (f'image.{extension}', file_content, f'image/{extension}')}
        upload_response = requests.post(
            upload_url,
            data=upload_fields,
            files=files
        )
        
        # Clean up the temporary file
        try:
            time.sleep(0.5)  # Give some time before deleting
            os.unlink(temp_file_path)
        except Exception as e:
            logger.warning(f"Could not delete temporary file {temp_file_path}: {str(e)}")
        
        upload_response.raise_for_status()
        
        # Return the image ID
        return {"id": image_id, "message": "Image uploaded successfully"}
    except Exception as e:
        logger.error(f"Error uploading image: {str(e)}")
        if hasattr(e, 'response') and e.response:
            error_detail = e.response.text
            status_code = e.response.status_code
            logger.error(f"Response details: {error_detail}")
        else:
            error_detail = str(e)
            status_code = 500
        
        raise HTTPException(status_code=status_code, detail=error_detail)

@app.get("/models")
async def get_models():
    """
    Get available Leonardo AI models
    """
    if not LEONARDO_API_KEY:
        raise HTTPException(status_code=500, detail="API key not configured")
    
    headers = {
        "Authorization": f"Bearer {LEONARDO_API_KEY}",
        "Content-Type": "application/json"
    }
    
    try:
        response = requests.get(
            f"{LEONARDO_API_BASE_URL}/models",
            headers=headers
        )
        
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        logger.error(f"Error calling Leonardo API: {str(e)}")
        if hasattr(e, 'response') and e.response:
            error_detail = e.response.text
            status_code = e.response.status_code
        else:
            error_detail = str(e)
            status_code = 500
        
        raise HTTPException(status_code=status_code, detail=error_detail)

@app.get("/image/test")
async def get_test_image():
    """
    Return the test.jpg image file
    """
    image_path = "test.jpg"
    
    if not os.path.exists(image_path):
        raise HTTPException(status_code=404, detail="Image not found")
        
    return FileResponse(
        image_path, 
        media_type="image/jpeg",
        filename="test.jpg"
    )

@app.get("/download-images/{generation_id}")
async def download_images(generation_id: str):
    """
    Download all images from a generation as a zip file
    """
    if generation_id not in generated_images:
        raise HTTPException(status_code=404, detail="Generated images not found")
    
    # Create a zip file in memory
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for i, image in enumerate(generated_images[generation_id]):
            try:
                # Download the image
                response = requests.get(image['url'])
                response.raise_for_status()
                
                # Add the image to the zip file
                filename = f"image_{i+1}.png"
                zip_file.writestr(filename, response.content)
            except Exception as e:
                logger.error(f"Error downloading image {image['url']}: {str(e)}")
                continue
    
    # Reset file pointer
    zip_buffer.seek(0)
    
    # Return the zip file
    return StreamingResponse(
        zip_buffer, 
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=generation_{generation_id}.zip"}
    )

# Add Ideogram API Key
# IDEOGRAM_API_KEY = "zshbRFLd-WJ_IYW0KdTRbBN_jbSUVRZF_yY64GMs6uTE-vwE24s6t59WWwIHaIMBU3unWOaSEhceSgc6q6kqvg"
IDEOGRAM_API_KEY = ""

# Add Ideogram model and style mappings
IDEOGRAM_MODELS = {
    "Version 2a": "V_2A",
    "Version 2a Turbo": "V_2A_TURBO",
    "Version 2": "V_2",
    "Version 2 Turbo": "V_2_TURBO",
    "Version 3": "V_3"
}

IDEOGRAM_STYLES = {
    "Auto": "AUTO",
    "Design": "DESIGN",
    "Render 3D": "RENDER_3D"
}

# Add a function to get the next available file number
def get_next_file_number(category_folder, theme, category, counter_override=None):
    """
    Get the next available file number for the given theme and category.
    
    The filename convention is: TTCCCNNNNN where:
    - TT: 2-digit theme code (e.g., 01 for Pets)
    - CCC: 3-digit category code (e.g., 002 for Cats)
    - NNNNN: 5-digit sequential number starting at 00001
    
    For example: 0100200001.png for the first Pets/Cats image
    
    Args:
        category_folder: Path to the category folder
        theme: Theme name
        category: Category name
        counter_override: Optional integer to override the starting counter (1-100)
    """
    # Get theme and category codes
    theme_code = THEME_MAPPING.get(theme, "00")
    category_code = CATEGORY_MAPPING.get(category, "000")
    
    # Pattern for the numeric filename format
    pattern = f"{theme_code}{category_code}"
    pattern_length = len(pattern)
    
    # Get existing files that match the pattern
    existing_files = []
    try:
        if os.path.exists(category_folder):
            existing_files = [f for f in os.listdir(category_folder) if f.startswith(pattern) and 
                             f.endswith(('.png', '.jpg', '.jpeg', '.avif'))]
    except Exception as e:
        logger.error(f"Error checking directory for files: {e}")
    
    # Find the highest number
    max_number = 0
    for file in existing_files:
        filename = os.path.basename(file)
        # Extract the numeric part (last 5 digits before extension)
        try:
            file_without_ext = os.path.splitext(filename)[0]
            if len(file_without_ext) >= pattern_length + 5:
                number_part = file_without_ext[pattern_length:pattern_length+5]
                if number_part.isdigit():
                    number = int(number_part)
                    if number > max_number:
                        max_number = number
        except Exception as e:
            logger.error(f"Error parsing file number from {filename}: {str(e)}")
            continue
    
    # Return the next number (always increment the largest number found)
    # If no files exist with the pattern, start from 1
    # However, if counter_override is provided, use that as the minimum starting number
    if counter_override is not None and isinstance(counter_override, int) and 1 <= counter_override <= 100:
        # Use the override value as the starting point, but ensure we don't conflict with existing files
        next_number = max(max_number + 1, counter_override) if max_number > 0 else counter_override
        logger.info(f"Next file number for {pattern} with override {counter_override}: {next_number:05d}")
    else:
        next_number = max_number + 1 if max_number > 0 else 1
        logger.info(f"Next file number for {pattern}: {next_number:05d}")
    return next_number

# Ideogram generation function
async def generate_with_ideogram(prompt, aspect_ratio, model, style, num_images, negative_prompt=None, seed=None, reference_image_path=None, rendering_speed="DEFAULT"):
    """Generate images using Ideogram API"""
    try:
        # Check if using V_3 model which has different API endpoint and structure
        is_v3 = model == "V_3"
        
        if is_v3:
            # Use V_3 API endpoint
            url = "https://api.ideogram.ai/v1/ideogram-v3/generate"
            logger.info("Using Ideogram V_3 API endpoint")
            
            # Determine if we should use style reference image for V_3
            use_style_reference = False
            style_reference_base64 = None
            
            if reference_image_path and os.path.exists(reference_image_path):
                try:
                    # Validate the file is accessible and readable
                    with open(reference_image_path, 'rb') as test_file:
                        test_file.read(1)  # Read first byte to test accessibility
                    
                    logger.info(f"✅ Using style reference image for V_3: {reference_image_path}")
                    
                    # For V_3 with style reference - encode image to base64
                    import base64
                    with open(reference_image_path, 'rb') as image_file:
                        style_reference_base64 = base64.b64encode(image_file.read()).decode('utf-8')
                    
                    # Get file size for debugging
                    file_size = os.path.getsize(reference_image_path)
                    filename = os.path.basename(reference_image_path)
                    logger.info(f"📁 Style reference image size: {file_size} bytes, filename: {filename}")
                    logger.info(f"📁 Base64 encoded length: {len(style_reference_base64)} characters")
                    
                    use_style_reference = True
                    
                except Exception as style_error:
                    logger.warning(f"❌ Failed to access style reference image: {style_error}")
                    logger.info("Proceeding without style reference image")
                    use_style_reference = False
                    style_reference_base64 = None
            else:
                logger.info("ℹ️ No style reference image provided for V_3")
            
            # Prepare JSON payload for V_3 API
            payload = {
                'prompt': prompt,
                'rendering_speed': rendering_speed
            }
            
            # Add aspect ratio - convert format for V3
            if aspect_ratio and aspect_ratio == "ASPECT_1_1":
                payload['aspect_ratio'] = "1x1"  # Convert ASPECT_1_1 to 1x1 format for V3
            elif aspect_ratio and ":" in aspect_ratio:
                payload['aspect_ratio'] = aspect_ratio.replace(":", "x")  # Convert 1:1 to 1x1 format
            else:
                payload['aspect_ratio'] = "1x1"  # Default to 1x1
            
            # Add style type - for V3, always use "GENERAL"
            payload['style_type'] = "GENERAL"
            
            # Add number of images - ensure it's an integer
            payload['num_images'] = int(num_images) if num_images else 1
            
            # Add negative prompt if provided
            if negative_prompt and negative_prompt.strip():
                payload['negative_prompt'] = negative_prompt.strip()
                logger.info(f"Using negative prompt for V3: {negative_prompt.strip()}")
            
            # Add seed if provided
            if seed is not None and seed != "":
                try:
                    seed_value = int(seed) if isinstance(seed, str) else seed
                    payload['seed'] = seed_value
                    logger.info(f"Using seed value: {seed_value} for Ideogram V3 generation")
                except (ValueError, TypeError):
                    logger.warning(f"Invalid seed value provided: {seed}, using random seed instead")
            
            # Add style reference image if available
            if use_style_reference and style_reference_base64:
                payload['style_reference_images'] = [style_reference_base64]
                logger.info(f"✅ Added base64 style reference to payload")
            
            # Prepare headers for JSON request
            headers = {
                "Api-Key": IDEOGRAM_API_KEY,
                "Content-Type": "application/json"
            }
            
            # Send JSON request
            if use_style_reference and style_reference_base64:
                # PRINT IDEOGRAM V_3 PAYLOAD WITH STYLE REFERENCE
                print("\n===== IDEOGRAM V_3 GENERATE API - WITH STYLE REFERENCE =====")
                print(f"URL: {url}")
                import json
                payload_copy = payload.copy()
                # Truncate base64 for display
                if 'style_reference_images' in payload_copy:
                    payload_copy['style_reference_images'] = [f"<base64_image_{len(style_reference_base64)}_chars>"]
                print(f"Payload: {json.dumps(payload_copy, indent=2)}")
                print("=============================================================\n")
                
                logger.info(f"🚀 Sending V_3 JSON request WITH style reference")
                response = requests.post(url, headers=headers, json=payload)
                
            else:
                # PRINT IDEOGRAM V_3 PAYLOAD WITHOUT STYLE REFERENCE
                print("\n===== IDEOGRAM V_3 GENERATE API - NO STYLE REFERENCE =====")
                print(f"URL: {url}")
                import json
                print(f"Payload: {json.dumps(payload, indent=2)}")
                print("==========================================================\n")
                
                logger.info(f"🚀 Sending V_3 JSON request WITHOUT style reference")
                response = requests.post(url, headers=headers, json=payload)
                
        else:
            # Use original API endpoint for non-V_3 models
            url = "https://api.ideogram.ai/generate"
            
            # Create payload with the given parameters for older models
            payload = {
                "image_request": {
                    "prompt": prompt,
                    "model": model,
                    "style_type": style,
                    "resolution": "RESOLUTION_1024_1024",
                    "num_images": num_images,
                    "magic_prompt_option": "OFF"
                }
            }
            
            # Only include seed if it has a valid value (not None and not empty)
            if seed is not None and seed != "":
                try:
                    # Convert to int if it's a string number
                    seed_value = int(seed) if isinstance(seed, str) else seed
                    payload["image_request"]["seed"] = seed_value
                    logger.info(f"Using seed value: {seed_value} for Ideogram generation")
                except (ValueError, TypeError):
                    logger.warning(f"Invalid seed value provided: {seed}, using random seed instead")
            else:
                logger.info("No seed value provided for Ideogram generation, using random seed")
            
            # PRINT IDEOGRAM PAYLOAD
            print("\n===== IDEOGRAM API PAYLOAD =====")
            import json
            print(json.dumps(payload, indent=2))
            print("================================\n")
            
            # Prepare headers for older models - try both header formats for compatibility
            headers = {
                "Api-Key": IDEOGRAM_API_KEY,
                "Authorization": f"Bearer {IDEOGRAM_API_KEY}",
                "Content-Type": "application/json"
            }
            
            logger.info(f"Sending request to Ideogram API with payload: {payload}")
            logger.info(f"Using headers: {dict((k, v[:20] + '...' if len(v) > 20 else v) for k, v in headers.items())}")
            
            try:
                response = requests.post(url, headers=headers, json=payload)
                logger.info(f"Ideogram API response status: {response.status_code}")
                
                # Log response headers for debugging
                logger.info(f"Response headers: {dict(response.headers)}")
                
                # Log partial response for debugging (if not successful)
                if response.status_code != 200:
                    logger.error(f"API Error Response: {response.text[:500]}...")
                    
            except requests.exceptions.RequestException as req_error:
                logger.error(f"Request error for older models: {str(req_error)}")
                raise
        
        # Note: Negative prompt is supported for V3, but not for older models
        if negative_prompt and negative_prompt.strip() and not is_v3:
            logger.info(f"Negative prompt provided but will NOT be sent to Ideogram API (not supported in {model}): {negative_prompt.strip()}")
        
        response.raise_for_status()
        
        response_data = response.json()
        logger.info(f"Ideogram response: {response_data}")
        
        # Extract URLs directly from the response data
        result_images = []
        if "data" in response_data and isinstance(response_data["data"], list):
            for item in response_data["data"]:
                if "url" in item:
                    result_images.append(item["url"])
        
        if not result_images:
            return None, "No images generated in the response.", None
        
        # Log the number of images generated
        model_display = "V_3" if is_v3 else model
        logger.info(f"Ideogram {model_display} generated {len(result_images)} images out of {num_images if not is_v3 else 1} requested")
        
        # Store the generated images in the global dictionary
        generation_id = f"ideogram_{int(time.time())}"
        generated_images[generation_id] = [{"url": url} for url in result_images]
        
        return result_images, f"Generated {len(result_images)} images with Ideogram {model_display}.", generation_id
            
    except Exception as e:
        logger.error(f"Error in Ideogram generation: {str(e)}")
        return None, f"Error: {str(e)}", None

# Update the uploading feature to Google Drive, whereby IF the file naming of the generation and existing file on that specific subfolder are EXACTLY SIMILAR, then add (1) or (2) at the end of the file naming for each file that has similar file naming
def upload_to_google_drive(file_path, parent_folder_id=None, theme=None, category=None, subcategory=None):
    """Upload a file to Google Drive with appropriate folder structure."""
    logger.info(f"--- Google Drive Upload Start ---")
    logger.info(f"File: {file_path}")
    logger.info(f"Theme: {theme}, Category: {category}, Subcategory: {subcategory}")
    try:        
        # Check if file exists first        
        if not os.path.exists(file_path):            
            logger.error(f"File does not exist: {file_path}")            
            print(f"Google Drive upload failed: File not found: {file_path}")            
            return None                
        # Create Google Drive service        
        service = create_google_drive_service()        
        if not service:            
            logger.error("Failed to create Google Drive service")            
            print("Google Drive upload failed: Could not initialize Drive service")            
            return None
        
        # Default to specific folder ID if no parent_folder_id is provided
        if not parent_folder_id:
            # Use the root folder directly
            parent_folder_id = '1rGP_5RIbYkqrJ1KEgXH8SHN8bFaBtkpm' # User specified folder ID
            logger.info(f"Using root folder as parent: {parent_folder_id}")
            print(f"[Google Drive Debug] Using root folder as parent: {parent_folder_id}")
        
        # Create folder structure based on theme and category
        if theme:
            logger.info(f"Processing theme: {theme}")
            theme_folder_id = get_or_create_folder(service, parent_folder_id, theme)
            if not theme_folder_id:
                logger.error(f"Failed to get or create theme folder '{theme}'")
                return None
            parent_folder_id = theme_folder_id
            logger.info(f"Using theme folder ID: {parent_folder_id}")
            if category:
                logger.info(f"Processing category: {category}")
                category_folder_id = get_or_create_folder(service, theme_folder_id, category)
                if not category_folder_id:
                    logger.error(f"Failed to get or create category folder '{category}'")
                    return None
                parent_folder_id = category_folder_id
                logger.info(f"Using category folder ID: {parent_folder_id}")
                if subcategory:
                    logger.info(f"Processing subcategory: {subcategory}")
                    subcategory_folder_id = get_or_create_folder(service, category_folder_id, subcategory)
                    if not subcategory_folder_id:
                        logger.error(f"Failed to get or create subcategory folder '{subcategory}'")
                        return None
                    parent_folder_id = subcategory_folder_id
                    logger.info(f"Using subcategory folder ID: {parent_folder_id}")
        
        # Get file name from path
        file_name = os.path.basename(file_path)
        
        # Check if file contains "_card" in filename OR is a ZIP file and create Card subfolder if needed
        if category and ("_card" in file_name.lower() or file_name.lower().endswith('.zip')):
            card_folder_id = get_or_create_folder(service, parent_folder_id, "Card")
            if not card_folder_id:
                logger.error(f"Failed to get or create Card folder in category '{category}'")
                return None
            parent_folder_id = card_folder_id
            if "_card" in file_name.lower():
                logger.info(f"File '{file_name}' contains '_card', placing in Card subfolder: {card_folder_id}")
                print(f"[Google Drive Debug] File '{file_name}' will be placed in Card subfolder (contains '_card')")
            elif file_name.lower().endswith('.zip'):
                logger.info(f"File '{file_name}' is a ZIP file, placing in Card subfolder: {card_folder_id}")
                print(f"[Google Drive Debug] ZIP file '{file_name}' will be placed in Card subfolder")
        elif category:
            logger.info(f"File '{file_name}' does not contain '_card' and is not a ZIP file, placing directly in category folder")
            print(f"[Google Drive Debug] File '{file_name}' will be placed directly in '{category}' folder")
        
        # Check if file already exists to avoid duplicates
        # Escape single quotes in file_name for the query
        safe_file_name = file_name.replace("'", "\\'")
        query = f"name='{safe_file_name}' and '{parent_folder_id}' in parents and trashed=false"
        response = service.files().list(q=query, spaces='drive', fields='files(id, name)', supportsAllDrives=True, includeItemsFromAllDrives=True).execute()
        if response.get('files'):
            existing_file_id = response['files'][0]['id']
            logger.info(f"File '{file_name}' already exists with ID: {existing_file_id}")
            return existing_file_id
        
        # Determine MIME type using mimetypes
        mime_type, _ = mimetypes.guess_type(file_path)
        if not mime_type:
            mime_type = 'application/octet-stream' # Default MIME type
        
        # Upload file        
        file_metadata = {            
            'name': file_name,            
            'parents': [parent_folder_id]
        }                
        try:            
            # Import MediaFileUpload here to ensure it's available   
            print("[DRIVE DEBUG] About to upload into parent_folder_id =", parent_folder_id)         
            media = MediaFileUpload(file_path, mimetype=mime_type)                        
            file = service.files().create(                
                body=file_metadata,                
                media_body=media,                
                fields='id',
                supportsAllDrives=True,
                ).execute()
            file_id = file.get('id')            
            if file_id:                
                logger.info(f"File '{file_name}' uploaded to Google Drive with ID: {file_id}")
                print(f"Successfully uploaded '{file_name}' to Google Drive")                
                return file_id            
            else:                
                logger.error(f"Upload succeeded but no file ID returned for '{file_name}'")  
                print(f"Upload issue: No file ID returned for '{file_name}'")                
                return None                        
        except Exception as upload_error: # This handles errors specifically from the create/upload process           
            logger.error(f"Error during file upload operation for '{file_name}': {str(upload_error)}")            
            print(f"Google Drive upload error for '{file_name}': {str(upload_error)}")            
            return None        
    except Exception as e: # This is the general error handler for the whole function       
        logger.error(f"Error in upload_to_google_drive for '{file_path}': {str(e)}")        
        print(f"Failed to upload file to Google Drive ('{file_path}'): {str(e)}")        
        return None

# Note: Removed incomplete upload_and_generate_image function that was causing issues
# The complete implementation is below

def convert_avif(input_path, output_path, output_format='PNG'):
    """Convert AVIF image to a supported format (PNG by default)"""
    try:
        # Import the working AVIF conversion function
        import sys
        import os
        sys.path.append(os.path.dirname(__file__))
        from avif_fix import convert_avif_simple
        return convert_avif_simple(input_path, output_path, output_format)
    except ImportError:
        # Fallback implementation if avif_fix is not available
        try:
            from PIL import Image
            import pillow_heif
            pillow_heif.register_heif_opener()
            
            with Image.open(input_path) as img:
                if img.mode in ('RGBA', 'LA'):
                    if output_format.upper() == 'PNG':
                        img = img.convert('RGBA')
                    else:
                        background = Image.new('RGB', img.size, (255, 255, 255))
                        if img.mode == 'RGBA':
                            try:
                                img_bands = img.split()
                                background.paste(img, mask=img_bands[-1])
                            except Exception as e:
                                logger.warning(f"Error handling RGBA channels in convert_avif: {str(e)}, using fallback paste")
                                background.paste(img)
                        else:
                            background.paste(img)
                        img = background
                elif img.mode != 'RGB':
                    img = img.convert('RGB')
                
                if output_format.upper() == 'PNG':
                    img.save(output_path, 'PNG', optimize=True)
                elif output_format.upper() in ['JPG', 'JPEG']:
                    img.save(output_path, 'JPEG', quality=95, optimize=True)
                else:
                    img.save(output_path, output_format)
            
            logger.info(f"Successfully converted AVIF image to {output_format}: {output_path}")
            return output_path
        except Exception as e:
            logger.error(f"Error converting AVIF image: {str(e)}")
            logger.error("Consider installing pillow-heif: pip install pillow-heif")
            return input_path
    except Exception as e:
        logger.error(f"Error in convert_avif: {str(e)}")
        return input_path

def remove_background_photoroom(input_path):
    """Remove background from image using PhotoRoom API"""
    try:
        from PIL import Image
        from io import BytesIO
        import tempfile
        
        if input_path.lower().endswith('.avif'):
            input_path = convert_avif(input_path, input_path.rsplit('.', 1)[0] + '.png', 'PNG')    
        if not PHOTOROOM_API_KEY:
            raise ValueError("Photoroom API key missing.")
        url = "https://sdk.photoroom.com/v1/segment"
        headers = {"Accept": "image/png, application/json", "x-api-key": PHOTOROOM_API_KEY}
        with open(input_path, "rb") as f:
            resp = requests.post(url, headers=headers, files={"image_file": f})
        if resp.status_code != 200:
            raise Exception(f"PhotoRoom API error: {resp.status_code} - {resp.text}")
        
        # Get the result from PhotoRoom and convert to RGBA
        result_image = Image.open(BytesIO(resp.content)).convert("RGBA")
        
        # Apply the same alpha edge processing as Birefnet to ensure consistent quality
        # This ensures both methods produce images with similar alpha channel characteristics
        result_image = improve_alpha_edges(result_image, threshold=180, edge_feather=4, use_gaussian_blur=True, feather_intensity=0.6)
        
        # Save to a temporary file and reload to ensure proper alpha channel consistency
        # This matches the same process used in Birefnet for consistent behavior
        with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp:
            result_image.save(tmp.name, 'PNG')
            # Load it back to ensure proper alpha channel data
            final_result = Image.open(tmp.name).convert('RGBA')
            
            # Clean up the temporary file
            try:
                os.unlink(tmp.name)
            except:
                pass
                
        logger.info("✅ PHOTOROOM: Applied alpha edge processing for consistency with Birefnet")
        return final_result
        
    except Exception as e:
        logger.error(f"Error in remove_background_photoroom: {str(e)}")
        logger.error(traceback.format_exc())
        return None

# Inline definition of process_generated_images function to replace import
def process_generated_images(images, status, zip_file, ref_image, prompt, counter, variation_nums, theme=None, category=None, filename_convention="numeric", encode_to_base64=False):
    """Process generated images to add metadata for display"""
    # Import necessary libraries
    import os
    import tempfile
    import zipfile
    from PIL import Image
    
    if not images:
        return [], status, zip_file, [], variation_nums
    
    # Log the raw images received - safely handle different types
    try:
        # Check if images is a GalleryData object
        if hasattr(images, '__class__') and images.__class__.__name__ == 'GalleryData':
            logger.info(f"Processing GalleryData object with images")
        else:
            # Try to log the first few images if it's a list-like object
            preview = str(images)[:100] + "..." if len(str(images)) > 100 else str(images)
            logger.info(f"Processing images: {preview}")
    except Exception as e:
        logger.info(f"Unable to preview images object: {type(images)}")
    
    # Convert any image objects to their file paths and ensure jpg/png format
    image_paths = []
    
    # Handle GalleryData object from Gradio
    if hasattr(images, '__class__') and images.__class__.__name__ == 'GalleryData':
        try:
            # Try to extract paths from GalleryData
            if hasattr(images, 'paths') and images.paths:
                for path in images.paths:
                    if path and os.path.exists(path):
                        image_paths.append(path)
            # If no paths attribute, try to iterate through the object
            elif hasattr(images, '__iter__'):
                for img in images:
                    if isinstance(img, str) and os.path.exists(img):
                        image_paths.append(img)
                    elif isinstance(img, dict) and 'path' in img and os.path.exists(img['path']):
                        image_paths.append(img['path'])
        except Exception as e:
            logger.error(f"Error extracting paths from GalleryData: {str(e)}")
    else:
        # Handle other image formats
        try:
            # Try to iterate through images if it's iterable
            for img in images:
                img_path = None
                
                # Extract the actual path from different formats
                if isinstance(img, str) and os.path.exists(img):
                    # Direct file path
                    img_path = img
                elif isinstance(img, tuple) and len(img) > 0 and isinstance(img[0], str) and os.path.exists(img[0]):
                    # Tuple format from Gradio
                    img_path = img[0]
                elif hasattr(img, 'filename') and os.path.exists(img.filename):
                    # File-like object
                    img_path = img.filename
                elif hasattr(img, 'name') and os.path.exists(img.name):
                    # Another file-like object format
                    img_path = img.name
                elif isinstance(img, Image.Image):
                    # PIL Image object - save to temp file
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.jpg') as tmp:
                        img.convert('RGB').save(tmp.name, 'JPEG')
                        img_path = tmp.name
                else:
                    logger.warning(f"Unsupported image format: {type(img)}")
                    continue
                
                # If we have a valid path, ensure it's jpg/png format
                if img_path:
                    if img_path.lower().endswith('.webp'):
                        try:
                            # Convert webp to jpg
                            image = Image.open(img_path)
                            jpg_path = img_path.rsplit('.', 1)[0] + '.jpg'
                            image.convert('RGB').save(jpg_path, 'JPEG')
                            logger.info(f"Converted webp to jpg: {jpg_path}")
                            image_paths.append(jpg_path)
                        except Exception as e:
                            logger.error(f"Error converting webp to jpg: {str(e)}")
                            image_paths.append(img_path)  # Use original if conversion fails
                    else:
                        image_paths.append(img_path)
                else:
                    logger.warning(f"Skipping unsupported image format: {type(img)}")
        except Exception as e:
            logger.error(f"Error processing images: {str(e)}")
            return [], status, zip_file, image_paths, variation_nums
    
    # Create reference paths array - use the same reference image for all
    ref_paths = [ref_image] * len(image_paths)
    
    # Get reference filename
    ref_filename = None
    if isinstance(ref_image, str):
        ref_filename = os.path.basename(ref_image)
    
    # Create display images with metadata
    try:
        display_images = create_display_images_with_metadata(image_paths, ref_paths, variation_nums, ref_filename)
        
        # Ensure zip_file is a valid path
        if zip_file and isinstance(zip_file, str) and os.path.exists(zip_file):
            # Check if ZIP contains webp files and convert them
            try:
                with zipfile.ZipFile(zip_file, 'r') as existing_zip:
                    webp_files_found = any(name.lower().endswith('.webp') for name in existing_zip.namelist())
                
                if webp_files_found:
                    # Create a new ZIP with converted files
                    new_zip_path = zip_file.rsplit('.', 1)[0] + '_jpg.zip'
                    with zipfile.ZipFile(existing_zip.filename, 'r') as src_zip:
                        with zipfile.ZipFile(new_zip_path, 'w') as dest_zip:
                            for item in src_zip.infolist():
                                data = src_zip.read(item.filename)
                                if item.filename.lower().endswith('.webp'):
                                    # Convert to jpg
                                    with tempfile.NamedTemporaryFile(delete=False, suffix='.webp') as tmp:
                                        tmp.write(data)
                                        tmp.flush()
                                        
                                        image = Image.open(tmp.name)
                                        jpg_filename = item.filename.rsplit('.', 1)[0] + '.jpg'
                                        
                                        with tempfile.NamedTemporaryFile(delete=False, suffix='.jpg') as jpg_tmp:
                                            image.convert('RGB').save(jpg_tmp.name, 'JPEG')
                                            jpg_tmp.flush()
                                            
                                            with open(jpg_tmp.name, 'rb') as jpg_file:
                                                jpg_data = jpg_file.read()
                                                dest_zip.writestr(jpg_filename, jpg_data)
                                            
                                            os.unlink(jpg_tmp.name)
                                        
                                        os.unlink(tmp.name)
                                else:
                                    dest_zip.writestr(item, data)
                    
                    # Use the new ZIP
                    if os.path.exists(new_zip_path) and os.path.getsize(new_zip_path) > 0:
                        zip_file = new_zip_path
                        logger.info(f"Created new ZIP with jpg files: {new_zip_path}")
            except Exception as e:
                logger.error(f"Error converting ZIP contents: {str(e)}")
            
            logger.info(f"Valid ZIP file for download: {zip_file}")
        else:
            logger.warning(f"Invalid or missing ZIP file: {zip_file}")
            # Don't return an invalid value to the File component
            zip_file = None
        
        # Generate base64 files if requested
        base64_files = []
        if encode_to_base64:
            logger.info("Base64 encoding enabled - generating base64 files for all images")
            try:
                base64_files = batch_encode_images_to_base64(image_paths)
                logger.info(f"Successfully created {len(base64_files)} base64 files")
                
                # Add base64 files to ZIP if it exists
                if zip_file and os.path.exists(zip_file) and base64_files:
                    logger.info("Adding base64 files to existing ZIP")
                    import zipfile
                    import tempfile
                    
                    # Create new ZIP with base64 files included
                    new_zip_path = zip_file.rsplit('.', 1)[0] + '_with_base64.zip'
                    
                    with zipfile.ZipFile(zip_file, 'r') as existing_zip:
                        with zipfile.ZipFile(new_zip_path, 'w') as new_zip:
                            # Copy existing files
                            for item in existing_zip.infolist():
                                data = existing_zip.read(item.filename)
                                new_zip.writestr(item.filename, data)
                            
                            # Add base64 files
                            for base64_file in base64_files:
                                if os.path.exists(base64_file):
                                    with open(base64_file, 'r', encoding='utf-8') as f:
                                        base64_content = f.read()
                                    base64_filename = os.path.basename(base64_file)
                                    new_zip.writestr(base64_filename, base64_content)
                    
                    # Replace the original ZIP with the new one
                    if os.path.exists(new_zip_path) and os.path.getsize(new_zip_path) > 0:
                        zip_file = new_zip_path
                        logger.info(f"Updated ZIP file with base64 files: {new_zip_path}")
                    
            except Exception as base64_error:
                logger.error(f"Error generating base64 files: {str(base64_error)}")
                # Continue without base64 files - don't fail the entire process
        
        return display_images, status, zip_file, image_paths, variation_nums
    except Exception as e:
        logger.error(f"Error in process_generated_images: {str(e)}")
        return image_paths, f"Error processing images for display: {str(e)}", zip_file, image_paths, variation_nums

# Ideogram generation function
async def generate_with_ideogram(prompt, aspect_ratio, model, style, num_images, negative_prompt=None, seed=None, reference_image_path=None, rendering_speed="DEFAULT"):
    """Generate images using Ideogram API"""
    try:
        # Check if using V_3 model which has different API endpoint and structure
        is_v3 = model == "V_3"
        
        if is_v3:
            # Use V_3 API endpoint
            url = "https://api.ideogram.ai/v1/ideogram-v3/generate"
            logger.info("Using Ideogram V_3 API endpoint")
            
            # Determine if we should use style reference image for V_3
            use_style_reference = False
            style_reference_base64 = None
            
            if reference_image_path and os.path.exists(reference_image_path):
                try:
                    # Validate the file is accessible and readable
                    with open(reference_image_path, 'rb') as test_file:
                        test_file.read(1)  # Read first byte to test accessibility
                    
                    logger.info(f"✅ Using style reference image for V_3: {reference_image_path}")
                    
                    # For V_3 with style reference - encode image to base64
                    import base64
                    with open(reference_image_path, 'rb') as image_file:
                        style_reference_base64 = base64.b64encode(image_file.read()).decode('utf-8')
                    
                    # Get file size for debugging
                    file_size = os.path.getsize(reference_image_path)
                    filename = os.path.basename(reference_image_path)
                    logger.info(f"📁 Style reference image size: {file_size} bytes, filename: {filename}")
                    logger.info(f"📁 Base64 encoded length: {len(style_reference_base64)} characters")
                    
                    use_style_reference = True
                    
                except Exception as style_error:
                    logger.warning(f"❌ Failed to access style reference image: {style_error}")
                    logger.info("Proceeding without style reference image")
                    use_style_reference = False
                    style_reference_base64 = None
            else:
                logger.info("ℹ️ No style reference image provided for V_3")
            
            # Prepare JSON payload for V_3 API
            payload = {
                'prompt': prompt,
                'rendering_speed': rendering_speed
            }
            
            # Add aspect ratio - convert format for V3
            if aspect_ratio and aspect_ratio == "ASPECT_1_1":
                payload['aspect_ratio'] = "1x1"  # Convert ASPECT_1_1 to 1x1 format for V3
            elif aspect_ratio and ":" in aspect_ratio:
                payload['aspect_ratio'] = aspect_ratio.replace(":", "x")  # Convert 1:1 to 1x1 format
            else:
                payload['aspect_ratio'] = "1x1"  # Default to 1x1
            
            # Add style type - for V3, always use "GENERAL"
            payload['style_type'] = "GENERAL"
            
            # Add number of images - ensure it's an integer
            payload['num_images'] = int(num_images) if num_images else 1
            
            # Add negative prompt if provided
            if negative_prompt and negative_prompt.strip():
                payload['negative_prompt'] = negative_prompt.strip()
                logger.info(f"Using negative prompt for V3: {negative_prompt.strip()}")
            
            # Add seed if provided
            if seed is not None and seed != "":
                try:
                    seed_value = int(seed) if isinstance(seed, str) else seed
                    payload['seed'] = seed_value
                    logger.info(f"Using seed value: {seed_value} for Ideogram V3 generation")
                except (ValueError, TypeError):
                    logger.warning(f"Invalid seed value provided: {seed}, using random seed instead")
            
            # Add style reference image if available
            if use_style_reference and style_reference_base64:
                payload['style_reference_images'] = [style_reference_base64]
                logger.info(f"✅ Added base64 style reference to payload")
            
            # Prepare headers for JSON request
            headers = {
                "Api-Key": IDEOGRAM_API_KEY,
                "Content-Type": "application/json"
            }
            
            # Send JSON request
            if use_style_reference and style_reference_base64:
                # PRINT IDEOGRAM V_3 PAYLOAD WITH STYLE REFERENCE
                print("\n===== IDEOGRAM V_3 GENERATE API - WITH STYLE REFERENCE =====")
                print(f"URL: {url}")
                import json
                payload_copy = payload.copy()
                # Truncate base64 for display
                if 'style_reference_images' in payload_copy:
                    payload_copy['style_reference_images'] = [f"<base64_image_{len(style_reference_base64)}_chars>"]
                print(f"Payload: {json.dumps(payload_copy, indent=2)}")
                print("=============================================================\n")
                
                logger.info(f"🚀 Sending V_3 JSON request WITH style reference")
                response = requests.post(url, headers=headers, json=payload)
                
            else:
                # PRINT IDEOGRAM V_3 PAYLOAD WITHOUT STYLE REFERENCE
                print("\n===== IDEOGRAM V_3 GENERATE API - NO STYLE REFERENCE =====")
                print(f"URL: {url}")
                import json
                print(f"Payload: {json.dumps(payload, indent=2)}")
                print("==========================================================\n")
                
                logger.info(f"🚀 Sending V_3 JSON request WITHOUT style reference")
                response = requests.post(url, headers=headers, json=payload)
                
        else:
            # Use original API endpoint for non-V_3 models
            url = "https://api.ideogram.ai/generate"
            
            # Create payload with the given parameters for older models
            payload = {
                "image_request": {
                    "prompt": prompt,
                    "model": model,
                    "style_type": style,
                    "resolution": "RESOLUTION_1024_1024",
                    "num_images": num_images,
                    "magic_prompt_option": "OFF"
                }
            }
            
            # Only include seed if it has a valid value (not None and not empty)
            if seed is not None and seed != "":
                try:
                    # Convert to int if it's a string number
                    seed_value = int(seed) if isinstance(seed, str) else seed
                    payload["image_request"]["seed"] = seed_value
                    logger.info(f"Using seed value: {seed_value} for Ideogram generation")
                except (ValueError, TypeError):
                    logger.warning(f"Invalid seed value provided: {seed}, using random seed instead")
            else:
                logger.info("No seed value provided for Ideogram generation, using random seed")
            
            # PRINT IDEOGRAM PAYLOAD
            print("\n===== IDEOGRAM API PAYLOAD =====")
            import json
            print(json.dumps(payload, indent=2))
            print("================================\n")
            
            # Prepare headers for older models - try both header formats for compatibility
            headers = {
                "Api-Key": IDEOGRAM_API_KEY,
                "Authorization": f"Bearer {IDEOGRAM_API_KEY}",
                "Content-Type": "application/json"
            }
            
            logger.info(f"Sending request to Ideogram API with payload: {payload}")
            logger.info(f"Using headers: {dict((k, v[:20] + '...' if len(v) > 20 else v) for k, v in headers.items())}")
            
            try:
                response = requests.post(url, headers=headers, json=payload)
                logger.info(f"Ideogram API response status: {response.status_code}")
                
                # Log response headers for debugging
                logger.info(f"Response headers: {dict(response.headers)}")
                
                # Log partial response for debugging (if not successful)
                if response.status_code != 200:
                    logger.error(f"API Error Response: {response.text[:500]}...")
                    
            except requests.exceptions.RequestException as req_error:
                logger.error(f"Request error for older models: {str(req_error)}")
                raise
        
        # Note: Negative prompt is supported for V3, but not for older models
        if negative_prompt and negative_prompt.strip() and not is_v3:
            logger.info(f"Negative prompt provided but will NOT be sent to Ideogram API (not supported in {model}): {negative_prompt.strip()}")
        
        response.raise_for_status()
        
        response_data = response.json()
        logger.info(f"Ideogram response: {response_data}")
        
        # Extract URLs directly from the response data
        result_images = []
        if "data" in response_data and isinstance(response_data["data"], list):
            for item in response_data["data"]:
                if "url" in item:
                    result_images.append(item["url"])
        
        if not result_images:
            return None, "No images generated in the response.", None
        
        # Log the number of images generated
        model_display = "V_3" if is_v3 else model
        logger.info(f"Ideogram {model_display} generated {len(result_images)} images out of {num_images if not is_v3 else 1} requested")
        
        # Store the generated images in the global dictionary
        generation_id = f"ideogram_{int(time.time())}"
        generated_images[generation_id] = [{"url": url} for url in result_images]
        
        return result_images, f"Generated {len(result_images)} images with Ideogram {model_display}.", generation_id
            
    except Exception as e:
        logger.error(f"Error in Ideogram generation: {str(e)}")
        return None, f"Error: {str(e)}", None

# Update the uploading feature to Google Drive, whereby IF the file naming of the generation and existing file on that specific subfolder are EXACTLY SIMILAR, then add (1) or (2) at the end of the file naming for each file that has similar file naming
def upload_to_google_drive(file_path, parent_folder_id=None, theme=None, category=None, subcategory=None):   
    """Upload a file to Google Drive with appropriate folder structure."""    
    try:        
        # Check if file exists first        
        if not os.path.exists(file_path):            
            logger.error(f"File does not exist: {file_path}")            
            print(f"Google Drive upload failed: File not found: {file_path}")            
            return None                
        # Create Google Drive service        
        service = create_google_drive_service()        
        if not service:            
            logger.error("Failed to create Google Drive service")            
            print("Google Drive upload failed: Could not initialize Drive service")            
            return None
        
        # Default to specific folder ID if no parent_folder_id is provided
        if not parent_folder_id:
            # Use the root folder directly
            parent_folder_id = '1rGP_5RIbYkqrJ1KEgXH8SHN8bFaBtkpm' # User specified folder ID
            logger.info(f"Using root folder as parent: {parent_folder_id}")
            print(f"[Google Drive Debug] Using root folder as parent: {parent_folder_id}")
        
        # Create folder structure based on theme and category
        if theme:
            theme_folder_id = get_or_create_folder(service, parent_folder_id, theme)
            if not theme_folder_id:
                logger.error(f"Failed to get or create theme folder '{theme}'")
                return None
            if category:
                category_folder_id = get_or_create_folder(service, theme_folder_id, category)
                if not category_folder_id:
                    logger.error(f"Failed to get or create category folder '{category}'")
                    return None
                if subcategory:
                    # Resolve subcategory folder name using SUBCATEGORY_FOLDER_MAPPING
                    subcategory_folder_name = SUBCATEGORY_FOLDER_MAPPING.get(category, {}).get(subcategory)
                    if subcategory_folder_name:
                        logger.info(f"Resolved subcategory '{subcategory}' to folder name '{subcategory_folder_name}' for category '{category}'")
                        subcategory_folder_id = get_or_create_folder(service, category_folder_id, subcategory_folder_name)
                    else:
                        logger.warning(f"No mapping found for subcategory '{subcategory}' in category '{category}', using raw subcategory name")
                        subcategory_folder_id = get_or_create_folder(service, category_folder_id, subcategory)
                    
                    if not subcategory_folder_id:
                        logger.error(f"Failed to get or create subcategory folder '{subcategory_folder_name or subcategory}'")
                        return None
                    parent_folder_id = subcategory_folder_id
                    logger.info(f"Using subcategory folder: {subcategory_folder_name or subcategory} (ID: {subcategory_folder_id})")
                else:
                    parent_folder_id = category_folder_id
            else:
                parent_folder_id = theme_folder_id
        
        # Get file name from path
        file_name = os.path.basename(file_path)
        
        # Check if file contains "_card" in filename OR is a ZIP file and create Card subfolder if needed
        if category and ("_card" in file_name.lower() or file_name.lower().endswith('.zip')):
            card_folder_id = get_or_create_folder(service, parent_folder_id, "Card")
            if not card_folder_id:
                logger.error(f"Failed to get or create Card folder in category '{category}'")
                return None
            parent_folder_id = card_folder_id
            if "_card" in file_name.lower():
                logger.info(f"File '{file_name}' contains '_card', placing in Card subfolder: {card_folder_id}")
                print(f"[Google Drive Debug] File '{file_name}' will be placed in Card subfolder (contains '_card')")
            elif file_name.lower().endswith('.zip'):
                logger.info(f"File '{file_name}' is a ZIP file, placing in Card subfolder: {card_folder_id}")
                print(f"[Google Drive Debug] ZIP file '{file_name}' will be placed in Card subfolder")
        elif category:
            logger.info(f"File '{file_name}' does not contain '_card' and is not a ZIP file, placing directly in category folder")
            print(f"[Google Drive Debug] File '{file_name}' will be placed directly in '{category}' folder")
        
        # Check if file already exists to avoid duplicates
        # Escape single quotes in file_name for the query
        safe_file_name = file_name.replace("'", "\\'")
        query = f"name='{safe_file_name}' and '{parent_folder_id}' in parents and trashed=false"
        response = service.files().list(q=query, spaces='drive', fields='files(id, name)').execute()
        if response.get('files'):
            existing_file_id = response['files'][0]['id']
            logger.info(f"File '{file_name}' already exists with ID: {existing_file_id}")
            return existing_file_id
        
        # Determine MIME type using mimetypes
        mime_type, _ = mimetypes.guess_type(file_path)
        if not mime_type:
            mime_type = 'application/octet-stream' # Default MIME type
        
        # Upload file        
        file_metadata = {            
            'name': file_name,            
            'parents': [parent_folder_id]
        }                
        try:            
            # Import MediaFileUpload here to ensure it's available            
            media = MediaFileUpload(file_path, mimetype=mime_type)                        
            file = service.files().create(                
                body=file_metadata,                
                media_body=media,                
                fields='id',
                supportsAllDrives=True).execute()
            file_id = file.get('id')            
            if file_id:                
                logger.info(f"File '{file_name}' uploaded to Google Drive with ID: {file_id}")
                print(f"Successfully uploaded '{file_name}' to Google Drive")                
                return file_id            
            else:                
                logger.error(f"Upload succeeded but no file ID returned for '{file_name}'")  
                print(f"Upload issue: No file ID returned for '{file_name}'")                
                return None                        
        except Exception as upload_error: # This handles errors specifically from the create/upload process           
            logger.error(f"Error during file upload operation for '{file_name}': {str(upload_error)}")            
            print(f"Google Drive upload error for '{file_name}': {str(upload_error)}")            
            return None        
    except Exception as e: # This is the general error handler for the whole function       
        logger.error(f"Error in upload_to_google_drive for '{file_path}': {str(e)}")        
        print(f"Failed to upload file to Google Drive ('{file_path}'): {str(e)}")        
        return None

# Note: Removed incomplete upload_and_generate_image function that was causing issues
# The complete implementation is below

def convert_avif(input_path, output_path, output_format='PNG'):
    """Convert AVIF image to a supported format (PNG by default)"""
    try:
        # Import the working AVIF conversion function
        import sys
        import os
        sys.path.append(os.path.dirname(__file__))
        from avif_fix import convert_avif_simple
        return convert_avif_simple(input_path, output_path, output_format)
    except ImportError:
        # Fallback implementation if avif_fix is not available
        try:
            from PIL import Image
            import pillow_heif
            pillow_heif.register_heif_opener()
            
            with Image.open(input_path) as img:
                if img.mode in ('RGBA', 'LA'):
                    if output_format.upper() == 'PNG':
                        img = img.convert('RGBA')
                    else:
                        background = Image.new('RGB', img.size, (255, 255, 255))
                        if img.mode == 'RGBA':
                            try:
                                img_bands = img.split()
                                background.paste(img, mask=img_bands[-1])
                            except Exception as e:
                                logger.warning(f"Error handling RGBA channels in convert_avif: {str(e)}, using fallback paste")
                                background.paste(img)
                        else:
                            background.paste(img)
                        img = background
                elif img.mode != 'RGB':
                    img = img.convert('RGB')
                
                if output_format.upper() == 'PNG':
                    img.save(output_path, 'PNG', optimize=True)
                elif output_format.upper() in ['JPG', 'JPEG']:
                    img.save(output_path, 'JPEG', quality=95, optimize=True)
                else:
                    img.save(output_path, output_format)
            
            logger.info(f"Successfully converted AVIF image to {output_format}: {output_path}")
            return output_path
        except Exception as e:
            logger.error(f"Error converting AVIF image: {str(e)}")
            logger.error("Consider installing pillow-heif: pip install pillow-heif")
            return input_path
    except Exception as e:
        logger.error(f"Error in convert_avif: {str(e)}")
        return input_path

def remove_background_photoroom(input_path):
    """Remove background from image using PhotoRoom API"""
    try:
        from PIL import Image
        from io import BytesIO
        import tempfile
        
        if input_path.lower().endswith('.avif'):
            input_path = convert_avif(input_path, input_path.rsplit('.', 1)[0] + '.png', 'PNG')    
        if not PHOTOROOM_API_KEY:
            raise ValueError("Photoroom API key missing.")
        url = "https://sdk.photoroom.com/v1/segment"
        headers = {"Accept": "image/png, application/json", "x-api-key": PHOTOROOM_API_KEY}
        with open(input_path, "rb") as f:
            resp = requests.post(url, headers=headers, files={"image_file": f})
        if resp.status_code != 200:
            raise Exception(f"PhotoRoom API error: {resp.status_code} - {resp.text}")
        
        # Get the result from PhotoRoom and convert to RGBA
        result_image = Image.open(BytesIO(resp.content)).convert("RGBA")
        
        # Apply the same alpha edge processing as Birefnet to ensure consistent quality
        # This ensures both methods produce images with similar alpha channel characteristics
        result_image = improve_alpha_edges(result_image, threshold=180, edge_feather=4, use_gaussian_blur=True, feather_intensity=0.6)
        
        # Save to a temporary file and reload to ensure proper alpha channel consistency
        # This matches the same process used in Birefnet for consistent behavior
        with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp:
            result_image.save(tmp.name, 'PNG')
            # Load it back to ensure proper alpha channel data
            final_result = Image.open(tmp.name).convert('RGBA')
            
            # Clean up the temporary file
            try:
                os.unlink(tmp.name)
            except:
                pass
                
        logger.info("✅ PHOTOROOM: Applied alpha edge processing for consistency with Birefnet")
        return final_result
        
    except Exception as e:
        logger.error(f"Error in remove_background_photoroom: {str(e)}")
        logger.error(traceback.format_exc())
        return None

# Initialize OpenAI client for Qwen
client = OpenAI(
    api_key=os.getenv("DASHSCOPE_API_KEY", ''),
    base_url=""
)

# Update the upload_and_generate_image function to handle multiple reference images and card templates
async def upload_and_generate_image(
    provider, reference_images, card_template, theme, category, subcategory, 
    # Leonardo parameters
    model_name=None, width=1024, height=1024, guidance_scale=7,  # Changed from magic_strength to guidance_scale
    generated_prompt="", negative_prompt="", 
    # Image processing modes
    image_process_mode=None, strength_type=None,
    # Legacy parameters removed - now using multi-reference system
    preset_style=None, num_images=1, output_format="png",
    # Ideogram parameters
    ideogram_model=None, ideogram_style=None, ideogram_num_images=1,
    # Filename settings
    filename_convention="numeric",
    # S3 upload settings
    upload_to_s3_bucket=True,
    # Optional seed for reproducibility
    seed=None,
    # Activity and expression parameters
    activity=None, facial_expression=None, fur_color=None, ethnicity="Auto",
    # Stop flag to cancel generation
    stop_flag=False,
    # Google Drive upload settings
    upload_to_gdrive: bool = False,
    # Background removal is now always active - both original and processed images are saved
    # Base64 encoding option
    encode_to_base64: bool = False,
    # Generation type parameter to distinguish between standard and activity generation
    generation_type: str = "standard",
    # Multi-reference image support for Leonardo
    reference_image_1=None, ref_type_1="None", ref_strength_1="Mid",
    reference_image_2=None, ref_type_2="None", ref_strength_2="Mid",
    reference_image_3=None, ref_type_3="None", ref_strength_3="Mid",
    # Ideogram style reference control
    ideogram_disable_style_reference: bool = False,
    # Ideogram rendering speed for V3 model
    ideogram_rendering_speed: str = "DEFAULT",
    # Modified prompt parameter for Excel generation
    modified_prompt: str = "",

    # Imagen-4 parameters
    imagen4_model: str = "google/imagen-4",
    imagen4_aspect_ratio: str = "1:1",
    imagen4_safety_filter: str = "block_only_high",
    imagen4_num_images: int = 1,
    
    # Counter override parameter
    counter_override: int = None,
):
    """Generate images using provider API and save them with proper naming and organization."""
    try:
        # Import required modules
        from PIL import Image, ImageDraw, ImageFont
        import tempfile
        import io
        import os
        import requests
        import zipfile
        import shutil
        import time
        from datetime import datetime
        
        # CRITICAL FIX: Handle reference_images array ambiguity at the very start
        if isinstance(reference_images, np.ndarray):
            if reference_images.size == 0:
                reference_images = []
            elif reference_images.size == 1:
                # Convert single item array to list
                reference_images = [reference_images.item()]
            else:
                # Convert multi-item array to list
                reference_images = reference_images.tolist()
            logger.info(f"Converted numpy array reference_images to list: {type(reference_images)}")
        elif reference_images is not None and not isinstance(reference_images, (list, str, dict)):
            # Convert any other type to string
            reference_images = str(reference_images)
            logger.info(f"Converted reference_images to string: {reference_images}")
        
        logger.info(f"Starting image generation with {provider}, theme: {theme}, category: {category}")
        print(f"Starting image generation with {provider}, theme: {theme}, category: {category}")
        # Add more detailed console output
        print(f"[GENERATION] Provider: {provider}, Theme: {theme}, Category: {category}")
        print(f"[PROMPT] {generated_prompt[:100]}..." if generated_prompt and len(generated_prompt) > 100 else f"[PROMPT] {generated_prompt}")
        if activity:
            print(f"[ACTIVITY] {activity}")
        if facial_expression:
            print(f"[EXPRESSION] {facial_expression}")
        if fur_color:
            print(f"[FUR COLOR] {fur_color}")

        # Reset stop flag at the beginning of generation and check if it was previously set
        if stop_flag:
            logger.info("Stop flag was set from previous operation, but starting new generation anyway")
            print("[INFO] Stop flag reset - Starting new generation")
        
        # Note: The stop_flag is now only used for cancelling during generation, not before it starts
        
        # Use the modified prompt if available, otherwise use generated prompt
        if modified_prompt and modified_prompt.strip() != "":
            prompt = modified_prompt
            logger.info(f"Using modified prompt: {prompt}")
            print(f"Using modified prompt: {prompt[:100]}..." if len(prompt) > 100 else f"Using modified prompt: {prompt}")
        elif generated_prompt and generated_prompt.strip() != "":
            prompt = generated_prompt
            logger.info(f"Using generated prompt: {prompt}")
            print(f"Using generated prompt: {prompt[:100]}..." if len(prompt) > 100 else f"Using generated prompt: {prompt}")
        else:
            # Fallback prompt if both are empty
            prompt = f"Pixar Style, Plain White Background, a cartoon-style {category} character with vibrant colors"
            logger.warning(f"Using fallback prompt: {prompt}")
            print(f"WARNING: Using fallback prompt: {prompt}")
        
        # Base64 encoding status check
        if encode_to_base64:
            logger.info("Base64 encoding is enabled - images will be encoded to base64 files.")
        else:
            logger.info("Base64 encoding is disabled - only regular image files will be generated.")
        
        # Directory setup for saving output
        theme_str = theme if isinstance(theme, str) else str(theme)
        category_str = category if isinstance(category, str) else str(category)
        theme_folder = os.path.join(IMAGES_DIR, theme_str.lower())
        category_folder = os.path.join(theme_folder, category_str.lower())
        os.makedirs(category_folder, exist_ok=True)
        
        # Note: Stop flag check removed - flag is only used during generation, not before it starts
        if stop_flag:
            logger.info("Stop flag was set but continuing with generation (directory setup complete)")
            print("[INFO] Stop flag detected but proceeding with generation")
        
        # Get the next file number based on theme/category, with optional counter override
        # Convert counter_override from "Auto" string to None if needed
        counter_override_int = None
        if counter_override is not None and counter_override != "Auto":
            try:
                counter_override_int = int(counter_override)
                logger.info(f"Counter override provided: {counter_override_int}")
            except (ValueError, TypeError):
                logger.warning(f"Invalid counter override value: {counter_override}, using auto numbering")
                counter_override_int = None
        
        next_file_number = get_next_file_number(category_folder, theme, category, counter_override_int)
        
        # Define base filename pattern for naming the output files
        # Get theme and category codes from the mappings
        theme_code = THEME_MAPPING.get(theme, "00")
        category_code = CATEGORY_MAPPING.get(category, "000")
        
        # Create the base filename pattern using the numeric convention
        # Format: ThemeCodeCategoryCodeImageNumber (e.g., 0100200001)
        base_filename_pattern = f"{theme_code}{category_code}{next_file_number:05d}"
        if counter_override_int is not None:
            logger.info(f"Using base filename pattern with counter override: {base_filename_pattern} (starting from {counter_override_int})")
        else:
            logger.info(f"Using base filename pattern: {base_filename_pattern}")
        
        # Process reference images if provided
        init_image_ids = []
        temp_dirs_to_cleanup = []
        
        # Process card template if provided
        card_template_img = None
        if card_template is not None:
            try:
                logger.info(f"Processing card template: {card_template} (type: {type(card_template).__name__})")
                
                # Handle different input types for card_template
                if isinstance(card_template, np.ndarray):
                    if card_template.size > 0:
                        # Try to convert first element if it's a string path
                        if isinstance(card_template[0], (str, bytes)):
                            card_template = str(card_template[0])
                            logger.info(f"Converted numpy array to path string: {card_template}")
                        else:
                            # It's actual image data, convert to a temporary file
                            temp_dir = tempfile.mkdtemp()
                            temp_dirs_to_cleanup.append(temp_dir)
                            temp_file_path = os.path.join(temp_dir, "card_template.png")
                            
                            # Save the numpy array as an image
                            Image.fromarray(card_template.astype(np.uint8)).save(temp_file_path)
                            card_template = temp_file_path
                            logger.info(f"Saved numpy image data to temporary file: {card_template}")
                
                # Handle tuple input (sometimes files are passed as tuples from Gradio)
                if isinstance(card_template, tuple):
                    logger.info(f"Card template is a tuple: {card_template}")
                    # Check if first element is a file path string
                    if len(card_template) > 0 and isinstance(card_template[0], str):
                        card_template = card_template[0]
                        logger.info(f"Extracted file path from tuple: {card_template}")
                    else:
                        logger.warning(f"Cannot extract valid file path from tuple: {card_template}")
                        card_template = None
                
                # Handle dictionary input (common from Gradio interfaces)
                if isinstance(card_template, dict):
                    if 'name' in card_template:
                        card_template = card_template['name']
                        logger.info(f"Extracted file path from dictionary: {card_template}")
                    else:
                        logger.warning(f"Card template dictionary doesn't contain 'name' key: {card_template}")
                        card_template = None
                
                # Handle list input
                if isinstance(card_template, list):
                    if len(card_template) > 0:
                        if isinstance(card_template[0], dict) and 'name' in card_template[0]:
                            card_template = card_template[0]['name']
                            logger.info(f"Extracted file path from list item: {card_template}")
                        elif isinstance(card_template[0], (str, bytes)):
                            card_template = str(card_template[0])
                            logger.info(f"Used first item from list as path: {card_template}")
                        else:
                            logger.warning(f"Card template list contains invalid item type: {type(card_template[0]).__name__}")
                            card_template = None
                    else:
                        logger.warning("Card template list is empty")
                        card_template = None
                
                # Final validation of the card template path
                if card_template is not None and isinstance(card_template, str):
                    # Normalize path separators for the platform
                    card_template = os.path.normpath(card_template)
                    
                    # Check if file exists
                    if os.path.exists(card_template) and os.path.isfile(card_template):
                        # Verify it's an image file by extension
                        valid_extensions = ['.jpg', '.jpeg', '.png', '.webp', '.bmp', '.tiff', '.gif']
                        if any(card_template.lower().endswith(ext) for ext in valid_extensions):
                            try:
                                # Try to open the image to verify it's valid
                                card_template_img = Image.open(card_template).convert('RGBA')
                                logger.info(f"Successfully loaded card template: {card_template}")
                                print(f"Using card template: {card_template}")
                            except Exception as img_error:
                                logger.error(f"Failed to open card template image: {str(img_error)}")
                                card_template_img = None
                        else:
                            logger.warning(f"Card template file doesn't have a valid image extension: {card_template}")
                            card_template_img = None
                    else:
                        logger.warning(f"Card template file doesn't exist or is not a file: {card_template}")
                        card_template_img = None
                else:
                    logger.warning(f"Invalid card template value type: {type(card_template).__name__}")
                    card_template_img = None
                    
            except Exception as e:
                logger.error(f"Error processing card template: {str(e)}")
                import traceback
                logger.error(traceback.format_exc())
                card_template_img = None
                
        # Log card template status
        if card_template_img is None:
            logger.warning("No card template provided or card template loading failed. Generated images will not be applied to card template.")
            if card_template is not None:
                logger.warning(f"Card template was provided as '{card_template}' (type: {type(card_template).__name__}) but could not be loaded.")
            print("WARNING: No card template provided. Generated images will not be applied to card template.")
        else:
            logger.info(f"Card template is available. Generated images will be automatically applied to it.")
            print(f"INFO: Using card template. Each generated image will be applied to the template.")

        # Global breed detection for all providers from the first available reference image
        def get_first_reference_image_path():
            """Helper function to get the first available reference image path for breed detection and Ideogram style reference"""
            try:
                # Check reference_image_1 first (priority) - this is what Ideogram V3 should use for style reference
                if reference_image_1 is not None:
                    logger.info(f"🔍 REFERENCE IMAGE EXTRACTION: Processing reference_image_1 of type: {type(reference_image_1)}")
                    
                    if isinstance(reference_image_1, np.ndarray):
                        # Handle numpy array - save to temp file
                        import tempfile
                        temp_dir = tempfile.mkdtemp()
                        temp_dirs_to_cleanup.append(temp_dir)
                        temp_file_path = os.path.join(temp_dir, "temp_ref_image.png")
                        Image.fromarray(reference_image_1.astype(np.uint8)).save(temp_file_path)
                        logger.info(f"✅ REFERENCE IMAGE EXTRACTION: Converted numpy array to temp file: {temp_file_path}")
                        return temp_file_path
                    elif isinstance(reference_image_1, list) and len(reference_image_1) > 0:
                        # Handle list format - extract first item
                        first_item = reference_image_1[0]
                        if isinstance(first_item, dict) and 'name' in first_item:
                            file_path = first_item['name']
                        elif isinstance(first_item, str):
                            file_path = first_item
                        else:
                            file_path = str(first_item)
                        logger.info(f"🔍 REFERENCE IMAGE EXTRACTION: Extracted from list: {file_path}")
                    elif isinstance(reference_image_1, dict):
                        # Handle dictionary format (Gradio file upload)
                        file_path = reference_image_1.get('name', reference_image_1.get('path', str(reference_image_1)))
                        logger.info(f"🔍 REFERENCE IMAGE EXTRACTION: Extracted from dict: {file_path}")
                    elif isinstance(reference_image_1, str):
                        # Direct string path
                        file_path = reference_image_1
                        logger.info(f"🔍 REFERENCE IMAGE EXTRACTION: Direct string path: {file_path}")
                    else:
                        # Try to convert to string
                        file_path = str(reference_image_1)
                        logger.info(f"🔍 REFERENCE IMAGE EXTRACTION: Converted to string: {file_path}")
                    
                    # Validate the extracted path
                    if isinstance(file_path, str) and file_path.strip():
                        if os.path.exists(file_path) and os.path.isfile(file_path):
                            logger.info(f"✅ REFERENCE IMAGE EXTRACTION: Valid file found: {file_path}")
                            return file_path
                        else:
                            logger.warning(f"❌ REFERENCE IMAGE EXTRACTION: File doesn't exist: {file_path}")
                    else:
                        logger.warning(f"❌ REFERENCE IMAGE EXTRACTION: Invalid file path: {file_path}")
                
                # Check legacy reference_images if reference_image_1 not available
                if reference_images is not None:
                    logger.info(f"🔍 REFERENCE IMAGE EXTRACTION: Falling back to legacy reference_images")
                    if isinstance(reference_images, list) and len(reference_images) > 0:
                        first_item = reference_images[0]
                        if isinstance(first_item, dict) and 'name' in first_item:
                            file_path = first_item['name']
                        elif isinstance(first_item, str):
                            file_path = first_item
                        else:
                            file_path = str(first_item)
                    elif isinstance(reference_images, dict):
                        file_path = reference_images.get('name', reference_images.get('path', str(reference_images)))
                    elif isinstance(reference_images, str):
                        file_path = reference_images
                    else:
                        file_path = str(reference_images)
                    
                    if isinstance(file_path, str) and file_path.strip() and os.path.exists(file_path) and os.path.isfile(file_path):
                        logger.info(f"✅ REFERENCE IMAGE EXTRACTION: Valid legacy file found: {file_path}")
                        return file_path
                    else:
                        logger.warning(f"❌ REFERENCE IMAGE EXTRACTION: Invalid legacy file path: {file_path}")
                        
            except Exception as e:
                logger.error(f"❌ REFERENCE IMAGE EXTRACTION ERROR: {str(e)}")
                import traceback
                logger.error(traceback.format_exc())
            
            logger.warning("❌ REFERENCE IMAGE EXTRACTION: No valid reference image found")
            return None
        
        # Perform breed detection for all providers
        first_ref_image_path = get_first_reference_image_path()
        if first_ref_image_path:
            detected_breed = detect_animal_breed(first_ref_image_path)
            logger.info(f"🐾 GLOBAL BREED DETECTION: Detected breed from first reference image: {detected_breed}")
            print(f"🐾 Detected Animal Breed (All Providers): {detected_breed}")
        else:
            logger.info("No reference image available for breed detection")

        # Process both legacy single reference and new multi-reference systems
        if provider == "Leonardo":
            # Initialize reference image data list
            reference_image_data = []
            
            # Helper function to validate and process a single reference image
            async def process_single_reference_image(ref_image, ref_name="reference"):
                if ref_image is None:
                    logger.info(f"No {ref_name} image provided")
                    return None
                    
                logger.info(f"Processing {ref_name} image: {type(ref_image)}")
                
                # Handle numpy array directly
                if isinstance(ref_image, np.ndarray):
                    try:
                        import tempfile
                        from PIL import Image
                        
                        # Create temporary directory and file
                        temp_dir = tempfile.mkdtemp()
                        temp_dirs_to_cleanup.append(temp_dir)
                        temp_file_path = os.path.join(temp_dir, f"{ref_name}_image.png")
                        
                        # Save the numpy array as an image
                        Image.fromarray(ref_image.astype(np.uint8)).save(temp_file_path)
                        logger.info(f"Converted numpy array to temporary image file: {temp_file_path}")
                        file_path = temp_file_path
                    except Exception as np_error:
                        logger.error(f"Failed to process numpy array image for {ref_name}: {str(np_error)}")
                        return None
                # Normalize the uploaded file into a file path string
                elif isinstance(ref_image, list):
                    file_path = ref_image[0]['name'] if (len(ref_image) > 0 and isinstance(ref_image[0], dict)) else ref_image[0]
                elif isinstance(ref_image, dict):
                    file_path = ref_image.get('name', ref_image)
                else:
                    file_path = ref_image
                    
                # Enhanced validation for reference image files
                valid_file_path = False
                
                if file_path is not None:
                    if isinstance(file_path, str) and file_path.strip():
                        if os.path.exists(file_path):
                            if os.path.isfile(file_path):
                                # Check if it's actually an image file by extension
                                valid_extensions = ['.jpg', '.jpeg', '.png', '.webp', '.bmp', '.tiff', '.gif']
                                if any(file_path.lower().endswith(ext) for ext in valid_extensions):
                                    valid_file_path = True
                                    logger.info(f"Valid {ref_name} image file found: {file_path}")
                                else:
                                    logger.warning(f"{ref_name} file exists but doesn't have a valid image extension: {file_path}")
                            else:
                                logger.warning(f"{ref_name} path exists but is not a file: {file_path}")
                        else:
                            logger.warning(f"{ref_name} file path doesn't exist: {file_path}")
                    else:
                        logger.warning(f"Invalid {ref_name} file path string: {file_path}")
                else:
                    logger.warning(f"{ref_name} image is None")
                    
                # Only proceed if we have a valid file path
                if valid_file_path:
                    try:
                        # Detect animal breed for the first reference image (reference_image_1 or legacy reference)
                        if ref_name in ["reference_image_1", "legacy reference"]:
                            detected_breed = detect_animal_breed(file_path)
                            logger.info(f"🐾 BREED DETECTION: Detected breed for {ref_name}: {detected_breed}")
                            print(f"🐾 Detected Animal Breed: {detected_breed}")
                        
                        # Upload the reference image to Leonardo
                        image_id = await upload_image_to_leonardo(file_path)
                        if image_id:
                            logger.info(f"Successfully uploaded {ref_name} image to Leonardo: {file_path}, ID: {image_id}")
                            return image_id
                        else:
                            logger.warning(f"Failed to get image ID for {ref_name} image: {file_path}")
                            return None
                    except Exception as upload_error:
                        logger.error(f"Error uploading {ref_name} image to Leonardo: {str(upload_error)}")
                        return None
                else:
                    logger.warning(f"Invalid {ref_name} image path: {file_path}")
                    return None
            
            # Process legacy single reference system for backward compatibility
            if reference_images is not None:
                image_id = await process_single_reference_image(reference_images, "legacy reference")
                if image_id:
                    init_image_ids.append(image_id)
            
            # Process individual reference images (1, 2, 3)
            
            # Process reference image 1
            if reference_image_1 is not None:
                logger.info(f"🔍 DEBUGGING: Processing reference_image_1, ref_type_1='{ref_type_1}', ref_strength_1='{ref_strength_1}'")
                image_id = await process_single_reference_image(reference_image_1, "reference_image_1")
                if image_id:
                    logger.info(f"✅ Successfully got image_id for reference_image_1: {image_id}")
                    reference_image_data.append({
                        'id': image_id,
                        'type': ref_type_1 if ref_type_1 != "None" else None,
                        'strength': ref_strength_1,
                        'name': 'reference_image_1'
                    })
                    init_image_ids.append(image_id)
                    logger.info(f"📊 Added reference_image_1 to data: type='{ref_type_1}', strength='{ref_strength_1}', id='{image_id}'")
                else:
                    logger.error(f"❌ Failed to get image_id for reference_image_1")
            else:
                logger.info(f"ℹ️ reference_image_1 is None, skipping")
            
            # Process reference image 2
            if reference_image_2 is not None:
                logger.info(f"🔍 DEBUGGING: Processing reference_image_2, ref_type_2='{ref_type_2}', ref_strength_2='{ref_strength_2}'")
                image_id = await process_single_reference_image(reference_image_2, "reference_image_2")
                if image_id:
                    logger.info(f"✅ Successfully got image_id for reference_image_2: {image_id}")
                    reference_image_data.append({
                        'id': image_id,
                        'type': ref_type_2 if ref_type_2 != "None" else None,
                        'strength': ref_strength_2,
                        'name': 'reference_image_2'
                    })
                    init_image_ids.append(image_id)
                    logger.info(f"📊 Added reference_image_2 to data: type='{ref_type_2}', strength='{ref_strength_2}', id='{image_id}'")
                else:
                    logger.error(f"❌ Failed to get image_id for reference_image_2")
            else:
                logger.info(f"ℹ️ reference_image_2 is None, skipping")
            
            # Process reference image 3
            if reference_image_3 is not None:
                logger.info(f"🔍 DEBUGGING: Processing reference_image_3, ref_type_3='{ref_type_3}', ref_strength_3='{ref_strength_3}'")
                image_id = await process_single_reference_image(reference_image_3, "reference_image_3")
                if image_id:
                    logger.info(f"✅ Successfully got image_id for reference_image_3: {image_id}")
                    reference_image_data.append({
                        'id': image_id,
                        'type': ref_type_3 if ref_type_3 != "None" else None,
                        'strength': ref_strength_3,
                        'name': 'reference_image_3'
                    })
                    init_image_ids.append(image_id)
                    logger.info(f"📊 Added reference_image_3 to data: type='{ref_type_3}', strength='{ref_strength_3}', id='{image_id}'")
                else:
                    logger.error(f"❌ Failed to get image_id for reference_image_3")
            else:
                logger.info(f"ℹ️ reference_image_3 is None, skipping")
                
            # Enhanced debugging for image IDs
            if init_image_ids:
                logger.info(f"🎯 FINAL SUMMARY: Available image IDs for controlnet: {init_image_ids}")
                logger.info(f"🎯 FINAL SUMMARY: Reference image data count: {len(reference_image_data)}")
                for i, ref_data in enumerate(reference_image_data):
                    logger.info(f"🎯 FINAL SUMMARY: Ref {i+1}: name='{ref_data['name']}', type='{ref_data['type']}', strength='{ref_data['strength']}', id='{ref_data['id']}'")
            else:
                logger.warning("⚠️ FINAL SUMMARY: No image IDs available for controlnet configuration")

        # Generate images based on selected provider
        result_images = []
        image_objects = []
        card_image_objects = []
        image_paths = []
        card_image_paths = []
        ref_image_paths = []
        generation_id = None
        
        # Provider code block
        if provider == "Leonardo":
            # Use Leonardo for generation
            # Get first available model as fallback if model_name not found
            fallback_model = list(MODEL_NAMES.values())[0] if MODEL_NAMES else "b2614463-296c-462a-9586-aafdb8f00e36"
            model_id = MODEL_NAMES.get(model_name, fallback_model)
            
            # Note: Stop flag check removed - proceeding with Leonardo API call
            if stop_flag:
                logger.info("Stop flag detected but proceeding with Leonardo API call")
                print("[INFO] Stop flag detected but continuing with Leonardo generation")
            
            # Initialize payload for Leonardo API
            payload = {
                "prompt": prompt,
                "modelId": model_id,
                "width": width,
                "height": height,
                "num_images": num_images,
                "guidance_scale": guidance_scale,  # Add guidance_scale parameter
                "promptMagic": False,  # Default to False
            }
            
            # Convert possible NumPy arrays to Python types for safe boolean checks
            if isinstance(negative_prompt, (list, np.ndarray)):
                negative_prompt = str(negative_prompt[0]) if len(negative_prompt) > 0 else ""
                
            if isinstance(preset_style, (list, np.ndarray)):
                preset_style = str(preset_style[0]) if len(preset_style) > 0 else None
                
            if isinstance(image_process_mode, (list, np.ndarray)):
                image_process_mode = str(image_process_mode[0]) if len(image_process_mode) > 0 else None
                
            # Check if negative_prompt is a valid string and not empty before using it
            if negative_prompt is not None and isinstance(negative_prompt, str) and negative_prompt.strip():
                payload["negative_prompt"] = negative_prompt.strip()
                print(f"Using negative prompt: {negative_prompt.strip()}")
            else:
                print("No negative prompt provided or negative prompt is empty")
            
            # Add seed parameter if provided
            if seed is not None:
                try:
                    # Directly handle integer seed values first (most common case from generate_wrapper)
                    if isinstance(seed, int):
                        payload["seed"] = seed
                        logger.info(f"Using integer seed value directly: {seed}")
                        print(f"Using seed value: {seed} for generation")
                    # Handle potential string inputs by checking if it's a valid integer string
                    elif isinstance(seed, str):
                        # Check if the string is a valid number
                        if seed.strip().replace('-', '').isdigit():
                            seed_value = int(seed.strip())
                            # Add seed to payload
                            payload["seed"] = seed_value
                            logger.info(f"Using seed value converted from string: {seed_value}")
                            print(f"Using seed value: {seed_value} for generation")
                        else:
                            logger.warning(f"Non-numeric seed value provided: {seed}")
                            print(f"Non-numeric seed value ignored: {seed}")
                    # Handle float (convert to int)
                    elif isinstance(seed, float):
                        seed_value = int(seed)
                        payload["seed"] = seed_value
                        logger.info(f"Using seed value converted from float: {seed_value}")
                        print(f"Using seed value: {seed_value} for generation")
                    else:
                        logger.warning(f"Unsupported seed type: {type(seed)}")
                        print(f"Unsupported seed type ignored: {type(seed)}")
                except (ValueError, TypeError) as e:
                    logger.warning(f"Invalid seed value: {seed}, error: {str(e)}")
                    print(f"Invalid seed value ignored: {seed}")
            
            # Check if preset_style is a valid string and not "None" before using it
            preset_style_is_valid = preset_style is not None and isinstance(preset_style, str) and preset_style != "None"
            if preset_style_is_valid:
                style_uuid = PRESET_STYLES.get(preset_style)
                if style_uuid:
                    # Add styleUUID parameter for compatibility
                    payload["styleUUID"] = style_uuid
                    logger.info(f"Added styleUUID to payload: {style_uuid} from preset: {preset_style}")
                else:
                    logger.warning(f"No UUID found for preset style: {preset_style}. Available presets: {list(PRESET_STYLES.keys())}")
            else:
                logger.info(f"No preset style selected or preset is 'None': {preset_style}")
            
            # Prompt Magic is disabled - always set to False
            payload["promptMagic"] = False
            logger.info("Prompt Magic disabled in Leonardo payload")
            
            # PRINT LEONARDO PAYLOAD
            print("\n===== LEONARDO API PAYLOAD =====")
            import json
            print(json.dumps(payload, indent=2))
            print("================================\n")
            
            # Handle multiple image processing modes with controlnets
            controlnets = []
            
            # Helper function to get preprocessor ID for a reference type
            def get_preprocessor_id(ref_type, model_name):
                # Map short names to full names for backward compatibility
                ref_type_mapping = {
                    "Style": "Style Reference",
                    "Character": "Character Reference", 
                    "Content": "Content Reference"
                }
                
                # Convert short name to full name if needed
                if ref_type in ref_type_mapping:
                    ref_type = ref_type_mapping[ref_type]
                
                if ref_type not in ["Style Reference", "Character Reference", "Content Reference"]:
                    return None
                
                # Check for Character Reference incompatibility
                if ref_type == "Character Reference" and model_name in ["Phoenix 1.0", "Flux Dev", "Flux Schnell"]:
                    logger.warning(f"Character Reference is not supported with {model_name} model.")
                    return None
                
                # Get the appropriate preprocessor ID
                preprocessor_id = IMAGE_PROCESS_MODES.get(ref_type)
                
                # Check for model-specific overrides
                if ref_type == "Style Reference":
                    if model_name == "Phoenix 1.0":
                        preprocessor_id = 166
                        logger.info(f"Using Phoenix 1.0 specific Style Reference preprocessor ID: {preprocessor_id}")
                    elif model_name == "Flux Dev":
                        preprocessor_id = 299
                        logger.info(f"Using Flux Dev specific Style Reference preprocessor ID: {preprocessor_id}")
                    elif model_name == "Flux Schnell":
                        preprocessor_id = 298
                        logger.info(f"Using Flux Schnell specific Style Reference preprocessor ID: {preprocessor_id}")
                elif ref_type == "Content Reference":
                    if model_name == "Phoenix 1.0":
                        preprocessor_id = 364
                        logger.info(f"Using Phoenix 1.0 specific Content Reference preprocessor ID: {preprocessor_id}")
                    elif model_name == "Flux Dev":
                        preprocessor_id = 233
                        logger.info(f"Using Flux Dev specific Content Reference preprocessor ID: {preprocessor_id}")
                    elif model_name == "Flux Schnell":
                        preprocessor_id = 232
                        logger.info(f"Using Flux Schnell specific Content Reference preprocessor ID: {preprocessor_id}")
                
                return preprocessor_id
            
            # Use ref_type_1 and ref_strength_1 as legacy fallback values for backward compatibility
            legacy_reference_type = ref_type_1 if ref_type_1 and ref_type_1 != "None" else "None"
            legacy_reference_strength = ref_strength_1 if ref_strength_1 else "Mid"
            
            # Handle legacy reference type conversion
            if isinstance(legacy_reference_type, (list, np.ndarray)) and len(legacy_reference_type) > 0:
                legacy_reference_type = str(legacy_reference_type[0])
            
            # Handle legacy reference strength conversion
            if isinstance(legacy_reference_strength, (list, np.ndarray)) and len(legacy_reference_strength) > 0:
                legacy_reference_strength = str(legacy_reference_strength[0])
            
            # Check for legacy single mode (backward compatibility)
            image_process_mode_is_valid = image_process_mode is not None and isinstance(image_process_mode, str) and image_process_mode != "None"
            reference_type_is_valid = legacy_reference_type is not None and isinstance(legacy_reference_type, str) and legacy_reference_type != "None"
            
            # Process legacy single reference system if available and no multi-reference data
            if (reference_type_is_valid or image_process_mode_is_valid) and init_image_ids and len(reference_image_data) == 0:
                init_image_id = init_image_ids[0]
                logger.info(f"Using legacy single reference system with image ID: {init_image_id}")
                
                # Legacy single mode processing
                if image_process_mode_is_valid:
                    preprocessor_id = IMAGE_PROCESS_MODES.get(image_process_mode)
                    if preprocessor_id is not None:
                        controlnet = {
                            "initImageId": init_image_id,
                            "initImageType": "UPLOADED",
                            "preprocessorId": preprocessor_id,
                            "strengthType": strength_type
                        }
                        controlnets.append(controlnet)
                        logger.info(f"Added legacy image_process_mode controlnet: {controlnet}")
                
                # Legacy unified reference type processing
                if reference_type_is_valid:
                    preprocessor_id = get_preprocessor_id(legacy_reference_type, model_name)
                    if preprocessor_id is not None:
                        controlnet = {
                            "initImageId": init_image_id,
                            "initImageType": "UPLOADED",
                            "preprocessorId": preprocessor_id,
                            "strengthType": legacy_reference_strength
                        }
                        controlnets.append(controlnet)
                        logger.info(f"Added legacy {legacy_reference_type} controlnet: {controlnet}")
            
            # Process multi-reference system (new approach)
            if reference_image_data:
                logger.info(f"🎯 CONTROLNET PROCESSING: Processing {len(reference_image_data)} individual reference images for controlnets")
                
                for ref_data in reference_image_data:
                    ref_id = ref_data['id']
                    ref_type = ref_data['type']
                    ref_strength = ref_data['strength']
                    ref_name = ref_data['name']
                    
                    logger.info(f"🔍 CONTROLNET: Processing {ref_name}: id='{ref_id}', type='{ref_type}', strength='{ref_strength}'")
                    
                    if ref_type and ref_type != "None":
                        preprocessor_id = get_preprocessor_id(ref_type, model_name)
                        logger.info(f"🔍 CONTROLNET: Preprocessor ID for {ref_type} with model {model_name}: {preprocessor_id}")
                        
                        if preprocessor_id is not None:
                            controlnet = {
                                "initImageId": ref_id,
                                "initImageType": "UPLOADED", 
                                "preprocessorId": preprocessor_id,
                                "strengthType": ref_strength
                            }
                            controlnets.append(controlnet)
                            logger.info(f"✅ CONTROLNET: Added {ref_name} controlnet ({ref_type}): {controlnet}")
                            print(f"✅ CONTROLNET ADDED: {ref_name} - Type: {ref_type}, Strength: {ref_strength}, Image ID: {ref_id}")
                        else:
                            logger.warning(f"❌ CONTROLNET: Could not get preprocessor ID for {ref_name} with type {ref_type}")
                            print(f"❌ CONTROLNET FAILED: {ref_name} - No preprocessor ID found for type {ref_type}")
                    else:
                        logger.info(f"⏭️ CONTROLNET: Skipping {ref_name} - no reference type specified (type='{ref_type}')")
                        print(f"⏭️ CONTROLNET SKIPPED: {ref_name} - No reference type selected")
            else:
                logger.info(f"ℹ️ CONTROLNET: No reference image data available for multi-reference processing")
            
            # Note: Stop flag check removed during controlnet setup
            if stop_flag:
                logger.info("Stop flag detected but continuing with controlnet setup")
                print("[INFO] Stop flag detected but proceeding with controlnet setup")
            
            # Only add controlnets if we have any
            if controlnets:
                payload["controlnets"] = controlnets
                logger.info(f"🎯 PAYLOAD: Final controlnet configuration added to payload: {payload['controlnets']}")
                
                # Print controlnet information for user
                print("\n===== CONTROLNET CONFIGURATION =====")
                print(f"✅ TOTAL CONTROLNETS ADDED TO LEONARDO PAYLOAD: {len(controlnets)}")
                for i, controlnet in enumerate(controlnets):
                    controlnet_type = "Unknown"
                    preprocessor_id = controlnet.get("preprocessorId")
                    strength = controlnet.get("strengthType")
                    image_id = controlnet.get("initImageId")
                    
                    # Identify controlnet type based on preprocessor ID
                    for name, pid in IMAGE_PROCESS_MODES.items():
                        if pid == preprocessor_id:
                            controlnet_type = name
                            break
                    
                    # Check for special Phoenix/Flux models
                    if preprocessor_id == 166 or preprocessor_id == 299 or preprocessor_id == 298:
                        controlnet_type = "Style Reference"
                    elif preprocessor_id == 364 or preprocessor_id == 233 or preprocessor_id == 232:
                        controlnet_type = "Content Reference"
                        
                    print(f"✅ Controlnet #{i+1}: {controlnet_type}, Strength: {strength}, Preprocessor ID: {preprocessor_id}, Leonardo Image ID: {image_id}")
                print("===================================\n")
            else:
                # Log why controlnets weren't added
                any_controlnet_requested = reference_type_is_valid or image_process_mode_is_valid or len(reference_image_data) > 0
                print("\n===== CONTROLNET CONFIGURATION =====")
                print("❌ NO CONTROLNETS ADDED TO LEONARDO PAYLOAD")
                if any_controlnet_requested:
                    if not init_image_ids:
                        logger.warning("❌ PAYLOAD: No controlnets added because no valid reference image IDs were obtained. Check that the reference images were uploaded successfully.")
                        print("❌ REASON: No valid reference images were uploaded.")
                    elif len(reference_image_data) > 0:
                        logger.info("❌ PAYLOAD: Reference images were uploaded but no controlnets were configured (no valid reference types specified)")
                        print("❌ REASON: Reference images uploaded but no controlnet types were specified.")
                    else:
                        logger.info(f"❌ PAYLOAD: No controlnets were configured despite having image IDs: {init_image_ids}")
                        print(f"❌ REASON: Unknown issue - have image IDs: {init_image_ids}")
                elif init_image_ids:
                    logger.info("ℹ️ PAYLOAD: Reference images uploaded but no controlnet configuration requested")
                    print("ℹ️ REASON: Reference images uploaded but no controlnet options selected.")
                else:
                    print("ℹ️ REASON: No reference images provided.")
                print("===================================\n")
            
            # Add "Do not hallucinate" to the end of the prompt
            if prompt and isinstance(prompt, str):
                # Check if the prompt already ends with "Do not hallucinate"
                if not prompt.strip().endswith("Do not hallucinate"):
                    # Remove any ", Current Filename Setting" that might be at the end
                    if prompt.strip().endswith(", Current Filename Setting"):
                        prompt = prompt.strip()[:-len(", Current Filename Setting")]
                    
                    # Add the instruction at the end
                    prompt = f"{prompt.strip()}, Do not hallucinate"
                    logger.info(f"Added 'Do not hallucinate' to prompt: {prompt[:100]}...")
            
            # Call Leonardo API for image generation
            headers_gen = {
                "Authorization": f"Bearer {LEONARDO_API_KEY}",
                "Content-Type": "application/json"
            }
            
            # Ensure valid payload parameters
            # Validate width and height (must be multiples of 8 for most models)
            payload["width"] = int(width - (width % 8))
            payload["height"] = int(height - (height % 8))
            
            # Ensure model_id is valid
            if not payload.get("modelId"):
                fallback_model = list(MODEL_NAMES.values())[0] if MODEL_NAMES else "b2614463-296c-462a-9586-aafdb8f00e36"
                payload["modelId"] = fallback_model
            
            # Log full payload for debugging
            logger.info(f"Sending generation request to Leonardo with payload: {json.dumps(payload)}")
            
            try:
                response = requests.post(
                    f"{LEONARDO_API_BASE_URL}/generations",
                    headers=headers_gen,
                    json=payload
                )
                
                # Improved error handling
                if not response.ok:
                    error_info = f"Status: {response.status_code}"
                    try:
                        error_json = response.json()
                        error_info += f", Details: {json.dumps(error_json)}"
                    except:
                        error_info += f", Response: {response.text[:200]}"
                    logger.error(f"Leonardo API error: {error_info}")
                    raise Exception(f"Leonardo API error: {error_info}")
                
                response.raise_for_status()
                generation_data = response.json()
                generation_id = generation_data['sdGenerationJob']['generationId']
                logger.info(f"Generation initiated with ID: {generation_id}")
                
                # Poll for generation completion
                status = "PENDING"
                max_tries = 15
                tries = 0
                
                try:
                    while status != "COMPLETE" and tries < max_tries:
                        tries += 1
                        time.sleep(5)
                        generation_result = await get_generation(generation_id, wait=False)
                        if 'generations_by_pk' in generation_result:
                            status = generation_result['generations_by_pk']['status']
                            logger.info(f"Generation status: {status}, attempt {tries}/{max_tries}")
                            if status == "COMPLETE" and 'generated_images' in generation_result['generations_by_pk']:
                                for img_data in generation_result['generations_by_pk']['generated_images']:
                                    img_url = img_data.get('url')
                                    if img_url:
                                        result_images.append(img_url)
                                # Remove the break statement that was causing only one image to be processed
                                # The code should now process all generated images
                except Exception as gen_error:
                    logger.error(f"Error during Leonardo API call: {str(gen_error)}")
                    if hasattr(gen_error, 'response') and gen_error.response:
                        logger.error(f"Response: {gen_error.response.status_code} - {gen_error.response.text}")
                    raise
            except Exception as e:
                logger.error(f"Error calling Leonardo API: {str(e)}")
                return [], f"Error calling Leonardo API: {str(e)}", None, None, None, None
                
        elif provider == "Ideogram":
            # Use Ideogram for generation
            # Make sure we have valid parameters
            if not ideogram_model:
                ideogram_model = "Version 2a"  # Default to Version 2a
            
            if not ideogram_style:
                ideogram_style = "Auto"  # Default to Auto style
            
            # Handle case when ideogram_model is a list-like object
            if isinstance(ideogram_model, (list, np.ndarray)):
                ideogram_model = str(ideogram_model[0]) if len(ideogram_model) > 0 else "Version 2a"
                
            # Handle case when ideogram_style is a list-like object
            if isinstance(ideogram_style, (list, np.ndarray)):
                ideogram_style = str(ideogram_style[0]) if len(ideogram_style) > 0 else "Auto"
                
            # Ensure ideogram_num_images is a valid integer
            if isinstance(ideogram_num_images, (list, np.ndarray)):
                try:
                    if len(ideogram_num_images) > 0:
                        if isinstance(ideogram_num_images[0], (int, float, str)) and str(ideogram_num_images[0]).isdigit():
                            ideogram_num_images = int(ideogram_num_images[0])
                        else:
                            ideogram_num_images = 1
                    else:
                        ideogram_num_images = 1
                except Exception as e:
                    logger.warning(f"Error parsing ideogram_num_images: {str(e)}, using default value 1")
                    ideogram_num_images = 1
            elif isinstance(ideogram_num_images, str):
                if ideogram_num_images.isdigit():
                    ideogram_num_images = int(ideogram_num_images)
                else:
                    ideogram_num_images = 1
            elif not isinstance(ideogram_num_images, int):
                ideogram_num_images = 1
            
            # Direct lookup from the IDEOGRAM_MODELS and IDEOGRAM_STYLES dictionaries
            ideogram_model_val = IDEOGRAM_MODELS.get(ideogram_model, "V_2A")
            ideogram_style_val = IDEOGRAM_STYLES.get(ideogram_style, "AUTO")
            
            # Print Ideogram parameters
            print("\n===== IDEOGRAM PARAMETERS =====")
            print(f"Selected Model: {ideogram_model}")
            print(f"Matched API Value: {ideogram_model_val}")
            print(f"Raw Style from UI: '{ideogram_style}'")
            print(f"Matched Style: '{ideogram_style}'")
            print(f"Final API Style: '{ideogram_style_val}'")
            print(f"Number of images: {ideogram_num_images}")
            
            # Add "Do not hallucinate" to the end of the prompt
            if prompt and isinstance(prompt, str):
                # Check if the prompt already ends with "Do not hallucinate"
                if not prompt.strip().endswith("Do not hallucinate"):
                    # Remove any ", Current Filename Setting" that might be at the end
                    if prompt.strip().endswith(", Current Filename Setting"):
                        prompt = prompt.strip()[:-len(", Current Filename Setting")]
                    
                    # Add the instruction at the end
                    prompt = f"{prompt.strip()}, Do not hallucinate"
                    logger.info(f"Added 'Do not hallucinate' to prompt: {prompt[:100]}...")
            
            print(f"Prompt: {prompt[:100]}..." if len(prompt) > 100 else f"Prompt: {prompt}")
            
            # Convert negative_prompt for generate_with_ideogram
            if isinstance(negative_prompt, (list, np.ndarray)):
                negative_prompt_str = str(negative_prompt[0]) if len(negative_prompt) > 0 else None
            elif negative_prompt is not None and isinstance(negative_prompt, str) and negative_prompt.strip():
                negative_prompt_str = negative_prompt
            else:
                negative_prompt_str = None
                
            # Print the negative prompt safely
            if negative_prompt_str is not None:
                print(f"Negative prompt: {negative_prompt_str[:100]}..." if len(negative_prompt_str) > 100 else f"Negative prompt: {negative_prompt_str}")
            else:
                print("Negative prompt: None")
                
            # Process seed - only pass if it's a valid integer
            valid_seed = None
            if seed is not None:
                try:
                    # Handle different types of seed inputs
                    if isinstance(seed, (int, np.integer)):
                        valid_seed = int(seed)
                        print(f"Seed: {valid_seed}")
                    elif isinstance(seed, str) and seed.strip().isdigit():
                        valid_seed = int(seed.strip())
                        print(f"Seed: {valid_seed} (converted from string)")
                    elif isinstance(seed, float) and seed.is_integer():
                        valid_seed = int(seed)
                        print(f"Seed: {valid_seed} (converted from float)")
                    else:
                        logger.warning(f"Invalid seed format (not passing to API): {seed}")
                        print(f"Invalid seed format (not passing to API): {seed}")
                except Exception as e:
                    logger.warning(f"Error processing seed: {str(e)}")
                    print(f"Seed: None (error processing: {str(e)})")
            else:
                print("Seed: None (using random seed)")
                
            print("==============================\n")
            
            # Get reference image path for V_3 style reference support (same as Leonardo Reference Image 1)
            logger.info(f"🎯 IDEOGRAM V3 STYLE REFERENCE: Checking for reference image (model: {ideogram_model_val})")
            print(f"🎯 Checking for style reference image for Ideogram V3...")
            
            reference_image_path = None
            if ideogram_model_val == "V_3":
                if ideogram_disable_style_reference:
                    logger.info(f"🚫 IDEOGRAM V3 STYLE REFERENCE: Disabled by user checkbox")
                    print(f"🚫 Style reference disabled by user - using prompt-only generation")
                else:
                    reference_image_path = get_first_reference_image_path()
                    logger.info(f"🎯 IDEOGRAM V3 STYLE REFERENCE: get_first_reference_image_path() returned: {reference_image_path}")
                    print(f"🎯 Style reference path extraction result: {reference_image_path}")
                    
                    if reference_image_path:
                        if os.path.exists(reference_image_path):
                            logger.info(f"✅ IDEOGRAM V3 STYLE REFERENCE: Using style reference image: {reference_image_path}")
                            print(f"🎨 Ideogram V3: Using style reference image: {os.path.basename(reference_image_path)}")
                            print(f"🎨 Full path: {reference_image_path}")
                        else:
                            logger.warning(f"❌ IDEOGRAM V3 STYLE REFERENCE: Path exists but file missing: {reference_image_path}")
                            print(f"❌ Style reference path invalid: {reference_image_path}")
                            reference_image_path = None
                    else:
                        logger.info(f"ℹ️ IDEOGRAM V3 STYLE REFERENCE: No style reference image found")
                        print(f"ℹ️ No style reference image available for Ideogram V3")
            else:
                logger.info(f"ℹ️ IDEOGRAM V3 STYLE REFERENCE: Not using V3 model, no style reference needed")
                print(f"ℹ️ Not using Ideogram V3, style reference not applicable")
            
            # Generate images with Ideogram
            logger.info(f"🚀 IDEOGRAM GENERATION: Calling generate_with_ideogram with reference_image_path: {reference_image_path}")
            generation_result = await generate_with_ideogram(
                prompt=prompt,
                aspect_ratio="ASPECT_1_1",
                model=ideogram_model_val,
                style=ideogram_style_val,
                num_images=ideogram_num_images,
                negative_prompt=negative_prompt_str,
                seed=valid_seed,
                reference_image_path=reference_image_path,
                rendering_speed=ideogram_rendering_speed
            )
            
            if generation_result and generation_result[0]:
                result_images = generation_result[0]
            else:
                # Handle generation failure
                error_message = "Ideogram generation failed"
                if generation_result and len(generation_result) > 1:
                    error_message = generation_result[1]  # Use the error message from the generation result
                
                logger.error(f"❌ IDEOGRAM GENERATION FAILED: {error_message}")
                
                # Clean up temporary directories
                for temp_dir in temp_dirs_to_cleanup:
                    try:
                        shutil.rmtree(temp_dir, ignore_errors=True)
                    except Exception as e:
                        logger.warning(f"Error cleaning up temp dir {temp_dir}: {str(e)}")
                
                return [], error_message, None, None, None, None
        
        elif provider == "Imagen-4":
            try:
                logger.info("Initiating image generation with IMAGEN 4 using Replicate")
                print(f"🎨 Imagen-4: Generating {imagen4_num_images} image(s) with settings: aspect_ratio={imagen4_aspect_ratio}, safety_filter={imagen4_safety_filter}")
                
                # Convert reference images to bytes for the generation function
                primary_image_bytes = None
                if reference_images is not None:
                    try:
                        # Handle numpy array case
                        if isinstance(reference_images, np.ndarray):
                            if reference_images.size > 0:
                                ref_path = reference_images.flatten()[0] if reference_images.size > 0 else None
                            else:
                                ref_path = None
                        # Handle list case
                        elif isinstance(reference_images, list) and len(reference_images) > 0:
                            ref_path = reference_images[0]
                        else:
                            ref_path = None
                        
                        # Load the reference image if we have a valid path
                        if ref_path and isinstance(ref_path, str) and ref_path.strip():
                            with open(ref_path, "rb") as f:
                                primary_image_bytes = f.read()
                            logger.info(f"Loaded reference image: {ref_path}")
                        else:
                            logger.info("No valid reference image path found")
                    except Exception as ref_error:
                        logger.warning(f"Failed to load reference image: {ref_error}")
                        primary_image_bytes = None

                # Call the existing helper function for Imagen-4 generation
                generated_urls = generate_single_image_imagen4(
                    primary_image_bytes=primary_image_bytes,
                    reference_images_bytes_list=[],  # Not used in the basic version
                    prompt=prompt,
                    aspect_ratio=imagen4_aspect_ratio,
                    num_images=imagen4_num_images,
                    safety_filter_level=imagen4_safety_filter,
                    model=imagen4_model
                )

                if generated_urls and len(generated_urls) > 0:
                    result_images = generated_urls
                    logger.info(f"✅ Imagen-4: Successfully generated {len(result_images)} image(s)")
                    print(f"✅ Imagen-4: Generated {len(result_images)} image(s) successfully")
                else:
                    logger.error("❌ Imagen-4: No valid images generated")
                    return [], "Imagen-4: No valid images generated", None, None, None, None

            except Exception as e:
                logger.error(f"Error during IMAGEN 4 generation: {e}")
                return [], f"Error with IMAGEN 4: {e}", None, None, None, None
        
        else:
            logger.error(f"Unsupported provider: {provider}")
            return [], f"Unsupported provider: {provider}", None, None, None, None
        
        for i, img_data in enumerate(result_images):
            try:
                # Handle different data types from different providers
                if isinstance(img_data, str):
                    # Check if it's a URL or base64 data
                    if img_data.startswith('http') or img_data.startswith('data:'):
                        # It's a URL - download it (Leonardo, Ideogram)
                        resp = requests.get(img_data)
                        resp.raise_for_status()
                        img_content = resp.content
                        img = Image.open(io.BytesIO(img_content)).convert('RGBA')
                    else:
                        # It's base64 data (Imagen 4)
                        try:
                            img_content = base64.b64decode(img_data)
                            img = Image.open(io.BytesIO(img_content)).convert('RGBA')
                        except Exception as b64_error:
                            logger.error(f"Failed to decode base64 data: {b64_error}")
                            continue
                else:
                    # Handle other data types (lists, etc.)
                    logger.error(f"Unexpected image data type: {type(img_data)}")
                    continue
                
                # Use sequential numbering for multiple images
                if i > 0:
                    file_num = next_file_number + i
                    # Always use the numeric convention format: ThemeCodeCategoryCodeImageNumber
                    base_filename = f"{theme_code}{category_code}{file_num:05d}"
                else:
                    # For the first image, use the base filename pattern directly
                    base_filename = base_filename_pattern
                
                file_path = os.path.join(category_folder, base_filename + ".png")
                with open(file_path, 'wb') as f:
                    f.write(img_content)
                logger.info(f"Saved image to {file_path}")
                
                # Always save the original image first
                original_img = Image.open(file_path)
                original_filename = f"{base_filename}_original.png"
                original_path = os.path.join(category_folder, original_filename)
                original_img.save(original_path)
                logger.info(f"Saved original image to {original_path}")
                
                # Apply background removal using birefnet_hr
                logger.info(f"Applying automatic background removal with birefnet_hr to {file_path}")
                processed_img = remove_background_birefnet_hr(file_path)
                
                if processed_img is not None:
                    # Apply improved alpha edge processing for better appearance
                    logger.info(f"========== APPLYING ALPHA EDGE IMPROVEMENT ==========")
                    print(f"[PROCESSING] Improving alpha channel edges after background removal")
                    original_size = processed_img.size
                    
                    # Apply the improved alpha edge processing
                    processed_img = improve_alpha_edges(processed_img, threshold=180, edge_feather=4, use_gaussian_blur=True, feather_intensity=0.6)
                    
                    logger.info(f"Alpha edge improvement complete - maintaining image size: {original_size}")
                    print(f"[SUCCESS] Alpha edges improved, removing greyish-white artifacts and creating clean boundaries")
                    
                    # Save the transparent background version for card template use
                    transparent_img = processed_img.copy()
                    
                    # Apply to card template if provided
                    if card_template is not None:
                        try:
                            logger.info(f"🎴 Applying to card template for {base_filename}")
                            # Load card template
                            if isinstance(card_template, str):
                                card_template_img = Image.open(card_template)
                            else:
                                card_template_img = card_template
                            
                            # Apply the transparent image to the card template
                            card_with_image = place_image_on_card(card_template_img.copy(), transparent_img, preserve_original_alpha=True)
                            
                            # Save card image
                            card_filename = f"{base_filename}_card.png"
                            card_path = os.path.join(category_folder, card_filename)
                            card_with_image.save(card_path)
                            logger.info(f"✅ Card template applied and saved to {card_path}")
                            
                            # Add card image to collections for gallery display
                            card_image_objects.append(card_with_image)
                            card_image_paths.append(card_path)
                            
                        except Exception as card_error:
                            logger.error(f"❌ Error applying to card template for {base_filename}: {str(card_error)}")
                    
                    # ALWAYS apply white background for the main output image
                    canvas = Image.new("RGBA", processed_img.size, "WHITE")
                    # Only access alpha channel if the image has one
                    try:
                        img_bands = processed_img.split()
                        if processed_img.mode == 'RGBA' and len(img_bands) == 4:
                            canvas.paste(processed_img, mask=img_bands[3])
                        else:
                            canvas.paste(processed_img)
                    except Exception as e:
                        logger.warning(f"Error handling image channels: {str(e)}, using fallback paste")
                        canvas.paste(processed_img)
                    processed_img = canvas
                    
                    # Save in chosen output format
                    if isinstance(output_format, str) and output_format.lower() == "jpg":
                        processed_img = processed_img.convert("RGB")
                        new_file_path = os.path.join(category_folder, base_filename + ".jpg")
                    else:
                        new_file_path = os.path.join(category_folder, base_filename + ".png")
                    processed_img.save(new_file_path)
                    file_path = new_file_path
                    img = processed_img
                    
                else:
                    logger.warning(f"Background removal failed for {file_path}, using original image")
                
                # Add both original and processed images to collections
                image_objects.append(original_img)  # Add original image
                image_objects.append(img)           # Add processed image
                image_paths.append(original_path)   # Add original path
                image_paths.append(file_path)       # Add processed path
                
                # Add reference image paths for both (same reference for both versions)
                if reference_images is not None and isinstance(reference_images, list) and len(reference_images) > 0:
                    ref_image_paths.append(reference_images[0])  # For original
                    ref_image_paths.append(reference_images[0])  # For processed
                else:
                    ref_image_paths.append(None)  # For original
                    ref_image_paths.append(None)  # For processed
            except Exception as e:
                logger.error(f"Error processing image {img_url}: {str(e)}")
        
        # Function to add image to cell while maintaining aspect ratio
        def add_image_to_cell_with_aspect_ratio(worksheet, img_path, cell_reference, max_width=200, max_height=150):
            """Add an image to a specific cell while maintaining aspect ratio"""
            try:
                # Open and process the image
                with Image.open(img_path) as img:
                    # Get original dimensions
                    orig_width, orig_height = img.size
                    aspect_ratio = orig_width / orig_height
                    
                    # Calculate new dimensions while maintaining aspect ratio
                    if aspect_ratio > 1:  # Landscape
                        new_width = min(max_width, orig_width)
                        new_height = int(new_width / aspect_ratio)
                        if new_height > max_height:
                            new_height = max_height
                            new_width = int(new_height * aspect_ratio)
                    else:  # Portrait or square
                        new_height = min(max_height, orig_height)
                        new_width = int(new_height * aspect_ratio)
                        if new_width > max_width:
                            new_width = max_width
                            new_height = int(new_width / aspect_ratio)
                    
                    # Resize image while maintaining aspect ratio
                    img_resized = img.resize((new_width, new_height), Image.LANCZOS)
                    
                    # Create an in-memory file-like object for the image
                    img_buffer = io.BytesIO()
                    img_resized.save(img_buffer, format='PNG')
                    img_buffer.seek(0)
                    
                    # Create an openpyxl image object
                    xl_img = XLImage(img_buffer)
                    
                    # Get the cell to position the image properly
                    cell = worksheet[cell_reference]
                    
                    # Position the image in the cell
                    xl_img.anchor = cell_reference
                    worksheet.add_image(xl_img)
                    
                    return True
                    
            except Exception as e:
                logger.error(f"Error adding image to cell {cell_reference}: {str(e)}")
                return False

        # Generate Excel file with image details - use a timestamp to create a unique filename
        timestamp = get_gmt7_filename_timestamp()
        excel_filename = f"{theme_code}{category_code}{timestamp}.xlsx"
        excel_path = os.path.join(category_folder, excel_filename)
        wb = Workbook()
        ws = wb.active
        ws.title = "Generated Images"
        
        # Add comprehensive headers including all metadata
        headers = [
            "Generated Prompt",     # A
            "Modified Prompt",      # B  
            "Output Filename",      # C
            "Reference Image",      # D
            "Generated Image",      # E
            "Card Image",          # F
            "Activity",            # G
            "Facial Expression",   # H
            "Fur Color",           # I
            "Theme",               # J
            "Category",            # K
            "Provider",            # L
            "Model",               # M
            "Timestamp"            # N
        ]
        ws.append(headers)
        
        # Set column widths for better visibility
        column_widths = {
            'A': 60,  # Generated Prompt
            'B': 60,  # Modified Prompt
            'C': 35,  # Output Filename
            'D': 25,  # Reference Image
            'E': 25,  # Generated Image
            'F': 25,  # Card Image
            'G': 20,  # Activity
            'H': 20,  # Facial Expression
            'I': 20,  # Fur Color
            'J': 15,  # Theme
            'K': 15,  # Category
            'L': 15,  # Provider
            'M': 20,  # Model
            'N': 20   # Timestamp
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Get current timestamp for metadata
        current_timestamp = get_gmt7_timestamp()
        
        # Process each generated image and add to the Excel file
        for i, file_path in enumerate(image_paths):
            output_filename = os.path.basename(file_path)
            
            # Add a new row
            row_num = i + 2  # Start from row 2
            ws.row_dimensions[row_num].height = 150
            
            # Add generated prompt to column A
            ws.cell(row=row_num, column=1, value=prompt)
            
            # Add modified prompt to column B
            ws.cell(row=row_num, column=2, value=modified_prompt or prompt)
            
            # Add filename to column C
            ws.cell(row=row_num, column=3, value=output_filename)
            
            # Add reference image to column D if available
            if reference_images is not None:
                try:
                    ref_path = None
                    if isinstance(reference_images, str):
                        ref_path = reference_images
                    elif isinstance(reference_images, list) and len(reference_images) > 0:
                        # Handle list format - extract first item properly
                        first_item = reference_images[0]
                        if isinstance(first_item, dict) and 'name' in first_item:
                            ref_path = first_item['name']
                        elif isinstance(first_item, str):
                            ref_path = first_item
                        else:
                            ref_path = str(first_item)
                    elif hasattr(reference_images, 'name'):
                        ref_path = reference_images.name
                    
                    if ref_path and isinstance(ref_path, str) and os.path.exists(ref_path):
                        add_image_to_cell_with_aspect_ratio(ws, ref_path, f'D{row_num}')
                except Exception as e:
                    logger.error(f"Error adding reference image to Excel: {str(e)}")
            
            # Add generated image to column E
            try:
                add_image_to_cell_with_aspect_ratio(ws, file_path, f'E{row_num}')
            except Exception as e:
                logger.error(f"Error adding generated image to Excel: {str(e)}")
            
            # Add card image to column F if available
            if i < len(card_image_paths):
                try:
                    add_image_to_cell_with_aspect_ratio(ws, card_image_paths[i], f'F{row_num}')
                except Exception as e:
                    logger.error(f"Error adding card image to Excel: {str(e)}")
            
            # Add metadata columns
            ws.cell(row=row_num, column=7, value=activity or '')  # Activity
            ws.cell(row=row_num, column=8, value=facial_expression or '')  # Facial Expression
            ws.cell(row=row_num, column=9, value=fur_color or '')  # Fur Color
            ws.cell(row=row_num, column=10, value=theme or '')  # Theme
            ws.cell(row=row_num, column=11, value=category or '')  # Category
            ws.cell(row=row_num, column=12, value=provider or '')  # Provider
            ws.cell(row=row_num, column=13, value=model_name or '')  # Model
            ws.cell(row=row_num, column=14, value=current_timestamp)  # Timestamp
        
        # Save the Excel file
        wb.save(excel_path)
        logger.info(f"Excel file generated at {excel_path}")
        
        # Collect reference image metadata
        ref_image_metadata = {}
        if reference_images is not None:
            # Handle different reference image input types
            ref_path = None
            if isinstance(reference_images, str):
                ref_path = reference_images
            elif isinstance(reference_images, list) and len(reference_images) > 0:
                # Handle list format - extract first item properly
                first_item = reference_images[0]
                if isinstance(first_item, dict) and 'name' in first_item:
                    ref_path = first_item['name']
                elif isinstance(first_item, str):
                    ref_path = first_item
                else:
                    ref_path = str(first_item)
            elif hasattr(reference_images, 'name'):
                ref_path = reference_images.name
            
            if ref_path and isinstance(ref_path, str):
                ref_image_metadata = get_image_metadata(ref_path)
        
        # Collect card template metadata
        card_template_metadata = {}
        if card_template and card_template_img:
            if isinstance(card_template, str):
                card_template_metadata = get_image_metadata(card_template)
            elif hasattr(card_template, 'name'):
                card_template_metadata = get_image_metadata(card_template.name)
        
        # Text metadata files are no longer created - all metadata is now in Excel file
        individual_metadata_files = []
        
        s3_image_status = ""
        if upload_to_s3_bucket:
            theme_str = theme if isinstance(theme, str) else str(theme)
            category_str = category if isinstance(category, str) else str(category)
            
            # Include Excel file and exclude text metadata files from S3 upload
            all_files_for_s3 = image_paths + card_image_paths + [excel_path]
            
            # Add base64 files to S3 upload if enabled
            if encode_to_base64:
                try:
                    logger.info("Base64 encoding enabled - creating base64 files for S3 upload")
                    all_image_paths = image_paths + card_image_paths
                    base64_files = batch_encode_images_to_base64(all_image_paths)
                    all_files_for_s3.extend(base64_files)
                    logger.info(f"✅ Successfully created {len(base64_files)} base64 files for S3 upload")
                    print(f"✅ Successfully created {len(base64_files)} base64 files for S3 upload")
                except Exception as base64_error:
                    logger.error(f"❌ Error preparing base64 files for S3 upload: {str(base64_error)}")
                    print(f"❌ Error preparing base64 files for S3 upload: {str(base64_error)}")
            
            s3_image_urls = await asyncio.to_thread(
                upload_multiple_files_to_s3,
                all_files_for_s3,
                bucket_folder=f"{theme_str.lower()}/{category_str.lower()}"
            )
            if s3_image_urls:
                s3_image_status = f" Uploaded {len(s3_image_urls)} files (images + Excel{' + base64' if encode_to_base64 else ''}) to S3."
            else:
                s3_image_status = " Failed to upload files to S3."
        
        # Google Drive upload is now automatic - always enabled
        gdrive_image_status = ""
        # Upload regular images and Excel file using hierarchical folder structure
        regular_files_for_gdrive = image_paths + card_image_paths + [excel_path]
        
        gdrive_image_urls = await asyncio.to_thread(
            upload_multiple_files_to_google_drive_hierarchical,
            regular_files_for_gdrive,
            theme=theme,
            category=category,
            subcategory=subcategory
        )
            
        # Upload base64 files separately to dedicated Base64 folder if enabled
        gdrive_base64_urls = []
        if encode_to_base64:
            try:
                logger.info("Base64 encoding enabled - creating and uploading base64 files to dedicated Base64 folder")
                all_image_paths = image_paths + card_image_paths
                base64_files = batch_encode_images_to_base64(all_image_paths)
                logger.info(f"Successfully created {len(base64_files)} base64 files")
                
                # Upload base64 files to dedicated Base64 folder using hierarchical structure
                gdrive_base64_urls = await asyncio.to_thread(
                    upload_multiple_base64_to_google_drive_hierarchical,
                    base64_files,
                    theme=theme,
                    category=category,
                    subcategory=subcategory
                )
                logger.info(f"✅ Successfully uploaded {len(gdrive_base64_urls)} base64 files to Base64 folder")
                print(f"✅ Successfully uploaded {len(gdrive_base64_urls)} base64 files to Base64 folder")
            except Exception as base64_error:
                logger.error(f"❌ Error preparing/uploading base64 files to Google Drive: {str(base64_error)}")
                print(f"❌ Error preparing/uploading base64 files to Google Drive: {str(base64_error)}")
            
            # Combine status messages
            regular_count = len(gdrive_image_urls) if gdrive_image_urls else 0
            base64_count = len(gdrive_base64_urls) if gdrive_base64_urls else 0
            
            if regular_count > 0 or base64_count > 0:
                status_parts = []
                if regular_count > 0:
                    status_parts.append(f"{regular_count} files (images + Excel)")
                if base64_count > 0:
                    status_parts.append(f"{base64_count} base64 files to Base64 folder")
                gdrive_image_status = f" Uploaded {' and '.join(status_parts)} on Google Drive."
            else:
                gdrive_image_status = " Failed to upload files to Google Drive."
        
        for temp_dir in temp_dirs_to_cleanup:
            try:
                shutil.rmtree(temp_dir, ignore_errors=True)
            except Exception as e:
                logger.warning(f"Error cleaning up temp dir {temp_dir}: {str(e)}")
        all_image_objects = image_objects + card_image_objects
        provider_name = "Leonardo" if provider == "Leonardo" else "Ideogram"
        try:
            logger.info(f"Creating ZIP file of all generated images...")
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            zip_filename = f"{theme_code}{category_code}_{timestamp}.zip"
            zip_filepath = os.path.join(category_folder, zip_filename)
            
            # Text metadata files are no longer created - all metadata is now in Excel file
            metadata_files = []
            
            with zipfile.ZipFile(zip_filepath, 'w', compression=zipfile.ZIP_DEFLATED) as zipf:
                file_counter = 1  # Counter for sequential numbering
                
                # Add generated images with proper naming
                for img_path in image_paths:
                    if os.path.exists(img_path):
                        # Get file extension
                        file_ext = os.path.splitext(img_path)[1].lower()
                        if not file_ext:
                            file_ext = '.png'  # Default to PNG if no extension
                        
                        # Create proper filename using TTCCCNNNNN convention
                        proper_filename = f"{theme_code}{category_code}{file_counter:05d}{file_ext}"
                        
                        # Add file to ZIP with proper filename
                        zipf.write(img_path, proper_filename)
                        logger.info(f"Added to ZIP with proper naming: {proper_filename}")
                        file_counter += 1
                
                # Add card images with proper naming
                for card_path in card_image_paths:
                    if os.path.exists(card_path):
                        # Get file extension
                        file_ext = os.path.splitext(card_path)[1].lower()
                        if not file_ext:
                            file_ext = '.png'  # Default to PNG if no extension
                        
                        # Create proper filename using TTCCCNNNNN convention for card images
                        proper_filename = f"{theme_code}{category_code}{file_counter:05d}_card{file_ext}"
                        
                        # Add file to ZIP with proper filename
                        zipf.write(card_path, proper_filename)
                        logger.info(f"Added card to ZIP with proper naming: {proper_filename}")
                        file_counter += 1
                
                # Add base64 encoded files if requested
                if encode_to_base64:
                    logger.info("Base64 encoding enabled - adding base64 files to ZIP")
                    try:
                        # Combine all image paths for base64 encoding
                        all_image_paths = image_paths + card_image_paths
                        base64_files = batch_encode_images_to_base64(all_image_paths)
                        for base64_file in base64_files:
                            if os.path.exists(base64_file):
                                zipf.write(base64_file, os.path.basename(base64_file))
                                logger.info(f"✅ Added base64 file to ZIP: {os.path.basename(base64_file)}")
                                print(f"✅ Added base64 file to ZIP: {os.path.basename(base64_file)}")
                        logger.info(f"✅ Successfully added {len(base64_files)} base64 files to ZIP")
                        print(f"✅ Successfully added {len(base64_files)} base64 files to ZIP")
                    except Exception as base64_error:
                        logger.error(f"❌ Error adding base64 files to ZIP: {str(base64_error)}")
                        print(f"❌ Error adding base64 files to ZIP: {str(base64_error)}")
                        # Continue without base64 files - don't fail the entire ZIP creation
                if os.path.exists(excel_path):
                    zipf.write(excel_path, os.path.basename(excel_path))
                    logger.info(f"Added to ZIP: {excel_path}")
                # Excel file already contains all metadata - no separate metadata files needed
            logger.info(f"Created ZIP file at: {zip_filepath}")
            print(f"Created ZIP file at: {zip_filepath}")
            s3_zip_status = ""
            if upload_to_s3_bucket:
                s3_zip_url = await asyncio.to_thread(
                    upload_zip_to_s3,
                    zip_filepath,
                    theme=theme_str,
                    category=category_str
                )
                if s3_zip_url:
                    s3_zip_status = f" Uploaded ZIP to S3: {s3_zip_url}"
                    logger.info(f"Uploaded ZIP to S3: {s3_zip_url}")
                    print(f"Uploaded ZIP to S3: {s3_zip_url}")
                else:
                    s3_zip_status = " Failed to upload ZIP file to S3."
                    print("Failed to upload ZIP file to S3.")
            s3_status = s3_image_status + s3_zip_status
            gdrive_zip_status = ""
            if upload_to_gdrive:
                gdrive_zip_url = await asyncio.to_thread(
                    upload_to_google_drive,
                    zip_filepath,
                    parent_folder_id=None,
                    theme=theme,
                    category=category,
                    subcategory=subcategory
                )
                if gdrive_zip_url:
                    gdrive_zip_status = f" Uploaded ZIP to Google Drive: {gdrive_zip_url}"
                    logger.info(f"Uploaded ZIP to Google Drive: {gdrive_zip_url}")
                    print(f"Uploaded ZIP to Google Drive: {gdrive_zip_url}")
                else:
                    gdrive_zip_status = " Failed to upload ZIP file to Google Drive."
                    print("Failed to upload ZIP file to Google Drive.")
            gdrive_status = gdrive_image_status + gdrive_zip_status
            if card_template_img:
                print(f"Returning {len(all_image_objects)} images, ZIP file: {zip_filepath}")
                return all_image_objects, f"Generation complete with {provider_name}! Generated {len(image_objects)} images and {len(card_image_objects)} card images.{s3_status}{gdrive_status}\nExcel file: {excel_path}", zip_filepath, None, None, None
            else:
                print(f"Returning {len(image_objects)} images, ZIP file: {zip_filepath}")
                return image_objects, f"Generation complete with {provider_name}! Generated {len(image_objects)} images with automatic background removal and saved to {category_folder}.{s3_status}{gdrive_status}\nExcel file: {excel_path}", zip_filepath, None, None, None
        except Exception as zip_error:
            logger.error(f"Error creating ZIP file: {str(zip_error)}")
            print(f"Error creating ZIP file: {str(zip_error)}")
            s3_status = s3_image_status
            gdrive_status = gdrive_image_status
            if card_template_img:
                print(f"Returning {len(all_image_objects)} images without ZIP file")
                return all_image_objects, f"Generation complete with {provider_name}! Generated {len(image_objects)} images and {len(card_image_objects)} card images.{s3_status}{gdrive_status}\nExcel file: {excel_path}", None, None, None, None
            else:
                print(f"Returning {len(image_objects)} images without ZIP file")
                return image_objects, f"Generation complete with {provider_name}! Generated {len(image_objects)} images with automatic background removal and saved to {category_folder}.{s3_status}{gdrive_status}\nExcel file: {excel_path}", None, None, None, None
        
        finally:
            # Clean up any temporary directories
            for temp_dir in temp_dirs_to_cleanup:
                try:
                    import shutil
                    if os.path.exists(temp_dir):
                        shutil.rmtree(temp_dir)
                        logger.info(f"Cleaned up temporary directory in upload_and_generate_image: {temp_dir}")
                except Exception as cleanup_error:
                    logger.warning(f"Failed to clean up temporary directory in upload_and_generate_image {temp_dir}: {str(cleanup_error)}")

        return result_images, f"Generated {len(result_images)} images with {provider}", generation_id, image_objects, card_image_objects, s3_status + gdrive_status
    except Exception as e:
        logger.error(f"Error in upload_and_generate_image: {str(e)}")
        return [], f"Error: {str(e)}", None, None, None, None

# Helper function to upload an image to Leonardo
async def upload_image_to_leonardo(file_path):
    """Upload an image to Leonardo and return the image ID"""
    try:
        # Check if file exists
        if not os.path.exists(file_path):
            logger.error(f"File not found: {file_path}")
            raise FileNotFoundError(f"File not found: {file_path}")
            
        # Get file extension
        filename = os.path.basename(file_path)
        extension = filename.split('.')[-1].lower()
        
        logger.info(f"Uploading image to Leonardo: {file_path}")
        
        # Prepare headers and payload for getting presigned URL
        headers_req = {
            "Authorization": f"Bearer {LEONARDO_API_KEY}",
            "Content-Type": "application/json",
            "accept": "application/json"
        }
        pres_payload = {"extension": extension}
        
        # Get presigned URL
        presigned_response = requests.post(
            f"{LEONARDO_API_BASE_URL}/init-image",
            json=pres_payload,
            headers=headers_req
        )
        
        if not presigned_response.ok:
            error_msg = f"Failed to get presigned URL: {presigned_response.status_code} - {presigned_response.text}"
            logger.error(error_msg)
            raise Exception(error_msg)
            
        presigned_data = presigned_response.json()
        
        logger.info(f"Received presigned URL response")
        
        # Extract upload information
        upload_url = presigned_data['uploadInitImage']['url']
        upload_fields = json.loads(presigned_data['uploadInitImage']['fields'])
        image_id = presigned_data['uploadInitImage']['id']
        
        logger.info(f"Image ID assigned: {image_id}")
        
        # Upload the file directly from the provided path
        with open(file_path, 'rb') as file_data:
            files = {'file': file_data}
            upload_response = requests.post(
                upload_url,
                data=upload_fields,
                files=files
            )
        
        if not upload_response.ok:
            error_msg = f"Failed to upload image: {upload_response.status_code} - {upload_response.text}"
            logger.error(error_msg)
            raise Exception(error_msg)
            
        logger.info(f"Image uploaded successfully with ID: {image_id}")
        return image_id
        
    except Exception as e:
        if hasattr(e, 'response') and e.response:
            logger.error(f"Upload error details: {e.response.text}")
        logger.error(f"Error uploading image: {str(e)}")
        return None  # Return None instead of raising to allow the function to continue

async def upload_canvas_images_to_leonardo(init_image_path, mask_image_path):
    """
    Upload both init image and mask image to Leonardo for canvas inpainting.
    Returns init_image_id and mask_image_id needed for inpainting.
    """
    try:
        logger.info(f"🎨 CANVAS UPLOAD: Starting canvas image upload for inpainting")
        logger.info(f"🎨 CANVAS UPLOAD: Init image: {init_image_path}")
        logger.info(f"🎨 CANVAS UPLOAD: Mask image: {mask_image_path}")
        
        # Request presigned URLs for both init and mask images
        url = f"{LEONARDO_API_BASE_URL}/canvas-init-image"
        payload = {
            "initExtension": "png",  # Using PNG for better quality
            "maskExtension": "png"
        }
        
        headers = {
            "accept": "application/json",
            "content-type": "application/json",
            "authorization": f"Bearer {LEONARDO_API_KEY}"
        }
        
        logger.info("🎨 CANVAS UPLOAD: Requesting presigned URLs...")
        response = requests.post(url, json=payload, headers=headers)
        
        if not response.ok:
            error_msg = f"Failed to get canvas presigned URLs: {response.status_code} - {response.text}"
            logger.error(f"🎨 CANVAS ERROR: {error_msg}")
            raise Exception(error_msg)
        
        canvas_data = response.json()
        upload_canvas_data = canvas_data['uploadCanvasInitImage']
        
        # Extract upload details
        init_image_id = upload_canvas_data['initImageId']
        mask_image_id = upload_canvas_data['masksImageId']
        
        # Init image upload details
        init_fields = json.loads(upload_canvas_data['initFields'])
        init_url = upload_canvas_data['initUrl']
        init_key = upload_canvas_data['initKey']
        
        # Mask image upload details
        mask_fields = json.loads(upload_canvas_data['masksFields'])
        mask_url = upload_canvas_data['masksUrl']
        mask_key = upload_canvas_data['masksKey']
        
        logger.info(f"🎨 CANVAS UPLOAD: Got presigned URLs - Init ID: {init_image_id}, Mask ID: {mask_image_id}")
        
        # Upload init image
        logger.info("🎨 CANVAS UPLOAD: Uploading init image...")
        with open(init_image_path, 'rb') as init_file:
            init_files = {'file': init_file}
            init_response = requests.post(init_url, data=init_fields, files=init_files)
        
        if not init_response.ok:
            error_msg = f"Failed to upload init image: {init_response.status_code}"
            logger.error(f"🎨 CANVAS ERROR: {error_msg}")
            raise Exception(error_msg)
        
        logger.info(f"🎨 CANVAS UPLOAD: Init image uploaded successfully")
        logger.info(f"🎨 CANVAS UPLOAD: Init image URL: https://cdn.leonardo.ai/{init_key}")
        
        # Upload mask image
        logger.info("🎨 CANVAS UPLOAD: Uploading mask image...")
        with open(mask_image_path, 'rb') as mask_file:
            mask_files = {'file': mask_file}
            mask_response = requests.post(mask_url, data=mask_fields, files=mask_files)
        
        if not mask_response.ok:
            error_msg = f"Failed to upload mask image: {mask_response.status_code}"
            logger.error(f"🎨 CANVAS ERROR: {error_msg}")
            raise Exception(error_msg)
        
        logger.info(f"🎨 CANVAS UPLOAD: Mask image uploaded successfully")
        logger.info(f"🎨 CANVAS UPLOAD: Mask image URL: https://cdn.leonardo.ai/{mask_key}")
        
        logger.info(f"✅ CANVAS UPLOAD: Both images uploaded successfully for inpainting")
        return init_image_id, mask_image_id
        
    except Exception as e:
        logger.error(f"❌ CANVAS UPLOAD ERROR: {str(e)}")
        raise e


async def generate_inpainting_with_leonardo(
    prompt,
    init_image_id,
    mask_image_id,
    model_id="1e60896f-3c26-4296-8ecc-53e2afecc132",  # Leonardo Diffusion XL
    num_images=4,
    init_strength=0.13,  # Inpaint strength (0.87 actual strength)
    guidance_scale=7
):
    """
    Generate inpainted images using Leonardo Canvas API
    """
    try:
        logger.info(f"🎨 INPAINTING: Starting inpainting generation")
        logger.info(f"🎨 INPAINTING: Prompt: {prompt}")
        logger.info(f"🎨 INPAINTING: Model: {model_id}")
        logger.info(f"🎨 INPAINTING: Init strength: {init_strength}")
        logger.info(f"🎨 INPAINTING: Guidance scale: {guidance_scale}")
        logger.info(f"🎨 INPAINTING: Number of images: {num_images}")
        
        url = f"{LEONARDO_API_BASE_URL}/generations"
        
        payload = {
            "prompt": prompt,
            "canvasRequest": True,
            "num_images": num_images,
            "init_strength": init_strength,
            "canvasRequestType": "INPAINT",
            "guidance_scale": guidance_scale,
            "modelId": model_id,
            "canvasInitId": init_image_id,
            "canvasMaskId": mask_image_id
        }
        
        headers = {
            "accept": "application/json",
            "content-type": "application/json",
            "authorization": f"Bearer {LEONARDO_API_KEY}"
        }
        
        logger.info("🎨 INPAINTING: Sending generation request...")
        response = requests.post(url, json=payload, headers=headers)
        
        if not response.ok:
            error_msg = f"Inpainting generation failed: {response.status_code} - {response.text}"
            logger.error(f"🎨 INPAINTING ERROR: {error_msg}")
            raise Exception(error_msg)
        
        generation_data = response.json()
        generation_id = generation_data['sdGenerationJob']['generationId']
        
        logger.info(f"🎨 INPAINTING: Generation started with ID: {generation_id}")
        
        # Wait for completion and get results
        logger.info("🎨 INPAINTING: Waiting for generation to complete...")
        await asyncio.sleep(45)  # Wait for processing
        
        # Get generation results
        get_url = f"{LEONARDO_API_BASE_URL}/generations/{generation_id}"
        get_response = requests.get(get_url, headers=headers)
        
        if not get_response.ok:
            error_msg = f"Failed to get inpainting results: {get_response.status_code} - {get_response.text}"
            logger.error(f"🎨 INPAINTING ERROR: {error_msg}")
            raise Exception(error_msg)
        
        result_data = get_response.json()
        logger.info(f"✅ INPAINTING: Generation completed successfully")
        
        return result_data
        
    except Exception as e:
        logger.error(f"❌ INPAINTING ERROR: {str(e)}")
        raise e


async def generate_inpainting_with_ideogram(prompt, init_image_path, mask_image_path, style_reference_path=None, num_images=1):
    """
    Generate inpainted images using Ideogram V3 Edit API
    
    Args:
        prompt: The inpainting prompt
        init_image_path: Path to the original image
        mask_image_path: Path to the mask image
        style_reference_path: Optional path to style reference image
        num_images: Number of images to generate (default 1)
    
    Returns:
        Response data from Ideogram API
    """
    try:
        url = "https://api.ideogram.ai/v1/ideogram-v3/edit"
        
        logger.info(f"🎨 IDEOGRAM INPAINTING: Starting inpainting generation")
        logger.info(f"🎨 IDEOGRAM INPAINTING: Prompt: {prompt}")
        logger.info(f"🎨 IDEOGRAM INPAINTING: Init image: {init_image_path}")
        logger.info(f"🎨 IDEOGRAM INPAINTING: Mask image: {mask_image_path}")
        logger.info(f"🎨 IDEOGRAM INPAINTING: Style reference: {style_reference_path}")
        logger.info(f"🎨 IDEOGRAM INPAINTING: Number of images: {num_images}")
        
        # Prepare form data - EXACT format from user's example  
        data = {
            'prompt': prompt,
            'num_images': int(num_images) if num_images else 1,
            'rendering_speed': "DEFAULT"
        }
        
        # Prepare files - using exact format from Ideogram V3 API documentation
        files = {}
        
        # Add the original image (required)
        init_filename = os.path.basename(init_image_path)
        files['image'] = (init_filename, open(init_image_path, 'rb'))
        
        # Add the mask image (required)
        mask_filename = os.path.basename(mask_image_path)
        files['mask'] = (mask_filename, open(mask_image_path, 'rb'))
        
        # Handle style reference image (optional) with clean flag-based logic
        use_style_reference = False
        style_filename = None
        
        if style_reference_path and os.path.exists(style_reference_path):
            try:
                # Validate the file is accessible and readable
                with open(style_reference_path, 'rb') as test_file:
                    test_file.read(1)  # Read first byte to test accessibility
                
                style_filename = os.path.basename(style_reference_path)
                # Use exact same format as V3 generate API: 'style_reference_images': (filename, file_object)
                files['style_reference_images'] = (style_filename, open(style_reference_path, 'rb'))
                
                # Get file size for debugging
                file_size = os.path.getsize(style_reference_path)
                logger.info(f"✅ IDEOGRAM INPAINTING: Using style reference image: {style_filename}")
                logger.info(f"📁 IDEOGRAM INPAINTING: Style reference path: {style_reference_path}")
                logger.info(f"📊 IDEOGRAM INPAINTING: Style reference file size: {file_size} bytes")
                
                use_style_reference = True
                
            except Exception as style_error:
                logger.warning(f"❌ IDEOGRAM INPAINTING: Failed to access style reference image: {style_error}")
                logger.info("🎨 IDEOGRAM INPAINTING: Proceeding without style reference")
                use_style_reference = False
        else:
            logger.info("ℹ️ IDEOGRAM INPAINTING: No style reference image provided")
        
        # PRINT IDEOGRAM INPAINTING PAYLOAD - EXACT FORMAT AS USER'S EXAMPLE
        print("\n===== IDEOGRAM V3 EDIT API - EXACT USER FORMAT =====")
        print(f"URL: {url}")
        print(f"Data: {data}")
        print(f"Files: {list(files.keys())}")
        for key, (filename, _) in files.items():
            print(f"  - {key}: {filename}")
        if use_style_reference and style_filename:
            print(f"✅ Style reference included: {style_filename}")
        else:
            print("ℹ️ No style reference image")
        print("=====================================================\n")
        
        # Prepare headers - EXACT format from user's example
        headers = {
            "Api-Key": IDEOGRAM_API_KEY
        }
        
        logger.info("🎨 IDEOGRAM INPAINTING: Sending edit request (EXACT FORMAT)")
        
        # Send request using EXACT format from user's example - style reference files included in files dict
        response = requests.post(url, headers=headers, data=data, files=files)
        
        # Close file handles
        for file_tuple in files.values():
            file_tuple[1].close()
        
        if not response.ok:
            error_msg = f"Ideogram inpainting failed: {response.status_code} - {response.text}"
            logger.error(f"🎨 IDEOGRAM INPAINTING ERROR: {error_msg}")
            return None
        
        response_data = response.json()
        logger.info(f"🎨 IDEOGRAM INPAINTING: Generation completed successfully")
        logger.info(f"🎨 IDEOGRAM INPAINTING: Response: {response_data}")
        
        return response_data
        
    except Exception as e:
        logger.error(f"❌ IDEOGRAM INPAINTING ERROR: {str(e)}")
        return None


def create_mask_from_drawing(drawing_data, mask_mode="default"):
    """
    Convert Gradio drawing/sketching data to a proper inpainting mask.
    
    Args:
        drawing_data: The drawing data from Gradio ImageEditor
        mask_mode: Either "default" (Black: Unbrushed, White: Brushed) or "inverted" (White: Unbrushed, Black: Brushed)
    
    Returns:
        Path to the created mask file
    """
    try:
        import tempfile
        from PIL import Image, ImageDraw
        import numpy as np
        
        logger.info("🎨 MASK CREATION: Creating mask from drawing data")
        
        # Handle the drawing data from Gradio
        if drawing_data is None:
            raise ValueError("No drawing data provided")
        
        logger.info(f"🎨 MASK CREATION DEBUG: drawing_data type: {type(drawing_data)}")
        if isinstance(drawing_data, dict):
            logger.info(f"🎨 MASK CREATION DEBUG: drawing_data keys: {list(drawing_data.keys())}")
        
        # Extract ONLY the brush strokes, not the composite/background
        mask_image = None
        background_image = None
        
        if isinstance(drawing_data, dict):
            # Priority order: layers > composite > background > image
            # 'layers' contains just the brush strokes without background
            if 'layers' in drawing_data and drawing_data['layers'] is not None:
                image_data = drawing_data['layers']
                logger.info("🎨 MASK CREATION DEBUG: Using 'layers' data (brush strokes only)")
                
                # Layers might be a list or a single image
                if isinstance(image_data, list) and len(image_data) > 0:
                    # Use the last layer (most recent brush strokes)
                    image_data = image_data[-1]
                    logger.info(f"🎨 MASK CREATION DEBUG: Using layer {len(drawing_data['layers'])-1} from {len(drawing_data['layers'])} layers")
            
            elif 'composite' in drawing_data and 'background' in drawing_data:
                # We have both composite and background - subtract background to get brush strokes
                logger.info("🎨 MASK CREATION DEBUG: Extracting brush strokes by subtracting background from composite")
                composite_data = drawing_data['composite']
                background_data = drawing_data['background']
                
                # Convert both to PIL Images
                if isinstance(composite_data, Image.Image):
                    composite_img = composite_data
                elif isinstance(composite_data, np.ndarray):
                    composite_img = Image.fromarray(composite_data.astype(np.uint8))
                else:
                    composite_img = Image.open(composite_data) if isinstance(composite_data, str) else None
                
                if isinstance(background_data, Image.Image):
                    background_img = background_data
                elif isinstance(background_data, np.ndarray):
                    background_img = Image.fromarray(background_data.astype(np.uint8))
                else:
                    background_img = Image.open(background_data) if isinstance(background_data, str) else None
                
                if composite_img and background_img:
                    # Convert to arrays and subtract to get brush strokes
                    composite_array = np.array(composite_img.convert('RGB'))
                    background_array = np.array(background_img.convert('RGB'))
                    
                    # Calculate difference - brush strokes will have significant differences
                    diff_array = np.abs(composite_array.astype(int) - background_array.astype(int))
                    
                    # Sum across RGB channels to get intensity of changes
                    brush_intensity = np.sum(diff_array, axis=2)
                    
                    # Create mask where there's significant difference (brush strokes)
                    # Use a higher threshold to be more selective
                    brush_threshold = 30  # Minimum difference to consider as brush stroke
                    brush_mask = (brush_intensity > brush_threshold).astype(np.uint8) * 255
                    
                    mask_image = Image.fromarray(brush_mask, 'L')
                    logger.info("🎨 MASK CREATION DEBUG: Successfully extracted brush strokes from composite-background difference")
                else:
                    # Fallback to composite
                    image_data = composite_data
                    logger.info("🎨 MASK CREATION DEBUG: Fallback to composite data")
            
            else:
                # Fallback to available keys
                possible_keys = ['composite', 'image', 'background', 'data']
                for key in possible_keys:
                    if key in drawing_data and drawing_data[key] is not None:
                        image_data = drawing_data[key]
                        logger.info(f"🎨 MASK CREATION DEBUG: Using fallback key '{key}'")
                        break
                
                if image_data is None:
                    logger.warning(f"🎨 MASK CREATION DEBUG: No valid data found. Available keys: {list(drawing_data.keys())}")
                    raise ValueError(f"No valid image data found in keys: {list(drawing_data.keys())}")
        else:
            image_data = drawing_data
            logger.info("🎨 MASK CREATION DEBUG: Using direct drawing data")
        
        # Convert to PIL Image if we haven't already processed it
        if mask_image is None:
            logger.info(f"🎨 MASK CREATION DEBUG: Converting image_data type: {type(image_data)}")
            
            if isinstance(image_data, np.ndarray):
                logger.info("🎨 MASK CREATION DEBUG: Converting numpy array to PIL Image")
                # Check if it has alpha channel
                if image_data.shape[-1] == 4:  # RGBA
                    # Use alpha channel as mask - areas with alpha > 0 are brush strokes
                    alpha_channel = image_data[:, :, 3]
                    mask_image = Image.fromarray(alpha_channel, 'L')
                    logger.info("🎨 MASK CREATION DEBUG: Using alpha channel from RGBA data")
                else:
                    mask_image = Image.fromarray(image_data.astype(np.uint8))
            elif isinstance(image_data, str):
                logger.info(f"🎨 MASK CREATION DEBUG: Opening image from file path: {image_data}")
                mask_image = Image.open(image_data)
            elif isinstance(image_data, Image.Image):
                logger.info("🎨 MASK CREATION DEBUG: Already a PIL Image")
                mask_image = image_data
            elif isinstance(image_data, dict):
                # Handle nested dictionary structures
                logger.info("🎨 MASK CREATION DEBUG: Handling nested dictionary structure")
                if 'path' in image_data:
                    mask_image = Image.open(image_data['path'])
                elif 'name' in image_data:
                    mask_image = Image.open(image_data['name'])
                else:
                    raise ValueError(f"Unable to extract image from dictionary structure: {list(image_data.keys())}")
            else:
                # Last resort: try to treat as file-like object
                logger.info(f"🎨 MASK CREATION DEBUG: Attempting to open as file-like object: {type(image_data)}")
                try:
                    mask_image = Image.open(image_data)
                except Exception as e:
                    raise ValueError(f"Unable to process image data of type {type(image_data)}: {str(e)}")
        
        # Ensure we have a valid mask image
        if mask_image is None:
            raise ValueError("Failed to create mask image from drawing data")
        
        # Convert to grayscale if not already
        if mask_image.mode == 'RGBA':
            # Use alpha channel for mask if available
            mask_array = np.array(mask_image)[:, :, 3]
            logger.info("🎨 MASK CREATION DEBUG: Using alpha channel from RGBA image")
        else:
            mask_image = mask_image.convert('L')
            mask_array = np.array(mask_image)
            logger.info("🎨 MASK CREATION DEBUG: Using grayscale conversion")
        
        # Create binary mask with stricter threshold for accuracy
        # Only areas with significant pixel values should be considered brushed
        threshold = 128  # Higher threshold for more accuracy - only strong brush strokes
        binary_mask = np.where(mask_array > threshold, 255, 0).astype(np.uint8)
        
        # Additional cleanup: remove isolated pixels and small areas
        from scipy import ndimage
        try:
            # Remove small isolated pixels (noise)
            binary_mask = ndimage.binary_opening(binary_mask > 0, structure=np.ones((3,3))).astype(np.uint8) * 255
            logger.info("🎨 MASK CREATION DEBUG: Applied morphological opening to clean up mask")
        except ImportError:
            logger.warning("🎨 MASK CREATION: scipy not available, skipping morphological cleanup")
        
        # Apply mask mode (invert if needed)
        if mask_mode == "inverted":
            logger.info("🎨 MASK CREATION: Applying inverted mask mode")
            binary_mask = 255 - binary_mask  # Invert the mask
        
        # Create final mask image
        final_mask = Image.fromarray(binary_mask, 'L')
        
        # Log mask statistics for debugging
        unique_values = np.unique(binary_mask)
        white_pixels = np.sum(binary_mask == 255)
        black_pixels = np.sum(binary_mask == 0)
        total_pixels = binary_mask.size
        
        logger.info(f"🎨 MASK CREATION: Mask statistics (Mode: {mask_mode}):")
        logger.info(f"  - Unique values: {unique_values}")
        if mask_mode == "default":
            logger.info(f"  - White pixels (to inpaint): {white_pixels} ({white_pixels/total_pixels*100:.1f}%)")
            logger.info(f"  - Black pixels (to preserve): {black_pixels} ({black_pixels/total_pixels*100:.1f}%)")
        else:
            logger.info(f"  - Black pixels (to inpaint): {black_pixels} ({black_pixels/total_pixels*100:.1f}%)")
            logger.info(f"  - White pixels (to preserve): {white_pixels} ({white_pixels/total_pixels*100:.1f}%)")
        logger.info(f"  - Total pixels: {total_pixels}")
        
        # Save to temporary file
        temp_dir = tempfile.mkdtemp()
        mask_path = os.path.join(temp_dir, "inpaint_mask.png")
        final_mask.save(mask_path, "PNG")
        
        logger.info(f"🎨 MASK CREATION: Mask created and saved to: {mask_path}")
        logger.info(f"🎨 MASK CREATION: Mask size: {final_mask.size}")
        
        return mask_path
        
    except Exception as e:
        logger.error(f"❌ MASK CREATION ERROR: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise e


# --- Google Drive Folder Mappings ---

# 1. Theme to Folder Name Mapping
THEME_FOLDER_MAPPING = {
    "Pets": "01. PETS",
    "Sports": "02. SPORTS",
    "Hobbies": "03. HOBBIES",
    "Zodiac": "04. ZODIAC"
}

# 2. Category to Folder Name Mapping (by Theme)
CATEGORY_FOLDER_MAPPING = {
    "Pets": {
        "Dogs": "001. DOGS",
        "Cats": "002. CATS",
        "Alpaca": "003. ALPACA",
        "Rabbit": "004. RABBIT",
        "Polar Bear": "005. POLAR BEAR",
        "Panda": "006. PANDA",
        "Hamster": "007. HAMSTER",
        "Tiger": "008. TIGER"
    },
    "Sports": {
        "Football": "014. FOOTBALL",
        "Basketball": "015. BASKETBALL",
        "Tennis": "016. TENNIS",
        "Running": "017. RUNNING",
        "Padel": "018. PADEL",
        "Gym": "019. GYM",
        "Pilates": "020. PILATES",
        "Yoga": "021. YOGA",
        "Cycling": "022. CYCLING",
        "Hiking": "023. HIKING",
        "Badminton": "024. BADMINTON",
        "Golf": "025. GOLF"
    },
    "Hobbies": {
        "Car": "026. CAR",
        "Motorbike": "027. MOTORBIKE",
        "Game": "028. GAME",
        "Travelling": "029. TRAVELLING",
        "Movies": "030. MOVIES",
        "Music": "031. MUSIC",
        "Foodies": "032. FOODIES",
        "Martial Arts": "033. MARTIAL ARTS"
    },
    "Zodiac": {
        "Aries": "034. ARIES",
        "Taurus": "035. TAURUS",
        "Gemini": "036. GEMINI",
        "Cancer": "037. CANCER",
        "Leo": "038. LEO",
        "Virgo": "039. VIRGO",
        "Libra": "040. LIBRA",
        "Scorpio": "041. SCORPIO",
        "Sagittarius": "042. SAGITTARIUS",
        "Capricorn": "043. CAPRICORN",
        "Aquarius": "044. AQUARIUS",
        "Pisces": "045. PISCES"
    }
}

# --- UI Helper Functions ---

def get_categories_for_theme(theme):
    """Returns the list of categories for a given theme."""
    return THEME_CATEGORIES.get(theme, [])

def get_subcategories_for_category(theme, category):
    """Returns the list of subcategories for a given theme and category."""
    return THEME_CATEGORY_SUBCATEGORIES.get(theme, {}).get(category, [])

def update_category_dropdown(theme):
    """Updates the category dropdown choices based on the selected theme."""
    categories = get_categories_for_theme(theme)
    return gr.Dropdown(
        choices=categories,
        value=categories[0] if categories else None,
        interactive=True
    )



# Removed duplicate THEME_CATEGORY_SUBCATEGORIES definition - using the complete one above

def get_subcategories(theme, category):
    """Get subcategories for a given theme and category."""
    return THEME_CATEGORY_SUBCATEGORIES.get(theme, {}).get(category, [])

def update_subcategory_dropdown(theme, category):
    """Update the subcategory dropdown choices based on theme and category."""
    subcategories = get_subcategories(theme, category)
    logger.info(f"Updating subcategories for theme='{theme}', category='{category}': {subcategories}")
    
    if subcategories:
        # If subcategories exist, show the dropdown with choices
        return gr.Dropdown(
            choices=subcategories,
            value=subcategories[0],
            interactive=True,
            visible=True
        )
    else:
        # If no subcategories, hide the dropdown and set its value to None
        return gr.Dropdown(
            choices=[],
            value=None,
            interactive=False,
            visible=False
        )

# Create display images with metadata - moved outside create_gradio_ui to be available globally
def create_display_images_with_metadata(image_paths, ref_image_paths, variation_numbers, reference_filename=None):
    """Create display images without adding metadata text (as requested)"""
    try:
        if not image_paths:
            logger.warning("No image paths provided to create_display_images_with_metadata")
            return [], None
        
        display_images = []
        ref_image_path = None
        
        logger.info(f"Creating display images from {len(image_paths)} images")
        
        # If we have reference image paths, use the first one
        if ref_image_paths and len(ref_image_paths) > 0:
            ref_path = ref_image_paths[0]
            if isinstance(ref_path, dict) and 'name' in ref_path:
                ref_image_path = ref_path['name']
            else:
                ref_image_path = ref_path
        
        # Process each image to include in the gallery
        for i, img_path in enumerate(image_paths):
            # Just add the image path directly to the display list
            # This simpler approach avoids metadata rendering issues
            display_images.append(img_path)
            logger.info(f"Added image {i+1}/{len(image_paths)} to gallery: {img_path}")
        
        logger.info(f"Total images added to gallery: {len(display_images)}")
        return display_images, ref_image_path
        
    except Exception as e:
        logger.error(f"Error in create_display_images_with_metadata: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return [], None

# Define custom CSS for the Gradio interface
custom_css = """
    .gradio-container {
        font-family: 'Arial', sans-serif;
    }
    #modified_prompt_display {
        background-color: #f0f8ff;
        border-left: 3px solid #0066cc;
        padding: 8px;
    }
    .image-preview img {
        object-fit: contain;
        max-height: 100%;
        max-width: 100%;
    }
    .warning-text {
        color: #cc0000;
        font-weight: bold;
    }
"""

def bg_removal_wrapper(reference_preview, card_template, bg_method, remove_watermark_checkbox, original_zip_file_path, current_image_index=None, extracted_images_state=None):
    """Synchronous wrapper for background removal function - supports both single images and ZIP files"""
    try:
        import asyncio
        import numpy as np
        
        # Handle None inputs and numpy arrays properly
        has_reference = False
        if reference_preview is not None:
            if isinstance(reference_preview, np.ndarray):
                has_reference = reference_preview.size > 0
            elif isinstance(reference_preview, str):
                has_reference = len(reference_preview.strip()) > 0
            else:
                has_reference = bool(reference_preview)
        
        if not has_reference:
            return [], "No image provided for background removal", None
        
        # Handle card_template parameter safely
        card_template_path = None
        if card_template is not None:
            if isinstance(card_template, np.ndarray):
                # For numpy arrays, we need to save it as a temporary file
                import tempfile
                temp_dir = tempfile.mkdtemp()
                card_template_path = os.path.join(temp_dir, "temp_card_template.png")
                from PIL import Image
                Image.fromarray(card_template).save(card_template_path)
            elif isinstance(card_template, str) and card_template.strip():
                card_template_path = card_template
        
        # Handle reference_preview parameter safely
        reference_path = None
        if isinstance(reference_preview, np.ndarray):
            # For numpy arrays, we need to save it as a temporary file
            import tempfile
            temp_dir = tempfile.mkdtemp()
            
            # Extract original filename from the current image path if available
            original_filename = "temp_reference.png"  # Default fallback
            
            # If we have extracted_images_state and current_image_index, get the original filename
            if (extracted_images_state and current_image_index is not None and 
                current_image_index < len(extracted_images_state)):
                current_image_path = extracted_images_state[current_image_index]
                if isinstance(current_image_path, str):
                    original_filename = os.path.basename(current_image_path)
                    # Remove '_original' from filename if present
                    if '_original' in original_filename:
                        name_part, ext = os.path.splitext(original_filename)
                        cleaned_name = name_part.replace('_original', '')
                        original_filename = f"{cleaned_name}{ext}"
                        logger.info(f"Removed '_original' from filename: {original_filename}")
                    logger.info(f"Using processed filename: {original_filename}")
            
            reference_path = os.path.join(temp_dir, original_filename)
            from PIL import Image
            Image.fromarray(reference_preview).save(reference_path)
            logger.info(f"Saved NumPy array to temporary file with original name: {reference_path}")
        elif isinstance(reference_preview, str):
            reference_path = reference_preview
        
        # Check if we have an original ZIP file to process
        is_zip_file = False
        zip_path_to_process = None
        
        if original_zip_file_path and isinstance(original_zip_file_path, str) and original_zip_file_path.lower().endswith('.zip'):
            is_zip_file = True
            zip_path_to_process = original_zip_file_path
        
        if is_zip_file:
            # Process ZIP file with multiple images
            logger.info(f"Processing ZIP file for background removal: {zip_path_to_process}")
            result = asyncio.run(
                process_zip_with_bg_removal(
                    zip_path=zip_path_to_process,
                    card_template_path=card_template_path,
                    bg_method=bg_method,
                    should_remove_watermark=remove_watermark_checkbox
                )
            )
            
            # Result format: (processed_paths, status_message, zip_filepath)
            processed_paths, status_message, zip_filepath = result
            
            if processed_paths:
                # Return gallery-compatible format
                return processed_paths, status_message, zip_filepath
            else:
                return [], status_message, None
        else:
            # Process single image
            result = asyncio.run(
                process_image_with_birefnet(
                    image_path=reference_path,
                    card_template_path=card_template_path,
                    bg_method=bg_method,
                    should_remove_watermark=remove_watermark_checkbox
                )
            )
            
            # Result format: (output_paths_list, status_message)
            output_paths, status_message = result
            
            if output_paths:  # If successful
                # Create a ZIP file for single image download
                output_dir = os.path.join("generated_output", "removed_backgrounds")
                os.makedirs(output_dir, exist_ok=True)
                
                zip_filename = f"processed_bg_removal_{int(time.time())}.zip"
                zip_filepath = os.path.join(output_dir, zip_filename)
                
                with zipfile.ZipFile(zip_filepath, 'w', compression=zipfile.ZIP_DEFLATED) as zipf:
                    for path in output_paths:
                        if os.path.exists(path):
                            arcname = os.path.basename(path)
                            zipf.write(path, arcname)
                            logger.info(f"Added {arcname} to single-image ZIP file")
                
                return output_paths, status_message, zip_filepath
            else:
                return [], status_message, None
            
    except Exception as e:
        logger.error(f"Error in bg_removal_wrapper: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return [], f"Error processing image: {str(e)}", None

def create_gradio_ui():
    """Create the Gradio UI for the application"""
    # Initialize the demo
    demo = gr.Blocks(
        title="Bank AI Image Generator",
        theme=gr.themes.Soft(),
        css=custom_css
    )
    
    # Initialize inpainting state variables
    inpainting_enabled = gr.State(False)
    generated_images_state = gr.State([])
    inpaint_results = gr.State([])

        with gr.Row():
            # with gr.Column(scale=2):
            # --- Multi-reference image upload for Leonardo ---
            reference_image_1 = gr.File(
                label="Reference Image 1 (triggers prompt generation, supports ZIP)",
                file_types=["image", ".jpg", ".jpeg", ".png", ".avif", ".webp", ".zip"],
                interactive=True
            )
            reference_image_2 = gr.File(
                label="Reference Image 2 (Optional, single image only)",
                file_types=["image", ".jpg", ".jpeg", ".png", ".avif", ".webp"],
                interactive=True
            )

            reference_image_3 = gr.File(
                label="Reference Image 3 (Optional, single image only)",
                file_types=["image", ".jpg", ".jpeg", ".png", ".avif", ".webp"],
                interactive=True
            )       
            
            # Card template upload (optional)
            card_template = gr.File(
                label="Upload Card Template (Optional, JPG or PNG only)",
                file_types=["image", ".jpg", ".jpeg", ".png"],
                interactive=True
            )
        with gr.Row():        
            ref_type_1 = gr.Dropdown(
                label="Reference Type 1",
                choices=["None", "Style", "Character", "Content"],
                value="None",
                interactive=True
            )
            ref_strength_1 = gr.Dropdown(
                label="Strength 1",
                choices=["Low", "Mid", "High"],
                value="Mid",
                interactive=True
            )    
            ref_type_2 = gr.Dropdown(
                label="Reference Type 2",
                choices=["None", "Style", "Character", "Content"],
                value="None",
                interactive=True
            )
            ref_strength_2 = gr.Dropdown(
                label="Strength 2",
                choices=["Low", "Mid", "High"],
                value="Mid",
                interactive=True
            )
            ref_type_3 = gr.Dropdown(
                label="Reference Type 3",
                choices=["None", "Style", "Character", "Content"],
                value="None",
                interactive=True
            )
            ref_strength_3 = gr.Dropdown(
                label="Strength 3",
                choices=["Low", "Mid", "High"],
                value="Mid",
                interactive=True
            ) 
            
        with gr.Row():
            # Reference image 1 preview (supports ZIP and multiple images)
            reference_preview = gr.Image(
                label="Reference Image 1 Preview",
                interactive=False,
                height=400
            )
            
            # Reference image 2 preview (single image only)
            reference_2_preview = gr.Image(
                label="Reference Image 2 Preview",
                interactive=False,
                height=400
            )
            
            # Reference image 3 preview (single image only)
            reference_3_preview = gr.Image(
                label="Reference Image 3 Preview", 
                interactive=False,
                height=400
            )
    
            card_template_preview = gr.Image(
                label="Card Template",
                interactive=False,
                height=400
            )
            
        with gr.Row():
            # Generated prompt display from current image
            generated_prompt_display = gr.Textbox(
                label="Generated Prompt (Current Image)",
                lines=3,
                interactive=False,
                placeholder="Upload an image to generate a prompt automatically..."
            )
            
            # Create hidden variables to maintain compatibility with existing code
            prompt_modification_details = gr.Textbox(visible=False, interactive=False)
            
            # Make the modified prompt display visible and properly styled
            modified_prompt_display = gr.Textbox(
                label="Modified Prompt",
                lines=3,
                interactive=True,  # Allow manual editing
                visible=True,
                placeholder="Modified prompt with activity/expression will appear here...",
                elem_id="modified_prompt_display"
            )
            # Negative prompt input (still editable)
            negative_prompt = gr.Textbox(
                label="Negative Prompt",
                lines=2,
                value="Cropped hair, rainbow hair, ombre hair, cropped face, cropped hand, cropped legs, defect eyes, defect hands, defect feet, defect fingers, extra fingers, extra feet, extra objects, black and white, sketch-like, 2D",
                interactive=True,
                info="Default negative prompts are pre-filled but can be edited or cleared as needed."
            )                   
            
        # Navigation controls for multiple images - positioned below prompts
        with gr.Row():
            # Navigation controls for Reference Image 1 only (supports ZIP)
            with gr.Row(visible=False) as image_nav_controls:
                prev_button = gr.Button("← Previous Image")
                image_counter = gr.Markdown("Image 0/0")
                next_button = gr.Button("Next Image →")
                
            # Navigation controls for Reference Image 2 (single images only, no ZIP)
            with gr.Row(visible=False) as image_2_nav_controls:
                prev_button_2 = gr.Button("← Previous Ref 2")
                image_counter_2 = gr.Markdown("Ref 2: No images")
                next_button_2 = gr.Button("Next Ref 2 →")
                
            # Navigation controls for Reference Image 3 (single images only, no ZIP)  
            with gr.Row(visible=False) as image_3_nav_controls:
                prev_button_3 = gr.Button("← Previous Ref 3")
                image_counter_3 = gr.Markdown("Ref 3: No images")
                next_button_3 = gr.Button("Next Ref 3 →")
            
        # Manual Variation Generation Section
        with gr.Group():
            gr.Markdown("### 🎨 Manual Variation Generation")
            with gr.Row():
                manual_variation_input = gr.Textbox(
                    label="Manual Variation Request",
                    placeholder="e.g., 'Generate 3 variations of ice cream', 'Create 5 different dog poses', etc.",
                    interactive=True,
                    scale=3
                )
                ethnicity_dropdown = gr.Dropdown(
                    label="Ethnicity",
                    choices=list(ETHNIC_TRAITS.keys()),
                    value="Auto",
                    interactive=True,
                    info="Select ethnicity to apply specific physical traits.",
                    scale=1
                )
                generate_variations_button = gr.Button("Submit", variant="primary", scale=1)
        
        # Combined Prompt Utilities Section
        with gr.Row():
            gr.Markdown("### 📋 Prompt Utilities")
            
        with gr.Row():
            with gr.Column(scale=1):
                gr.Markdown("**Copy Generated Prompt**")
                gr.Markdown("_Copy the automatically generated prompt from uploaded image_")
                copy_generated_prompt_button = gr.Button(
                    "📋 Copy Generated Prompt",
                    variant="secondary",
                    elem_id="copy_generated_prompt_btn"
                )
            
            with gr.Column(scale=1):
                gr.Markdown("**Copy Modified Prompt**")
                gr.Markdown("_Copy the modified prompt with activity/expression/fur color_")
                copy_modified_prompt_button = gr.Button(
                    "📋 Copy Modified Prompt",
                    variant="secondary",
                    elem_id="copy_modified_prompt_btn"
                )
                
            with gr.Column(scale=1):
                gr.Markdown("**Regenerate Current Prompt**")
                gr.Markdown("_Regenerate prompt and values for currently selected image only_")
                regenerate_prompt_button = gr.Button(
                    "🔄 Regenerate Current Prompt",
                    variant="secondary",
                    elem_id="regenerate_prompt_btn"
                )
                
        with gr.Row():   
            # Add checkboxes for pre-defined vs Qwen-generated options
            with gr.Group():
                gr.Markdown("### Generation Method Selection")
                with gr.Row():
                    use_predefined_options = gr.Checkbox(
                        label="Use Pre-defined Options - If checked, re-iterate buttons will use pre-defined activities/expressions/fur colors",
                        value=True
                    )
                    use_qwen_generation = gr.Checkbox(
                        label="Use Qwen AI Generation - If checked, re-iterate buttons will use Qwen to generate new activities/expressions/fur colors",
                        value=False
                    )
                
                # Function to sync the checkboxes (when one is checked, uncheck the other)
                def sync_predefined_checkbox(checked):
                    return not checked
                    
                def sync_qwen_checkbox(checked):
                    return not checked
                    
                # Add event handlers to keep checkboxes in sync
                use_predefined_options.change(
                    fn=sync_qwen_checkbox,
                    inputs=[use_predefined_options],
                    outputs=[use_qwen_generation]
                )
                
                use_qwen_generation.change(
                    fn=sync_predefined_checkbox,
                    inputs=[use_qwen_generation],
                    outputs=[use_predefined_options]
                )
                
                # Add new disable checkbox for activity/expression/fur color
                with gr.Row():
                    disable_activity_expression_fur = gr.Checkbox(
                        label="Disable Activity/Expression/Fur Color - Skip automatic generation of activities, expressions, and fur colors",
                        value=False,
                        info="When checked, only the base prompt will be used without activity/expression/fur enhancements"
                    )

        with gr.Column(visible=True) as activity_section: # This group will be toggled
            # Character activity and facial expression inputs
            with gr.Row():
                with gr.Column():
                    activity_input = gr.Textbox(
                        label="Activity (Optional)",
                        placeholder="Describe an action or activity, e.g., 'Leaping gracefully over a tiny rain puddle'",
                        interactive=True
                    )
                    reiterate_activity_button = gr.Button("Re-Iterate", scale=1)
                
                with gr.Column():
                    facial_expression_input = gr.Textbox(
                        label="Facial Expression (Optional)",
                        placeholder="Describe the facial expression, e.g., 'Exuberant delight'",
                        interactive=True
                    )
                    reiterate_expression_button = gr.Button("Re-Iterate", scale=1)

                with gr.Column():
                    fur_color_input = gr.Textbox(
                        label="Fur Color (Optional) - For non-human subjects",
                        placeholder="Describe the fur color, e.g., 'Golden retriever with a reddish-gold coat'",
                        interactive=True
                    )
                    reiterate_fur_color_button = gr.Button("Re-Iterate", scale=1)

        # Function to toggle the visibility of the activity section
        def toggle_activity_section(checked):
            return gr.update(visible=not checked)

        # Event handler for the checkbox
        disable_activity_expression_fur.change(
            fn=toggle_activity_section,
            inputs=[disable_activity_expression_fur],
            outputs=[activity_section]
        )
                

                
        # Main generation controls - provider, theme, category, etc.
        with gr.Row():
            # Output format option 
            output_format = gr.Radio(
                choices=["png", "jpg"], 
                value="png", 
                label="Output Format"
            )
            
            # Google Drive upload is now automatic - no checkbox needed
            

            
            # Filename convention is now hardcoded to use numeric naming
            # No UI element needed - always uses "Current Filename Setting" (numeric)
            
            # Add Base64 encoding checkbox
            base64_encode_checkbox = gr.Checkbox(
                label="Encode Images to Base64",
                value=False,
                interactive=True,
                info="When checked, generated images will be encoded to base64 and saved alongside regular images in ZIP files, S3, and Google Drive."
            )
            
            # Provider selection radio
            provider_tabs = gr.Radio(
                choices=["Leonardo", "Ideogram", "Imagen-4"],
                label="Select Provider",
                value="Leonardo"
            ) 
                        # Separated theme and category dropdowns
            theme_dropdown = gr.Dropdown(
                label="Theme",
                choices=list(THEME_CATEGORIES.keys()),
                value=list(THEME_CATEGORIES.keys())[0]
            )
            initial_theme = list(THEME_CATEGORIES.keys())[0]
            initial_categories = get_categories_for_theme(initial_theme)
            category_dropdown = gr.Dropdown(
                label="Category",
                choices=initial_categories,
                value=initial_categories[0] if initial_categories else None,
                interactive=True
            )

            initial_subcategories = get_subcategories(initial_theme, initial_categories[0] if initial_categories else None)
            subcategory_dropdown = gr.Dropdown(
                label="Subcategory",
                choices=initial_subcategories,
                value=initial_subcategories[0] if initial_subcategories else None,
                interactive=True,
                visible=True if initial_subcategories else False  # Show if subcategories exist
            )
            
            # Image counter override dropdown
            counter_override_dropdown = gr.Dropdown(
                label="Image Counter Override (Optional)",
                choices=["Auto"] + list(range(1, 101)),
                value="Auto",
                interactive=True,
                info="Override the starting image counter. Select 'Auto' for automatic numbering or choose 1-100 to start from that number."
            )

        # Event handlers for dropdowns
        def update_category_and_subcategory(theme):
            categories = get_categories_for_theme(theme)
            first_category = categories[0] if categories else None
            subcategories = get_subcategories(theme, first_category)
            return (
                gr.Dropdown(choices=categories, value=first_category, interactive=True),
                gr.Dropdown(choices=subcategories, value=subcategories[0] if subcategories else None, interactive=True)
            )

        theme_dropdown.change(
            fn=update_category_and_subcategory,
            inputs=[theme_dropdown],
            outputs=[category_dropdown, subcategory_dropdown]
        )

        category_dropdown.change(
            fn=update_subcategory_dropdown,
            inputs=[theme_dropdown, category_dropdown],
            outputs=[subcategory_dropdown]
        )

        def log_subcategory_change(theme, category, subcategory):
            logger.info(f"[UI Change] Theme: {theme}, Category: {category}, Subcategory: {subcategory}")

        subcategory_dropdown.change(
            fn=log_subcategory_change,
            inputs=[theme_dropdown, category_dropdown, subcategory_dropdown],
            outputs=[]
        )

        with gr.Row():
            remove_bg_button = gr.Button("Remove Background & Apply to Card Template")
            generate_with_activity_button = gr.Button("Generate With Activity/Expression")
            generate_button = gr.Button("Generate", variant="primary")
            # Status is now shown on the image itself - keeping this hidden for backwards compatibility
            bg_removal_status = gr.Textbox(label="Background Removal Status", interactive=False, visible=False)
        
        # Add background removal method selection
        with gr.Row():
            bg_method = gr.Radio(["birefnet_hr", "photoroom"],
                            label="Background Removal Method", 
                            value="birefnet_hr",
                            interactive=True)
            remove_watermark_checkbox = gr.Checkbox(
                label="Remove Watermarks", 
                value=False,
                info="Check this box to remove watermarks from processed images"
            )

            # Background removal is now always active - original images are also saved alongside processed ones
         
        # Add stop button in a new row, directly below the generate buttons
        with gr.Row():
            # Create an empty column to align the stop button with the generate button
            with gr.Column(scale=1):
                pass
            with gr.Column(scale=2):
                stop_button = gr.Button("Stop Generation", variant="stop", size="lg")
            with gr.Column(scale=1):
                pass
         
        # Leonardo and Ideogram settings
        with gr.Row():           
            # Split the screen into two columns for side-by-side parameter display
            with gr.Column(scale=2):
                gr.Markdown("### Leonardo Settings")
                
                # Warning message for Leonardo settings
                leonardo_warning = gr.Markdown(
                    "⚠️ These settings only apply when Leonardo is selected as the provider.",
                    visible=True
                )

                # Note: Image Processing Mode removed - now handled by Reference Type 1 and Strength 1
                
                # Preset style selection
                preset_style = gr.Dropdown(
                    label="Preset Style",
                    choices=list(PRESET_STYLES.keys()),
                    value="3D Render",
                    interactive=True
                )
                
                # Model selection
                leonardo_model_dropdown = gr.Dropdown(
                    label="Model",
                    choices=list(MODEL_NAMES.keys()),
                    value=list(MODEL_NAMES.keys())[3] if MODEL_NAMES else "Phoenix 1.0",
                    interactive=True
                )
                
                # Number of images
                leonardo_num_images = gr.Dropdown(
                    label="Number of Images",
                    choices=list(range(1, 9)),  # 1 to 8 images
                    value=1,
                    interactive=True
                )
                
                # Add guidance scale slider
                guidance_scale_slider = gr.Slider(
                    label="Guidance Scale",
                    minimum=1,
                    maximum=20,
                    step=1.0,
                    value=7,
                    interactive=True,
                    info="Controls how closely the image follows the prompt (1-20). Higher values = more prompt adherence."
                )
                
                # Update the Leonardo seed input maximum
                leonardo_seed = gr.Number(
                    label="Seed (optional)",
                    value=None,
                    precision=0,
                    minimum=0,
                    maximum=9223372036854775807,  # Max value for int64 (2^63 - 1)
                    interactive=True,
                    info="Enter a number for reproducible generations. Leave empty for random seed.",
                    elem_id="leonardo_seed_input"  # Add explicit element ID for better tracking
                )
                
            # Ideogram settings
            with gr.Column(scale=2):
                gr.Markdown("### Ideogram Settings")

                # Warning message for Ideogram settings
                ideogram_warning = gr.Markdown(
                    "⚠️ These settings only apply when Ideogram is selected as the provider.",
                    visible=True
                )               

                # Model
                ideogram_model = gr.Dropdown(
                    label="Model",
                    choices=list(IDEOGRAM_MODELS.keys()),
                    value=list(IDEOGRAM_MODELS.keys())[0] if IDEOGRAM_MODELS else "Version 2a",
                    interactive=True,
                    info="Select Ideogram model version"
                )
                
                # Style
                ideogram_style = gr.Dropdown(
                    label="Style",
                    choices=list(IDEOGRAM_STYLES.keys()),
                    value=list(IDEOGRAM_STYLES.keys())[0] if IDEOGRAM_STYLES else "Auto",
                    interactive=True,
                    info="Select style for the generated images"
                )
                
                # Number of images
                ideogram_num_images = gr.Dropdown(
                    label="Number of Images",
                    choices=list(range(1, 9)),  # 1 to 8 images
                    value=1,
                    interactive=True
                )
                
                # Update the Ideogram seed input maximum
                ideogram_seed = gr.Number(
                    label="Seed (optional)",
                    value=None,
                    precision=0,
                    minimum=0,
                    maximum=9223372036854775807,  # Max value for int64 (2^63 - 1)
                    interactive=True,
                    info="Enter a number for reproducible generations. Leave empty for random seed."
                )
                
                # Checkbox to disable style reference
                ideogram_disable_style_reference = gr.Checkbox(
                    label="🚫 Disable Style Reference",
                    value=False,
                    interactive=True,
                    info="Check to disable using uploaded reference images as style reference for Ideogram V3 (prompt-only generation)"
                )
                
                # Rendering speed (Ideogram V3 only)
                ideogram_rendering_speed = gr.Dropdown(
                    label="⚡ Rendering Speed (V3 Only)",
                    choices=["DEFAULT", "TURBO", "QUALITY"],
                    value="DEFAULT",
                    interactive=True,
                    info="Only applicable for Ideogram V3: TURBO (fastest), DEFAULT (balanced), QUALITY (best quality)",
                    visible=False  # Initially hidden, will be shown when V3 is selected
                )
                
            # Imagen-4 settings
            with gr.Column(scale=2):
                gr.Markdown("### Imagen-4 Settings")
                
                # Warning message for Imagen-4 settings
                imagen4_warning = gr.Markdown(
                    "⚠️ These settings only apply when Imagen-4 is selected as the provider.",
                    visible=True
                )
                
                # Imagen-4 model toggle
                imagen4_model = gr.Radio(
                    choices=["google/imagen-4", "google/imagen-4-fast", "google/imagen-4-ultra"],
                    label="🤖 Model Version",
                    value="google/imagen-4",
                    info="Choose between regular Imagen-4, Fast version (faster generation), or Ultra version (Ultra: higher quality, max 1 image)"
                )
                
                # Imagen-4 number of images setting
                imagen4_num_images = gr.Dropdown(
                    choices=[1, 2, 3, 4],
                    label="🔢 Number of Images",
                    value=1,
                    info="Number of images to generate (Ultra version limited to 1)"
                )
                
                # Imagen-4 aspect ratio setting
                imagen4_aspect_ratio = gr.Dropdown(
                    choices=["1:1", "16:9", "9:16", "3:4", "4:3"], 
                    label="📐 Aspect Ratio", 
                    value="1:1",
                    info="Aspect ratio for generated images"
                )
                
                # Imagen-4 safety filter setting
                imagen4_safety_filter = gr.Dropdown(
                    choices=["block_low_and_above", "block_medium_and_above", "block_only_high"],
                    label="🛡️ Safety Filter",
                    value="block_only_high",
                    info="Safety filter level for generated images"
                )
        
        # Add a separate gallery for processed images
        with gr.Row():
            with gr.Column(scale=2):
                output_gallery = gr.Gallery(
                    label="Generated Images",
                    columns=[2],
                    rows=[3],  # Increased from 2 to 3 rows to display more images
                    object_fit="contain",
                    height=400,  # Increased height for better visibility
                    show_label=True,
                    elem_id="output_gallery",
                    preview=True,  # Add preview capability for better viewing
                    interactive=False  # Disable interactivity - use button for inpainting instead
                )
                
                # Add select button for inpainting
                with gr.Row():
                    select_for_inpainting_button = gr.Button(
                        "🎨 Select Image for Inpainting",
                        variant="secondary",
                        size="lg",
                        interactive=False
                    )

        
        gr.Markdown("### Download")
        
        # Status Section - Main status bar used for all status updates
        with gr.Row():
            download_zip = gr.File(
                label="Download Images as ZIP",
                interactive=False,
                file_count="single",
                type="filepath"
            )  
            status_text = gr.Textbox(
                label="Status",
                interactive=False,
                value="Ready for image generation. Upload a reference image to begin."
            )
        
        # Generated images state for inpainting button
        generated_images_state = gr.State([])
        
        # Function to update number of images dropdown based on model selection
        def update_imagen4_num_images(model):
            if model == "google/imagen-4-ultra":
                return gr.Dropdown(
                    choices=[1],
                    value=1,
                    label="🔢 Number of Images",
                    info="Ultra version limited to 1 image only"
                )
            else:
                return gr.Dropdown(
                    choices=[1, 2, 3, 4],
                    value=1,
                    label="🔢 Number of Images",
                    info="Number of images to generate (1-4 for regular version)"
                )
        
        # Toggle interactivity of provider-specific UI components
        def toggle_provider_settings(provider):
            # Update provider state
            warning_msg = f"Note: You are currently using {provider}. The settings for the other providers will be ignored during generation."
            leonardo_warning = "⚠️ These settings only apply when Leonardo is selected as the provider."
            ideogram_warning = "⚠️ These settings only apply when Ideogram is selected as the provider."
            imagen4_warning = "⚠️ These settings only apply when Imagen-4 is selected as the provider."
            
            return [
                provider,
                warning_msg,
                leonardo_warning,
                ideogram_warning,
                imagen4_warning
            ]
        
        # When Imagen-4 model changes, update number of images dropdown
        imagen4_model.change(
            fn=update_imagen4_num_images,
            inputs=[imagen4_model],
            outputs=[imagen4_num_images]
        )
        
        # When provider selection changes, toggle component interactivity
        provider_tabs.change(
            fn=toggle_provider_settings,
            inputs=[provider_tabs],
            outputs=[
                provider_state,
                status_text,  # Use status text to show warning
                leonardo_warning,
                ideogram_warning,
                imagen4_warning
            ]
        )
        
        # === INPAINTING HELPER FUNCTIONS ===
        # Function to load selected image into editor
        def load_image_for_inpainting(selected_image, generated_images):
            """Load the selected image into the image editor for mask creation"""
            try:
                if not selected_image or not generated_images:
                    return gr.update(value=None), "No image selected"
                
                # Extract index from selection (e.g., "Image 1: filename.png" -> 0)
                image_index = int(selected_image.split(":")[0].replace("Image ", "")) - 1
                
                if 0 <= image_index < len(generated_images):
                    image_path = generated_images[image_index]
                    logger.info(f"🎨 INPAINTING: Loading image for editing: {image_path}")
                    
                    # Return the image path and success message
                    return gr.update(value=image_path), f"Image loaded: {os.path.basename(image_path)}. Draw areas you want to modify."
                else:
                    return gr.update(value=None), "Invalid image selection"
                    
            except Exception as e:
                logger.error(f"❌ INPAINTING ERROR: Failed to load image: {str(e)}")
                return gr.update(value=None), f"Error loading image: {str(e)}"
        
        # Function to update inpainting download visibility
        def update_inpaint_download(gallery_images, status_text, zip_file):
            """Show/hide inpainting download based on results"""
            try:
                if zip_file and "Successfully generated" in status_text:
                    return gr.update(value=zip_file, visible=True)
                else:
                    return gr.update(value=None, visible=False)
            except:
                return gr.update(value=None, visible=False)
        
        # Function to update UI text based on mask mode
        def update_mask_mode_text(mask_mode):
            """Update UI text based on selected mask mode"""
            if "Inverted" in mask_mode:
                instructions = """
                **Instructions for Inpainting (Inverted Mode):**
                1. Select an image from your generated results
                2. Adjust the brush size slider to your preferred size (5-100 pixels)
                3. Use the brush tool to paint over areas you want to modify (painted areas will be BLACK/inpainted)
                4. Enter your inpainting prompt describing what you want in the painted areas
                5. Click "Generate Inpainting" to create new variations
                """
                mask_preview_update = gr.update(label="Mask Preview (Black=Inpaint, White=Keep)")
            else:
                instructions = """
                **Instructions for Inpainting:**
                1. Select an image from your generated results
                2. Adjust the brush size slider to your preferred size (5-100 pixels)
                3. Use the brush tool to paint over areas you want to modify (painted areas will be WHITE/inpainted)
                4. Enter your inpainting prompt describing what you want in the painted areas
                5. Click "Generate Inpainting" to create new variations
                """
                mask_preview_update = gr.update(label="Mask Preview (White=Inpaint, Black=Keep)")
            
            return instructions, mask_preview_update
        
        # Function to update brush size in ImageEditor
        def update_brush_size(brush_size):
            """Update the ImageEditor with new brush size"""
            logger.info(f"🖌️ BRUSH SIZE: Updating brush size to {brush_size}")
            
            # Update the ImageEditor with new brush configuration
            # Note: Due to Gradio limitations, brush size changes may require reloading the image
            return gr.update(
                brush=gr.Brush(default_size=int(brush_size), colors=["white"])
            )
        
        # Function to toggle inpainting provider-specific controls
        def toggle_inpaint_provider_controls(provider):
            """Toggle UI controls based on selected inpainting provider"""
            logger.info(f"🎯 INPAINTING PROVIDER: Switching to {provider}")
            
            if provider == "Ideogram":
                # For Ideogram: Show style reference, hide Leonardo controls, set inverted mask mode
                return [
                    gr.update(visible=True),  # Show Ideogram style reference group
                    gr.update(visible=False),  # Hide inpaint strength (Leonardo only)
                    gr.update(visible=False),  # Hide guidance scale (Leonardo only)
                    gr.update(value="White: Unbrushed, Black: Brushed (Inverted)")  # Set inverted mask mode for Ideogram
                ]
            else:
                # For Leonardo: Hide style reference, show Leonardo controls, set default mask mode
                return [
                    gr.update(visible=False),  # Hide Ideogram style reference group
                    gr.update(visible=True),   # Show inpaint strength (Leonardo only)
                    gr.update(visible=True),   # Show guidance scale (Leonardo only)
                    gr.update(value="Black: Unbrushed, White: Brushed (Default)")  # Set default mask mode for Leonardo
                ]
        
        # Function to handle gallery selection and enable inpainting button
        def enable_inpainting_button(generated_images, status_text):
            """Enable the inpainting button when images are available"""
            # Log debug information
            logger.info(f"🎨 INPAINTING BUTTON DEBUG: generated_images={type(generated_images)}, length={len(generated_images) if generated_images is not None else 'None'}")
            logger.info(f"🎨 INPAINTING BUTTON DEBUG: status_text='{status_text[:100] if status_text else None}'")
            
            # Check if we have actual images (handle both PIL objects and file paths)
            has_images = False
            if generated_images is not None:
                if isinstance(generated_images, list):
                    if len(generated_images) > 0:
                        # Debug: Log all items in the list
                        logger.info(f"🎨 INPAINTING BUTTON DEBUG: List contents: {generated_images}")
                        
                        # Check if first item is a valid path or PIL Image
                        first_item = generated_images[0]
                        logger.info(f"🎨 INPAINTING BUTTON DEBUG: First item type: {type(first_item)}, value: '{first_item}'")
                        
                        # Handle PIL Image objects
                        from PIL import Image
                        if isinstance(first_item, Image.Image):
                            has_images = True
                            logger.info(f"🎨 INPAINTING BUTTON: ✅ Found {len(generated_images)} PIL Image objects")
                        
                        # Handle string file paths
                        elif isinstance(first_item, str) and len(first_item.strip()) > 0:
                            # Additional check to see if the file actually exists
                            import os
                            first_item_clean = first_item.strip()
                            logger.info(f"🎨 INPAINTING BUTTON DEBUG: Checking existence of: '{first_item_clean}'")
                            
                            if os.path.exists(first_item_clean):
                                has_images = True
                                logger.info(f"🎨 INPAINTING BUTTON: ✅ Found {len(generated_images)} valid file paths, first: {first_item_clean[:50]}...")
                            else:
                                logger.warning(f"🎨 INPAINTING BUTTON: ❌ Image path doesn't exist: '{first_item_clean}'")
                                logger.info(f"🎨 INPAINTING BUTTON DEBUG: Current working directory: {os.getcwd()}")
                                logger.info(f"🎨 INPAINTING BUTTON DEBUG: Absolute path would be: {os.path.abspath(first_item_clean)}")
                                
                                # FALLBACK: Enable anyway if we have valid string paths (for debugging)
                                logger.info(f"🎨 INPAINTING BUTTON: 🔄 FALLBACK - Enabling despite file check failure")
                                has_images = True
                        else:
                            logger.warning(f"🎨 INPAINTING BUTTON: ❌ First item is neither PIL Image nor valid string: {type(first_item)} - '{first_item}'")
                    else:
                        logger.info(f"🎨 INPAINTING BUTTON: Empty images list")
                else:
                    logger.warning(f"🎨 INPAINTING BUTTON: generated_images is not a list: {type(generated_images)}")
            else:
                logger.warning(f"🎨 INPAINTING BUTTON: generated_images is None")
            
            # More lenient status check - look for "generated" or success indicators
            status_ok = False
            if status_text:
                status_lower = status_text.lower()
                success_keywords = ["generated", "successfully", "complete", "finished", "done", "success", "batch processing complete"]
                # More specific error keywords to avoid false positives
                error_keywords = ["error:", "failed to", "critical error", "fatal error", "❌", "error generating", "error in"]
                
                has_success = any(word in status_lower for word in success_keywords)
                # Check for actual error indicators, not just the word "error" anywhere
                has_error = any(word in status_lower for word in error_keywords)
                
                # Special handling for batch processing complete
                if "batch processing complete" in status_lower:
                    status_ok = True
                    logger.info(f"🎨 INPAINTING BUTTON: Status indicates batch completion: '{status_text[:50]}...'")
                elif has_success and not has_error:
                    status_ok = True
                    logger.info(f"🎨 INPAINTING BUTTON: Status indicates success: '{status_text[:50]}...'")
                elif has_error:
                    logger.info(f"🎨 INPAINTING BUTTON: Status indicates error: '{status_text[:50]}...'")                    
                else:
                    # If we have success keywords but no clear error indicators, treat as success
                    if has_success:
                        status_ok = True
                        logger.info(f"🎨 INPAINTING BUTTON: Status indicates success (no error indicators): '{status_text[:50]}...'")
                    else:
                        logger.info(f"🎨 INPAINTING BUTTON: Status is neutral: '{status_text[:50]}...'")
            else:
                logger.warning(f"🎨 INPAINTING BUTTON: No status text provided")
            
            if has_images and status_ok:
                logger.info(f"🎨 INPAINTING BUTTON: ✅ ENABLING - {len(generated_images)} images available")
                return gr.update(interactive=True, value="🎨 Select Image for Inpainting")
            else:
                logger.info(f"🎨 INPAINTING BUTTON: ❌ DISABLED - has_images={has_images}, status_ok={status_ok}")
                return gr.update(interactive=False, value="🎨 Select Image for Inpainting")
        
        # Function to handle "Select for Inpainting" button click - connects to main inpainting tab
        def select_image_for_inpainting(generated_images):
            """Select an image from the generated images and populate the main inpainting tab"""
            try:
                if not generated_images or len(generated_images) == 0:
                    return [
                        gr.update(choices=[], value=None),  # inpaint_image_selector in main tab
                        [],  # generated_images_state
                        "No images available for inpainting. Generate some images first, then click this button.",  # inpaint_status in main tab
                        gr.update(value=None)  # inpaint_image_editor in main tab
                    ]
                
                # Handle PIL Image objects by converting them to temporary files
                from PIL import Image
                import tempfile
                
                processed_images = []
                choices = []
                
                for i, img in enumerate(generated_images):
                    if isinstance(img, Image.Image):
                        # Convert PIL Image to temporary file
                        temp_dir = tempfile.mkdtemp()
                        temp_file = os.path.join(temp_dir, f"temp_image_{i+1}.png")
                        img.save(temp_file, 'PNG')
                        processed_images.append(temp_file)
                        choices.append(f"Image {i+1}: temp_image_{i+1}.png")
                        logger.info(f"🎨 INPAINTING: Converted PIL Image {i+1} to temp file: {temp_file}")
                    elif isinstance(img, str):
                        # Already a file path
                        processed_images.append(img)
                        choices.append(f"Image {i+1}: {os.path.basename(img)}")
                    else:
                        logger.warning(f"🎨 INPAINTING: Unknown image type {type(img)} at index {i}")
                        continue
                
                if not processed_images:
                    return [
                        gr.update(choices=[], value=None),  # inpaint_image_selector in main tab
                        [],  # generated_images_state
                        "No valid images found for inpainting",  # inpaint_status in main tab
                        gr.update(value=None)  # inpaint_image_editor in main tab
                    ]
                
                logger.info(f"🎨 INPAINTING: Setting up main inpainting tab with {len(processed_images)} images")
                
                # Load the first image into the editor automatically
                first_image_path = processed_images[0]
                logger.info(f"🎨 INPAINTING: Auto-loading first image into main tab editor: {first_image_path}")
                
                return [
                    gr.update(choices=choices, value=choices[0] if choices else None),  # inpaint_image_selector in main tab
                    processed_images,  # Store the file paths for inpainting
                    f"✅ Images loaded in main inpainting tab! Go to 'Image Inpainting' tab to draw mask and generate. First image auto-loaded: {os.path.basename(first_image_path)}",  # inpaint_status in main tab
                    gr.update(value=first_image_path)  # inpaint_image_editor in main tab - load first image automatically
                ]
                    
            except Exception as e:
                logger.error(f"❌ INPAINTING ERROR: Failed to setup main inpainting tab: {str(e)}")
                import traceback
                logger.error(traceback.format_exc())
                return [
                    gr.update(choices=[], value=None),  # inpaint_image_selector in main tab
                    [],  # generated_images_state
                    f"Error setting up inpainting: {str(e)}",  # inpaint_status in main tab
                    gr.update(value=None)  # inpaint_image_editor in main tab
                ]

        # Function to perform inpainting (for generated images)
        async def perform_inpainting(selected_image, editor_data, prompt, mask_mode, strength, guidance, num_images, generated_images, provider, style_reference=None):
            """Perform inpainting using selected provider (Leonardo or Ideogram)"""
            try:
                if not selected_image or not editor_data or not prompt.strip():
                    return [], "Please select an image, create a mask, and enter a prompt", None
                
                # Extract image index and get the original image path
                image_index = int(selected_image.split(":")[0].replace("Image ", "")) - 1
                if not (0 <= image_index < len(generated_images)):
                    return [], "Invalid image selection", None
                
                init_image_path = generated_images[image_index]
                logger.info(f"🎨 INPAINTING: Starting inpainting for {init_image_path} using {provider}")
                
                # Convert mask mode string to parameter
                mask_mode_param = "inverted" if "Inverted" in mask_mode else "default"
                logger.info(f"🎨 INPAINTING: Using mask mode: {mask_mode_param}")
                
                # Create mask from drawing data
                mask_path = create_mask_from_drawing(editor_data, mask_mode_param)
                
                if provider == "Leonardo":
                    # Leonardo inpainting workflow
                    # Upload both images to Leonardo
                    init_image_id, mask_image_id = await upload_canvas_images_to_leonardo(init_image_path, mask_path)
                
                    # Generate inpainted images
                    result_data = await generate_inpainting_with_leonardo(
                        prompt=prompt,
                        init_image_id=init_image_id,
                        mask_image_id=mask_image_id,
                        num_images=int(num_images),
                        init_strength=strength,
                        guidance_scale=int(guidance)
                    )
                    
                    # Process Leonardo results
                    if 'generations_by_pk' in result_data and result_data['generations_by_pk']:
                        generation = result_data['generations_by_pk']
                        
                        if 'generated_images' in generation and generation['generated_images']:
                            inpainted_images = []
                            
                            for img_data in generation['generated_images']:
                                if 'url' in img_data:
                                    # Download the image
                                    img_url = img_data['url']
                                    logger.info(f"🎨 LEONARDO INPAINTING: Downloading result image: {img_url}")
                                    
                                    response = requests.get(img_url)
                                    if response.ok:
                                        # Save to temporary file
                                        import tempfile
                                        temp_dir = tempfile.mkdtemp()
                                        temp_file = os.path.join(temp_dir, f"leonardo_inpainted_{len(inpainted_images)+1}.png")
                                        
                                        with open(temp_file, 'wb') as f:
                                            f.write(response.content)
                                        
                                        inpainted_images.append(temp_file)
                            
                            if inpainted_images:
                                logger.info(f"✅ LEONARDO INPAINTING: Successfully generated {len(inpainted_images)} inpainted images")
                                
                                # Create ZIP file for download
                                zip_path = create_inpainting_zip(inpainted_images, prompt)
                                
                                return inpainted_images, f"Successfully generated {len(inpainted_images)} inpainted images with Leonardo!", zip_path
                            else:
                                return [], "No images were generated by Leonardo", None
                        else:
                            return [], "No images found in Leonardo generation results", None
                    else:
                        return [], "Invalid Leonardo generation response format", None
                        
                elif provider == "Ideogram":
                    # Ideogram inpainting workflow
                    # Generate inpainted images using Ideogram V3 Edit
                    result_data = await generate_inpainting_with_ideogram(
                        prompt=prompt,
                        init_image_path=init_image_path,
                        mask_image_path=mask_path,
                        style_reference_path=style_reference,
                        num_images=int(num_images)
                    )
                    
                    if result_data and 'data' in result_data:
                        inpainted_images = []
                        
                        for img_data in result_data['data']:
                            if 'url' in img_data:
                                # Download the image
                                img_url = img_data['url']
                                logger.info(f"🎨 IDEOGRAM INPAINTING: Downloading result image: {img_url}")
                                
                                response = requests.get(img_url)
                                if response.ok:
                                    # Save to temporary file
                                    import tempfile
                                    temp_dir = tempfile.mkdtemp()
                                    temp_file = os.path.join(temp_dir, f"ideogram_inpainted_{len(inpainted_images)+1}.png")
                                    
                                    with open(temp_file, 'wb') as f:
                                        f.write(response.content)
                                    
                                    inpainted_images.append(temp_file)
                        
                        if inpainted_images:
                            logger.info(f"✅ IDEOGRAM INPAINTING: Successfully generated {len(inpainted_images)} inpainted images")
                            
                            # Create ZIP file for download
                            zip_path = create_inpainting_zip(inpainted_images, prompt)
                            
                            return inpainted_images, f"Successfully generated {len(inpainted_images)} inpainted images with Ideogram!", zip_path
                        else:
                            return [], "No images were generated by Ideogram", None
                    else:
                        return [], "Invalid Ideogram generation response", None
                        
                else:
                    return [], f"Unsupported provider: {provider}", None
                
            except Exception as e:
                error_msg = f"Inpainting failed: {str(e)}"
                logger.error(f"❌ INPAINTING ERROR: {error_msg}")
                return [], error_msg, None
        
        # === TAB 2: IMAGE INPAINTING ===
        with gr.Tab("🖌️ Image Inpainting"):
            gr.Markdown("### 🎨 Image Inpainting")
            gr.Markdown("Modify specific areas of your generated images or upload any image to inpaint.")
            
            # Combined inpainting instructions
            inpainting_instructions = gr.Markdown("""
            **Instructions for Image Inpainting:**
            1. **For Generated Images:** Select an image from your generated results
            2. **For Any Image:** Upload any image you want to modify
            3. Adjust the brush size slider to your preferred size (5-100 pixels)
            4. Use the brush tool to paint over areas you want to modify (painted areas will be WHITE/inpainted)
            5. Enter your inpainting prompt describing what you want in the painted areas
            6. Click "Generate Inpainting" to create new variations
            """)
            
            with gr.Row():
                with gr.Column(scale=1):
                    # Provider selection for inpainting
                    inpaint_provider = gr.Radio(
                        label="🎯 Inpainting Provider",
                        choices=["Leonardo", "Ideogram"],
                        value="Leonardo",
                        interactive=True,
                        info="Choose which AI provider to use for inpainting"
                    )
                    
                    # Image selection dropdown (for generated images)
                    inpaint_image_selector = gr.Dropdown(
                        label="Select Generated Image to Inpaint",
                        choices=[],
                        value=None,
                        interactive=True,
                        info="Choose from your generated images"
                    )
                    
                    # Alternative: Upload any image
                    standalone_image_upload = gr.File(
                        label="OR Upload Any Image to Inpaint",
                        file_types=["image", ".jpg", ".jpeg", ".png", ".avif", ".webp"],
                        interactive=True,
                        type="filepath"
                    )
                    
                    # Inpainting prompt
                    inpaint_prompt = gr.Textbox(
                        label="Inpainting Prompt",
                        lines=3,
                        placeholder="Describe what you want to add/change in the painted areas...",
                        interactive=True,
                        info="Describe what should appear in the areas you painted"
                    )
                    
                    # Style reference image upload for Ideogram inpainting
                    with gr.Group(visible=False) as ideogram_inpaint_group:
                        inpaint_style_reference = gr.File(
                            label="🎨 Style Reference Image (Ideogram Only) - Optional upload for Ideogram inpainting",
                            file_types=["image"],
                            type="filepath",
                            interactive=True
                        )
                    
                    # Mask mode selection
                    mask_mode = gr.Radio(
                        label="🎭 Mask Mode",
                        choices=[
                            "Black: Unbrushed, White: Brushed (Default)",
                            "White: Unbrushed, Black: Brushed (Inverted)"
                        ],
                        value="Black: Unbrushed, White: Brushed (Default)",
                        interactive=True,
                        info="Auto-switches based on provider: Leonardo uses Default, Ideogram uses Inverted"
                    )
                    
                    # Inpainting controls (Leonardo specific)
                    with gr.Row():
                        inpaint_strength = gr.Slider(
                            label="🎯 Inpaint Strength (Leonardo Only)",
                            minimum=0.1,
                            maximum=0.9,
                            value=0.3,
                            step=0.01,
                            interactive=True,
                            info="Leonardo Only: Lower = more like original, Higher = more creative"
                        )
                        
                        inpaint_guidance = gr.Slider(
                            label="🎯 Guidance Scale (Leonardo Only)",
                            minimum=1,
                            maximum=20,
                            value=7,
                            step=1,
                            interactive=True,
                            info="Leonardo Only: How closely to follow the prompt"
                        )
                    
                    with gr.Row():
                        inpaint_num_images = gr.Slider(
                            label="Number of Images",
                            minimum=1,
                            maximum=8,
                            value=4,
                            step=1,
                            interactive=True
                        )
                        
                        brush_size_slider = gr.Slider(
                            label="🖌️ Brush Size",
                            minimum=5,
                            maximum=100,
                            value=20,
                            step=1,
                            interactive=True,
                            info="Adjust brush size for painting the mask"
                        )
                        
                        # Inpainting button
                        inpaint_button = gr.Button(
                            "🎨 Generate Inpainting",
                            variant="primary",
                            size="lg",
                            interactive=True
                        )
                
                with gr.Column(scale=2):
                    # Image editor for creating masks
                    inpaint_image_editor = gr.ImageEditor(
                        label="Draw Mask (Paint areas to modify)",
                        type="pil",
                        interactive=True,
                        brush=gr.Brush(default_size=20, colors=["white"]),
                        height=400
                    )
                    
                    # Preview of the mask
                    inpaint_mask_preview = gr.Image(
                        label="Mask Preview",
                        interactive=False,
                        height=200,
                        visible=False
                    )
            
            # Inpainting results
            with gr.Row():
                inpaint_gallery = gr.Gallery(
                    label="Inpainting Results",
                    columns=[2],
                    rows=[2],
                    object_fit="contain",
                    height=400,
                    show_label=True,
                    preview=True,
                    interactive=False
                )
            
            # Inpainting download and status
            with gr.Row():
                with gr.Column(scale=1):
                    inpaint_download_zip = gr.File(
                        label="Download Inpainted Images as ZIP",
                        interactive=False,
                        file_count="single",
                        type="filepath",
                        visible=False
                    )
                with gr.Column(scale=2):
                    inpaint_status = gr.Textbox(
                        label="Inpainting Status",
                        interactive=False,
                        value="Select an image and create a mask to start inpainting"
                    )
            
            # Event handlers for inpainting
            

            
            # Handle inpainting provider change
            inpaint_provider.change(
                fn=toggle_inpaint_provider_controls,
                inputs=[inpaint_provider],
                outputs=[
                    ideogram_inpaint_group,
                    inpaint_strength,
                    inpaint_guidance,
                    mask_mode
                ]
            )
            
            # Handle mask mode change
            mask_mode.change(
                fn=update_mask_mode_text,
                inputs=[mask_mode],
                outputs=[inpainting_instructions, inpaint_mask_preview]
            )
            
            # Handle brush size change
            brush_size_slider.change(
                fn=update_brush_size,
                inputs=[brush_size_slider],
                outputs=[inpaint_image_editor]
            )
            
                        # Handle uploaded image (alternative to selecting from generated images)
            def handle_uploaded_image_for_inpainting(uploaded_file):
                """Handle when user uploads an image directly for inpainting"""
                if uploaded_file is None:
                    return None, "No image uploaded"
                
                try:
                    # Clear the dropdown selection when user uploads an image
                    return uploaded_file, "Image uploaded successfully. You can now create a mask and inpaint."
                except Exception as e:
                    return None, f"Error loading uploaded image: {str(e)}"
            
            # When user uploads an image, load it in the editor and update status
            standalone_image_upload.change(
                fn=handle_uploaded_image_for_inpainting,
                inputs=[standalone_image_upload],
                outputs=[inpaint_image_editor, inpaint_status]
            )
            
            # Create a combined inpainting function that handles both generated and uploaded images
            def perform_combined_inpainting(selector_value, uploaded_image, editor_data, prompt, mask_mode, strength, guidance, num_images, generated_images, provider, style_reference):
                """Handle inpainting for both generated and uploaded images"""
                import asyncio
                
                # Determine which image to use: uploaded or selected from generated
                if uploaded_image is not None:
                    # Use uploaded image
                    return asyncio.run(perform_standalone_inpainting(
                        uploaded_image, editor_data, prompt, mask_mode, strength, guidance, num_images, provider, style_reference
                    ))
                elif selector_value is not None:
                    # Use selected generated image
                    return asyncio.run(perform_inpainting(
                        selector_value, editor_data, prompt, mask_mode, strength, guidance, num_images, generated_images, provider, style_reference
                    ))
                else:
                    return [], "Please either select a generated image or upload an image to inpaint", None
            
            # Update the inpainting button to use the combined function
            inpaint_button.click(
                fn=perform_combined_inpainting,
                inputs=[
                    inpaint_image_selector,
                    standalone_image_upload,
                    inpaint_image_editor,
                    inpaint_prompt,
                    mask_mode,
                    inpaint_strength,
                    inpaint_guidance,
                    inpaint_num_images,
                    generated_images_state,
                    inpaint_provider,
                    inpaint_style_reference
                ],
                outputs=[inpaint_gallery, inpaint_status, inpaint_download_zip]
            ).then(
                fn=update_inpaint_download,
                inputs=[inpaint_gallery, inpaint_status, inpaint_download_zip],
                outputs=[inpaint_download_zip]
            )
            
            # Load image when selection changes
            inpaint_image_selector.change(
                fn=load_image_for_inpainting,
                inputs=[inpaint_image_selector, generated_images_state],
                outputs=[inpaint_image_editor, inpaint_status]
            )
            
            # Update instructions and mask preview label when mask mode changes
            mask_mode.change(
                fn=update_mask_mode_text,
                inputs=[mask_mode],
                outputs=[inpainting_instructions, inpaint_mask_preview]
            )
            
            # Update brush size when slider changes
            brush_size_slider.change(
                fn=update_brush_size,
                inputs=[brush_size_slider],
                outputs=[inpaint_image_editor]
            )
            
            # Update UI when inpainting provider changes
            inpaint_provider.change(
                fn=toggle_inpaint_provider_controls,
                inputs=[inpaint_provider],
                outputs=[
                    ideogram_inpaint_group,  # Show/hide Ideogram style reference
                    inpaint_strength,        # Show/hide inpaint strength (Leonardo only)
                    inpaint_guidance,        # Show/hide guidance scale (Leonardo only)
                    mask_mode               # Auto-set mask mode based on provider
                ]
            )
        
        # Event handlers
        
        # Helper function to get the selected category based on theme
        def get_selected_category(theme, category, *args):
            # Directly use the selected category from the dropdown
            return category
        
        # Function to update category dropdown based on selected theme
        def update_category_dropdown(theme):
            categories = get_categories_for_theme(theme)
            return gr.Dropdown(choices=categories, value=categories[0] if categories else None)
        
        # Function to handle manual variation generation
        def handle_manual_variation_generation(variation_request, generated_prompt_state, extracted_images_state, current_image_index, all_prompts_state, all_modified_prompts_state, all_activities_state, all_expressions_state, all_fur_colors_state):
            """
            Handle manual variation generation based on user request
            """
            try:
                if not variation_request or not variation_request.strip():
                    return ("Please enter a variation request", 
                            all_prompts_state, all_modified_prompts_state, all_activities_state, all_expressions_state, all_fur_colors_state,
                            "Image 0/0", False, extracted_images_state, current_image_index)
                
                if not generated_prompt_state or not generated_prompt_state.strip():
                    return ("No base prompt available. Please upload an image first.", 
                            all_prompts_state, all_modified_prompts_state, all_activities_state, all_expressions_state, all_fur_colors_state,
                            "Image 0/0", False, extracted_images_state, current_image_index)
                
                logger.info(f"🎯 Processing manual variation request: {variation_request}")
                
                # Extract number of variations from request (default to 1)
                import re
                num_match = re.search(r'(\d+)', variation_request)
                num_variations = int(num_match.group(1)) if num_match else 1
                
                # Check if user requested more than 5 variations
                if num_variations > 5:
                    return ("❌ Error: Maximum number of variations allowed is 5. Please reduce your request.", 
                            all_prompts_state, all_modified_prompts_state, all_activities_state, all_expressions_state, all_fur_colors_state,
                            "Image 0/0", False, extracted_images_state, current_image_index)
                
                # Ensure minimum of 1 variation
                num_variations = max(num_variations, 1)
                
                logger.info(f"📊 Generating {num_variations} variations")
                
                # Generate variations using Qwen
                variations = generate_manual_variations(generated_prompt_state, variation_request, num_variations)
                
                if not variations:
                    return ("Failed to generate variations. Please try again.", 
                            all_prompts_state, all_modified_prompts_state, all_activities_state, all_expressions_state, all_fur_colors_state,
                            "Image 0/0", False, extracted_images_state, current_image_index)
                
                # Update states with variations
                new_all_prompts = variations.copy()
                new_all_modified_prompts = variations.copy()  # For variations, use the same as prompts since no activity/expression
                new_all_activities = ["" for _ in variations]  # Empty activities for variations
                new_all_expressions = ["" for _ in variations]  # Empty expressions for variations
                new_all_fur_colors = ["" for _ in variations]  # Empty fur colors for variations
                
                # For variations, we don't have actual images, so create placeholder paths
                new_extracted_images = [f"variation_{i+1}" for i in range(len(variations))]
                
                # Set current index to 0 and show navigation
                new_current_index = 0
                new_image_counter = f"Variation {new_current_index + 1}/{len(variations)}"
                nav_visible = len(variations) > 1
                
                success_message = f"✅ Successfully generated {len(variations)} variations based on your request"
                logger.info(f"🎉 {success_message}")
                
                return (success_message,
                        new_all_prompts, new_all_modified_prompts, new_all_activities, new_all_expressions, new_all_fur_colors,
                        new_image_counter, nav_visible, new_extracted_images, new_current_index)
                
            except Exception as e:
                error_msg = f"❌ Error generating variations: {str(e)}"
                logger.error(error_msg)
                return (error_msg,
                        all_prompts_state, all_modified_prompts_state, all_activities_state, all_expressions_state, all_fur_colors_state,
                        "Image 0/0", False, extracted_images_state, current_image_index)
            
        # Event handlers for dropdowns
        def update_category_and_subcategory(theme):
            categories = get_categories_for_theme(theme)
            first_category = categories[0] if categories else None
            subcategories = get_subcategories(theme, first_category)
            return (
                gr.Dropdown(choices=categories, value=first_category, interactive=True),
                gr.Dropdown(choices=subcategories, value=subcategories[0] if subcategories else None, interactive=True)
            )

        theme_dropdown.change(
            fn=update_category_and_subcategory,
            inputs=[theme_dropdown],
            outputs=[category_dropdown, subcategory_dropdown]
        )

        category_dropdown.change(
            fn=update_subcategory_dropdown,
            inputs=[theme_dropdown, category_dropdown],
            outputs=[subcategory_dropdown]
        )
        
        # When Ideogram model changes, toggle rendering speed visibility
        ideogram_model.change(
            fn=toggle_ideogram_rendering_speed,
            inputs=[ideogram_model],
            outputs=[ideogram_rendering_speed]
        )
        
        # Function to extract images from ZIP and prepare for sequential processing
        def process_uploaded_file(file_path, disable_activity_expression_fur=False):
            if file_path is None:
                return None, "No file uploaded.", "", [], 0, [], False, None, "Image 0/0", None
            
            try:
                logger.info(f"Processing uploaded file: {file_path}")
                
                # Handle different file formats:
                # 1. String path to file
                # 2. Dictionary with 'name' key (Gradio upload)
                # 3. File object with 'name' attribute
                # 4. ZIP file
                
                actual_file_path = None
                
                # Check if it's a dictionary (Gradio file upload format)
                if isinstance(file_path, dict) and 'name' in file_path:
                    actual_file_path = file_path['name']
                    logger.info(f"Using path from Gradio file upload dict: {actual_file_path}")
                # Check if it's a file-like object with a name attribute
                elif hasattr(file_path, 'name'):
                    actual_file_path = file_path.name
                    logger.info(f"Using path from file object: {actual_file_path}")
                # Otherwise assume it's a string path
                elif isinstance(file_path, str):
                    actual_file_path = file_path
                    logger.info(f"Using direct string path: {actual_file_path}")
                else:
                    logger.warning(f"Unsupported file format: {type(file_path)}")
                    return None, f"Unsupported file format: {type(file_path)}", "", [], 0, [], False, None, "Image 0/0", None
                
                # Verify the file exists
                if not os.path.exists(actual_file_path):
                    logger.warning(f"File does not exist: {actual_file_path}")
                    return None, f"File does not exist: {actual_file_path}", "", [], 0, [], False, None, "Image 0/0", None
                
                # Handle AVIF conversion for single files (before ZIP check)
                if actual_file_path.lower().endswith('.avif'):
                    logger.info(f"Converting AVIF file: {actual_file_path}")
                    try:
                        # Convert AVIF to PNG
                        png_path = actual_file_path.rsplit('.', 1)[0] + '.png'
                        converted_path = convert_avif(actual_file_path, png_path, 'PNG')
                        
                        if converted_path != actual_file_path:
                            # Conversion successful, use the converted file
                            actual_file_path = converted_path
                            logger.info(f"Successfully converted AVIF to PNG: {actual_file_path}")
                        else:
                            # Conversion failed
                            logger.error(f"Failed to convert AVIF file: {actual_file_path}")
                            return None, "Failed to convert AVIF file. Please try a different format.", "", [], 0, [], False, None, "Image 0/0", None
                    except Exception as e:
                        logger.error(f"Error converting AVIF file: {str(e)}")
                        return None, f"Error converting AVIF file: {str(e)}", "", [], 0, [], False, None, "Image 0/0", None
                
                # Check if it's a ZIP file
                if actual_file_path.lower().endswith('.zip'):
                    try:
                        # Extract images from ZIP
                        extracted_images, temp_dir = extract_images_from_zip(actual_file_path)
                        if not extracted_images:
                            logger.warning(f"No valid images found in ZIP file: {actual_file_path}")
                            return None, f"No valid images found in ZIP file: {actual_file_path}", "", [], 0, [], False, None, "Image 0/0", None
                        
                        # Convert AVIF images to PNG
                        converted_images = []
                        for img_path in extracted_images:
                            if img_path.lower().endswith('.avif'):
                                logger.info(f"Converting AVIF file in ZIP: {img_path}")
                                try:
                                    # Convert AVIF to PNG
                                    png_path = img_path.rsplit('.', 1)[0] + '.png'
                                    converted_path = convert_avif(img_path, png_path, 'PNG')
                                    
                                    if converted_path != img_path:
                                        # Conversion successful, use the converted file
                                        converted_images.append(converted_path)
                                        logger.info(f"Successfully converted AVIF to PNG: {converted_path}")
                                    else:
                                        # Conversion failed
                                        logger.error(f"Failed to convert AVIF file in ZIP: {img_path}")
                                        return None, f"Failed to convert AVIF file in ZIP: {img_path}", "", [], 0, [], False, None, "Image 0/0", None
                                except Exception as e:
                                    logger.error(f"Error converting AVIF file in ZIP: {str(e)}")
                                    return None, f"Error converting AVIF file in ZIP: {str(e)}", "", [], 0, [], False, None, "Image 0/0", None
                            else:
                                converted_images.append(img_path)
                        
                        # Update the extracted images with the converted ones
                        extracted_images = converted_images
                        
                        # Generate prompts and enhanced values for each image
                        all_prompts = []
                        all_modified_prompts = []
                        all_activities = []
                        all_expressions = []
                        all_fur_colors = []
                        
                        for img_path in extracted_images:
                            prompt = generate_prompt_from_image(img_path)
                            all_prompts.append(prompt)
                            
                            # Detect if the image contains humans
                            contains_human = detect_human_in_image(img_path)
                            
                            # Auto-generate activity, expression, and fur color for each image (unless disabled)
                            if disable_activity_expression_fur:
                                # Skip activity/expression/fur generation
                                activity = ""
                                facial_expression = ""
                                fur_color = ""
                                modified_prompt = prompt  # Use base prompt without enhancements
                                logger.info(f"🚫 Skipping activity/expression/fur generation for image {len(all_prompts)} (disabled)")
                            else:
                                activity, facial_expression = generate_activity_expression_from_prompt(prompt)
                                fur_color = generate_fur_color_for_prompt(prompt)
                                modified_prompt = enhance_prompt_with_activity_expression(prompt, activity, facial_expression, fur_color, "Auto", contains_human)
                            
                            all_modified_prompts.append(modified_prompt)
                            all_activities.append(activity)
                            all_expressions.append(facial_expression)
                            all_fur_colors.append(fur_color)
                        
                        # Set the current image index to 0
                        current_index = 0
                        
                        # Show navigation controls
                        nav_controls_visible = True
                        
                        # Update the reference preview with the first image
                        reference_preview = extracted_images[current_index]
                        
                        # Update the image counter
                        image_counter = f"Image {current_index+1}/{len(extracted_images)}"
                        
                        # Return the results
                        return (
                            reference_preview,
                            "Ready for image generation. Upload a reference image to begin.",
                            all_prompts[current_index],
                            extracted_images,
                            current_index,
                            all_prompts,
                            all_modified_prompts,
                            all_activities,
                            all_expressions,
                            all_fur_colors,
                            nav_controls_visible,
                            reference_preview,
                            image_counter,
                            actual_file_path if actual_file_path.lower().endswith('.zip') else None  # Original ZIP path
                        )
                    except Exception as e:
                        error_msg = f"Error processing ZIP file: {str(e)}"
                        logger.error(error_msg)
                        import traceback
                        logger.error(traceback.format_exc())
                        return (
                            None,  # reference_preview
                            error_msg,  # status_text
                            "",  # generated_prompt_state
                            [],  # extracted_images_state
                            0,  # current_image_index
                            [],  # all_prompts_state
                            [],  # all_modified_prompts_state
                            [],  # all_activities_state
                            [],  # all_expressions_state
                            [],  # all_fur_colors_state
                            False,  # nav_controls_visible
                            None,  # reference_preview (again)
                            "Image 0/0",  # image_counter
                            None  # original_zip_file_path
                        )
                else:
                    try:
                        # Single image processing
                        prompt = generate_prompt_from_image(actual_file_path)
                        
                        # Detect if the image contains humans
                        contains_human = detect_human_in_image(actual_file_path)
                        
                        # Auto-generate activity, expression, and fur color for single image (unless disabled)
                        if disable_activity_expression_fur:
                            # Skip activity/expression/fur generation
                            activity = ""
                            facial_expression = ""
                            fur_color = ""
                            modified_prompt = prompt  # Use base prompt without enhancements
                            logger.info(f"🚫 Skipping activity/expression/fur generation for single image (disabled)")
                        else:
                            activity, facial_expression = generate_activity_expression_from_prompt(prompt)
                            fur_color = generate_fur_color_for_prompt(prompt)
                            modified_prompt = enhance_prompt_with_activity_expression(prompt, activity, facial_expression, fur_color, "Auto", contains_human)
                        
                        return (
                            actual_file_path,
                            "Ready for image generation. Upload a reference image to begin.",
                            prompt,
                            [actual_file_path],
                            0,
                            [prompt],
                            [modified_prompt],
                            [activity],
                            [facial_expression],
                            [fur_color],
                            False,
                            actual_file_path,
                            "Image 1/1",
                            None  # original_zip_file_path (None for single images)
                        )
                    except Exception as e:
                        error_msg = f"Error processing image file: {str(e)}"
                        logger.error(error_msg)
                        import traceback
                        logger.error(traceback.format_exc())
                        return (
                            None,  # reference_preview
                            error_msg,  # status_text
                            "",  # generated_prompt_state
                            [],  # extracted_images_state
                            0,  # current_image_index
                            [],  # all_prompts_state
                            [],  # all_modified_prompts_state
                            [],  # all_activities_state
                            [],  # all_expressions_state
                            [],  # all_fur_colors_state
                            False,  # nav_controls_visible
                            None,  # reference_preview (again)
                            "Image 0/0",  # image_counter
                            None  # original_zip_file_path
                        )
            except Exception as e:
                error_msg = f"Error processing file: {str(e)}"
                logger.error(error_msg)
                import traceback
                logger.error(traceback.format_exc())
                return (
                    None,  # reference_preview
                    error_msg,  # status_text
                    "",  # generated_prompt_state
                    [],  # extracted_images_state
                    0,  # current_image_index
                    [],  # all_prompts_state
                    [],  # all_modified_prompts_state
                    [],  # all_activities_state
                    [],  # all_expressions_state
                    [],  # all_fur_colors_state
                    False,  # nav_controls_visible
                    None,  # reference_preview (again)
                    "Image 0/0"  # image_counter
                )
        
        # Function to show the previous image in the sequence
        def show_previous_image(current_index, extracted_images, all_prompts, all_modified_prompts, all_activities, all_expressions, all_fur_colors):
            if not extracted_images or len(extracted_images) <= 1:
                # For single image, don't allow navigation - return current state unchanged
                return current_index, gr.update(), gr.update(), "", "", "", "", "Image 0/0", False, False
            
            # Calculate new index (with wraparound)
            new_index = (current_index - 1) % len(extracted_images)
            
            # Check if this is variation mode (placeholder paths start with "variation_")
            is_variation_mode = extracted_images[new_index].startswith("variation_")
            
            # Get the image at the new index (or None for variations)
            image_path = None if is_variation_mode else extracted_images[new_index]
            
            # Get the stored values for this index (if available) - BUT NOT the generated prompt
            # The generated prompt should remain unchanged (as it's the default from reference image)
            modified_prompt = all_modified_prompts[new_index] if new_index < len(all_modified_prompts) else ""
            activity = all_activities[new_index] if new_index < len(all_activities) else ""
            expression = all_expressions[new_index] if new_index < len(all_expressions) else ""
            fur_color = all_fur_colors[new_index] if new_index < len(all_fur_colors) else ""
            
            # Update counter text based on mode
            if is_variation_mode:
                counter_text = f"Variation {new_index+1}/{len(extracted_images)}"
            else:
                counter_text = f"Image {new_index+1}/{len(extracted_images)}"
            
            # Reset manual edit flag when navigating to different item
            # IMPORTANT: Do NOT return current_prompt - use gr.update() to keep generated prompt unchanged
            return new_index, image_path, gr.update(), modified_prompt, activity, expression, fur_color, counter_text, True, False
        
        # Function to show the next image in the sequence
        def show_next_image(current_index, extracted_images, all_prompts, all_modified_prompts, all_activities, all_expressions, all_fur_colors):
            if not extracted_images or len(extracted_images) <= 1:
                # For single image, don't allow navigation - return current state unchanged
                return current_index, gr.update(), gr.update(), "", "", "", "", "Image 0/0", False, False
            
            # Calculate new index (with wraparound)
            new_index = (current_index + 1) % len(extracted_images)
            
            # Check if this is variation mode (placeholder paths start with "variation_")
            is_variation_mode = extracted_images[new_index].startswith("variation_")
            
            # Get the image at the new index (or None for variations)
            image_path = None if is_variation_mode else extracted_images[new_index]
            
            # Get the stored values for this index (if available) - BUT NOT the generated prompt
            # The generated prompt should remain unchanged (as it's the default from reference image)
            modified_prompt = all_modified_prompts[new_index] if new_index < len(all_modified_prompts) else ""
            activity = all_activities[new_index] if new_index < len(all_activities) else ""
            expression = all_expressions[new_index] if new_index < len(all_expressions) else ""
            fur_color = all_fur_colors[new_index] if new_index < len(all_fur_colors) else ""
            
            # Update counter text based on mode
            if is_variation_mode:
                counter_text = f"Variation {new_index+1}/{len(extracted_images)}"
            else:
                counter_text = f"Image {new_index+1}/{len(extracted_images)}"
            
            # Reset manual edit flag when navigating to different item
            # IMPORTANT: Do NOT return current_prompt - use gr.update() to keep generated prompt unchanged
            return new_index, image_path, gr.update(), modified_prompt, activity, expression, fur_color, counter_text, True, False
        
        # Function to update stored prompts and values for current image
        def update_stored_values(current_index, all_modified_prompts, all_activities, all_expressions, all_fur_colors, 
                                new_modified_prompt, new_activity, new_expression, new_fur_color):
            """Update stored values for the current image"""
            if current_index >= 0:
                # Ensure arrays are long enough
                while len(all_modified_prompts) <= current_index:
                    all_modified_prompts.append("")
                while len(all_activities) <= current_index:
                    all_activities.append("")
                while len(all_expressions) <= current_index:
                    all_expressions.append("")
                while len(all_fur_colors) <= current_index:
                    all_fur_colors.append("")
                
                # Update values
                all_modified_prompts[current_index] = new_modified_prompt if new_modified_prompt else ""
                all_activities[current_index] = new_activity if new_activity else ""
                all_expressions[current_index] = new_expression if new_expression else ""
                all_fur_colors[current_index] = new_fur_color if new_fur_color else ""
            
            return all_modified_prompts, all_activities, all_expressions, all_fur_colors
        
        # Function to regenerate prompts for current image only
        def regenerate_current_prompt(current_index, extracted_images, all_prompts, all_modified_prompts, all_activities, all_expressions, all_fur_colors):
            """Regenerate prompt and values for current image only"""
            if not extracted_images or current_index >= len(extracted_images):
                return all_prompts, all_modified_prompts, all_activities, all_expressions, all_fur_colors, "", "", "", "", "", False
            
            # Generate new prompt for current image
            img_path = extracted_images[current_index]
            new_prompt = generate_prompt_from_image(img_path)
            
            # Detect if the image contains humans
            contains_human = detect_human_in_image(img_path)
            
            # Auto-generate activity, expression, and fur color
            activity, facial_expression = generate_activity_expression_from_prompt(new_prompt)
            fur_color = generate_fur_color_for_prompt(new_prompt)
            modified_prompt = enhance_prompt_with_activity_expression(new_prompt, activity, facial_expression, fur_color, "Auto", contains_human)
            
            # Update stored values
            all_prompts[current_index] = new_prompt
            all_modified_prompts[current_index] = modified_prompt
            all_activities[current_index] = activity
            all_expressions[current_index] = facial_expression
            all_fur_colors[current_index] = fur_color
            
            logger.info(f"Regenerated prompt for image {current_index + 1}: {new_prompt[:50]}...")
            
            # Reset manual edit flag when regenerating prompt
            return all_prompts, all_modified_prompts, all_activities, all_expressions, all_fur_colors, new_prompt, modified_prompt, activity, facial_expression, fur_color, False
            
        # ===== NEW FUNCTIONS FOR REFERENCE IMAGES 2 & 3 =====
        
        def process_reference_2_upload(file_path):
            """Process single image upload for Reference Image 2 (no ZIP support)"""
            if file_path is None:
                return None, [], 0, False, "Ref 2: No images"
            
            try:
                logger.info(f"Processing Reference Image 2: {file_path}")
                
                # Handle different file formats similar to reference image 1
                actual_file_path = None
                if isinstance(file_path, dict) and 'name' in file_path:
                    actual_file_path = file_path['name']
                elif hasattr(file_path, 'name'):
                    actual_file_path = file_path.name
                elif isinstance(file_path, str):
                    actual_file_path = file_path
                else:
                    logger.warning(f"Unsupported file format for Ref 2: {type(file_path)}")
                    return None, [], 0, False, "Ref 2: Error"
                
                if not os.path.exists(actual_file_path):
                    logger.warning(f"Reference Image 2 file does not exist: {actual_file_path}")
                    return None, [], 0, False, "Ref 2: File not found"
                
                # Handle AVIF conversion if needed
                if actual_file_path.lower().endswith('.avif'):
                    try:
                        png_path = actual_file_path.rsplit('.', 1)[0] + '.png'
                        converted_path = convert_avif(actual_file_path, png_path, 'PNG')
                        if converted_path != actual_file_path:
                            actual_file_path = converted_path
                    except Exception as e:
                        logger.error(f"Error converting AVIF file for Ref 2: {str(e)}")
                        return None, [], 0, False, "Ref 2: AVIF error"
                
                # Reject ZIP files for reference images 2 and 3
                if actual_file_path.lower().endswith('.zip'):
                    logger.warning("Reference Image 2 does not support ZIP files")
                    return None, [], 0, False, "Ref 2: ZIP not supported"
                
                # Return single image in list format
                return actual_file_path, [actual_file_path], 0, len([actual_file_path]) > 1, f"Ref 2: Image 1/1"
                
            except Exception as e:
                logger.error(f"Error processing Reference Image 2: {str(e)}")
                return None, [], 0, False, "Ref 2: Error"
        
        def process_reference_3_upload(file_path):
            """Process single image upload for Reference Image 3 (no ZIP support)"""
            if file_path is None:
                return None, [], 0, False, "Ref 3: No images"
            
            try:
                logger.info(f"Processing Reference Image 3: {file_path}")
                
                # Handle different file formats similar to reference image 1
                actual_file_path = None
                if isinstance(file_path, dict) and 'name' in file_path:
                    actual_file_path = file_path['name']
                elif hasattr(file_path, 'name'):
                    actual_file_path = file_path.name
                elif isinstance(file_path, str):
                    actual_file_path = file_path
                else:
                    logger.warning(f"Unsupported file format for Ref 3: {type(file_path)}")
                    return None, [], 0, False, "Ref 3: Error"
                
                if not os.path.exists(actual_file_path):
                    logger.warning(f"Reference Image 3 file does not exist: {actual_file_path}")
                    return None, [], 0, False, "Ref 3: File not found"
                
                # Handle AVIF conversion if needed
                if actual_file_path.lower().endswith('.avif'):
                    try:
                        png_path = actual_file_path.rsplit('.', 1)[0] + '.png'
                        converted_path = convert_avif(actual_file_path, png_path, 'PNG')
                        if converted_path != actual_file_path:
                            actual_file_path = converted_path
                    except Exception as e:
                        logger.error(f"Error converting AVIF file for Ref 3: {str(e)}")
                        return None, [], 0, False, "Ref 3: AVIF error"
                
                # Reject ZIP files for reference images 2 and 3
                if actual_file_path.lower().endswith('.zip'):
                    logger.warning("Reference Image 3 does not support ZIP files")
                    return None, [], 0, False, "Ref 3: ZIP not supported"
                
                # Return single image in list format
                return actual_file_path, [actual_file_path], 0, len([actual_file_path]) > 1, f"Ref 3: Image 1/1"
                
            except Exception as e:
                logger.error(f"Error processing Reference Image 3: {str(e)}")
                return None, [], 0, False, "Ref 3: Error"
                
        def navigate_reference_2(direction, current_index, images_list):
            """Navigate through Reference Image 2 gallery (single images only)"""
            if not images_list or len(images_list) <= 1:
                return current_index, images_list[0] if images_list else None, f"Ref 2: Image 1/1" if images_list else "Ref 2: No images"
            
            if direction == "prev":
                new_index = (current_index - 1) % len(images_list)
            else:  # direction == "next"
                new_index = (current_index + 1) % len(images_list)
            
            image_path = images_list[new_index]
            counter_text = f"Ref 2: Image {new_index+1}/{len(images_list)}"
            
            return new_index, image_path, counter_text
            
        def navigate_reference_3(direction, current_index, images_list):
            """Navigate through Reference Image 3 gallery (single images only)"""
            if not images_list or len(images_list) <= 1:
                return current_index, images_list[0] if images_list else None, f"Ref 3: Image 1/1" if images_list else "Ref 3: No images"
            
            if direction == "prev":
                new_index = (current_index - 1) % len(images_list)
            else:  # direction == "next"
                new_index = (current_index + 1) % len(images_list)
            
            image_path = images_list[new_index]
            counter_text = f"Ref 3: Image {new_index+1}/{len(images_list)}"
            
            return new_index, image_path, counter_text
        
        # Connect file upload to process and show previews 
        reference_image_1.change(
            fn=process_uploaded_file,
            inputs=[reference_image_1, disable_activity_expression_fur],
            outputs=[
                reference_preview,
                status_text, 
                generated_prompt_state, 
                extracted_images_state, 
                current_image_index, 
                all_prompts_state, 
                all_modified_prompts_state,
                all_activities_state,
                all_expressions_state,
                all_fur_colors_state,
                nav_controls_visible, 
                reference_preview, 
                image_counter,
                original_zip_file_path  # Add the new state variable
            ]
        ).then(
            # Update navigation controls visibility
            fn=lambda x: gr.update(visible=x),
            inputs=[nav_controls_visible],
            outputs=[image_nav_controls]
        ).then(
            # Update prompt display with stored values
            fn=lambda prompt, modified_list, activity_list, expr_list, fur_list: (
                prompt, 
                modified_list[0] if modified_list and len(modified_list) > 0 else "", 
                activity_list[0] if activity_list and len(activity_list) > 0 else "", 
                expr_list[0] if expr_list and len(expr_list) > 0 else "", 
                fur_list[0] if fur_list and len(fur_list) > 0 else "", 
                gr.update(visible=True)
            ),
            inputs=[generated_prompt_state, all_modified_prompts_state, all_activities_state, all_expressions_state, all_fur_colors_state],
            outputs=[generated_prompt_display, modified_prompt_display, activity_input, facial_expression_input, fur_color_input, generate_with_activity_button]
        )
        
        # Connect navigation buttons
        prev_button.click(
            fn=show_previous_image,
            inputs=[
                current_image_index,
                extracted_images_state,
                all_prompts_state,
                all_modified_prompts_state,
                all_activities_state,
                all_expressions_state,
                all_fur_colors_state
            ],
            outputs=[
                current_image_index,
                reference_preview,
                generated_prompt_state,
                modified_prompt_display,
                activity_input,
                facial_expression_input,
                fur_color_input,
                image_counter,
                nav_controls_visible,
                manual_prompt_edit_state
            ]
        ).then(
            # Update visibility after navigation
            fn=lambda x: gr.update(visible=x),
            inputs=[nav_controls_visible],
            outputs=[image_nav_controls]
        ).then(
            # Update displayed values
            fn=lambda prompt, modified, activity, expr, fur: (prompt, modified, activity, expr, fur, gr.update(visible=True)),
            inputs=[generated_prompt_state, modified_prompt_display, activity_input, facial_expression_input, fur_color_input],
            outputs=[generated_prompt_display, modified_prompt_display, activity_input, facial_expression_input, fur_color_input, generate_with_activity_button]
        )
        
        # Synchronize navigation with current image index
        next_button.click(
            fn=show_next_image,
            inputs=[
                current_image_index,
                extracted_images_state,
                all_prompts_state,
                all_modified_prompts_state,
                all_activities_state,
                all_expressions_state,
                all_fur_colors_state
            ],
            outputs=[
                current_image_index,
                reference_preview,
                generated_prompt_state,
                modified_prompt_display,
                activity_input,
                facial_expression_input,
                fur_color_input,
                image_counter,
                nav_controls_visible,
                manual_prompt_edit_state
            ]
        ).then(
            # Update visibility after navigation
            fn=lambda x: gr.update(visible=x),
            inputs=[nav_controls_visible],
            outputs=[image_nav_controls]
        ).then(
            # Update displayed values
            fn=lambda prompt, modified, activity, expr, fur: (prompt, modified, activity, expr, fur, gr.update(visible=True)),
            inputs=[generated_prompt_state, modified_prompt_display, activity_input, facial_expression_input, fur_color_input],
            outputs=[generated_prompt_display, modified_prompt_display, activity_input, facial_expression_input, fur_color_input, generate_with_activity_button]
        )
        
        # ===== NEW EVENT HANDLERS FOR REFERENCE IMAGES 2 & 3 =====
        
        # Connect Reference Image 2 upload
        reference_image_2.change(
            fn=process_reference_2_upload,
            inputs=[reference_image_2],
            outputs=[
                reference_2_preview,
                ref_2_images_state,
                ref_2_current_index,
                ref_2_nav_visible,
                image_counter_2
            ]
        ).then(
            # Update navigation controls visibility for Ref 2
            fn=lambda x: gr.update(visible=x),
            inputs=[ref_2_nav_visible],
            outputs=[image_2_nav_controls]
        )
        
        # Connect Reference Image 3 upload
        reference_image_3.change(
            fn=process_reference_3_upload,
            inputs=[reference_image_3],
            outputs=[
                reference_3_preview,
                ref_3_images_state,
                ref_3_current_index,
                ref_3_nav_visible,
                image_counter_3
            ]
        ).then(
            # Update navigation controls visibility for Ref 3
            fn=lambda x: gr.update(visible=x),
            inputs=[ref_3_nav_visible],
            outputs=[image_3_nav_controls]
        )
        
        # Connect Reference Image 2 navigation buttons
        prev_button_2.click(
            fn=lambda idx, imgs: navigate_reference_2("prev", idx, imgs),
            inputs=[ref_2_current_index, ref_2_images_state],
            outputs=[ref_2_current_index, reference_2_preview, image_counter_2]
        )
        
        next_button_2.click(
            fn=lambda idx, imgs: navigate_reference_2("next", idx, imgs),
            inputs=[ref_2_current_index, ref_2_images_state],
            outputs=[ref_2_current_index, reference_2_preview, image_counter_2]
        )
        
        # Connect Reference Image 3 navigation buttons
        prev_button_3.click(
            fn=lambda idx, imgs: navigate_reference_3("prev", idx, imgs),
            inputs=[ref_3_current_index, ref_3_images_state],
            outputs=[ref_3_current_index, reference_3_preview, image_counter_3]
        )
        
        next_button_3.click(
            fn=lambda idx, imgs: navigate_reference_3("next", idx, imgs),
            inputs=[ref_3_current_index, ref_3_images_state],
            outputs=[ref_3_current_index, reference_3_preview, image_counter_3]
        )
        
        # Function to sync the checkboxes (when one is checked, uncheck the other)
        def sync_predefined_checkbox(checked):
            return not checked
            
        def sync_qwen_checkbox(checked):
            return not checked
            
        # Add event handlers to keep checkboxes in sync
        use_predefined_options.change(
            fn=sync_qwen_checkbox,
            inputs=[use_predefined_options],
            outputs=[use_qwen_generation]
        )
        
        use_qwen_generation.change(
            fn=sync_predefined_checkbox,
            inputs=[use_qwen_generation],
            outputs=[use_predefined_options]
        )
        
        # Connect the reiterate buttons to their respective functions
        reiterate_activity_button.click(
            fn=update_with_new_activity_and_prompt,
            inputs=[generated_prompt_display, activity_input, facial_expression_input, fur_color_input, use_predefined_options],
            outputs=[activity_input, modified_prompt_display]
        )
        
        reiterate_expression_button.click(
            fn=update_with_new_expression_and_prompt,
            inputs=[generated_prompt_display, activity_input, facial_expression_input, fur_color_input, use_predefined_options],
            outputs=[facial_expression_input, modified_prompt_display]
        )
        
        reiterate_fur_color_button.click(
            fn=update_with_new_fur_color_and_prompt,
            inputs=[generated_prompt_display, activity_input, facial_expression_input, fur_color_input, use_predefined_options],
            outputs=[fur_color_input, modified_prompt_display]
        )
        
        # Connect the stop button
        stop_button.click(
            lambda: True,
            outputs=stop_generation_flag
        )
        
        # Function to handle manual prompt edits and set the manual edit flag
        def handle_manual_prompt_edit(current_index, all_modified_prompts, all_activities, all_expressions, all_fur_colors, 
                                     new_modified_prompt, new_activity, new_expression, new_fur_color):
            """Handle manual prompt edits and set manual edit flag"""
            logger.info("🖋️ User manually edited the modified prompt - preserving changes")
            # Update stored values
            updated_prompts, updated_activities, updated_expressions, updated_fur_colors = update_stored_values(
                current_index, all_modified_prompts, all_activities, all_expressions, all_fur_colors,
                new_modified_prompt, new_activity, new_expression, new_fur_color
            )
            # Return updated values and set manual edit flag to True
            return updated_prompts, updated_activities, updated_expressions, updated_fur_colors, True

        # Add event handlers to save modified prompts and values when user changes them
        modified_prompt_display.change(
            fn=handle_manual_prompt_edit,
            inputs=[
                current_image_index,
                all_modified_prompts_state,
                all_activities_state,
                all_expressions_state,
                all_fur_colors_state,
                modified_prompt_display,
                activity_input,
                facial_expression_input,
                fur_color_input
            ],
            outputs=[
                all_modified_prompts_state,
                all_activities_state,
                all_expressions_state,
                all_fur_colors_state,
                manual_prompt_edit_state
            ]
        )
        
        activity_input.change(
            fn=update_stored_values,
            inputs=[
                current_image_index,
                all_modified_prompts_state,
                all_activities_state,
                all_expressions_state,
                all_fur_colors_state,
                modified_prompt_display,
                activity_input,
                facial_expression_input,
                fur_color_input
            ],
            outputs=[
                all_modified_prompts_state,
                all_activities_state,
                all_expressions_state,
                all_fur_colors_state
            ]
        )
        
        facial_expression_input.change(
            fn=update_stored_values,
            inputs=[
                current_image_index,
                all_modified_prompts_state,
                all_activities_state,
                all_expressions_state,
                all_fur_colors_state,
                modified_prompt_display,
                activity_input,
                facial_expression_input,
                fur_color_input
            ],
            outputs=[
                all_modified_prompts_state,
                all_activities_state,
                all_expressions_state,
                all_fur_colors_state
            ]
        )
        
        fur_color_input.change(
            fn=update_stored_values,
            inputs=[
                current_image_index,
                all_modified_prompts_state,
                all_activities_state,
                all_expressions_state,
                all_fur_colors_state,
                modified_prompt_display,
                activity_input,
                facial_expression_input,
                fur_color_input
            ],
            outputs=[
                all_modified_prompts_state,
                all_activities_state,
                all_expressions_state,
                all_fur_colors_state
            ]
        )
        
        # Connect regenerate prompt button
        regenerate_prompt_button.click(
            fn=regenerate_current_prompt,
            inputs=[
                current_image_index,
                extracted_images_state,
                all_prompts_state,
                all_modified_prompts_state,
                all_activities_state,
                all_expressions_state,
                all_fur_colors_state
            ],
            outputs=[
                all_prompts_state,
                all_modified_prompts_state,
                all_activities_state,
                all_expressions_state,
                all_fur_colors_state,
                generated_prompt_display,
                modified_prompt_display,
                activity_input,
                facial_expression_input,
                fur_color_input,
                manual_prompt_edit_state
            ]
        )
        
        # Copy button functionality for Generated Prompt
        copy_generated_prompt_button.click(
            fn=lambda prompt: (prompt, "Copied to clipboard!"),
            inputs=[generated_prompt_display],
            outputs=[generated_prompt_display, status_text],
            js="""
            function(prompt) {
                // Copy the prompt text to clipboard
                navigator.clipboard.writeText(prompt).then(function() {
                    console.log('Generated prompt copied to clipboard');
                }).catch(function(err) {
                    console.error('Failed to copy generated prompt: ', err);
                });
                return [prompt, "Generated prompt copied to clipboard!"];
            }
            """
        )
        
        # Copy button functionality for Modified Prompt
        copy_modified_prompt_button.click(
            fn=lambda prompt: (prompt, "Copied to clipboard!"),
            inputs=[modified_prompt_display],
            outputs=[modified_prompt_display, status_text],
            js="""
            function(prompt) {
                // Copy the prompt text to clipboard
                navigator.clipboard.writeText(prompt).then(function() {
                    console.log('Modified prompt copied to clipboard');
                }).catch(function(err) {
                    console.error('Failed to copy modified prompt: ', err);
                });
                return [prompt, "Modified prompt copied to clipboard!"];
            }
            """
        )
        
        # Manual variation generation button handler
        generate_variations_button.click(
            fn=handle_manual_variation_generation,
            inputs=[
                manual_variation_input,
                generated_prompt_state,
                extracted_images_state,
                current_image_index,
                all_prompts_state,
                all_modified_prompts_state,
                all_activities_state,
                all_expressions_state,
                all_fur_colors_state
            ],
            outputs=[
                status_text,
                all_prompts_state,
                all_modified_prompts_state,
                all_activities_state,
                all_expressions_state,
                all_fur_colors_state,
                image_counter,
                nav_controls_visible,
                extracted_images_state,
                current_image_index
            ]
        ).then(
            # Update prompt display with first variation
            fn=lambda prompts, modified, activity, expr, fur: (
                prompts[0] if prompts and len(prompts) > 0 else "",
                modified[0] if modified and len(modified) > 0 else "",
                activity[0] if activity and len(activity) > 0 else "",
                expr[0] if expr and len(expr) > 0 else "",
                fur[0] if fur and len(fur) > 0 else ""
            ),
            inputs=[all_prompts_state, all_modified_prompts_state, all_activities_state, all_expressions_state, all_fur_colors_state],
            outputs=[generated_prompt_display, modified_prompt_display, activity_input, facial_expression_input, fur_color_input]
        ).then(
            # Update navigation controls visibility
            fn=lambda x: gr.update(visible=x),
            inputs=[nav_controls_visible],
            outputs=[image_nav_controls]
        )

        
        # Function to handle activity/expression changes without overwriting manual edits
        def update_activity_preserve_manual_prompt(prompt, activity, expression, fur_color, ethnicity, current_index, extracted_images, modified_prompt_current, manual_edit_flag):
            """Update activity but preserve manually edited prompts"""
            activity_cleaned = activity.strip()
            
            # If user has manually edited the prompt, preserve it
            if manual_edit_flag:
                logger.info("🔒 Manual prompt edit detected - preserving user's modified prompt")
                return activity_cleaned, modified_prompt_current
            else:
                # Auto-generate modified prompt
                new_modified_prompt = enhance_prompt_with_activity_expression(
                    prompt, activity_cleaned, expression, fur_color, ethnicity, 
                    detect_human_in_image(extracted_images[current_index]) if extracted_images and current_index < len(extracted_images) else False
                )
                return activity_cleaned, new_modified_prompt
        
        def update_expression_preserve_manual_prompt(prompt, activity, expression, fur_color, ethnicity, current_index, extracted_images, modified_prompt_current, manual_edit_flag):
            """Update expression but preserve manually edited prompts"""
            expression_cleaned = expression.strip()
            
            # If user has manually edited the prompt, preserve it
            if manual_edit_flag:
                logger.info("🔒 Manual prompt edit detected - preserving user's modified prompt")
                return expression_cleaned, modified_prompt_current
            else:
                # Auto-generate modified prompt
                new_modified_prompt = enhance_prompt_with_activity_expression(
                    prompt, activity, expression_cleaned, fur_color, ethnicity,
                    detect_human_in_image(extracted_images[current_index]) if extracted_images and current_index < len(extracted_images) else False
                )
                return expression_cleaned, new_modified_prompt
        
        def update_fur_color_preserve_manual_prompt(prompt, activity, expression, fur_color, ethnicity, current_index, extracted_images, modified_prompt_current, manual_edit_flag):
            """Update fur color but preserve manually edited prompts"""
            fur_color_cleaned = fur_color.strip()
            
            # If user has manually edited the prompt, preserve it
            if manual_edit_flag:
                logger.info("🔒 Manual prompt edit detected - preserving user's modified prompt")
                return fur_color_cleaned, modified_prompt_current
            else:
                # Auto-generate modified prompt
                new_modified_prompt = enhance_prompt_with_activity_expression(
                    prompt, activity, expression, fur_color_cleaned, ethnicity,
                    detect_human_in_image(extracted_images[current_index]) if extracted_images and current_index < len(extracted_images) else False
                )
                return fur_color_cleaned, new_modified_prompt
        
        def update_ethnicity_preserve_manual_prompt(prompt, activity, expression, fur_color, ethnicity, current_index, extracted_images, modified_prompt_current, manual_edit_flag):
            """Update ethnicity but preserve manually edited prompts"""
            # If user has manually edited the prompt, preserve it
            if manual_edit_flag:
                logger.info("🔒 Manual prompt edit detected - preserving user's modified prompt")
                return ethnicity, modified_prompt_current
            else:
                # Auto-generate modified prompt
                new_modified_prompt = enhance_prompt_with_activity_expression(
                    prompt, activity, expression, fur_color, ethnicity,
                    detect_human_in_image(extracted_images[current_index]) if extracted_images and current_index < len(extracted_images) else False
                )
                return ethnicity, new_modified_prompt
        
        # Add event handlers to update the modified prompt when textboxes change
        activity_input.change(
            fn=update_activity_preserve_manual_prompt,
            inputs=[generated_prompt_display, activity_input, facial_expression_input, fur_color_input, ethnicity_dropdown, current_image_index, extracted_images_state, modified_prompt_display, manual_prompt_edit_state],
            outputs=[activity_input, modified_prompt_display]
        )
        
        facial_expression_input.change(
            fn=update_expression_preserve_manual_prompt,
            inputs=[generated_prompt_display, activity_input, facial_expression_input, fur_color_input, ethnicity_dropdown, current_image_index, extracted_images_state, modified_prompt_display, manual_prompt_edit_state],
            outputs=[facial_expression_input, modified_prompt_display]
        )
        
        fur_color_input.change(
            fn=update_fur_color_preserve_manual_prompt,
            inputs=[generated_prompt_display, activity_input, facial_expression_input, fur_color_input, ethnicity_dropdown, current_image_index, extracted_images_state, modified_prompt_display, manual_prompt_edit_state],
            outputs=[fur_color_input, modified_prompt_display]
        )
        
        ethnicity_dropdown.change(
            fn=update_ethnicity_preserve_manual_prompt,
            inputs=[generated_prompt_display, activity_input, facial_expression_input, fur_color_input, ethnicity_dropdown, current_image_index, extracted_images_state, modified_prompt_display, manual_prompt_edit_state],
            outputs=[ethnicity_dropdown, modified_prompt_display]
        )
        
        # Utility function to strip asterisks from the start and end of a string
        def strip_asterisks(text):
            if not isinstance(text, str):
                return text
            # Remove asterisks from start and end, and also clean up any markdown-style formatting
            cleaned = text.strip()
            # Remove leading and trailing asterisks
            while cleaned.startswith('*') or cleaned.startswith('**'):
                if cleaned.startswith('**'):
                    cleaned = cleaned[2:]
                elif cleaned.startswith('*'):
                    cleaned = cleaned[1:]
                cleaned = cleaned.strip()
            
            while cleaned.endswith('*') or cleaned.endswith('**'):
                if cleaned.endswith('**'):
                    cleaned = cleaned[:-2]
                elif cleaned.endswith('*'):
                    cleaned = cleaned[:-1]
                cleaned = cleaned.strip()
            
            # Remove any remaining double asterisks within the text (markdown bold formatting)
            cleaned = cleaned.replace('**', '')
            
            return cleaned

        # Function to initialize modified prompt when a new prompt is generated
        def update_modified_prompt_on_prompt_change(prompt, activity, expression, fur_color, ethnicity, current_index, extracted_images):
            """Update modified prompt whenever the main prompt changes - ensures gender is mentioned only for living creatures"""
            if prompt and prompt.strip():
                logger.info(f"🔄 Updating modified prompt")
                
                # Detect if the current image contains humans
                contains_human = False
                if extracted_images and current_index < len(extracted_images):
                    img_path = extracted_images[current_index]
                    contains_human = detect_human_in_image(img_path)
                
                # Detect if this is a living creature from the prompt
                is_living_creature = detect_living_creature_from_prompt(prompt)
                
                # Only ensure gender for living creatures
                if is_living_creature:
                    logger.info(f"🔄 Applying gender enforcement for living creature")
                    gender_enhanced_prompt = ensure_gender_in_prompt(prompt, contains_human=contains_human)
                else:
                    logger.info(f"🔄 Skipping gender enforcement for non-living object")
                    gender_enhanced_prompt = prompt
                
                # Then apply activity/expression/fur color/ethnicity modifications
                result = enhance_prompt_with_activity_expression(gender_enhanced_prompt, activity, expression, fur_color, ethnicity, contains_human)
                return strip_asterisks(result)
            return ""
        
        # Update modified prompt when the generated prompt changes
        generated_prompt_display.change(
            fn=update_modified_prompt_on_prompt_change,
            inputs=[generated_prompt_display, activity_input, facial_expression_input, fur_color_input, ethnicity_dropdown, current_image_index, extracted_images_state],
            outputs=[modified_prompt_display]
        )
        
        # Add handlers for the main action buttons
        # 1. Remove Background & Apply to Card Template button
        remove_bg_button.click(
            fn=bg_removal_wrapper,  # Use bg_removal_wrapper instead of process_image_with_birefnet directly
            inputs=[reference_preview, card_template, bg_method, remove_watermark_checkbox, original_zip_file_path, current_image_index, extracted_images_state],  # Add current_image_index and extracted_images_state
            outputs=[output_gallery, status_text, download_zip]  # Output to gallery instead of reference preview, include download ZIP
        )
        
        # Function to reset stop flag and then call generate_wrapper
        def generate_with_activity_and_reset_stop(provider_state, reference_preview, card_template, 
                                                  theme_dropdown, category_dropdown, subcategory_dropdown, leonardo_model_dropdown, 
                                                  guidance_scale_slider, modified_prompt_display, negative_prompt,
                                                  preset_style, leonardo_num_images, ideogram_model, ideogram_style,
                                                  ideogram_num_images, output_format, extracted_images_state, all_prompts_state,
                                                  current_image_index, modification_type, modification_details, modified_prompt,
                                                  reference_image_filename, filename_convention, s3_upload, leonardo_seed,
                                                  activity_input, facial_expression_input, fur_color_input, ethnicity_dropdown,
                                                  base64_encode_checkbox, stop_generation_flag, 
                                                  reference_image_1, ref_type_1, ref_strength_1,
                                                  reference_image_2, ref_type_2, ref_strength_2,
                                                  reference_image_3, ref_type_3, ref_strength_3,
                                                  ideogram_disable_style_reference, ideogram_rendering_speed,
                                                  imagen4_model, imagen4_aspect_ratio, imagen4_safety_filter, imagen4_num_images,
                                                  counter_override_dropdown):
            """Reset stop flag and start generation with activity"""
            logger.info("🔄 Resetting stop flag and starting generation with activity")
            logger.info(f"Received subcategory for generation: {subcategory_dropdown} (type: {type(subcategory_dropdown)})")
            if isinstance(subcategory_dropdown, list):
                subcategory = subcategory_dropdown[0] if subcategory_dropdown else None
            else:
                subcategory = subcategory_dropdown
            stop_flag_reset = False
            
            return generate_wrapper(provider_state, reference_preview, card_template, 
                                  theme_dropdown, category_dropdown, subcategory, leonardo_model_dropdown, 
                                  guidance_scale_slider, modified_prompt_display, negative_prompt,
                                  preset_style, leonardo_num_images, ideogram_model, ideogram_style,
                                  ideogram_num_images, output_format, extracted_images_state, all_prompts_state,
                                  current_image_index, modification_type, modification_details, modified_prompt,
                                  reference_image_filename, filename_convention, s3_upload, leonardo_seed,
                                  activity_input, facial_expression_input, fur_color_input, ethnicity_dropdown,
                                  base64_encode_checkbox, stop_flag_reset,
                                  reference_image_1=reference_image_1, ref_type_1=ref_type_1, ref_strength_1=ref_strength_1,
                                  reference_image_2=reference_image_2, ref_type_2=ref_type_2, ref_strength_2=ref_strength_2,
                                  reference_image_3=reference_image_3, ref_type_3=ref_type_3, ref_strength_3=ref_strength_3,
                                  ideogram_disable_style_reference=ideogram_disable_style_reference, ideogram_rendering_speed=ideogram_rendering_speed,
                                  imagen4_model=imagen4_model, imagen4_aspect_ratio=imagen4_aspect_ratio, 
                                  imagen4_safety_filter=imagen4_safety_filter, imagen4_num_images=imagen4_num_images,
                                  counter_override=counter_override_dropdown)

        # 2. Generate With Activity/Expression button
        generate_with_activity_button.click(
            fn=generate_with_activity_and_reset_stop,
            inputs=[
                provider_state, reference_preview, card_template, 
                theme_dropdown, category_dropdown, subcategory_dropdown, leonardo_model_dropdown, 
                guidance_scale_slider, modified_prompt_display, negative_prompt,
                preset_style, leonardo_num_images, ideogram_model, ideogram_style,
                ideogram_num_images, output_format, extracted_images_state, all_prompts_state,
                current_image_index, gr.State("activity"), gr.State(None), modified_prompt_display,
                reference_image_filename, gr.State("Current Filename Setting"), gr.State(True), leonardo_seed,
                activity_input, facial_expression_input, fur_color_input, ethnicity_dropdown,
                base64_encode_checkbox, stop_generation_flag, 
                reference_image_1, ref_type_1, ref_strength_1,
                reference_image_2, ref_type_2, ref_strength_2,
                reference_image_3, ref_type_3, ref_strength_3,
                ideogram_disable_style_reference, ideogram_rendering_speed,
                imagen4_model, imagen4_aspect_ratio, imagen4_safety_filter, imagen4_num_images,
                counter_override_dropdown
            ],
            outputs=[output_gallery, status_text, download_zip, generated_images_state, gr.State(None), gr.State(None)]
        ).then(
            fn=enable_inpainting_button,
            inputs=[generated_images_state, status_text],
            outputs=[select_for_inpainting_button]
        )

        # Function to reset stop flag and then call generate_wrapper for standard generation
        def generate_and_reset_stop(provider_state, reference_preview, card_template, 
                                     theme_dropdown, category_dropdown, subcategory_dropdown, leonardo_model_dropdown, 
                                     guidance_scale_slider, generated_prompt_display, negative_prompt,
                                     preset_style, leonardo_num_images, ideogram_model, ideogram_style,
                                     ideogram_num_images, output_format, extracted_images_state, all_prompts_state,
                                     current_image_index, modification_type, modification_details, modified_prompt,
                                     reference_image_filename, filename_convention, s3_upload, leonardo_seed,
                                     activity_input, facial_expression_input, fur_color_input, ethnicity_dropdown,
                                     base64_encode_checkbox, stop_generation_flag, 
                                     reference_image_1, ref_type_1, ref_strength_1,
                                     reference_image_2, ref_type_2, ref_strength_2,
                                     reference_image_3, ref_type_3, ref_strength_3,
                                     ideogram_disable_style_reference, ideogram_rendering_speed,
                                     imagen4_model, imagen4_aspect_ratio, imagen4_safety_filter, imagen4_num_images,
                                     counter_override_dropdown):
            """Reset stop flag and start standard generation"""
            logger.info("🔄 Resetting stop flag and starting standard generation")
            logger.info(f"Received subcategory for generation: {subcategory_dropdown} (type: {type(subcategory_dropdown)})")
            if isinstance(subcategory_dropdown, list):
                subcategory = subcategory_dropdown[0] if subcategory_dropdown else None
            else:
                subcategory = subcategory_dropdown
            stop_flag_reset = False
            
            return generate_wrapper(provider_state, reference_preview, card_template, 
                                     theme_dropdown, category_dropdown, subcategory, leonardo_model_dropdown, 
                                     guidance_scale_slider, generated_prompt_display, negative_prompt,
                                     preset_style, leonardo_num_images, ideogram_model, ideogram_style,
                                     ideogram_num_images, output_format, extracted_images_state, all_prompts_state,
                                     current_image_index, modification_type, modification_details, modified_prompt,
                                     reference_image_filename, filename_convention, s3_upload, leonardo_seed,
                                     activity_input, facial_expression_input, fur_color_input, ethnicity_dropdown,
                                     base64_encode_checkbox, stop_flag_reset,
                                     reference_image_1=reference_image_1, ref_type_1=ref_type_1, ref_strength_1=ref_strength_1,
                                     reference_image_2=reference_image_2, ref_type_2=ref_type_2, ref_strength_2=ref_strength_2,
                                     reference_image_3=reference_image_3, ref_type_3=ref_type_3, ref_strength_3=ref_strength_3,
                                     ideogram_disable_style_reference=ideogram_disable_style_reference, ideogram_rendering_speed=ideogram_rendering_speed,
                                     imagen4_model=imagen4_model, imagen4_aspect_ratio=imagen4_aspect_ratio, 
                                     imagen4_safety_filter=imagen4_safety_filter, imagen4_num_images=imagen4_num_images,
                                     counter_override=counter_override_dropdown)

        # 3. Generate button (standard generation without activity/expression)
        generate_button.click(
            fn=generate_and_reset_stop,
            inputs=[
                provider_state, reference_preview, card_template, 
                theme_dropdown, category_dropdown, subcategory_dropdown, leonardo_model_dropdown, 
                guidance_scale_slider, generated_prompt_display, negative_prompt,
                preset_style, leonardo_num_images, ideogram_model, ideogram_style,
                ideogram_num_images, output_format, extracted_images_state, all_prompts_state,
                current_image_index, gr.State(None), gr.State(None), gr.State(None),
                reference_image_filename, gr.State("Current Filename Setting"), gr.State(True), leonardo_seed,
                activity_input, facial_expression_input, fur_color_input, ethnicity_dropdown,
                base64_encode_checkbox, stop_generation_flag, 
                reference_image_1, ref_type_1, ref_strength_1,
                reference_image_2, ref_type_2, ref_strength_2,
                reference_image_3, ref_type_3, ref_strength_3,
                ideogram_disable_style_reference, ideogram_rendering_speed,
                imagen4_model, imagen4_aspect_ratio, imagen4_safety_filter, imagen4_num_images,
                counter_override_dropdown
            ],
            outputs=[output_gallery, status_text, download_zip, generated_images_state, gr.State(None), gr.State(None)]
        ).then(
            fn=enable_inpainting_button,
            inputs=[generated_images_state, status_text],
            outputs=[select_for_inpainting_button]
        )
        

        
        # ======= INPAINTING EVENT HANDLERS =======
        
        # Handle "Select for Inpainting" button click - connects to main inpainting tab
        select_for_inpainting_button.click(
            fn=select_image_for_inpainting,
            inputs=[generated_images_state],  # Use generated images state instead of gallery
            outputs=[
                inpaint_image_selector,  # main tab dropdown
                generated_images_state,  # state
                inpaint_status,  # main tab status
                inpaint_image_editor  # main tab editor
            ]
        )
        # Return the interface
        return interface

def add_image_to_cell(worksheet, img_path, cell_reference):
    """Add an image to a specific cell with proper sizing and positioning"""
    try:
        # Open and process the image
        with Image.open(img_path) as img:
            # Resize image to fit cell dimensions while maintaining aspect ratio
            img = img.resize((150, 150), Image.LANCZOS)
            
            # Create an in-memory file-like object for the image
            img_buffer = io.BytesIO()
            img.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            
            # Create an openpyxl image object directly from the buffer
            xl_img = XLImage(img_buffer)
            
            # Get the cell to position the image properly
            cell = worksheet[cell_reference]
            
            # Adjust width and height if needed
            col_letter = cell_reference[0]
            worksheet.column_dimensions[col_letter].width = 20
            row_num = int(cell_reference[1:])
            worksheet.row_dimensions[row_num].height = 120
            
            # Center the cell content
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add the image with proper anchoring to the cell
            xl_img.anchor = cell_reference
            worksheet.add_image(xl_img)
            
            return True
    except Exception as e:
        logger.error(f"Error adding image to cell {cell_reference}: {str(e)}")
        return False

# Add functions to handle ZIP file uploads and extract multiple reference images
def extract_images_from_zip(zip_file_path):
    """Extract images from a ZIP file and return a list of file paths to the extracted images"""
    extracted_images = []
    temp_dir = tempfile.mkdtemp()
    
    try:
        # Check if the file is actually a ZIP file
        if not zipfile.is_zipfile(zip_file_path):
            logger.error(f"File is not a valid ZIP file: {zip_file_path}")
            shutil.rmtree(temp_dir, ignore_errors=True)
            return [], None
            
        with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
            # Get list of files in the ZIP
            file_list = zip_ref.namelist()
            logger.info(f"ZIP contains {len(file_list)} files")
            
            # Extract all files to the temporary directory
            zip_ref.extractall(temp_dir)
            
            # Find all image files in the extracted content
            for root, _, files in os.walk(temp_dir):
                for file in files:
                    if file.lower().endswith(('.png', '.jpg', '.jpeg', '.avif', '.webp')):
                        file_path = os.path.join(root, file)
                        try:
                            # Verify it's a valid image by opening it
                            with Image.open(file_path) as img:
                                img.verify()  # Verify it's a valid image
                            extracted_images.append(file_path)
                        except Exception as img_error:
                            logger.warning(f"Skipping invalid image {file_path}: {str(img_error)}")
        
        if extracted_images:
            logger.info(f"Extracted {len(extracted_images)} valid images from ZIP file")
        else:
            logger.warning("No valid images found in ZIP file")
            
        return extracted_images, temp_dir
    except Exception as e:
        logger.error(f"Error extracting images from ZIP: {str(e)}")
        shutil.rmtree(temp_dir, ignore_errors=True)
        return [], None

# Function to find empty space on a card template
def find_empty_space(card_image, threshold=240):
    """Find the largest empty (white/light) area on a card template"""
    # Convert to grayscale
    if card_image.mode != 'L':
        gray = cv2.cvtColor(np.array(card_image), cv2.COLOR_RGB2GRAY)
    else:
        gray = np.array(card_image)
    
    # Threshold to find light areas
    _, binary = cv2.threshold(gray, threshold, 255, cv2.THRESH_BINARY)
    
    # Find contours of white regions
    contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    # Find the largest contour
    if not contours:
        # If no contours found, use the center of the image
        h, w = gray.shape
        return (w//4, h//4, w//2, h//2)
    
    largest_contour = max(contours, key=cv2.contourArea)
    x, y, w, h = cv2.boundingRect(largest_contour)
    
    # Return the bounding box of the largest empty space
    return (x, y, w, h)

def place_image_on_card(card_template, image_to_place, preserve_original_alpha=True):
    """Place an image on a card template in the largest empty space with transparent background
    
    Args:
        card_template: PIL Image of the card template
        image_to_place: PIL Image with transparent background to place on card
        preserve_original_alpha: If True, minimize alpha processing to preserve original background removal quality
    """
    try:
        # Make sure both images are in RGBA mode with full transparency preserved
        card = card_template.convert("RGBA")
        image = image_to_place.convert("RGBA")
        
        # Apply alpha improvement to the image before placing it on the card
        # Use minimal processing if preserve_original_alpha is True to maintain background removal quality
        if preserve_original_alpha:
            # Minimal processing - just ensure RGBA mode and basic cleanup
            if image.mode != 'RGBA':
                image = image.convert('RGBA')
            logger.info("🎨 CARD PLACEMENT: Using minimal alpha processing to preserve original background removal quality")
        else:
            # Full processing with same parameters as background removal
            image = improve_alpha_edges(image, threshold=180, edge_feather=4, use_gaussian_blur=True, feather_intensity=0.6)
            logger.info("🎨 CARD PLACEMENT: Applied full alpha improvement processing")
        
        # Create a completely transparent base image with the same size as the card
        base_width, base_height = card.size
        transparent_base = Image.new("RGBA", (base_width, base_height), (0, 0, 0, 0))
        
        # Find the empty space on the card
        empty_x, empty_y, empty_w, empty_h = find_empty_space(card)
        
        # Get original image dimensions
        orig_width, orig_height = image.size
        orig_aspect_ratio = orig_width / orig_height
        
        # Calculate target dimensions while maintaining aspect ratio
        # First try fitting to width
        target_w = int(empty_w * 0.9)  # Use 90% of empty space width
        target_h = int(target_w / orig_aspect_ratio)
        
        # If height exceeds available space, fit to height instead
        if target_h > empty_h * 0.9:
            target_h = int(empty_h * 0.9)  # Use 90% of empty space height
            target_w = int(target_h * orig_aspect_ratio)
        
        # Ensure minimum size
        target_w = max(target_w, 50)
        target_h = max(target_h, 50)
        
        # Resize the image to the calculated dimensions while preserving transparency
        # Use LANCZOS for high-quality resizing that preserves alpha channel
        resized_image = image.resize((target_w, target_h), Image.LANCZOS)
        
        # Apply alpha improvement again after resizing to clean up any artifacts from resizing
        # Use same parameters for consistency
        if preserve_original_alpha:
            # Minimal processing after resize - just ensure RGBA mode
            if resized_image.mode != 'RGBA':
                resized_image = resized_image.convert('RGBA')
        else:
            resized_image = improve_alpha_edges(resized_image, threshold=180, edge_feather=4, use_gaussian_blur=True, feather_intensity=0.6)
        
        # Calculate position to center the image in the empty space
        pos_x = empty_x + (empty_w - target_w) // 2
        pos_y = empty_y + (empty_h - target_h) // 2
        
        # First paste the card template onto the transparent base
        # Use the card's own alpha channel as the mask to preserve rounded corners
        transparent_base.paste(card, (0, 0), card)
        
        # Create a high-quality alpha composite for the resized image
        # This prevents edge artifacts during blending
        import numpy as np
        
        # Convert images to numpy arrays for precise alpha blending
        base_array = np.array(transparent_base)
        image_array = np.array(resized_image)
        
        # Extract the region where the image will be placed
        y1, y2 = pos_y, pos_y + target_h
        x1, x2 = pos_x, pos_x + target_w
        
        # Ensure we don't go out of bounds
        y1, y2 = max(0, y1), min(base_height, y2)
        x1, x2 = max(0, x1), min(base_width, x2)
        
        # Adjust image array size if needed
        actual_h, actual_w = y2 - y1, x2 - x1
        if actual_h != target_h or actual_w != target_w:
            image_array = image_array[:actual_h, :actual_w]
        
        # Perform alpha blending
        if image_array.size > 0:
            # Get alpha values
            image_alpha = image_array[:, :, 3:4] / 255.0
            base_alpha = base_array[y1:y2, x1:x2, 3:4] / 255.0
            
            # Calculate new alpha
            new_alpha = image_alpha + base_alpha * (1 - image_alpha)
            
            # Avoid division by zero
            new_alpha_safe = np.where(new_alpha == 0, 1, new_alpha)
            
            # Blend RGB channels
            image_rgb = image_array[:, :, :3]
            base_rgb = base_array[y1:y2, x1:x2, :3]
            
            new_rgb = (image_rgb * image_alpha + base_rgb * base_alpha * (1 - image_alpha)) / new_alpha_safe
            
            # Update the base array
            base_array[y1:y2, x1:x2, :3] = new_rgb.astype(np.uint8)
            base_array[y1:y2, x1:x2, 3:4] = (new_alpha * 255).astype(np.uint8)
        
        # Convert back to PIL Image
        result = Image.fromarray(base_array, 'RGBA')
        
        # Final alpha improvement to ensure clean edges - use same parameters for consistency
        if preserve_original_alpha:
            # Skip final alpha processing to preserve original quality
            logger.info("🎨 CARD PLACEMENT: Skipped final alpha processing to preserve original background removal quality")
        else:
            result = improve_alpha_edges(result, threshold=180, edge_feather=4, use_gaussian_blur=True, feather_intensity=0.6)
        
        logger.info("✅ CARD PLACEMENT: Successfully placed image on card with improved alpha blending")
        return result
        
    except Exception as e:
        logging.error(f"❌ CARD PLACEMENT ERROR: {str(e)}")
        import traceback
        logging.error(f"❌ CARD PLACEMENT TRACEBACK: {traceback.format_exc()}")
        # Return the original card if there's an error
        return card_template

# Store for card templates
card_templates = {}

# Helper functions for prompt modification
def toggle_prompt_modification(modification_type):
    if modification_type == "None":
        return gr.update(interactive=False, value="")
    else:
        return gr.update(interactive=True)
        
def clean_modified_prompt(prompt):
    """
    Clean the modified prompt to ensure it's a single continuous paragraph
    without asterisks and unwanted additional text
    """
    if not prompt:
        return prompt
    
    try:
        # Remove asterisks (both single and double)
        cleaned = re.sub(r'\*+', '', prompt)
        
        # Remove unwanted additional prompt text
        unwanted_patterns = [
            r'This prompt ensures that.*?while maintaining.*?\.',
            r'This prompt ensures that.*?from the original prompt\.',
            r'This ensures that.*?while maintaining.*?\.',
            r'This ensures that.*?from the original prompt\.',
            r'The prompt maintains.*?while adding.*?\.',
            r'The prompt maintains.*?from the original prompt\.',
            r'This modification.*?while preserving.*?\.',
            r'This modification.*?from the original prompt\.',
            r'This updated prompt.*?while keeping.*?\.',
            r'This updated prompt.*?from the original prompt\.'
        ]
        
        for pattern in unwanted_patterns:
            cleaned = re.sub(pattern, '', cleaned, flags=re.IGNORECASE | re.DOTALL)
        
        # Remove extra whitespace and normalize spacing
        cleaned = re.sub(r'\s+', ' ', cleaned)
        
        # Remove multiple commas and normalize comma spacing
        cleaned = re.sub(r',\s*,+', ',', cleaned)
        cleaned = re.sub(r'\s*,\s*', ', ', cleaned)
        
        # Remove leading/trailing commas and spaces
        cleaned = cleaned.strip(' ,')
        
        # Ensure it ends with proper punctuation if it doesn't already
        if cleaned and not cleaned.endswith(('.', '!', '?')):
            # If it ends with a comma, replace with period
            if cleaned.endswith(','):
                cleaned = cleaned[:-1] + '.'
            else:
                cleaned += '.'
        
        return cleaned
    except Exception as e:
        logger.error(f"Error cleaning modified prompt: {str(e)}")
        return prompt

def modify_prompt(original_prompt, modification_type, modification_details, reference_image_path=None):
    """
    Modify the original prompt based on the selected modification type and details.
    Uses Qwen to intelligently combine the original prompt with the new details.
    For non-human subjects, removes gender-specific language.
    """
    # Handle None values
    if modification_type is None or modification_details is None:
        return original_prompt
        
    if modification_type == "None" or not modification_details.strip():
        return original_prompt
        
    try:
        # Detect if the subject is human, animal, or object
        is_human_subject = False
        is_animal_subject = False
        
        if reference_image_path and os.path.exists(reference_image_path):
            is_human_subject = detect_human_in_image(reference_image_path)
            if not is_human_subject:
                # If not human, check if it's an animal
                detected_breed = detect_animal_breed(reference_image_path)
                is_animal_subject = detected_breed != "Unknown" and detected_breed != "Neither"
            logger.info(f"🔍 PROMPT MODIFICATION: Human: {is_human_subject}, Animal: {is_animal_subject}")
        else:
            # If no reference image, try to detect from the prompt content
            prompt_lower = original_prompt.lower()
            human_indicators = ['person', 'man', 'woman', 'boy', 'girl', 'human', 'character with', 'male', 'female']
            animal_indicators = ['cat', 'dog', 'animal', 'pet', 'kitten', 'puppy', '3d cartoon', 'creature']
            object_indicators = ['object', 'item', 'thing', 'product', 'building', 'car', 'landscape', 'scene']
            
            human_count = sum(1 for indicator in human_indicators if indicator in prompt_lower)
            animal_count = sum(1 for indicator in animal_indicators if indicator in prompt_lower)
            object_count = sum(1 for indicator in object_indicators if indicator in prompt_lower)
            
            # Determine subject type based on highest count
            if human_count > animal_count and human_count > object_count:
                is_human_subject = True
                is_animal_subject = False
            elif animal_count > human_count and animal_count > object_count:
                is_human_subject = False
                is_animal_subject = True
            else:
                is_human_subject = False
                is_animal_subject = False
            
            logger.info(f"🔍 PROMPT MODIFICATION: Prompt-based detection - Human: {is_human_subject}, Animal: {is_animal_subject} (human: {human_count}, animal: {animal_count}, object: {object_count})")
        
        # Create different system prompts based on subject type
        if is_human_subject:
            # For human subjects, use the original system prompt
            sys_prompt = f"""You are an assistant that modifies image generation prompts for HUMAN subjects.
            You need to combine the original prompt with new details about {modification_type.lower()}.
            
            IMPORTANT RULES:
            1. Integrate the new details naturally into the prompt without making it too long.
            2. Keep the essential elements of the original prompt but add or modify the {modification_type.lower()} related details.
            3. Any objects mentioned MUST BE appropriately sized and NEVER larger than the main subject.
            4. STRICTLY LIMIT objects to 1-2 maximum - DO NOT include more than 2 distinct objects.
            5. The main subject must always be the focal point and dominant element in the image.
            6. Objects should be proportionate and realistic in size compared to the character.
            7. If the original prompt or new details include too many objects, prioritize only the 1-2 most important ones.
            
            Only output the final modified prompt, nothing else."""
            
            logger.info("🧑 PROMPT MODIFICATION: Using human-friendly prompt modification")
        elif is_animal_subject:
            # For animal subjects, remove gender language but keep body part references
            sys_prompt = f"""You are an assistant that modifies image generation prompts for ANIMAL subjects (cats, dogs, pets, creatures).
            You need to combine the original prompt with new details about {modification_type.lower()}.
            
            CRITICAL ANIMAL RULES:
            1. DO NOT use any gender-specific language like "A female character with", "A male character with", "he", "she", "him", "her"
            2. DO NOT refer to the subject as a "character" - use appropriate terms like "cat", "dog", "animal", "creature", etc.
            3. Use neutral language like "the cat", "this dog", "the animal", "the creature"
            4. Replace any existing gender references with neutral terms
            5. Animal body parts (eyes, ears, mouth, paws, tail, etc.) are acceptable and should be preserved
            6. Focus on animal-specific characteristics and behaviors
            
            GENERAL RULES:
            7. Integrate the new details naturally into the prompt without making it too long.
            8. Keep the essential elements of the original prompt but add or modify the {modification_type.lower()} related details.
            9. Any objects mentioned MUST BE appropriately sized and NEVER larger than the main subject.
            10. STRICTLY LIMIT objects to 1-2 maximum - DO NOT include more than 2 distinct objects.
            11. The main subject must always be the focal point and dominant element in the image.
            12. Objects should be proportionate and realistic in size compared to the subject.
            13. If the original prompt or new details include too many objects, prioritize only the 1-2 most important ones.
            
            Only output the final modified prompt, nothing else."""
            
            logger.info("🐾 PROMPT MODIFICATION: Using animal-friendly prompt modification (removing gender language, keeping body parts)")
        else:
            # For non-living subjects (objects, landscapes, etc.), remove both gender and body part references
            sys_prompt = f"""You are an assistant that modifies image generation prompts for NON-LIVING subjects (objects, items, landscapes, buildings, products).
            You need to combine the original prompt with new details about {modification_type.lower()}.
            
            CRITICAL NON-LIVING RULES:
            1. DO NOT use any gender-specific language like "A female character with", "A male character with", "he", "she", "him", "her"
            2. DO NOT refer to the subject as a "character", "animal", or "creature" - use appropriate terms like "object", "item", "product", "building", "landscape", etc.
            3. DO NOT include body part references like "eyes", "mouth", "ears", "face", "body", "hands", "feet", etc.
            4. DO NOT include expressions like "sparkling eyes", "gentle mouth", "curved mouth", or any facial expressions
            5. Use neutral, descriptive language appropriate for inanimate objects
            6. Focus on physical properties like color, texture, shape, material, size, etc.
            7. Replace "Full Body Shot" with "Full Frame Shot" for non-living subjects
            
            GENERAL RULES:
            8. Integrate the new details naturally into the prompt without making it too long.
            9. Keep the essential elements of the original prompt but add or modify the {modification_type.lower()} related details.
            10. Any additional objects mentioned MUST BE appropriately sized and NEVER larger than the main subject.
            11. STRICTLY LIMIT objects to 1-2 maximum - DO NOT include more than 2 distinct objects.
            12. The main subject must always be the focal point and dominant element in the image.
            13. Objects should be proportionate and realistic in size compared to the subject.
            14. If the original prompt or new details include too many objects, prioritize only the 1-2 most important ones.
            
            Only output the final modified prompt, nothing else."""
            
            logger.info("🏗️ PROMPT MODIFICATION: Using object/non-living-friendly prompt modification (removing gender language and body parts)")
        
        user_prompt = f"""Original prompt: {original_prompt}
        Modify to specify {modification_type.lower()}: {modification_details}"""
        
        # Call Qwen to intelligently combine the prompts
        modified = inference_with_api(None, user_prompt, sys_prompt=sys_prompt, model_id="qwen2.5-72b-instruct")
        
        # Post-process the result based on subject type
        if not is_human_subject and modified:
            original_modified = modified
            
            if is_animal_subject:
                # For animals, only remove gender language but keep body parts
                gender_phrases_to_remove = [
                    "A female character with", "A male character with", 
                    "female character", "male character",
                    "she has", "he has", "she is", "he is",
                    "her ", "his ", "him ", "she ", " he "
                ]
                
                # Replace with neutral alternatives
                if "female character" in modified.lower():
                    modified = re.sub(r'\bfemale character\b', 'creature', modified, flags=re.IGNORECASE)
                if "male character" in modified.lower():
                    modified = re.sub(r'\bmale character\b', 'creature', modified, flags=re.IGNORECASE)
                if "A female character with" in modified:
                    modified = modified.replace("A female character with", "A creature with")
                if "A male character with" in modified:
                    modified = modified.replace("A male character with", "A creature with")
                
                # Remove pronouns
                modified = re.sub(r'\bshe has\b', 'it has', modified, flags=re.IGNORECASE)
                modified = re.sub(r'\bhe has\b', 'it has', modified, flags=re.IGNORECASE)
                modified = re.sub(r'\bshe is\b', 'it is', modified, flags=re.IGNORECASE)
                modified = re.sub(r'\bhe is\b', 'it is', modified, flags=re.IGNORECASE)
                modified = re.sub(r'\bher\b', 'its', modified, flags=re.IGNORECASE)
                modified = re.sub(r'\bhis\b', 'its', modified, flags=re.IGNORECASE)
                modified = re.sub(r'\bhim\b', 'it', modified, flags=re.IGNORECASE)
                modified = re.sub(r'\bshe\b', 'it', modified, flags=re.IGNORECASE)
                modified = re.sub(r'\bhe\b', 'it', modified, flags=re.IGNORECASE)
                
                if original_modified != modified:
                    logger.info(f"🐾 PROMPT MODIFICATION: Removed gender language from animal prompt")
                    logger.info(f"🐾 BEFORE: {original_modified}")
                    logger.info(f"🐾 AFTER: {modified}")
            else:
                # For non-living objects, remove both gender language AND body part references
                gender_phrases_to_remove = [
                    "A female character with", "A male character with", 
                    "female character", "male character",
                    "she has", "he has", "she is", "he is",
                    "her ", "his ", "him ", "she ", " he "
                ]
                
                body_part_phrases_to_remove = [
                    "sparkling eyes", "gentle mouth", "curved mouth", "up-curved mouth",
                    "bright eyes", "beautiful eyes", "expressive eyes", "large eyes",
                    "small mouth", "wide mouth", "smiling mouth", "open mouth",
                    "pointed ears", "floppy ears", "small ears", "large ears",
                    "face", "facial", "body", "hands", "feet", "arms", "legs"
                ]
                
                # Replace with neutral alternatives
                if "female character" in modified.lower():
                    modified = re.sub(r'\bfemale character\b', 'object', modified, flags=re.IGNORECASE)
                if "male character" in modified.lower():
                    modified = re.sub(r'\bmale character\b', 'object', modified, flags=re.IGNORECASE)
                if "A female character with" in modified:
                    modified = modified.replace("A female character with", "An object with")
                if "A male character with" in modified:
                    modified = modified.replace("A male character with", "An object with")
                
                # Replace "Full Body Shot" with "Full Frame Shot"
                modified = re.sub(r'\bFull Body Shot\b', 'Full Frame Shot', modified, flags=re.IGNORECASE)
                
                # Remove pronouns
                modified = re.sub(r'\bshe has\b', 'it has', modified, flags=re.IGNORECASE)
                modified = re.sub(r'\bhe has\b', 'it has', modified, flags=re.IGNORECASE)
                modified = re.sub(r'\bshe is\b', 'it is', modified, flags=re.IGNORECASE)
                modified = re.sub(r'\bhe is\b', 'it is', modified, flags=re.IGNORECASE)
                modified = re.sub(r'\bher\b', 'its', modified, flags=re.IGNORECASE)
                modified = re.sub(r'\bhis\b', 'its', modified, flags=re.IGNORECASE)
                modified = re.sub(r'\bhim\b', 'it', modified, flags=re.IGNORECASE)
                modified = re.sub(r'\bshe\b', 'it', modified, flags=re.IGNORECASE)
                modified = re.sub(r'\bhe\b', 'it', modified, flags=re.IGNORECASE)
                
                # Remove body part references
                for body_phrase in body_part_phrases_to_remove:
                    modified = re.sub(r'\b' + re.escape(body_phrase) + r'\b', '', modified, flags=re.IGNORECASE)
                
                # Clean up any double spaces or commas left by removals
                modified = re.sub(r'\s+', ' ', modified)  # Multiple spaces to single space
                modified = re.sub(r',\s*,', ',', modified)  # Double commas
                modified = re.sub(r',\s*$', '', modified)  # Trailing comma
                modified = re.sub(r'^\s*,', '', modified)  # Leading comma
                modified = modified.strip()
                
                if original_modified != modified:
                    logger.info(f"🏗️ PROMPT MODIFICATION: Removed gender language and body parts from object prompt")
                    logger.info(f"🏗️ BEFORE: {original_modified}")
                    logger.info(f"🏗️ AFTER: {modified}")
        
        # Fallback if API fails
        if not modified or modified.strip() == "":
            if modification_type == "Eye":
                modified = f"{original_prompt}, with {modification_details} eyes"
            elif modification_type == "Fur Color":
                modified = f"{original_prompt}, with {modification_details} fur"
        
        # Clean the modified prompt to ensure it's a single continuous paragraph
        if modified:
            modified = clean_modified_prompt(modified)
        
        return modified
    except Exception as e:
        logger.error(f"Error modifying prompt: {str(e)}")
        # Safe fallback that handles None values
        if modification_type is None:
            return original_prompt
        return f"{original_prompt}, with {modification_details} {modification_type.lower()}"
        
def update_modified_prompt(original_prompt, modification_type, modification_details, reference_image_path=None):
    # Handle None values
    if modification_type is None or modification_details is None:
        return ""
    if modification_type == "None" or not modification_details.strip():
        return ""
    return modify_prompt(original_prompt, modification_type, modification_details, reference_image_path)

# Helper function for ordinal suffixes
def get_ordinal_suffix(num):
    if 10 <= num % 100 <= 20:
        suffix = 'th'
    else:
        suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(num % 10, 'th')
    return suffix

# Add a function to get all categories from all themes
def get_all_categories():
    """Get a flattened list of all categories from all themes"""
    all_categories = []
    for theme, categories in THEME_CATEGORIES.items():
        all_categories.extend(categories)
    return sorted(all_categories)

# Add a function to check if negative prompt is supported for selected Ideogram model
def is_negative_prompt_supported(model_id):
    """Check if the selected Ideogram model supports negative prompts"""
    # Only V_2 and V_2_TURBO support negative prompts
    supported_models = ["V_2", "V_2_TURBO"]
    return model_id in supported_models

# Add a function to toggle rendering speed visibility for Ideogram V3
def toggle_ideogram_rendering_speed(ideogram_model):
    """Show or hide rendering speed dropdown based on Ideogram model selection"""
    # Get the API model ID
    model_id = IDEOGRAM_MODELS.get(ideogram_model, "")
    
    if model_id == "V_3":
        # Show rendering speed dropdown for V3 model
        return gr.update(visible=True)
    else:
        # Hide rendering speed dropdown for non-V3 models
        return gr.update(visible=False)

# Function to process ZIP file with multiple images for background removal
async def process_zip_with_bg_removal(zip_path, card_template_path=None, bg_method='birefnet_hr', should_remove_watermark=False):
    """Process all images in a ZIP file with background removal and generate two outputs per image: white background and card template"""
    try:
        from PIL import Image
        
        logger.info(f"Processing ZIP file with background removal method: {bg_method}")
        logger.info(f"Remove watermarks: {should_remove_watermark}")
        
        # Extract images from ZIP
        extracted_images, temp_dir = extract_images_from_zip(zip_path)
        if not extracted_images:
            return [], "No valid images found in ZIP file", None
        
        # Create output directory
        output_dir = os.path.join("generated_output", "removed_backgrounds")
        os.makedirs(output_dir, exist_ok=True)
        
        processed_images = []
        processed_paths = []
        
        # Process each image
        for i, image_path in enumerate(extracted_images):
            try:
                logger.info(f"Processing image {i+1}/{len(extracted_images)}: {os.path.basename(image_path)}")
                
                # Remove background
                if bg_method == 'birefnet_hr':
                    image_no_bg = remove_background_birefnet_hr(image_path)
                    logger.info(f"Successfully removed background using Birefnet HR for: {os.path.basename(image_path)}")
                elif bg_method == 'photoroom':
                    image_no_bg = remove_background_photoroom(image_path)
                    logger.info(f"Successfully removed background using PhotoRoom API for: {os.path.basename(image_path)}")
                else:
                    logger.error(f"Invalid background removal method: {bg_method}")
                    continue
                
                if image_no_bg is None:
                    logger.error(f"Failed to remove background for: {os.path.basename(image_path)}")
                    continue
                
                # Get original image dimensions
                original_img = Image.open(image_path)
                original_width, original_height = original_img.size
                
                filename = os.path.basename(image_path)
                base_name, _ = os.path.splitext(filename)
                
                # Remove '_original' from base_name if present
                if '_original' in base_name:
                    base_name = base_name.replace('_original', '')
                    logger.info(f"Removed '_original' from base filename: {base_name}")
                
                # Generate Output 1: White background version
                logger.info(f"Creating white background version for: {os.path.basename(image_path)}")
                white_bg_img = Image.new('RGBA', (original_width, original_height), (255, 255, 255, 255))
                
                # Paste the transparent image onto white background
                white_bg_img.paste(image_no_bg, (0, 0), image_no_bg)
                
                # Convert to RGB to remove alpha channel and ensure white background
                white_bg_final = Image.new('RGB', (original_width, original_height), (255, 255, 255))
                white_bg_final.paste(white_bg_img, (0, 0), white_bg_img)
                
                # Save white background version with EXACT same name as input
                white_bg_path = os.path.join(output_dir, f"{base_name}.png")
                white_bg_final.save(white_bg_path, format='PNG', optimize=True)
                logger.info(f"Saved white background image: {white_bg_path}")
                
                processed_images.append(white_bg_final)
                processed_paths.append(white_bg_path)
                
                # Generate Output 2: Apply to card template if provided
                if card_template_path and os.path.exists(card_template_path):
                    try:
                        logger.info(f"Applying to card template: {card_template_path}")
                        card_template = Image.open(card_template_path).convert("RGBA")
                        
                        # Use the transparent image for card template
                        transparent_img = image_no_bg.convert("RGBA")
                        
                        # Save the image to a temporary file to ensure full alpha preservation
                        with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp:
                            transparent_img.save(tmp.name, 'PNG')
                            # Load it back with full alpha channel data
                            transparent_img = Image.open(tmp.name).convert('RGBA')
                            
                            # Place the image on the card
                            # Use preserve_original_alpha=True to maintain background removal quality
                            result_with_card = place_image_on_card(card_template, transparent_img, preserve_original_alpha=True)
                            
                            # Clean up temp file
                            try:
                                os.unlink(tmp.name)
                            except:
                                pass
                        
                        # Process the image to remove any watermarks if requested
                        if should_remove_watermark:
                            logger.info("Removing watermarks as requested")
                            result_with_card = remove_watermark(result_with_card, is_photoroom=(bg_method == 'photoroom'))
                        
                        # Save the card template result
                        card_path = os.path.join(output_dir, f"{base_name}_card.png")
                        result_with_card.save(card_path, format='PNG', optimize=True)
                        logger.info(f"Saved image with card template: {card_path}")
                        
                        processed_images.append(result_with_card)
                        processed_paths.append(card_path)
                        
                    except Exception as e:
                        logger.error(f"Error applying to card template for {os.path.basename(image_path)}: {str(e)}")
                        # Continue without card template if application fails
                else:
                    logger.info(f"No card template provided for: {os.path.basename(image_path)}")
                
            except Exception as e:
                logger.error(f"Error processing image {os.path.basename(image_path)}: {str(e)}")
                continue
        
        if not processed_images:
            return [], "Failed to process any images from ZIP file", None
        
        # Create a ZIP file with all processed images
        zip_filename = f"processed_bg_removal_{int(time.time())}.zip"
        zip_filepath = os.path.join(output_dir, zip_filename)
        
        with zipfile.ZipFile(zip_filepath, 'w', compression=zipfile.ZIP_DEFLATED) as zipf:
            for path in processed_paths:
                if os.path.exists(path):
                    arcname = os.path.basename(path)
                    zipf.write(path, arcname)
                    logger.info(f"Added {arcname} to ZIP file")
        
        # Clean up temp directory
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)
        
        total_inputs = len(extracted_images)
        total_outputs = len(processed_images)
        success_msg = f"Successfully processed {total_inputs} input images, generated {total_outputs} output images from ZIP file"
        logger.info(success_msg)
        
        return processed_paths, success_msg, zip_filepath
        
    except Exception as e:
        logger.error(f"Error in process_zip_with_bg_removal: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return [], f"Error processing ZIP file: {str(e)}", None

# Function to process an image directly with Birefnet and apply to card template
async def process_image_with_birefnet(image_path, card_template_path=None, bg_method='birefnet_hr', should_remove_watermark=False):
    """Remove background from image using Birefnet HR or PhotoRoom and generate two outputs: white background and card template"""
    try:
        from PIL import Image, ImageDraw, ImageFont
        
        logger.info(f"Processing image with background removal method: {bg_method}")
        logger.info(f"Remove watermarks: {should_remove_watermark}")
        
        # Handle different input types for image_path
        if isinstance(image_path, tuple):
            if len(image_path) > 0 and isinstance(image_path[0], str):
                image_path = image_path[0]
                logger.info(f"Extracted image path from tuple: {image_path}")
        
        if isinstance(image_path, dict) and 'name' in image_path:
            image_path = image_path['name']
            logger.info(f"Extracted image path from dictionary: {image_path}")
            
        # Handle different input types for card_template_path
        if card_template_path:
            if isinstance(card_template_path, tuple):
                if len(card_template_path) > 0 and isinstance(card_template_path[0], str):
                    card_template_path = card_template_path[0]
                    logger.info(f"Extracted card template path from tuple: {card_template_path}")
                else:
                    logger.warning(f"Cannot extract valid card template path from tuple: {card_template_path}")
                    card_template_path = None
            
            if isinstance(card_template_path, dict) and 'name' in card_template_path:
                card_template_path = card_template_path['name']
                logger.info(f"Extracted card template path from dictionary: {card_template_path}")
        
        # Step 1: Remove background with Birefnet
        if not os.path.exists(image_path):
            return [], f"Error: Image file not found: {image_path}"

        if bg_method == 'birefnet_hr':
            image_no_bg = remove_background_birefnet_hr(image_path)
            logger.info(f"Successfully removed background using Birefnet HR for: {image_path}")
            print(f"[Success] Background removed using Birefnet HR for: {os.path.basename(image_path)}")
        elif bg_method == 'photoroom':
            image_no_bg = remove_background_photoroom(image_path)
            logger.info(f"Successfully removed background using PhotoRoom API for: {image_path}")
            print(f"[Success] Background removed using PhotoRoom API for: {os.path.basename(image_path)}")
        else:
            return [], f"Error: Invalid background removal method: {bg_method}"

        if image_no_bg is None:
            return [], "Error: Failed to remove background"
        
        # Create output directory
        output_dir = os.path.join("generated_output", "removed_backgrounds")
        os.makedirs(output_dir, exist_ok=True)

        # Get original image dimensions
        original_img = Image.open(image_path)
        original_width, original_height = original_img.size
        
        filename = os.path.basename(image_path)
        base_name, _ = os.path.splitext(filename)
        
        # Remove '_original' from base_name if present
        if '_original' in base_name:
            base_name = base_name.replace('_original', '')
            logger.info(f"Removed '_original' from base filename: {base_name}")
        
        output_paths = []
        
        # Step 2: Create white background version (First Output)
        logger.info("Creating white background version...")
        white_bg_img = Image.new('RGBA', (original_width, original_height), (255, 255, 255, 255))
        
        # Paste the transparent image onto white background
        white_bg_img.paste(image_no_bg, (0, 0), image_no_bg)
        
        # Convert to RGB to remove alpha channel and ensure white background
        white_bg_final = Image.new('RGB', (original_width, original_height), (255, 255, 255))
        white_bg_final.paste(white_bg_img, (0, 0), white_bg_img)
        
        # Save white background version with EXACT same name as input
        white_bg_path = os.path.join(output_dir, f"{base_name}.png")
        white_bg_final.save(white_bg_path, format='PNG', optimize=True)
        logger.info(f"Saved image with white background: {white_bg_path}")
        output_paths.append(white_bg_path)
        
        # Step 3: Apply to card template if provided (Second Output)
        if card_template_path and os.path.exists(card_template_path):
            try:
                logger.info(f"Applying to card template: {card_template_path}")
                card_template = Image.open(card_template_path).convert("RGBA")
                
                # Use the transparent image for card template
                transparent_img = image_no_bg.convert("RGBA")
                
                # Save the image to a temporary file to ensure full alpha preservation
                with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp:
                    transparent_img.save(tmp.name, 'PNG')
                    # Load it back with full alpha channel data
                    transparent_img = Image.open(tmp.name).convert('RGBA')
                    
                    # Place the image on the card
                    # Use preserve_original_alpha=True to maintain background removal quality
                    result_with_card = place_image_on_card(card_template, transparent_img, preserve_original_alpha=True)
                    
                    # Clean up temp file
                    try:
                        os.unlink(tmp.name)
                    except:
                        pass

                # Process the image to remove any watermarks if requested
                if should_remove_watermark:
                    logger.info("Removing watermarks as requested")
                    result_with_card = remove_watermark(result_with_card, is_photoroom=(bg_method == 'photoroom'))
                else:
                    logger.info("Watermark removal not requested - skipping")
                
                # Save the card template result
                card_path = os.path.join(output_dir, f"{base_name}_card.png")
                result_with_card.save(card_path, format='PNG', optimize=True)
                logger.info(f"Saved image with card template: {card_path}")
                output_paths.append(card_path)
                
                # Add success logs for card template application
                success_message = f"Successfully applied {os.path.basename(image_path)} to card template using {bg_method} method"
                logger.info(success_message)
                print(f"[Success] {success_message}")
                
            except Exception as e:
                logger.error(f"Error applying to card template: {str(e)}")
                import traceback
                logger.error(traceback.format_exc())
                # Continue without card template if application fails
        else:
            logger.info("No card template provided - only generating white background version")

        success_msg = f"Successfully processed image: {os.path.basename(image_path)} - Generated {len(output_paths)} output(s)"
        return output_paths, success_msg
    except Exception as e:
        logger.error(f"Error in process_image_with_birefnet: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return [], f"Error processing image: {str(e)}"

def get_aws_credentials():
    """Read AWS credentials from the credentials file"""
    config = configparser.ConfigParser()
    try:
        config.read('aws_credentials.txt')
        return {
            'bucket_name': config.get('AWS', 'bucket_name'),
            'region': config.get('AWS', 'region'),
            'access_key': config.get('AWS', 'access_key'),
            'secret_key': config.get('AWS', 'secret_key')
        }
    except Exception as e:
        logger.error(f"Error reading AWS credentials: {str(e)}")
        return None

def upload_to_s3(file_path, s3_object_name=None, bucket_folder=None):
    """Upload a file to an S3 bucket
    
    Args:
        file_path (str): Path to the local file
        s3_object_name (str): Name to give the file in S3. If not specified, uses the filename from file_path
        bucket_folder (str): Optional folder within the S3 bucket (e.g., 'images/', 'outputs/')
        
    Returns:
        str: S3 URL if upload was successful, None otherwise
    """
    # Validate the file_path
    if not file_path:
        logger.error("No file path provided for S3 upload")
        return None
        
    # Check file existence early
    if not os.path.exists(file_path):
        logger.error(f"File not found for S3 upload: {file_path}")
        return None
        
    # Skip Excel files explicitly
    if file_path.lower().endswith('.xlsx'):
        logger.info(f"Skipping upload of Excel file: {file_path}")
        return None
        
    # Log file details to aid debugging
    logger.info(f"Preparing to upload file: {file_path}")
    logger.info(f"File size: {os.path.getsize(file_path)} bytes")
    
    # Get AWS credentials
    credentials = get_aws_credentials()
    if not credentials:
        logger.error("Failed to get AWS credentials")
        return None
        
    # Validate credentials
    required_keys = ['bucket_name', 'region', 'access_key', 'secret_key']
    for key in required_keys:
        if not credentials.get(key):
            logger.error(f"Missing required AWS credential: {key}")
            return None
    
    # If S3 object name not specified, use local file name
    if s3_object_name is None:
        s3_object_name = Path(file_path).name
    
    # If bucket folder is specified, prepend it to the object name
    if bucket_folder:
        if not bucket_folder.endswith('/'):
            bucket_folder += '/'
        s3_object_name = f"{bucket_folder}{s3_object_name}"
    
    # Create S3 client
    try:
        logger.info(f"Creating S3 client for region {credentials['region']}")
        s3_client = boto3.client(
            service_name='s3',
            region_name=credentials['region'],
            aws_access_key_id=credentials['access_key'],
            aws_secret_access_key=credentials['secret_key']
        )
        
        logger.info(f"Uploading {file_path} to S3 bucket {credentials['bucket_name']} as {s3_object_name}")
        
        # Try to upload with a timeout
        s3_client.upload_file(
            Filename=file_path,
            Bucket=credentials['bucket_name'],
            Key=s3_object_name
        )
        
        # Verify the upload by checking if the object exists
        try:
            s3_client.head_object(Bucket=credentials['bucket_name'], Key=s3_object_name)
            logger.info(f"Verified object exists in S3: {s3_object_name}")
        except Exception as e:
            logger.warning(f"Could not verify object in S3: {str(e)}")
            # Continue execution even after warning
        
        # Generate S3 URL
        s3_url = f"https://{credentials['bucket_name']}.s3.{credentials['region']}.amazonaws.com/{s3_object_name}"
        logger.info(f"File uploaded successfully to {s3_url}")
        return s3_url
    except Exception as e:
        import traceback
        logger.error(f"Error uploading to S3: {str(e)}")
        logger.error(traceback.format_exc())
        # Don't stop execution after error
        return None

def upload_multiple_files_to_s3(file_paths, bucket_folder=None):
    """Upload multiple files to S3 bucket
    
    Args:
        file_paths (list): List of file paths to upload
        bucket_folder (str): Optional folder within the S3 bucket
        
    Returns:
        list: List of S3 URLs for successfully uploaded files
    """
    uploaded_urls = []
    
    for file_path in file_paths:
        url = upload_to_s3(file_path, bucket_folder=bucket_folder)
        if url:
            uploaded_urls.append(url)
    
    return uploaded_urls

def upload_zip_to_s3(zip_path, theme=None, category=None):
    """Upload a ZIP file to S3 with folder structure based on theme/category"""
    # Get AWS credentials
    credentials = get_aws_credentials()
    if not credentials:
        logger.warning("Failed to get AWS credentials. ZIP upload skipped.")
        return None
        
    # Create folder path with theme/category if provided
    folder = "outputs"
    
    if theme:
        # Convert theme to string if needed
        theme_str = theme if isinstance(theme, str) else str(theme)
        folder = f"{folder}/{theme_str.lower()}"
        if category:
            # Convert category to string if needed
            category_str = category if isinstance(category, str) else str(category)
            folder = f"{folder}/{category_str.lower()}"
    
    # Generate timestamp for unique filename
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    
    # Create S3 object name
    zip_filename = os.path.basename(zip_path)
    s3_object_name = f"{timestamp}_{zip_filename}"
    
    return upload_to_s3(zip_path, s3_object_name=s3_object_name, bucket_folder=folder)

def create_zip_file(image_paths, theme=None, category=None):
    """
    Create a ZIP file from image paths for download only
    
    Args:
        image_paths (list): List of image file paths to include in the ZIP
        theme (str): Theme name for proper naming convention
        category (str): Category name for proper naming convention
        
    Returns:
        str: Path to the created ZIP file or None if failed
    """
    import tempfile
    import zipfile
    
    if not image_paths:
        logger.warning("No images provided to create ZIP file")
        return None
        
    # Create timestamp for unique filename
    timestamp = get_gmt7_filename_timestamp()
    
    # Create proper ZIP filename using theme/category codes
    if theme and category:
        theme_code = THEME_MAPPING.get(theme, "00")
        category_code = CATEGORY_MAPPING.get(category, "000")
        zip_name = f"{theme_code}{category_code}_{timestamp}.zip"
    else:
        # Fallback to generic name if theme/category not provided
        zip_name = f"generated_images_{timestamp}.zip"
    
    # Create ZIP in temp directory
    temp_dir = tempfile.mkdtemp()
    zip_filepath = os.path.join(temp_dir, zip_name)
    
    try:
        # Create ZIP file
        file_counter = 1  # Counter for sequential numbering
        
        with zipfile.ZipFile(zip_filepath, 'w', compression=zipfile.ZIP_DEFLATED) as zipf:
            for img_path in image_paths:
                # Handle different types of image paths/objects
                actual_path = None
                
                # Check if it's a string path
                if isinstance(img_path, str):
                    actual_path = img_path
                # Check if it's a PIL Image object
                elif hasattr(img_path, 'save'):
                    # It's a PIL Image, save it to temporary file first
                    import tempfile
                    from PIL import Image
                    
                    temp_dir = tempfile.mkdtemp()
                    temp_filename = f"temp_image_{len(os.listdir(temp_dir))}.png"
                    actual_path = os.path.join(temp_dir, temp_filename)
                    
                    try:
                        img_path.save(actual_path)
                        logger.info(f"Saved PIL Image to temporary file: {actual_path}")
                    except Exception as save_error:
                        logger.error(f"Error saving PIL Image to file: {str(save_error)}")
                        continue
                # Check if it's a tuple/dict format
                elif isinstance(img_path, (tuple, list)) and len(img_path) > 0:
                    actual_path = img_path[0] if isinstance(img_path[0], str) else None
                elif isinstance(img_path, dict) and 'path' in img_path:
                    actual_path = img_path['path']
                else:
                    logger.warning(f"Unsupported image path format in ZIP creation: {type(img_path)}")
                    continue
                
                if actual_path and os.path.exists(actual_path):
                    # Get file extension
                    file_ext = os.path.splitext(actual_path)[1].lower()
                    if not file_ext:
                        file_ext = '.png'  # Default to PNG if no extension
                    
                    # Create proper filename using TTCCCNNNNN convention
                    if theme and category:
                        theme_code = THEME_MAPPING.get(theme, "00")
                        category_code = CATEGORY_MAPPING.get(category, "000")
                        # Format: TTCCCNNNNN.ext (e.g., 0100200001.png for Pets/Cats)
                        proper_filename = f"{theme_code}{category_code}{file_counter:05d}{file_ext}"
                    else:
                        # Fallback to generic naming if theme/category not provided
                        proper_filename = f"image_{file_counter:03d}{file_ext}"
                    
                    # Add file to ZIP with proper filename
                    zipf.write(actual_path, proper_filename)
                    logger.info(f"Added to ZIP with proper naming: {proper_filename}")
                    file_counter += 1
                else:
                    logger.warning(f"File not found for ZIP: {actual_path} (original: {type(img_path)})")
        
        logger.info(f"Created ZIP file: {zip_filepath}")
        return zip_filepath
        
    except Exception as e:
        logger.error(f"Error creating ZIP file: {str(e)}")
        return None

def generate_comprehensive_outputs_for_prompt_variations(
    reference_image, all_prompts, provider, card_template, theme, category, subcategory,
    leo_model, guidance_scale, neg_p, preset, leo_num_img,
    ideogram_model, ideogram_style, ideogram_num_img, output_f,
    filename_convention, safe_upload_to_s3, validated_seed, act_param, exp_param, 
    fur_param, ethnicity, safe_stop_flag, safe_upload_to_gdrive,
    safe_encode_to_base64, modification_type, has_activity, has_expression, has_fur_color, 
    has_ethnicity, reference_image_1, ref_type_1, ref_strength_1, reference_image_2, 
    ref_type_2, ref_strength_2, reference_image_3, ref_type_3, ref_strength_3,
    ideogram_disable_style_reference, ideogram_rendering_speed, modified_prompt,
    imagen4_model, imagen4_aspect_ratio, imagen4_safety_filter, imagen4_num_images,
    counter_override=None
):
    """
    Generate comprehensive outputs for multiple prompt variations using a single reference image
    This processes ALL prompt variations automatically when Generate with Activity is clicked
    """
    try:
        logger.info(f"🎯 PROMPT VARIATION BATCH: Starting comprehensive output generation for {len(all_prompts)} prompt variations")
        
        # Initialize collections for all output types
        all_generated_images = []
        all_nobg_images = []
        all_card_applied_images = []
        all_base64_files = []
        all_display_images = []
        excel_data = []
        successful_batches = 0
        
        # Create output directories
        timestamp = get_gmt7_filename_timestamp()
        output_dir = os.path.join("generated_output", "prompt_variations", f"{timestamp}")
        os.makedirs(output_dir, exist_ok=True)
        logger.info(f"📁 Created output directory: {output_dir}")

        # Process each prompt variation
        for variation_index, current_prompt in enumerate(all_prompts):
            try:
                logger.info(f"🎯 VARIATION {variation_index+1}/{len(all_prompts)}: Processing prompt variation")
                logger.info(f"🎯 VARIATION {variation_index+1}: Prompt: {current_prompt[:100]}...")
                
                # For prompt variations, we DON'T pass modified_prompt to ensure generated_prompt (the actual variation) is used
                # This fixes the issue where all 3 variations were using the same prompt instead of their unique variations
                
                # Step 1: Generate default images from provider for this variation
                logger.info(f"📸 VARIATION {variation_index+1}: Generating images with prompt variation")
                variation_results = sync_upload_and_generate_image(
                    provider=provider,
                    reference_images=reference_image,
                    card_template=card_template,
                    theme=theme,
                    category=category,
                    subcategory=subcategory,  # Pass actual subcategory value
                    model_name=leo_model,
                    width=1024,
                    height=1024,
                    guidance_scale=guidance_scale,
                    generated_prompt=current_prompt,  # Use the specific variation prompt
                    negative_prompt=neg_p,
                    preset_style=preset,
                    num_images=leo_num_img,
                    ideogram_model=ideogram_model,
                    ideogram_style=ideogram_style,
                    ideogram_num_images=ideogram_num_img,
                    output_format=output_f,
                    filename_convention=filename_convention,
                    upload_to_s3_bucket=safe_upload_to_s3,
                    seed=validated_seed,
                    activity=act_param,
                    facial_expression=exp_param,
                    fur_color=fur_param,
                    ethnicity=ethnicity,
                    stop_flag=safe_stop_flag,
                    upload_to_gdrive=safe_upload_to_gdrive,
                    encode_to_base64=False,  # We'll handle base64 separately
                    generation_type="activity" if (modification_type == "activity" or (has_activity or has_expression or has_fur_color or has_ethnicity)) else "standard",
                    reference_image_1=reference_image_1, ref_type_1=ref_type_1, ref_strength_1=ref_strength_1,
                    reference_image_2=reference_image_2, ref_type_2=ref_type_2, ref_strength_2=ref_strength_2,
                    reference_image_3=reference_image_3, ref_type_3=ref_type_3, ref_strength_3=ref_strength_3,
                    ideogram_disable_style_reference=ideogram_disable_style_reference,
                    ideogram_rendering_speed=ideogram_rendering_speed,
                    # NOTE: modified_prompt is intentionally NOT passed here to ensure generated_prompt is used
                    # Imagen-4 parameters
                    imagen4_model=imagen4_model,
                    imagen4_aspect_ratio=imagen4_aspect_ratio,
                    imagen4_safety_filter=imagen4_safety_filter,
                    imagen4_num_images=imagen4_num_images,
                    # Counter override parameter
                    counter_override=counter_override
                )
                
                if variation_results and len(variation_results) >= 4:
                    variation_images = variation_results[0] if variation_results[0] else []
                    
                    # Collect the generated images
                    if variation_images:
                        all_generated_images.extend(variation_images)
                        all_display_images.extend(variation_images)
                        successful_batches += 1
                        logger.info(f"✅ VARIATION {variation_index+1}: Generated {len(variation_images)} images")
                        
                        # Add Excel data for each generated image
                        for img_path in variation_images:
                            if isinstance(img_path, str) and os.path.exists(img_path):
                                excel_data.append({
                                    'reference_image': os.path.basename(reference_image) if isinstance(reference_image, str) else f"variation_{variation_index+1}",
                                    'generated_image': os.path.basename(img_path),
                                    'prompt': current_prompt,  # This is the actual variation prompt
                                    'modified_prompt': "",  # For variations, the prompt IS the variation, not a modification
                                    'provider': provider,
                                    'model': leo_model if provider == "Leonardo" else ideogram_model if provider == "Ideogram" else imagen4_model,
                                    'theme': theme,
                                    'category': category,
                                    'activity': act_param if act_param else "",
                                    'facial_expression': exp_param if exp_param else "",
                                    'fur_color': fur_param if fur_param else "",
                                    'ethnicity': ethnicity if ethnicity and ethnicity != "Auto" else "",
                                    'seed': validated_seed if validated_seed else "",
                                    'guidance_scale': guidance_scale,
                                    'timestamp': timestamp,
                                    'variation_index': variation_index + 1
                                })
                    else:
                        logger.warning(f"⚠️ VARIATION {variation_index+1}: No images generated")
                else:
                    logger.warning(f"⚠️ VARIATION {variation_index+1}: Invalid generation results")
                    
            except Exception as variation_error:
                logger.error(f"❌ VARIATION {variation_index+1}: Error - {str(variation_error)}")
                continue
        
        # Create comprehensive ZIP with all variations
        comprehensive_zip_path = None
        if all_generated_images:
            try:
                # Collect all files for ZIP
                all_files_for_zip = []
                all_files_for_zip.extend(all_generated_images)
                
                # Create Excel report if we have data
                excel_file_path = None
                if excel_data:
                    excel_file_path = create_comprehensive_excel_report(excel_data, theme, category)
                    if excel_file_path:
                        all_files_for_zip.append(excel_file_path)
                
                if all_files_for_zip:
                    comprehensive_zip_path = create_zip_file_with_metadata(
                        all_files_for_zip, 
                        metadata_files=None, 
                        encode_to_base64=False,
                        theme=theme,
                        category=category,
                        excel_path=excel_file_path,
                        original_images=all_generated_images,
                        nobg_images=all_nobg_images,
                        card_images=all_card_applied_images
                    )
                    logger.info(f"✅ Created comprehensive variations ZIP: {comprehensive_zip_path}")
            except Exception as zip_error:
                logger.error(f"❌ Error creating comprehensive variations ZIP: {str(zip_error)}")
        
        # Generate final status
        total_generated = len(all_generated_images)
        final_summary = f"🎯 PROMPT VARIATION BATCH PROCESSING COMPLETE!"
        final_stats = f"📊 RESULTS: {successful_batches}/{len(all_prompts)} prompt variations processed. Generated {total_generated} total images"
        
        if excel_data:
            final_stats += f", 1 Excel compilation"
        
        final_status = f"{final_summary}\n{final_stats}"
        
        logger.info(final_summary)
        logger.info(final_stats)
        
        return (
            all_display_images,      # For gallery display
            final_status,            # Status message
            comprehensive_zip_path,  # Download URL
            all_generated_images,    # For inpainting
            None,                    # No modified images for variation processing
            None                     # No modified ZIP for variation processing
        )
        
    except Exception as e:
        logger.error(f"❌ PROMPT VARIATION BATCH: Critical error: {str(e)}")
        return ([], f"Error in prompt variation batch processing: {str(e)}", None, [], None, None)

def generate_comprehensive_outputs_for_zip_batch(
    extracted_images, provider, card_template, theme, category, 
    leo_model, guidance_scale, generated_prompt, neg_p, preset, leo_num_img,
    ideogram_model, ideogram_style, ideogram_num_img, output_f,
    filename_convention, safe_upload_to_s3, validated_seed, act_param, exp_param, 
    fur_param, ethnicity, safe_stop_flag, safe_upload_to_gdrive,
    safe_encode_to_base64, modification_type, has_activity, has_expression, has_fur_color, 
    has_ethnicity, reference_image_1, ref_type_1, ref_strength_1, reference_image_2, 
    ref_type_2, ref_strength_2, reference_image_3, ref_type_3, ref_strength_3,
    ideogram_disable_style_reference, ideogram_rendering_speed, modified_prompt,
    all_prompts
):
    """
    Generate comprehensive outputs for multiple images in ZIP file including:
    - Default images from provider (non-removed BG)
    - Removed BG images
    - Applied images to card template
    - Base64 outputs
    - Excel file compilation
    """
    try:
        logger.info(f"🔄 COMPREHENSIVE ZIP BATCH: Starting comprehensive output generation for {len(extracted_images)} images")
        
        # Initialize collections for all output types
        all_generated_images = []
        all_nobg_images = []
        all_card_applied_images = []
        all_base64_files = []
        all_excel_data = []
        all_display_images = []
        all_download_urls = []
        batch_status_messages = []
        successful_batches = 0
        failed_batches = 0
        
        # Calculate expected total images
        images_per_reference = leo_num_img if provider == "Leonardo" else ideogram_num_img
        total_expected_images = len(extracted_images) * images_per_reference
        
        logger.info(f"🔄 COMPREHENSIVE ZIP BATCH: Expected to generate {total_expected_images} total images ({images_per_reference} per reference)")
        
        for batch_index, img_path in enumerate(extracted_images):
            try:
                logger.info(f"🔄 COMPREHENSIVE ZIP BATCH: Processing image {batch_index+1}/{len(extracted_images)}: {os.path.basename(img_path)}")
                
                # Get the prompt for this specific image
                current_prompt = all_prompts[batch_index] if batch_index < len(all_prompts) else generated_prompt
                
                # Clean the modified prompt if it exists
                if modified_prompt:
                    current_modified_prompt = clean_modified_prompt(modified_prompt)
                else:
                    current_modified_prompt = modified_prompt
                
                # Step 1: Generate default images from provider
                logger.info(f"📸 STEP 1: Generating default images for {os.path.basename(img_path)}")
                batch_results = sync_upload_and_generate_image(
                    provider=provider,
                    reference_images=img_path,
                    card_template=card_template,
                    theme=theme,
                    category=category,
                    subcategory=subcategory,  # Add subcategory argument
                    model_name=leo_model,
                    width=1024,
                    height=1024,
                    guidance_scale=guidance_scale,
                    generated_prompt=current_prompt,
                    negative_prompt=neg_p,
                    preset_style=preset,
                    num_images=leo_num_img,
                    ideogram_model=ideogram_model,
                    ideogram_style=ideogram_style,
                    ideogram_num_images=ideogram_num_img,
                    output_format=output_f,
                    filename_convention=filename_convention,
                    upload_to_s3_bucket=safe_upload_to_s3,
                    seed=validated_seed,
                    activity=act_param,
                    facial_expression=exp_param,
                    fur_color=fur_param,
                    ethnicity=ethnicity,
                    stop_flag=safe_stop_flag,
                    upload_to_gdrive=safe_upload_to_gdrive,
                    encode_to_base64=False,  # We'll handle base64 separately
                    generation_type="activity" if (modification_type == "activity" or (has_activity or has_expression or has_fur_color or has_ethnicity)) else "standard",
                    reference_image_1=reference_image_1, ref_type_1=ref_type_1, ref_strength_1=ref_strength_1,
                    reference_image_2=reference_image_2, ref_type_2=ref_type_2, ref_strength_2=ref_strength_2,
                    reference_image_3=reference_image_3, ref_type_3=ref_type_3, ref_strength_3=ref_strength_3,
                    ideogram_disable_style_reference=ideogram_disable_style_reference,
                    ideogram_rendering_speed=ideogram_rendering_speed,
                    modified_prompt=current_modified_prompt,
                    # Imagen-4 parameters (default values since not passed from batch)
                    imagen4_model="google/imagen-4",
                    imagen4_aspect_ratio="1:1",
                    imagen4_safety_filter="block_only_high",
                    imagen4_num_images=1,
                    # Counter override parameter
                    counter_override=counter_override
                )
                
                if not batch_results or len(batch_results) < 6:
                    failed_batches += 1
                    logger.error(f"❌ COMPREHENSIVE ZIP BATCH: Failed to generate images for {os.path.basename(img_path)}")
                    continue
                
                # Extract all returned values from batch_results
                batch_images, batch_status, batch_download_url, batch_original_images, batch_card_images, batch_other = batch_results[:6]
                
                if not batch_images and not batch_original_images:
                    failed_batches += 1
                    logger.error(f"❌ COMPREHENSIVE ZIP BATCH: No images generated for {os.path.basename(img_path)}")
                    continue
                
                # Step 2: Process original images (these are the generated images with background removed)
                if batch_original_images:
                    for img_idx, original_img in enumerate(batch_original_images):
                        try:
                            # Extract image path from different formats
                            if isinstance(original_img, str):
                                img_file_path = original_img
                            elif isinstance(original_img, tuple) and len(original_img) > 0:
                                img_file_path = original_img[0]
                            elif isinstance(original_img, dict) and 'path' in original_img:
                                img_file_path = original_img['path']
                            else:
                                logger.warning(f"⚠️ COMPREHENSIVE ZIP BATCH: Unsupported original image format: {type(original_img)}")
                                continue
                            
                            if not os.path.exists(img_file_path):
                                logger.warning(f"⚠️ COMPREHENSIVE ZIP BATCH: Original image file not found: {img_file_path}")
                                continue
                            
                            logger.info(f"🔧 STEP 2A: Processing original image {img_idx+1} (background removed)")
                            
                            # Add to original images collection (these are already background-removed)
                            all_generated_images.append(img_file_path)
                            all_display_images.append(original_img)  # Keep original format for display
                            
                            # Also add to no-bg images since they're already background-removed
                            all_nobg_images.append(img_file_path)
                            
                        except Exception as img_process_error:
                            logger.error(f"❌ COMPREHENSIVE ZIP BATCH: Error processing original image {img_idx+1}: {str(img_process_error)}")
                            continue
                
                # Step 2B: Process card images if available
                if batch_card_images:
                    for img_idx, card_img in enumerate(batch_card_images):
                        try:
                            # Extract image path from different formats
                            if isinstance(card_img, str):
                                card_file_path = card_img
                            elif isinstance(card_img, tuple) and len(card_img) > 0:
                                card_file_path = card_img[0]
                            elif isinstance(card_img, dict) and 'path' in card_img:
                                card_file_path = card_img['path']
                            else:
                                logger.warning(f"⚠️ COMPREHENSIVE ZIP BATCH: Unsupported card image format: {type(card_img)}")
                                continue
                            
                            if not os.path.exists(card_file_path):
                                logger.warning(f"⚠️ COMPREHENSIVE ZIP BATCH: Card image file not found: {card_file_path}")
                                continue
                            
                            logger.info(f"🔧 STEP 2B: Processing card image {img_idx+1}")
                            
                            # Add to card images collection
                            all_card_applied_images.append(card_file_path)
                            
                        except Exception as card_process_error:
                            logger.error(f"❌ COMPREHENSIVE ZIP BATCH: Error processing card image {img_idx+1}: {str(card_process_error)}")
                            continue
                
                # Step 2C: Process regular generated images (fallback if no original/card images)
                if not batch_original_images and batch_images:
                    for img_idx, generated_img_path in enumerate(batch_images):
                        try:
                            # Extract image path from different formats
                            if isinstance(generated_img_path, str):
                                img_file_path = generated_img_path
                            elif isinstance(generated_img_path, tuple) and len(generated_img_path) > 0:
                                img_file_path = generated_img_path[0]
                            elif isinstance(generated_img_path, dict) and 'path' in generated_img_path:
                                img_file_path = generated_img_path['path']
                            else:
                                logger.warning(f"⚠️ COMPREHENSIVE ZIP BATCH: Unsupported image format: {type(generated_img_path)}")
                                continue
                            
                            if not os.path.exists(img_file_path):
                                logger.warning(f"⚠️ COMPREHENSIVE ZIP BATCH: Generated image file not found: {img_file_path}")
                                continue
                            
                            logger.info(f"🔧 STEP 2C: Processing generated image {img_idx+1} for comprehensive outputs")
                            
                            # Add to default images collection
                            all_generated_images.append(img_file_path)
                            all_display_images.append(generated_img_path)  # Keep original format for display
                            
                            # 2d: Generate base64 if requested
                            if safe_encode_to_base64:
                                logger.info(f"📝 STEP 2D: Generating base64 for {os.path.basename(img_file_path)}")
                                try:
                                    base64_file = encode_image_to_base64_file(img_file_path)
                                    if base64_file:
                                        all_base64_files.append(base64_file)
                                        logger.info(f"✅ STEP 2D: Generated base64 file: {base64_file}")
                                except Exception as base64_error:
                                    logger.error(f"❌ STEP 2D: Error generating base64: {str(base64_error)}")
                            
                            # 2e: Collect data for Excel compilation
                            excel_row_data = {
                                'reference_image': os.path.basename(img_path),
                                'generated_image': os.path.basename(img_file_path),
                                'prompt': current_prompt,
                                'modified_prompt': current_modified_prompt if current_modified_prompt else '',
                                'provider': provider,
                                'model': leo_model if provider == "Leonardo" else ideogram_model,
                                'theme': theme,
                                'category': category,
                                'activity': act_param if act_param else '',
                                'facial_expression': exp_param if exp_param else '',
                                'fur_color': fur_param if fur_param else '',
                                'ethnicity': ethnicity if ethnicity else '',
                                'seed': validated_seed if validated_seed else '',
                                'guidance_scale': guidance_scale,
                                'timestamp': get_gmt7_timestamp()
                            }
                            all_excel_data.append(excel_row_data)
                            
                        except Exception as img_process_error:
                            logger.error(f"❌ COMPREHENSIVE ZIP BATCH: Error processing generated image {img_idx+1}: {str(img_process_error)}")
                            continue
                
                successful_batches += 1
                logger.info(f"✅ COMPREHENSIVE ZIP BATCH: Successfully processed {os.path.basename(img_path)}")
                
            except Exception as batch_error:
                failed_batches += 1
                logger.error(f"❌ COMPREHENSIVE ZIP BATCH: Error processing {os.path.basename(img_path)}: {str(batch_error)}")
                continue
        
        # Step 3: Generate Excel compilation
        excel_file_path = None
        if all_excel_data:
            logger.info(f"📊 STEP 3: Generating Excel compilation with {len(all_excel_data)} entries")
            try:
                excel_file_path = create_comprehensive_excel_report(all_excel_data, theme, category)
                logger.info(f"✅ STEP 3: Generated Excel file: {excel_file_path}")
            except Exception as excel_error:
                logger.error(f"❌ STEP 3: Error generating Excel file: {str(excel_error)}")
        
        # Step 4: Create comprehensive ZIP file
        logger.info(f"📦 STEP 4: Creating comprehensive ZIP file")
        comprehensive_zip_path = None
        try:
            # Collect all files for ZIP
            all_files_for_zip = []
            all_files_for_zip.extend(all_generated_images)  # Default generated images
            all_files_for_zip.extend(all_nobg_images)       # Background removed images
            all_files_for_zip.extend(all_card_applied_images)  # Card applied images
            all_files_for_zip.extend(all_base64_files)      # Base64 files
            
            if excel_file_path:
                all_files_for_zip.append(excel_file_path)   # Excel compilation
            
            if all_files_for_zip:
                comprehensive_zip_path = create_zip_file_with_metadata(
                    all_files_for_zip, 
                    metadata_files=None, 
                    encode_to_base64=False,  # Already handled above
                    theme=theme,
                    category=category,
                    excel_path=excel_file_path,
                    original_images=all_generated_images,
                    nobg_images=all_nobg_images,
                    card_images=all_card_applied_images
                )
                logger.info(f"✅ STEP 4: Created comprehensive ZIP: {comprehensive_zip_path}")
        except Exception as zip_error:
            logger.error(f"❌ STEP 4: Error creating comprehensive ZIP: {str(zip_error)}")
        
        # Generate final status
        total_generated = len(all_generated_images)
        final_summary = f"🎉 COMPREHENSIVE BATCH PROCESSING COMPLETE!"
        final_stats = f"📊 RESULTS: {successful_batches}/{len(extracted_images)} reference images processed. Generated {total_generated} default images, {len(all_nobg_images)} no-bg images, {len(all_card_applied_images)} card-applied images"
        
        if safe_encode_to_base64:
            final_stats += f", {len(all_base64_files)} base64 files"
        if excel_file_path:
            final_stats += f", 1 Excel compilation"
        
        final_status = f"{final_summary}\n{final_stats}"
        
        logger.info(final_summary)
        logger.info(final_stats)
        
        return (
            all_display_images,      # For gallery display
            final_status,            # Status message
            comprehensive_zip_path,  # Download URL
            all_generated_images,    # For inpainting
            None,                    # No modified images for batch processing
            None                     # No modified ZIP for batch processing
        )
        
    except Exception as e:
        logger.error(f"❌ COMPREHENSIVE ZIP BATCH: Critical error: {str(e)}")
        return ([], f"Error in comprehensive batch processing: {str(e)}", None, [], None, None)

def create_comprehensive_excel_report(excel_data, theme, category):
    """
    Create a comprehensive Excel report with all generation data
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Create output directory
        output_dir = os.path.join("generated_output", "comprehensive_batch")
        os.makedirs(output_dir, exist_ok=True)
        
        # Create Excel filename using theme/category codes for consistency
        timestamp = get_gmt7_filename_timestamp()
        theme_code = THEME_MAPPING.get(theme, "00")
        category_code = CATEGORY_MAPPING.get(category, "000")
        excel_filename = f"{theme_code}{category_code}_report_{timestamp}.xlsx"
        excel_path = os.path.join(output_dir, excel_filename)
        
        # Create DataFrame
        df = pd.DataFrame(excel_data)
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Generation Report"
        
        # Add headers
        headers = [
            'Reference Image', 'Generated Image', 'Prompt', 'Modified Prompt', 
            'Provider', 'Model', 'Theme', 'Category', 'Activity', 
            'Facial Expression', 'Fur Color', 'Ethnicity', 'Seed', 
            'Guidance Scale', 'Timestamp'
        ]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data
        for row_idx, data in enumerate(excel_data, 2):
            ws.cell(row=row_idx, column=1, value=data.get('reference_image', ''))
            ws.cell(row=row_idx, column=2, value=data.get('generated_image', ''))
            ws.cell(row=row_idx, column=3, value=data.get('prompt', ''))
            ws.cell(row=row_idx, column=4, value=data.get('modified_prompt', ''))
            ws.cell(row=row_idx, column=5, value=data.get('provider', ''))
            ws.cell(row=row_idx, column=6, value=data.get('model', ''))
            ws.cell(row=row_idx, column=7, value=data.get('theme', ''))
            ws.cell(row=row_idx, column=8, value=data.get('category', ''))
            ws.cell(row=row_idx, column=9, value=data.get('activity', ''))
            ws.cell(row=row_idx, column=10, value=data.get('facial_expression', ''))
            ws.cell(row=row_idx, column=11, value=data.get('fur_color', ''))
            ws.cell(row=row_idx, column=12, value=data.get('ethnicity', ''))
            ws.cell(row=row_idx, column=13, value=data.get('seed', ''))
            ws.cell(row=row_idx, column=14, value=data.get('guidance_scale', ''))
            ws.cell(row=row_idx, column=15, value=data.get('timestamp', ''))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(excel_path)
        logger.info(f"✅ Excel report saved: {excel_path}")
        return excel_path
        
    except Exception as e:
        logger.error(f"❌ Error creating Excel report: {str(e)}")
        return None



def create_zip_file_with_metadata(image_paths, metadata_files=None, encode_to_base64=False, theme=None, category=None, excel_path=None, 
                                 original_images=None, nobg_images=None, card_images=None):
    """
    Create a ZIP file from image paths and metadata files for download
    
    Args:
        image_paths (list): List of image file paths to include in the ZIP
        metadata_files (list): List of metadata file paths to include in the ZIP
        encode_to_base64 (bool): Whether to include base64 encoded files
        theme (str): Theme name for proper naming convention
        category (str): Category name for proper naming convention
        excel_path (str): Path to Excel file to include in ZIP
        original_images (list): List of original generated image paths
        nobg_images (list): List of background-removed image paths
        card_images (list): List of card-applied image paths
        
    Returns:
        str: Path to the created ZIP file or None if failed
    """
    import tempfile
    import zipfile
    
    # Check if we have comprehensive file types available
    if original_images or nobg_images or card_images:
        logger.info("Creating comprehensive ZIP with available file types (original, no-bg, card, Excel)")
        return create_comprehensive_zip_with_all_files(
            original_images=original_images or [],
            nobg_images=nobg_images or [],
            card_images=card_images or [],
            excel_path=excel_path,
            theme=theme,
            category=category,
            encode_to_base64=encode_to_base64
        )
    
    if not image_paths:
        logger.warning("No images provided to create ZIP file")
        return None
        
    # Create timestamp for unique filename
    timestamp = get_gmt7_filename_timestamp()
    
    # Create proper ZIP filename using theme/category codes
    if theme and category:
        theme_code = THEME_MAPPING.get(theme, "00")
        category_code = CATEGORY_MAPPING.get(category, "000")
        zip_name = f"{theme_code}{category_code}_{timestamp}.zip"
    else:
        # Fallback to generic name if theme/category not provided
        zip_name = f"generated_images_{timestamp}.zip"
    
    # Create ZIP in temp directory
    temp_dir = tempfile.mkdtemp()
    zip_filepath = os.path.join(temp_dir, zip_name)
    
    try:
        # Create ZIP file
        processed_image_paths = []  # Keep track of actual file paths for base64 processing
        file_counter = 1  # Counter for sequential numbering
        
        with zipfile.ZipFile(zip_filepath, 'w', compression=zipfile.ZIP_DEFLATED) as zipf:
            # Add image files with proper naming convention
            for img_path in image_paths:
                # Handle different types of image paths/objects
                actual_path = None
                
                # Check if it's a string path
                if isinstance(img_path, str):
                    actual_path = img_path
                # Check if it's a PIL Image object
                elif hasattr(img_path, 'save'):
                    # It's a PIL Image, save it to temporary file first
                    import tempfile
                    from PIL import Image
                    
                    temp_dir = tempfile.mkdtemp()
                    temp_filename = f"temp_image_{len(os.listdir(temp_dir))}.png"
                    actual_path = os.path.join(temp_dir, temp_filename)
                    
                    try:
                        img_path.save(actual_path)
                        logger.info(f"Saved PIL Image to temporary file: {actual_path}")
                    except Exception as save_error:
                        logger.error(f"Error saving PIL Image to file: {str(save_error)}")
                        continue
                # Check if it's a tuple/dict format
                elif isinstance(img_path, (tuple, list)) and len(img_path) > 0:
                    actual_path = img_path[0] if isinstance(img_path[0], str) else None
                elif isinstance(img_path, dict) and 'path' in img_path:
                    actual_path = img_path['path']
                else:
                    logger.warning(f"Unsupported image path format in ZIP creation: {type(img_path)}")
                    continue
                
                # Now process the actual file path with proper naming
                if actual_path and os.path.exists(actual_path):
                    # Get file extension
                    file_ext = os.path.splitext(actual_path)[1].lower()
                    if not file_ext:
                        file_ext = '.png'  # Default to PNG if no extension
                    
                    # Create proper filename using TTCCCNNNNN convention
                    if theme and category:
                        theme_code = THEME_MAPPING.get(theme, "00")
                        category_code = CATEGORY_MAPPING.get(category, "000")
                        # Format: TTCCCNNNNN.ext (e.g., 0100200001.png for Pets/Cats)
                        proper_filename = f"{theme_code}{category_code}{file_counter:05d}{file_ext}"
                    else:
                        # Fallback to generic naming if theme/category not provided
                        proper_filename = f"image_{file_counter:03d}{file_ext}"
                    
                    # Add file to ZIP with proper filename
                    zipf.write(actual_path, proper_filename)
                    processed_image_paths.append(actual_path)  # Track for base64 processing
                    logger.info(f"Added to ZIP with proper naming: {proper_filename}")
                    file_counter += 1
                else:
                    logger.warning(f"File not found for ZIP: {actual_path} (original: {type(img_path)})")
            
            # Add metadata files
            if metadata_files:
                for meta_path in metadata_files:
                    if os.path.exists(meta_path):
                        # Add metadata file to ZIP using just the filename
                        zipf.write(meta_path, os.path.basename(meta_path))
                        logger.info(f"Added metadata file to ZIP: {os.path.basename(meta_path)}")
                    else:
                        logger.warning(f"Metadata file not found for ZIP: {meta_path}")
            
            # Add base64 encoded files if requested
            if encode_to_base64:
                logger.info("Base64 encoding enabled - adding base64 files to ZIP")
                try:
                    base64_files = batch_encode_images_to_base64(processed_image_paths)
                    for base64_file in base64_files:
                        if os.path.exists(base64_file):
                            # Add base64 file to ZIP using just the filename
                            zipf.write(base64_file, os.path.basename(base64_file))
                            logger.info(f"Added base64 file to ZIP: {os.path.basename(base64_file)}")
                        else:
                            logger.warning(f"Base64 file not found for ZIP: {base64_file}")
                    logger.info(f"Successfully added {len(base64_files)} base64 files to ZIP")
                except Exception as base64_error:
                    logger.error(f"Error adding base64 files to ZIP: {str(base64_error)}")
                    # Continue without base64 files - don't fail the entire ZIP creation
            
            # Add Excel file if provided
            if excel_path and os.path.exists(excel_path):
                zipf.write(excel_path, os.path.basename(excel_path))
                logger.info(f"Added Excel file to ZIP: {os.path.basename(excel_path)}")
        
        logger.info(f"Created ZIP file with metadata{' and base64 files' if encode_to_base64 else ''}: {zip_filepath}")
        return zip_filepath
        
    except Exception as e:
        logger.error(f"Error creating ZIP file with metadata: {str(e)}")
        return None

def create_inpainting_zip(image_paths, prompt=None):
    """Create a ZIP file containing inpainted images with metadata"""
    if not image_paths:
        return None
    
    # Create a unique temporary directory for this ZIP
    temp_dir = tempfile.mkdtemp()
    zip_filename = f"inpainted_images_{get_gmt7_filename_timestamp()}.zip"
    zip_path = os.path.join(temp_dir, zip_filename)
    
    try:
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for i, image_path in enumerate(image_paths):
                if os.path.exists(image_path):
                    # Use a descriptive filename in the ZIP
                    filename_in_zip = f"inpainted_{i+1:03d}.png"
                    zipf.write(image_path, filename_in_zip)
            
            # Add metadata file if prompt is provided
            if prompt:
                metadata_content = f"Inpainting Details\n" \
                                 f"==================\n" \
                                 f"Generated: {get_gmt7_timestamp()}\n" \
                                 f"Prompt: {prompt}\n" \
                                 f"Images: {len(image_paths)}\n" \
                                 f"Provider: Leonardo AI (Canvas Inpainting)\n"
                
                metadata_path = os.path.join(temp_dir, "inpainting_metadata.txt")
                with open(metadata_path, 'w', encoding='utf-8') as f:
                    f.write(metadata_content)
                zipf.write(metadata_path, "metadata.txt")
                    
        logger.info(f"Created inpainting ZIP file: {zip_path} with {len(image_paths)} images")
        return zip_path
    except Exception as e:
        logger.error(f"Error creating inpainting ZIP file: {str(e)}")
        return None

def get_image_metadata(image_path):
    """
    Extract metadata from an image file
    
    Args:
        image_path (str): Path to the image file
        
    Returns:
        dict: Dictionary containing image metadata (dimensions, filename, etc.)
    """
    try:
        # Validate input type
        if not isinstance(image_path, str):
            logger.error(f"Error extracting image metadata from {image_path}: expected string path, got {type(image_path)}")
            return {}
            
        if not image_path or not os.path.exists(image_path):
            return {}
            
        with Image.open(image_path) as img:
            return {
                'filename': os.path.basename(image_path),
                'path': image_path,
                'dimensions': (img.width, img.height),
                'format': img.format,
                'mode': img.mode,
                'size_bytes': os.path.getsize(image_path)
            }
    except Exception as e:
        logger.error(f"Error extracting image metadata from {image_path}: {str(e)}")
        return {
            'filename': os.path.basename(image_path) if image_path else None,
            'path': image_path,
            'error': str(e)
        }

def create_text_metadata_files(
    prompt_content=None, 
    activity=None, 
    facial_expression=None, 
    fur_color=None, 
    theme=None, 
    category=None, 
    generation_type="standard",
    base_filename_pattern=None,
    # Additional metadata parameters
    reference_image_path=None,
    reference_image_filename=None,
    reference_image_dimensions=None,
    card_template_path=None,
    card_template_filename=None,
    card_template_dimensions=None,
    provider=None,
    model_name=None,
    seed=None,
    guidance_scale=None,
    num_images=None
):
    """
    Create text files with metadata information
    
    Args:
        prompt_content (str): The prompt content to save
        activity (str): Activity information
        facial_expression (str): Facial expression information  
        fur_color (str): Fur color information
        theme (str): Theme information
        category (str): Category information
        generation_type (str): Type of generation ("activity" or "standard")
        base_filename_pattern (str): Base filename pattern for consistent naming
        reference_image_path (str): Path to the reference image
        reference_image_filename (str): Filename of the reference image
        reference_image_dimensions (tuple): Width and height of reference image
        card_template_path (str): Path to the card template
        card_template_filename (str): Filename of the card template
        card_template_dimensions (tuple): Width and height of card template
        provider (str): AI provider used (Leonardo/Ideogram)
        model_name (str): Model name/ID used for generation
        seed (int): Seed value used for generation
        guidance_scale (float): Guidance scale used
        num_images (int): Number of images generated
        
    Returns:
        list: List of created text file paths
    """
    created_files = []
    
    try:
        # Create temp directory for text files
        temp_dir = tempfile.mkdtemp()
        timestamp = get_gmt7_filename_timestamp()
        
        # Use base filename pattern if provided, otherwise use timestamp
        if base_filename_pattern:
            file_prefix = base_filename_pattern
        else:
            file_prefix = f"generation_{timestamp}"
        
        # Always save Activity, Facial Expression, Fur Color
        metadata_content = []
        if activity and str(activity).strip():
            metadata_content.append(f"Activity: {activity}")
        if facial_expression and str(facial_expression).strip():
            metadata_content.append(f"Facial Expression: {facial_expression}")
        if fur_color and str(fur_color).strip():
            metadata_content.append(f"Fur Color: {fur_color}")
        if theme:
            metadata_content.append(f"Theme: {theme}")
        if category:
            metadata_content.append(f"Category: {category}")
        
        # Add generation technical metadata
        technical_metadata = []
        if provider:
            technical_metadata.append(f"Provider: {provider}")
        if model_name:
            technical_metadata.append(f"Model: {model_name}")
        if seed is not None:
            technical_metadata.append(f"Seed: {seed}")
        if guidance_scale is not None:
            technical_metadata.append(f"Guidance Scale: {guidance_scale}")
        if num_images is not None:
            technical_metadata.append(f"Number of Images: {num_images}")
        
        # Add reference image metadata
        reference_metadata = []
        if reference_image_filename:
            reference_metadata.append(f"Reference Image Filename: {reference_image_filename}")
        if reference_image_path:
            reference_metadata.append(f"Reference Image Path: {reference_image_path}")
            # Get additional metadata if available
            ref_metadata = get_image_metadata(reference_image_path)
            if ref_metadata.get('format'):
                reference_metadata.append(f"Reference Image Format: {ref_metadata.get('format')}")
            if ref_metadata.get('mode'):
                reference_metadata.append(f"Reference Image Mode: {ref_metadata.get('mode')}")
            if ref_metadata.get('size_bytes'):
                size_kb = ref_metadata.get('size_bytes') / 1024
                if size_kb > 1024:
                    size_mb = size_kb / 1024
                    reference_metadata.append(f"Reference Image File Size: {size_mb:.2f} MB")
                else:
                    reference_metadata.append(f"Reference Image File Size: {size_kb:.2f} KB")
        if reference_image_dimensions:
            try:
                width, height = reference_image_dimensions
                reference_metadata.append(f"Reference Image Dimensions: {width}x{height} pixels")
                reference_metadata.append(f"Reference Image Aspect Ratio: {width/height:.2f}")
            except (ValueError, TypeError):
                reference_metadata.append(f"Reference Image Dimensions: {reference_image_dimensions}")
        
        # Add card template metadata
        card_template_metadata = []
        if card_template_filename:
            card_template_metadata.append(f"Card Template Filename: {card_template_filename}")
        if card_template_path:
            card_template_metadata.append(f"Card Template Path: {card_template_path}")
            # Get additional metadata if available
            card_metadata = get_image_metadata(card_template_path)
            if card_metadata.get('format'):
                card_template_metadata.append(f"Card Template Format: {card_metadata.get('format')}")
            if card_metadata.get('mode'):
                card_template_metadata.append(f"Card Template Mode: {card_metadata.get('mode')}")
            if card_metadata.get('size_bytes'):
                size_kb = card_metadata.get('size_bytes') / 1024
                if size_kb > 1024:
                    size_mb = size_kb / 1024
                    card_template_metadata.append(f"Card Template File Size: {size_mb:.2f} MB")
                else:
                    card_template_metadata.append(f"Card Template File Size: {size_kb:.2f} KB")
        if card_template_dimensions:
            try:
                width, height = card_template_dimensions
                card_template_metadata.append(f"Card Template Dimensions: {width}x{height} pixels")
                card_template_metadata.append(f"Card Template Aspect Ratio: {width/height:.2f}")
            except (ValueError, TypeError):
                card_template_metadata.append(f"Card Template Dimensions: {card_template_dimensions}")
        
        # Create comprehensive metadata file
        all_metadata = metadata_content + technical_metadata + reference_metadata + card_template_metadata
        if all_metadata:
            metadata_file = os.path.join(temp_dir, f"{file_prefix}_metadata.txt")
            with open(metadata_file, 'w', encoding='utf-8') as f:
                f.write("=== GENERATION METADATA ===\n")
                f.write(f"Generation Type: {generation_type.upper()}\n")
                f.write(f"Timestamp: {get_gmt7_timestamp()} GMT+7\n\n")
                
                # Write creative metadata
                if metadata_content:
                    f.write("=== CREATIVE METADATA ===\n")
                    f.write("\n".join(metadata_content) + "\n\n")
                
                # Write technical metadata
                if technical_metadata:
                    f.write("=== TECHNICAL METADATA ===\n")
                    f.write("\n".join(technical_metadata) + "\n\n")
                
                # Write reference image metadata
                if reference_metadata:
                    f.write("=== REFERENCE IMAGE METADATA ===\n")
                    f.write("\n".join(reference_metadata) + "\n\n")
                
                # Write card template metadata
                if card_template_metadata:
                    f.write("=== CARD TEMPLATE METADATA ===\n")
                    f.write("\n".join(card_template_metadata) + "\n")
                    
            created_files.append(metadata_file)
            logger.info(f"Created comprehensive metadata file: {metadata_file}")
        
        # Save prompt content based on generation type
        if prompt_content and str(prompt_content).strip():
            if generation_type == "activity":
                prompt_file = os.path.join(temp_dir, f"{file_prefix}_modified_prompt.txt")
                prompt_label = "MODIFIED PROMPT (Generate with Activity)"
            else:
                prompt_file = os.path.join(temp_dir, f"{file_prefix}_generated_prompt.txt")
                prompt_label = "GENERATED PROMPT (Current Image)"
            
            with open(prompt_file, 'w', encoding='utf-8') as f:
                f.write(f"=== {prompt_label} ===\n")
                f.write(f"Timestamp: {get_gmt7_timestamp()} GMT+7\n")
                f.write(f"Generation Type: {generation_type.upper()}\n\n")
                f.write(str(prompt_content))
            created_files.append(prompt_file)
            logger.info(f"Created prompt file: {prompt_file}")
        
        logger.info(f"Created {len(created_files)} text metadata files")
        return created_files
        
    except Exception as e:
        logger.error(f"Error creating text metadata files: {str(e)}")
        return []

# Ethnicity trait definitions
ETHNIC_TRAITS = {
    "Auto": {
        "description": "No specific ethnic traits applied",
        "male_traits": "",
        "female_traits": ""
    },
    "Javanese – Indonesia": {
        "description": "Indonesian Javanese ethnic traits",
        "male_traits": "straight to mildly wavy hair with medium thickness and dark brown-black color in tidy short cuts with sparse beard growth, elongated oval face with gently prominent cheekbones, medium-broad nose with low bridge, almond eyes, softer brow ridge",
        "female_traits": "straight smooth dark brown-black hair traditionally long and often in low bun or braids for ceremonies, delicate oval face with subtle chin point and smooth cheeks, medium nose width, almond eyes with slightly arched brows"
    },
    "Sundanese – Indonesia": {
        "description": "Indonesian Sundanese ethnic traits",
        "male_traits": "straight fine-to-medium dark brown-black hair in medium-short styles with minimal facial hair, round-to-heart outline face with softer jaw, low nose bridge, bright expressive dark eyes",
        "female_traits": "straight and sleek dark brown-black hair often shoulder-length with layered fringe, rounder cheeks with gentle jawline, small low-bridge nose, wide dark eyes with naturally arched brows"
    },
    "Chinese Indonesians": {
        "description": "Chinese Indonesian (Tionghoa Indonesia) ethnic traits",
        "male_traits": "coarse straight dense dark brown-black hair frequently styled short or fade with sparse facial hair, oval-to-square face with low-to-medium nasal bridge, almond eyes with epicanthic fold, moderate cheekbones",
        "female_traits": "fine-to-medium straight and glossy dark brown-black hair in long styles with blunt or side fringe, soft oval contour with narrow chin and smooth mid-face, low nasal bridge, almond/inner double-lid eyes with long lashes"
    },
    "American – United States": {
        "description": "White/European American ethnic traits",
        "male_traits": "fine-to-medium thickness straight-to-wavy hair with natural colors spanning light blond, strawberry blond, brown shades to near-black, dense and coarse beards and moustaches, generally oval-to-rectangular face with medium-to-high nasal bridge and pronounced nasal dorsum, noticeable brow ridge, moderately defined cheekbones, eye hues varying widely including blue, green, hazel, grey, brown",
        "female_traits": "fine-to-medium straight, wavy or loosely curly hair with native palette ranging from platinum blond through auburn, chestnut, dark brown to black often worn long or shoulder-length, oval-to-heart outline face with medium/high nose bridge and softer jawline, fuller lips, prominent cheekbones, diverse eye colors including blue/green/hazel/brown"
    },
    "Hindi – India": {
        "description": "Indo-Aryan (Hindi-belt) ethnic traits",
        "male_traits": "medium-coarse wavy-to-straight dark brown-black hair with thick facial hair including beard/moustache, broad forehead with stronger jaw, high straight or slightly aquiline nose, deep-set large dark eyes",
        "female_traits": "fine-to-medium naturally wavy dark brown-black hair traditionally long and center-parted or braided, heart/oval shape face with fuller lips and smooth cheek contour, high nasal bridge with narrower tip, prominent expressive eyes"
    }
}

def detect_gender_from_prompt(prompt):
    """
    Enhanced gender detection from prompt content using comprehensive keyword analysis and context clues
    Returns 'male', 'female', or 'unknown'
    """
    prompt_lower = prompt.lower()
    logger.info(f"🔍 GENDER DETECTION: Analyzing prompt: {prompt[:50]}...")
    
    # Enhanced Male indicators with more comprehensive terms
    male_keywords = [
        # Direct gender terms
        'man', 'male', 'boy', 'guy', 'gentleman',
        # Family relations
        'father', 'dad', 'daddy', 'papa', 'son', 'brother', 'uncle', 'nephew', 'grandfather', 'grandpa', 'husband', 'boyfriend',
        # Occupations traditionally male
        'businessman', 'policeman', 'fireman', 'postman', 'chairman', 'spokesman',
        # Physical descriptors often male
        'bearded', 'mustache', 'moustache', 'beard', 'bald', 'balding',
        # Pronouns
        'he', 'his', 'him', 'himself'
    ]
    
    # Enhanced Female indicators with more comprehensive terms
    female_keywords = [
        # Direct gender terms
        'woman', 'female', 'girl', 'lady',
        # Family relations
        'mother', 'mom', 'mommy', 'mama', 'daughter', 'sister', 'aunt', 'niece', 'grandmother', 'grandma', 'wife', 'girlfriend',
        # Titles and roles
        'queen', 'princess', 'lady', 'mrs', 'miss', 'ms', 'ma\'am', 'madam', 'duchess', 'baroness', 'empress',
        # Occupations traditionally female
        'businesswoman', 'policewoman', 'chairwoman', 'spokeswoman', 'actress', 'waitress', 'hostess',
        # Physical descriptors often female
        'pregnant', 'expecting', 'lipstick', 'makeup', 'dress', 'skirt', 'heels', 'jewelry', 'necklace', 'earrings', 'bracelet',
        # Style and appearance
        'feminine', 'graceful', 'elegant', 'beautiful', 'pretty', 'lovely', 'gorgeous', 'stunning',
        # Pronouns
        'she', 'her', 'hers', 'herself'
    ]
    
    # Advanced pattern matching for compound words and phrases
    male_patterns = [
        r'\bmen\b', r'\bmales\b', r'\bboys\b', r'\bguys\b', r'\bgentlemen\b',
        r'\bmanly\b', r'\bmasc\b', r'\bhandsome\b', r'\brugged\b', r'\bstrong.{1,20}man\b',
        r'\byoung.{1,10}man\b', r'\bold.{1,10}man\b', r'\btall.{1,10}man\b'
    ]
    
    female_patterns = [
        r'\bwomen\b', r'\bfemales\b', r'\bgirls\b', r'\bladies\b',
        r'\bfeminine\b', r'\bbeautiful\b', r'\bpretty\b', r'\belegant\b', r'\bgraceful\b',
        r'\byoung.{1,10}woman\b', r'\bold.{1,10}woman\b', r'\btall.{1,10}woman\b'
    ]
    
    # Count keyword occurrences
    male_count = sum(1 for keyword in male_keywords if keyword in prompt_lower)
    female_count = sum(1 for keyword in female_keywords if keyword in prompt_lower)
    
    # Count pattern matches
    import re
    male_pattern_count = sum(1 for pattern in male_patterns if re.search(pattern, prompt_lower))
    female_pattern_count = sum(1 for pattern in female_patterns if re.search(pattern, prompt_lower))
    
    # Total scores
    total_male_score = male_count + male_pattern_count
    total_female_score = female_count + female_pattern_count
    
    logger.info(f"🔍 Gender scores - Male: {total_male_score} (keywords: {male_count}, patterns: {male_pattern_count})")
    logger.info(f"🔍 Gender scores - Female: {total_female_score} (keywords: {female_count}, patterns: {female_pattern_count})")
    
    # Determine gender with logging
    if total_male_score > total_female_score:
        logger.info(f"✅ GENDER DETECTED: MALE (score: {total_male_score} vs {total_female_score})")
        return 'male'
    elif total_female_score > total_male_score:
        logger.info(f"✅ GENDER DETECTED: FEMALE (score: {total_female_score} vs {total_male_score})")
        return 'female'
    else:
        logger.info(f"❓ GENDER UNKNOWN: Equal scores ({total_male_score} vs {total_female_score}) or no indicators found")
        return 'unknown'

def ensure_gender_in_prompt(prompt, contains_human=True):
    """
    Ensure that gender is explicitly mentioned in the prompt for HUMAN subjects only.
    If no gender is detected and the subject is human, use Qwen to add appropriate gender indication.
    For non-human subjects, this function returns the original prompt unchanged.
    """
    # Skip gender enforcement for non-human subjects
    if not contains_human:
        logger.info(f"🐾 Non-human subject detected - skipping gender enforcement")
        return prompt
        
    detected_gender = detect_gender_from_prompt(prompt)
    
    if detected_gender != 'unknown':
        logger.info(f"✅ Gender already clear in prompt: {detected_gender}")
        return prompt
    
    logger.info(f"⚠️ No clear gender in prompt for human subject, using Qwen to clarify gender")
    
    try:
        sys_prompt = """You are an expert at analyzing and enhancing image prompts to ensure gender clarity for HUMAN subjects.
Your task is to determine the likely intended gender of the HUMAN subject in the prompt and rewrite it to make the gender explicit and clear.

Rules:
1. Add explicit gender terms like "man", "woman", "male", or "female" 
2. Preserve ALL original content and characteristics
3. Make the gender reference natural and integrated
4. If the subject seems ambiguous, default to adding "person" but include gender-indicating traits
5. Do not change the core subject or characteristics, only clarify gender
6. Ensure the output is a complete, natural-sounding prompt
7. This should ONLY be applied to human subjects"""

        user_prompt = f"""Original prompt: {prompt}

This prompt describes a HUMAN subject but lacks clear gender indication. Please rewrite it to explicitly include gender (man/woman, male/female) while preserving all original characteristics and details. Make the gender reference natural and integrated into the description."""

        enhanced_prompt = inference_with_api(
            image_path=None,
            prompt=user_prompt,
            sys_prompt=sys_prompt,
            model_id="qwen2.5-72b-instruct"
        )
        
        if enhanced_prompt and len(enhanced_prompt.strip()) > 0:
            # Verify that gender was actually added
            new_gender = detect_gender_from_prompt(enhanced_prompt)
            if new_gender != 'unknown':
                logger.info(f"✅ Successfully added gender to prompt: {new_gender}")
                return enhanced_prompt
            else:
                logger.warning(f"⚠️ Qwen didn't add clear gender, falling back to manual addition")
                # Fallback: manually add gender indication
                if 'person' in prompt.lower() or 'character' in prompt.lower():
                    return prompt.replace('person', 'man or woman').replace('character', 'male or female character')
                else:
                    return f"{prompt}, depicting a man or woman"
        else:
            logger.warning(f"⚠️ Empty response from Qwen, using fallback")
            return f"{prompt}, depicting a man or woman"
            
    except Exception as e:
        logger.error(f"Error enhancing prompt with gender clarity: {str(e)}")
        # Fallback to simple addition
        return f"{prompt}, depicting a man or woman"

def enhance_prompt_with_ethnicity(prompt, ethnicity="Auto", detected_gender="unknown"):
    """
    Enhance prompt with ethnic traits based on selected ethnicity and detected gender
    
    Args:
        prompt (str): The original prompt
        ethnicity (str): Selected ethnicity from ETHNIC_TRAITS
        detected_gender (str): 'male', 'female', or 'unknown'
        
    Returns:
        str: Enhanced prompt with ethnic traits
    """
    if ethnicity == "Auto" or ethnicity not in ETHNIC_TRAITS:
        return prompt
        
    traits = ETHNIC_TRAITS[ethnicity]
    
    # Determine which traits to use
    if detected_gender == 'male' and traits['male_traits']:
        ethnic_traits = traits['male_traits']
    elif detected_gender == 'female' and traits['female_traits']:
        ethnic_traits = traits['female_traits']
    else:
        # If gender is unknown or no specific traits, combine both or use a general approach
        if traits['male_traits'] and traits['female_traits']:
            # Use Qwen to determine which traits to apply
            try:
                sys_prompt = """You are an expert at analyzing image prompts and determining appropriate ethnic traits. 
Based on the prompt content, determine if the subject appears to be male or female, then apply the appropriate ethnic traits.
If gender cannot be determined clearly, use neutral traits that could apply to either gender.
Integrate the ethnic traits naturally into the prompt while preserving all original content."""

                user_prompt = f"""Original prompt: {prompt}

Ethnicity: {ethnicity}
Male traits: {traits['male_traits']}
Female traits: {traits['female_traits']}

Please rewrite the prompt incorporating the appropriate ethnic traits based on the apparent gender of the subject, while preserving all original content."""

                enhanced_prompt = inference_with_api(
                    image_path=None,
                    prompt=user_prompt,
                    sys_prompt=sys_prompt,
                    model_id="qwen2.5-72b-instruct"
                )
                
                return enhanced_prompt if enhanced_prompt else prompt
                
            except Exception as e:
                logger.error(f"Error enhancing prompt with ethnicity via Qwen: {str(e)}")
                # Fallback to simple addition
                return f"{prompt}, with {traits['description']}"
        else:
            return prompt
    
    # Apply the specific traits
    try:
        sys_prompt = """You are an expert at enhancing image prompts with ethnic characteristics. 
Your task is to naturally integrate the provided ethnic traits into the original prompt while preserving all original content.
Make sure the ethnic traits are woven seamlessly into the description."""

        user_prompt = f"""Original prompt: {prompt}

Ethnic traits to integrate: {ethnic_traits}

Please rewrite the prompt naturally incorporating these ethnic traits while preserving all original content and characteristics."""

        enhanced_prompt = inference_with_api(
            image_path=None,
            prompt=user_prompt,
            sys_prompt=sys_prompt,
            model_id="qwen2.5-72b-instruct"
        )
        
        return enhanced_prompt if enhanced_prompt else prompt
        
    except Exception as e:
        logger.error(f"Error enhancing prompt with ethnicity: {str(e)}")
        # Fallback to simple addition
        return f"{prompt}, with {ethnic_traits}"

def get_context_appropriate_modifications(prompt, activity=None, facial_expression=None, fur_color=None, subject_type="unknown"):
    """
    Convert facial expressions and fur colors to context-appropriate modifications for different subject types.
    
    Args:
        prompt (str): The original prompt
        activity (str): Activity to apply
        facial_expression (str): Original facial expression
        fur_color (str): Original fur color
        subject_type (str): "human", "animal", or "object"
        
    Returns:
        tuple: (modified_activity, modified_expression, modified_color)
    """
    modified_activity = activity
    modified_expression = facial_expression
    modified_color = fur_color
    
    if subject_type == "object":
        # Convert facial expressions to appropriate object characteristics
        if facial_expression:
            expression_mappings = {
                "sparkling eyes": "glossy surface",
                "gentle up-curved mouth": "elegant curved design",
                "bright eyes": "reflective surface",
                "smiling": "curved design elements",
                "happy": "vibrant appearance",
                "sad": "muted tones",
                "angry": "sharp angular design",
                "surprised": "dynamic form",
                "excited": "energetic design",
                "calm": "smooth finish",
                "peaceful": "serene design"
            }
            
            # Try to find a mapping for the facial expression
            for face_expr, obj_equiv in expression_mappings.items():
                if face_expr.lower() in facial_expression.lower():
                    modified_expression = obj_equiv
                    logger.info(f"🔄 OBJECT ADAPTATION: Converted '{facial_expression}' to '{obj_equiv}' for object")
                    break
            else:
                # If no direct mapping, remove the facial expression
                modified_expression = None
                logger.info(f"🚫 OBJECT ADAPTATION: Removed facial expression '{facial_expression}' as it's not applicable to objects")
        
        # Convert fur colors to appropriate material colors/finishes
        if fur_color:
            color_mappings = {
                "silvery gray": "metallic silver finish",
                "golden": "golden metallic surface",
                "brown": "bronze finish",
                "black": "matte black coating",
                "white": "pristine white surface",
                "red": "glossy red finish",
                "blue": "deep blue coating",
                "green": "emerald green surface"
            }
            
            # Try to find a mapping for the fur color
            for fur, material in color_mappings.items():
                if fur.lower() in fur_color.lower():
                    modified_color = material
                    logger.info(f"🔄 OBJECT ADAPTATION: Converted '{fur_color}' to '{material}' for object")
                    break
            else:
                # Use the color but change the context
                modified_color = f"{fur_color} finish"
                logger.info(f"🔄 OBJECT ADAPTATION: Converted '{fur_color}' to '{fur_color} finish' for object")
    
    elif subject_type == "animal":
        # Keep facial expressions and fur colors as they are appropriate for animals
        logger.info(f"🐾 ANIMAL: Keeping original facial expression and fur color")
    
    else:  # human
        # Keep everything as is for humans
        logger.info(f"🧑 HUMAN: Keeping original facial expression and fur color")
    
    return modified_activity, modified_expression, modified_color

def detect_living_creature_from_prompt(prompt):
    """
    Detect if the prompt describes a living creature (human or animal) vs non-living object
    
    Args:
        prompt (str): The prompt to analyze
        
    Returns:
        bool: True if living creature, False if non-living object
    """
    if not prompt:
        return False
    
    prompt_lower = prompt.lower()
    
    # Living creature indicators
    living_indicators = [
        # Humans
        'man', 'woman', 'person', 'human', 'boy', 'girl', 'child', 'adult', 'character',
        # Animals
        'cat', 'dog', 'animal', 'pet', 'kitten', 'puppy', 'bird', 'fish', 'horse', 'cow',
        'lion', 'tiger', 'elephant', 'bear', 'wolf', 'fox', 'rabbit', 'mouse', 'hamster',
        'creature', 'beast', 'wildlife', 'mammal', 'reptile', 'amphibian',
        # Body parts that indicate living beings
        'face', 'eyes', 'mouth', 'nose', 'ears', 'hair', 'fur', 'paws', 'tail', 'wings',
        'legs', 'arms', 'hands', 'feet', 'head', 'body', 'neck'
    ]
    
    # Non-living object indicators
    non_living_indicators = [
        'object', 'item', 'thing', 'product', 'tool', 'device', 'machine', 'equipment',
        'building', 'house', 'car', 'vehicle', 'furniture', 'table', 'chair', 'book',
        'bottle', 'cup', 'plate', 'phone', 'computer', 'toy', 'ball', 'box', 'bag',
        'landscape', 'scenery', 'mountain', 'tree', 'flower', 'plant', 'rock', 'stone',
        'full frame shot',  # Indicator for non-living objects
        # Food-related indicators
        'food', 'meal', 'dish', 'bowl', 'noodles', 'soup', 'rice', 'bread', 'cake',
        'pizza', 'burger', 'sandwich', 'fruit', 'vegetable', 'meat', 'fish', 'chicken',
        'beef', 'pork', 'pasta', 'salad', 'dessert', 'snack', 'drink', 'beverage',
        'coffee', 'tea', 'juice', 'wine', 'beer', 'water', 'sauce', 'ingredient',
        'cooking', 'baked', 'fried', 'grilled', 'cooked', 'cuisine', 'recipe'
    ]
    
    # Count indicators
    living_count = sum(1 for indicator in living_indicators if indicator in prompt_lower)
    non_living_count = sum(1 for indicator in non_living_indicators if indicator in prompt_lower)
    
    logger.info(f"🔍 LIVING CREATURE DETECTION: Living indicators: {living_count}, Non-living indicators: {non_living_count}")
    
    # Return True if more living indicators than non-living
    return living_count > non_living_count

def enhance_prompt_with_activity_expression(prompt, activity=None, facial_expression=None, fur_color=None, ethnicity="Auto", contains_human=False):
    """
    Use Qwen API to enhance the prompt with activity, facial expression, fur color, and ethnicity,
    ensuring that activity and facial expression are mentioned before any traits.
    ALWAYS ensures that gender (man/woman, male/female) is explicitly mentioned in the final prompt.
    
    Args:
        prompt (str): The original generated prompt
        activity (str, optional): Activity to incorporate into the prompt
        facial_expression (str, optional): Facial expression to incorporate
        fur_color (str, optional): Fur color to incorporate
        ethnicity (str, optional): Ethnicity to incorporate (Auto, Javanese, etc.)
        contains_human (bool, optional): Whether the image contains humans (to skip fur color)
        
    Returns:
        str: Enhanced prompt with activity, expression, fur color, ethnicity, and explicit gender
    """
    # Log the beginning of the process with more visibility
    logger.info("======================================================")
    logger.info("STARTING PROMPT ENHANCEMENT WITH ACTIVITY/EXPRESSION/FUR COLOR/ETHNICITY + GENDER ENFORCEMENT")
    logger.info(f"Activity: '{activity if activity else 'None'}'")
    logger.info(f"Facial Expression: '{facial_expression if facial_expression else 'None'}'")
    logger.info(f"Fur Color: '{fur_color if fur_color else 'None'}'")
    logger.info(f"Ethnicity: '{ethnicity if ethnicity else 'Auto'}'")
    logger.info(f"Contains Human: {contains_human}")
    logger.info("======================================================")
    logger.info(f"Original prompt: {prompt[:100]}...")
    
    # Detect if this is a living creature or non-living object
    is_living_creature = detect_living_creature_from_prompt(prompt)
    logger.info(f"🔍 SUBJECT TYPE: {'Living Creature' if is_living_creature else 'Non-Living Object'}")
    
    # ALWAYS ensure gender is explicit in the prompt FIRST - but only for living creatures
    if is_living_creature:
        logger.info("🚀 STEP 1: Ensuring gender is explicitly mentioned in prompt for living creature")
        prompt = ensure_gender_in_prompt(prompt, contains_human=contains_human)
        logger.info(f"✅ Gender-enhanced prompt: {prompt[:100]}...")
    else:
        logger.info("🐾 STEP 1: Skipping gender enhancement for non-living object")
        # Clean any existing gender references that might have been incorrectly added
        import re
        original_prompt = prompt
        
        # Remove specific gender-related phrases that are inappropriate for non-living objects
        gender_patterns = [
            r',?\s*depicting a man or woman',
            r',?\s*depicting a male or female',
            r',?\s*with (his|her) ',
            r',?\s*(he|she) (has|is|was)',
            r'\b(man|woman)\s+with\s+',
            r'\bA\s+(man|woman)\s+with\s+'
        ]
        
        for pattern in gender_patterns:
            new_prompt = re.sub(pattern, '', prompt, flags=re.IGNORECASE)
            if new_prompt != prompt:
                prompt = new_prompt.strip()
                logger.info(f"🧹 Removed inappropriate gender reference from non-living object prompt")
        
        # Clean up any double spaces or commas that might result from the removal
        prompt = re.sub(r'\s+', ' ', prompt)  # Multiple spaces to single space
        prompt = re.sub(r',\s*,', ',', prompt)  # Double commas to single comma
        prompt = prompt.strip()
        
        if prompt != original_prompt:
            logger.info(f"🧹 Cleaned prompt: {prompt[:100]}...")
    
    # Simple string checks that work with both strings and numpy arrays
    has_activity = False
    if activity is not None:
        if isinstance(activity, str):
            has_activity = len(activity.strip()) > 0
        else:
            try:
                activity = str(activity)
                has_activity = len(activity.strip()) > 0
            except:
                has_activity = False
    
    has_expression = False
    if facial_expression is not None and is_living_creature:
        if isinstance(facial_expression, str):
            has_expression = len(facial_expression.strip()) > 0
        else:
            try:
                facial_expression = str(facial_expression)
                has_expression = len(facial_expression.strip()) > 0
            except:
                has_expression = False
    elif facial_expression is not None and not is_living_creature:
        logger.info(f"⚠️ NON-LIVING OBJECT DETECTED - Skipping facial expression '{facial_expression}' as non-living objects don't have facial features")
    
    has_fur_color = False
    if fur_color is not None and not contains_human:
        if isinstance(fur_color, str):
            has_fur_color = len(fur_color.strip()) > 0
        else:
            try:
                fur_color = str(fur_color)
                has_fur_color = len(fur_color.strip()) > 0
            except:
                has_fur_color = False
    elif contains_human and fur_color is not None:
        logger.info(f"⚠️ HUMAN DETECTED - Skipping fur color '{fur_color}' as humans don't have fur")
    
    # Check if ethnicity is provided and not Auto
    has_ethnicity = ethnicity and ethnicity != "Auto"
    
    # Enhanced logging for ethnicity
    if has_ethnicity:
        logger.info(f"✓ ETHNICITY DETECTED: {ethnicity} - Will apply ethnic traits")
    else:
        logger.info(f"✗ NO ETHNICITY: Using '{ethnicity}' - No ethnic traits will be applied")
    
    # If no modifications are provided, return the original prompt
    if not (has_activity or has_expression or has_fur_color or has_ethnicity):
        logger.info(f"No modifications provided, returning original prompt")
        return prompt
        
    # Prepare inputs for Qwen API
    modifications = []
    if has_activity:
        modifications.append(f"is performing the activity: {activity}")
        logger.info(f"Adding activity: {activity}")
    
    if has_expression:
        modifications.append(f"has a {facial_expression} facial expression")
        logger.info(f"Adding facial expression: {facial_expression}")
    elif facial_expression and not is_living_creature:
        logger.info(f"Skipping facial expression for non-living object: {facial_expression}")
        
    if has_fur_color:
        modifications.append(f"has {fur_color} fur")
        logger.info(f"Adding fur color: {fur_color}")
    
    # Create a system prompt that handles living creatures vs non-living objects appropriately
    if is_living_creature:
        sys_prompt = """
        You are an expert at rewriting image prompts for LIVING CREATURES (humans, animals, etc.). Your task is to rewrite the given prompt to seamlessly 
        incorporate new activities, facial expressions, fur colors, and ethnic characteristics, while ensuring that the activity 
        and facial expression are mentioned BEFORE any traits from the original prompt.

        CRITICAL GENDER REQUIREMENT:
        - The prompt MUST ALWAYS contain explicit gender indicators (man/woman, male/female, he/she, his/her)
        - If the original prompt already contains gender terms, PRESERVE them and ensure they remain prominent
        - Never remove or obscure gender indicators - they are essential
        - If you see gender terms like "man", "woman", "male", "female", keep them visible and clear

        Important rules:
        1. Start the prompt with the activity and facial expression (if provided).
        2. Follow with ALL traits, characteristics, and details from the original prompt.
        3. ALWAYS maintain and preserve explicit gender indicators throughout the prompt.
        4. Integrate ethnic characteristics naturally throughout the description, especially for facial features, hair, and physical traits.
        5. Do not add any new traits or characteristics not mentioned in the original prompt or the specified ethnic traits.
        6. Do not remove any details from the original prompt, especially gender indicators.
        7. Naturally integrate the new activity, facial expression, and/or ethnic traits into the prompt.
        8. Make the updated prompt read naturally and coherently while keeping gender explicit.
        9. Any objects mentioned in the activity must be appropriately sized and NEVER larger than the main subject.
        10. Limit the number of objects to 1-2 maximum.
        11. The main subject must always be the focal point and dominant element in the image.
        12. Objects should be proportionate and realistic in size compared to the character.
        13. When applying ethnic traits, ensure they are woven seamlessly into the existing description.
        14. VERIFY that gender terms are still clearly visible in your final output.
        15. Make the prompt as detailed as possible with rich descriptions of colors, textures, and characteristics.
        """
    else:
        sys_prompt = """
        You are an expert at rewriting image prompts for NON-LIVING OBJECTS (items, products, buildings, landscapes, etc.). Your task is to rewrite the given prompt to seamlessly 
        incorporate new activities while ensuring that the activity is mentioned BEFORE any traits from the original prompt.

        CRITICAL NON-LIVING OBJECT REQUIREMENTS:
        - DO NOT include any facial expressions, body traits, or living creature characteristics
        - DO NOT use terms like "eyes", "mouth", "face", "hands", "legs", "arms", "ears", "nose"
        - DO NOT use gender-specific language like "man", "woman", "male", "female", "he", "she", "him", "her"
        - Focus on physical properties like colors, textures, materials, shapes, and structural details
        - Use neutral, object-appropriate language

        Important rules:
        1. Start the prompt with the activity (if provided).
        2. Follow with ALL physical traits, characteristics, and details from the original prompt.
        3. Remove any inappropriate living creature references (facial features, body parts, or gender terms).
        4. Focus on materials, textures, colors, shapes, structural details, and physical properties.
        5. Do not add any new traits or characteristics not mentioned in the original prompt.
        6. Do not remove any appropriate physical details from the original prompt.
        7. Naturally integrate the new activity into the prompt.
        8. Make the updated prompt read naturally and coherently with object-appropriate language.
        9. Any additional objects mentioned in the activity must be appropriately sized and NEVER larger than the main subject.
        10. Limit the number of objects to 1-2 maximum.
        11. The main subject must always be the focal point and dominant element in the image.
        12. Objects should be proportionate and realistic in size.
        13. Make the prompt as detailed as possible with rich descriptions of materials, textures, colors, and structural characteristics.
        14. Use terms like "Full Frame Shot" instead of "Full Body Shot" for non-living objects.
        15. VERIFY that no inappropriate living creature terms remain in your final output.
        """
    
    # Detect gender for ethnic traits
    detected_gender = detect_gender_from_prompt(prompt) if has_ethnicity else "unknown"
    
    # Add ethnicity information if provided
    ethnicity_info = ""
    if has_ethnicity:
        traits = ETHNIC_TRAITS.get(ethnicity, {})
        logger.info(f"🎭 APPLYING {ethnicity} ETHNIC TRAITS for detected gender: {detected_gender}")
        
        if detected_gender == 'male' and traits.get('male_traits'):
            ethnicity_info = f"\n\nApply these {ethnicity} male ethnic traits: {traits['male_traits']}"
            logger.info(f"📋 Using MALE traits: {traits['male_traits'][:50]}...")
        elif detected_gender == 'female' and traits.get('female_traits'):
            ethnicity_info = f"\n\nApply these {ethnicity} female ethnic traits: {traits['female_traits']}"
            logger.info(f"📋 Using FEMALE traits: {traits['female_traits'][:50]}...")
        else:
            # Use both traits and let Qwen decide based on context
            if traits.get('male_traits') and traits.get('female_traits'):
                ethnicity_info = f"\n\nEthnicity: {ethnicity}\nMale traits: {traits['male_traits']}\nFemale traits: {traits['female_traits']}\n(Choose appropriate traits based on the apparent gender of the subject)"
                logger.info(f"📋 Using BOTH traits - letting Qwen decide based on gender context")
        
        if not ethnicity_info:
            logger.warning(f"⚠️ No traits found for ethnicity: {ethnicity}")
    else:
        logger.info(f"⚪ Skipping ethnicity traits - ethnicity is '{ethnicity}'")
    
    # Create a user prompt for the API
    if is_living_creature:
        user_prompt = f"""
        Original prompt: 
        {prompt}
        
        Please rewrite this prompt where the subject {' and '.join(modifications)}, 
        ensuring that the activity and facial expression are mentioned before any traits, 
        while preserving ALL traits, characteristics, and details from the original prompt.{ethnicity_info}
        """
    else:
        # For non-living objects, only include activity (no facial expressions)
        activity_modifications = [mod for mod in modifications if not mod.startswith("has a") or "facial expression" not in mod]
        if activity_modifications:
            user_prompt = f"""
            Original prompt: 
            {prompt}
            
            Please rewrite this prompt where the subject {' and '.join(activity_modifications)}, 
            ensuring that the activity is mentioned before any traits, 
            while preserving ALL appropriate physical traits, characteristics, and details from the original prompt.
            Remove any inappropriate living creature references like facial features, body parts, or gender terms.
            Focus on materials, textures, colors, shapes, and structural characteristics.
            """
        else:
            user_prompt = f"""
            Original prompt: 
            {prompt}
            
            Please rewrite this prompt to make it more detailed and object-appropriate,
            while preserving ALL appropriate physical traits, characteristics, and details from the original prompt.
            Remove any inappropriate living creature references like facial features, body parts, or gender terms.
            Focus on materials, textures, colors, shapes, and structural characteristics.
            """
    
    logger.info(f"Prepared modification request with {len(modifications)} modifications")
    
    try:
        # Call Qwen API for enhancing the prompt
        logger.info(f"Sending prompt modification request to Qwen API")
        enhanced_prompt = inference_with_api(
            image_path=None,
            prompt=user_prompt, 
            sys_prompt=sys_prompt,
            model_id="qwen2.5-72b-instruct"
        )
        
        logger.info(f"Received response from Qwen API: {enhanced_prompt[:100]}...")
        
        # Remove any additional text that might be generated before or after the actual prompt
        if enhanced_prompt:
            logger.info(f"Processing raw API response to extract clean prompt")
            import re
            patterns = [
                r'(?:Updated|Enhanced|Rewritten|New) prompt:\s*(.*?)(?:\n\n|$)',
                r'(?:Here\'s|Here is) the (?:updated|enhanced|rewritten|new) prompt:\s*(.*?)(?:\n\n|$)',
                r'"(.*?)"',  # Quoted text
            ]
            
            for pattern in patterns:
                match = re.search(pattern, enhanced_prompt, re.DOTALL)
                if match:
                    enhanced_prompt = match.group(1).strip()
                    logger.info(f"Extracted clean prompt using pattern match")
                    break
            
            logger.info(f"Final enhanced prompt: {enhanced_prompt[:100]}...")
            return enhanced_prompt
        else:
            # Fallback to original prompt if API response is empty
            logger.info(f"Empty response from API, falling back to original prompt")
            return prompt
            
    except Exception as e:
        logger.error(f"Error enhancing prompt with Qwen API: {str(e)}")
        # Fallback to simple combination only in case of API failure
        enhanced_prompt = prompt
        if has_activity:
            enhanced_prompt += f", {activity}"
        if has_expression:
            enhanced_prompt += f", with {facial_expression} facial expression"
        if has_fur_color:
            enhanced_prompt += f", with {fur_color} fur"
        
        logger.info(f"Fallback prompt after exception: {enhanced_prompt[:100]}...")
        return enhanced_prompt

def generate_manual_variations(base_prompt, variation_request, num_variations=3):
    """
    Generate multiple distinct variations of a prompt based on user request using Qwen
    
    Args:
        base_prompt (str): The original generated prompt to base variations on
        variation_request (str): User's manual request (e.g., "Generate 3 variations of ice cream")
        num_variations (int): Number of variations to generate (default: 3)
        
    Returns:
        list: List of generated variation prompts
    """
    try:
        logger.info(f"🎯 MANUAL VARIATION GENERATION STARTED")
        logger.info(f"Base Prompt: {base_prompt[:100]}...")
        logger.info(f"User Request: {variation_request}")
        logger.info(f"Number of Variations: {num_variations}")
        
        # Create system prompt for variation generation
        sys_prompt = f"""You are an expert prompt variation generator. Your task is to create {num_variations} distinct and unique variations of an image prompt based on the user's specific request.

REQUIREMENTS:
1. Generate EXACTLY {num_variations} different variations
2. Each variation must be intricate and detailed
3. All variations must be connected to the original base prompt
4. Each variation should be unique and distinct from the others
5. Maintain the core subject but vary specific details as requested
6. Each variation should be a complete, standalone prompt
7. Focus on the specific elements mentioned in the user's request

FORMAT:
Return {num_variations} numbered variations, each on a new line:
1. [First variation prompt]
2. [Second variation prompt]
3. [Third variation prompt]

Keep each variation detailed and specific, ensuring they are all different from each other while staying true to the base prompt and user request."""

        # Create user prompt
        user_prompt = f"""Base Prompt: {base_prompt}

User Request: {variation_request}

Please generate {num_variations} distinct variations based on this request. Each variation should be intricate, detailed, and unique while maintaining connection to the base prompt."""

        # Call Qwen API
        logger.info(f"📤 Sending variation generation request to Qwen API")
        response = inference_with_api(
            image_path=None,
            prompt=user_prompt,
            sys_prompt=sys_prompt,
            model_id="qwen2.5-72b-instruct"
        )
        
        logger.info(f"📥 Received response from Qwen API: {response[:200]}...")
        print(f"🔍 DEBUG: Full Qwen response for variations:\n{response}\n")
        
        # Parse the response to extract individual variations
        variations = []
        if response and response.strip():
            lines = response.strip().split('\n')
            logger.info(f"🔍 DEBUG: Split response into {len(lines)} lines")
            
            for i, line in enumerate(lines):
                line = line.strip()
                if not line:
                    continue
                    
                logger.info(f"🔍 DEBUG: Processing line {i+1}: {line[:100]}...")
                
                # Look for numbered variations (1., 2., 3., etc.)
                import re
                match = re.match(r'^\d+\.\s*(.+)', line)
                if match:
                    variation = match.group(1).strip()
                    if variation:
                        variations.append(variation)
                        logger.info(f"✅ Extracted variation {len(variations)}: {variation[:80]}...")
                        print(f"✅ EXTRACTED VARIATION {len(variations)}: {variation}")
                else:
                    # Also try to extract variations that might not be numbered correctly
                    if len(line) > 20:  # Minimum length for a meaningful variation
                        variations.append(line)
                        logger.info(f"✅ Extracted non-numbered variation {len(variations)}: {line[:80]}...")
                        print(f"✅ EXTRACTED NON-NUMBERED VARIATION {len(variations)}: {line}")
        else:
            logger.error(f"❌ Empty or None response from Qwen API")
            print(f"❌ EMPTY RESPONSE FROM QWEN API")
        
        # Ensure we have the requested number of variations
        if len(variations) < num_variations:
            logger.warning(f"⚠️ Only generated {len(variations)} variations instead of {num_variations}")
            # Fill missing variations with modified base prompt
            for i in range(len(variations), num_variations):
                fallback_variation = f"{base_prompt} (variation {i+1})"
                variations.append(fallback_variation)
                logger.info(f"🔄 Added fallback variation {i+1}")
        
        # Limit to requested number if we got more
        variations = variations[:num_variations]
        
        logger.info(f"🎉 Successfully generated {len(variations)} variations")
        return variations
        
    except Exception as e:
        logger.error(f"❌ Error generating manual variations: {str(e)}")
        # Return fallback variations based on base prompt
        fallback_variations = []
        for i in range(num_variations):
            fallback_variation = f"{base_prompt} (variation {i+1})"
            fallback_variations.append(fallback_variation)
        return fallback_variations

def generate_activity_expression_from_prompt(prompt):
    """
    Generate appropriate activity and facial expression based on the prompt using Qwen
    Returns a tuple of (activity, facial_expression)
    """
    try:
        # Extract the subject from the prompt
        subject_match = re.search(r'a\s+([\w\s-]+)', prompt, re.IGNORECASE)
        subject = subject_match.group(1).strip() if subject_match else "character"
        
        # Prepare system prompt for Qwen
        system_prompt = """You are a helpful AI assistant that suggests appropriate activities and facial expressions for characters.

For ACTIVITIES:
1. Generate activities that are INDEPENDENT of any specific image 
2. Any objects mentioned (balls, toys, pillows, etc.) MUST BE PROPORTIONATE in size to the subject (NEVER larger than the subject)
3. STRICTLY LIMIT objects to 1-2 maximum - DO NOT include more than 2 distinct objects
4. Activities should be natural and fitting for the subject
5. Keep activity descriptions concise (4-6 words) but vivid
6. The main subject must always be the dominant element - any objects should be secondary and appropriately sized
7. NEVER suggest activities with multiple or large objects that could visually overpower the main subject

For facial expressions, focus on creating expressive and detailed descriptions that fall into one of these five emotion categories:

1. Happiness: sparkling eyes, gentle up-curved mouth, joyful expression, beaming smile, etc.
2. Sadness: drooping eyes, down-turned mouth, teary gaze, melancholic expression, etc.
3. Anger: narrowed eyes, mouth pulled in snarl, furrowed brow, intense glare, etc.
4. Fear: huge round eyes, parted mouth, startled expression, trembling look, etc.
5. Surprise: popped wide eyes, o-shaped mouth, shocked expression, astonished face, etc.

Your facial expression suggestions should be detailed and include specific descriptions of eyes and mouth expressions.
"""
        
        # Prepare user prompt for Qwen
        user_prompt = f"""Based on this prompt description: "{prompt}"
        
The main subject appears to be: {subject}
        
Please suggest:
1. ONE appropriate activity this subject might be doing (a concise phrase, 4-6 words). If you include objects, make sure they are proportionate in size to the subject and not in excessive quantities. DO NOT base this on any specific image.

2. ONE fitting facial expression for this subject (a detailed phrase, max 8 words) that clearly conveys ONE of these emotions: happiness, sadness, anger, fear, or surprise. Include descriptions of both eyes and mouth.
        
Respond ONLY in this format:
Activity: [concise activity not based on any image]
Expression: [detailed expression including eyes and mouth]
"""
        
        # Call Qwen API
        response = inference_with_api(
            image_path=None,
            prompt=user_prompt, 
            sys_prompt=system_prompt,
            model_id="qwen2.5-72b-instruct"
        )
        
        # Parse the response
        activity = None
        facial_expression = None
        
        if response:
            # Extract activity using regex
            activity_match = re.search(r'Activity:\s*(.+)', response, re.IGNORECASE)
            if activity_match:
                activity = activity_match.group(1).strip()
                # Remove periods at end if any
                activity = re.sub(r'\.$', '', activity)
            
            # Extract expression using regex
            expression_match = re.search(r'Expression:\s*(.+)', response, re.IGNORECASE)
            if expression_match:
                facial_expression = expression_match.group(1).strip()
                # Remove periods at end if any
                facial_expression = re.sub(r'\.$', '', facial_expression)
        
        logger.info(f"Generated activity: {activity}, facial expression: {facial_expression} for prompt: {prompt[:50]}...")
        return activity, facial_expression
    except Exception as e:
        logger.error(f"Error generating activity and expression: {str(e)}")
        return None, None

def generate_fur_color_for_prompt(prompt):
    """
    Generate an appropriate fur color based on the subject in the prompt using Qwen
    or select from a diverse preset list of creative fur colors.
    
    Args:
        prompt (str): The original generated prompt
        
    Returns:
        str: Suggested fur color
    """
    try:
        # Subject-specific color mappings
        subject_color_map = {
            "cat": [
                # Common cat colors as requested by user
                "black", "white", "grey", "gray", "orange", "tabby", 
                "calico", "tortoiseshell", "blue", "cream", "chocolate",
                # Enhanced cat variations
                "tuxedo black and white", "silver tabby", "brown tabby",
                "orange tabby", "blue-cream", "dilute calico",
                "seal point siamese", "flame point siamese"
            ],
            "dog": [
                "golden retriever yellow", "chocolate brown", "black and tan",
                "dalmatian spots", "brindle", "merle blue", "tricolor",
                "fawn", "red", "liver", "blue roan", "sable"
            ],
            "bird": [
                "vibrant rainbow", "scarlet red", "royal blue", "emerald green", 
                "canary yellow", "iridescent black", "pink flamingo",
                "peacock blue", "cardinal red", "turquoise and orange"
            ],
            "fish": [
                "white red", "rainbow scales", "iridescent blue", "goldfish orange",
                "koi pattern", "cobalt blue", "betta red and blue",
                "silver with blue fins", "tiger striped orange", "tropical neon"
            ],
            "reptile": [
                "emerald scales", "red-eyed green", "albino white", 
                "leopard spotted", "tiger striped", "obsidian black",
                "desert sand", "forest moss green", "blue-tongued"
            ],
            "rabbit": [
                "snow white", "dutch pattern", "sandy brown", "spotted white", 
                "rex chocolate", "cottontail gray", "lop ear brown"
            ],
            "rodent": [
                "agouti brown", "hooded pattern", "himalayan white", "cinnamon", 
                "albino pink-eyed", "silver blue", "golden amber"
            ],
            "horse": [
                "chestnut", "bay", "dapple gray", "palomino gold", 
                "buckskin", "pinto pattern", "appaloosa spotted"
            ],
            "fox": [
                "rusty red", "arctic white", "silver", "cross fox pattern", 
                "marble fox mixed", "platinum", "shadow black"
            ],
            "wolf": [
                "timber gray", "arctic white", "black phase", "red wolf amber", 
                "agouti gray brown", "sable", "cream with gray mask"
            ],
            "bear": [
                "grizzly brown", "polar white", "american black", "cinnamon", 
                "panda black white", "kodiak brown", "spirit white"
            ],
            "panda": ["black and white", "red panda russet", "giant panda monochrome"],
            "unicorn": [
                "pearlescent white", "rainbow mane", "lavender shimmer", 
                "cotton candy pink", "celestial silver", "pastel rainbow"
            ],
            "dragon": [
                "ruby scales", "emerald green", "sapphire blue", "obsidian black", 
                "golden scales", "rainbow iridescent", "silver metallic"
            ]
        }
        
        # Define a diverse preset list of creative fur colors - expanded with user's examples
        preset_fur_colors = [
            # User provided examples
            "black", "white", "grey", "gray", "orange", "tabby", 
            "calico", "tortoiseshell", "blue", "cream", "chocolate",
            
            # Natural colors
            "warm golden brown", "deep chestnut", "soft cream", "rich chocolate", 
            "smoky gray", "rusty auburn", "midnight black", "pure white",
            "honey blonde", "ashy brown", "silvery gray", "sandy tan",
            
            # Fantasy colors
            "electric blue", "vibrant purple", "emerald green", "ruby red",
            "sunset orange", "frosty silver", "neon pink", "turquoise blue",
            "lavender purple", "golden yellow", "fiery copper", "rose gold",
            
            # Combinations
            "gray with silver tips", "black with golden streaks", "white with blue accents",
            "brown with copper highlights", "cream with caramel patches", "blue with purple undertones",
            "green with golden flecks", "red with orange gradient", "silver with blue shimmer",
            
            # Patterns & Textures
            "spotted amber", "striped silver", "marbled gray", "dappled gold",
            "speckled white", "ombré blue to purple", "gradient red to orange", "tipped with gold",
            "frosted white tips", "iridescent rainbow", "metallic copper", "pearlescent white"
        ]
        
        # Extract the subject from the prompt
        subject_match = re.search(r'a\s+([\w\s-]+)', prompt, re.IGNORECASE)
        subject = subject_match.group(1).strip() if subject_match else "character"
        subject = subject.lower()
        
        # Check if the subject is in our mapping
        matched_subject = None
        for key in subject_color_map.keys():
            if key in subject:
                matched_subject = key
                break
        
        # 30% chance to use subject-specific colors if we have a match
        import random
        if matched_subject and random.random() < 0.5:
            fur_color = random.choice(subject_color_map[matched_subject])
            logger.info(f"Selected subject-specific fur color: {fur_color} for {matched_subject} in prompt: {prompt[:50]}...")
            return fur_color
        # 30% chance to use a preset fur color for variety
        elif random.random() < 0.4:
            # Select a random fur color from the preset list
            fur_color = random.choice(preset_fur_colors)
            logger.info(f"Selected preset fur color: {fur_color} for prompt: {prompt[:50]}...")
            return fur_color
            
        # Prepare system prompt for Qwen with enhanced instructions for diversity and subject awareness
        system_prompt = """You are a creative AI assistant that suggests unique and visually distinctive fur colors for characters.

Your fur color suggestions should:
1. Be appropriate for the specific animal or creature in the prompt
2. Include breed-specific colors when applicable (like tabby for cats, brindle for dogs, etc.)
3. Be diverse and imaginative, not just common animal colors
4. Include striking combinations and patterns when appropriate
5. Use rich, evocative descriptors (e.g., "shimmering sapphire" rather than just "blue")
6. Vary between natural, fantasy, and unusual options
7. NOT be tied to the character's existing appearance in any reference image
8. Incorporate interesting textures and effects when suitable

Examples of subject-specific fur colors:
- For cats: tabby pattern with silver undertones, calico patches with cream highlights
- For dogs: brindle coat with golden tips, blue merle pattern with copper points  
- For aquatic creatures: iridescent scales with blue-green shimmer
- For fantasy creatures: crystalline fur that changes with the light

Examples of general great fur colors:
- Electric blue with silver sparkles
- Gradient sunset orange to deep crimson
- Frosted emerald green
- Marbled golden amber
- Deep violet with iridescent highlights
- Smoky charcoal with silver tips
- Russet copper with golden undertones"""
        
        # Prepare user prompt for Qwen with emphasis on diversity and subject awareness
        user_prompt = f"""Based on this prompt description: "{prompt}"
        
The main subject appears to be: {subject}
        
Create ONE truly unique and visually striking fur color that would be appropriate for this specific type of character or creature.
If the subject is a specific animal like a cat, dog, fish, etc., consider common and special colors or patterns for that animal.
Your suggestion should be creative, distinctive, and memorable.
        
For example:
- If it's a cat: consider tabby, calico, tortoiseshell, etc.
- If it's a fish: consider scales, patterns, and fin colors
- If it's a fantasy creature: consider unusual and magical colors

Respond ONLY with the fur color description, nothing else. Keep it to 6 words or less."""
        
        # Call Qwen API
        response = inference_with_api(
            image_path=None,
            prompt=user_prompt, 
            sys_prompt=system_prompt,
            model_id="qwen2.5-72b-instruct"
        )
        
        # Clean up the response
        if response:
            # Remove any periods, quotes or extra whitespace
            fur_color = response.strip().rstrip('.').strip('"\'')
            # If the response is too long, truncate it
            words = fur_color.split()
            if len(words) > 6:
                fur_color = ' '.join(words[:6])
            
            logger.info(f"Generated fur color: {fur_color} for prompt: {prompt[:50]}...")
            return fur_color
        
        # If API call fails, fall back to a preset fur color or subject-specific color
        if matched_subject:
            fur_color = random.choice(subject_color_map[matched_subject])
        else:
            fur_color = random.choice(preset_fur_colors)
        logger.info(f"Fallback to fur color: {fur_color} for prompt: {prompt[:50]}...")
        return fur_color
    except Exception as e:
        logger.error(f"Error generating fur color: {str(e)}")
        # Return a preset fur color on error with expanded options
        import random
        preset_fur_colors = [
            "black", "white", "gray", "orange", "tabby", "calico", 
            "tortoiseshell", "blue", "cream", "chocolate", "golden brown", 
            "electric blue", "silvery gray", "emerald green", "ruby red", "midnight black"
        ]
        return random.choice(preset_fur_colors)

# Activity, expression, and fur color re-iteration functions
def reiterate_activity(prompt, use_predefined=True):
    """Generate a new random activity for the prompt"""
    # First, try to determine what kind of animal or character this is
    try:
        subject = extract_subject_from_prompt(prompt)
        logger.info(f"Extracted subject: {subject}")
    except Exception as e:
        logger.error(f"Error extracting subject from prompt: {str(e)}")
        subject = "character"
    
    if use_predefined:
        # Get a random activity from the safe activities list
        return random.choice(ANIMAL_ACTIVITIES)
    else:
        # Use Qwen model to generate a custom activity
        try:
            # Create a more specific prompt to generate diverse activities
            activity_prompt = f"""
            I need you to create ONE creative and detailed activity for a {subject} character.
            
            MUST FOLLOW THESE INSTRUCTIONS:
            - Create ONLY ONE activity description (not a list)
            - Keep it 3-12 words, detailed but concise
            - NEVER start with "playing" or generic verbs like "doing"
            - Use varied, specific, imaginative verbs and vivid descriptions
            - Include context, setting, or props when appropriate
            - Make it age-appropriate, non-violent, and family-friendly
            - The activity should feel dynamic and engaging
            - DO NOT number your response or use bullet points
            - ONLY respond with the activity description, nothing else
            
            Example good activities:
            - "leaping gracefully over a tiny rain puddle"
            - "constructing an elaborate sand castle fortress"
            - "balancing precariously on a stack of colorful books"
            - "exploring an ancient treasure map with a magnifying glass"
            - "crafting a miniature boat from autumn leaves"
            
            Generate ONE activity for a {subject}:
            """
            
            # Generate activity with improved prompt
            response = inference_with_api(None, activity_prompt)
            if response:
                # Clean up response to get just the activity
                activity = response.strip().strip('"').strip("'")
                # Remove any bullet points or numbering that might have been added
                activity = re.sub(r'^\s*[-•*]\s*', '', activity)
                activity = re.sub(r'^\s*\d+\.\s*', '', activity)
                return activity
            else:
                logger.error("Failed to generate activity with Qwen, using predefined activity")
                return random.choice(ANIMAL_ACTIVITIES)
        except Exception as e:
            logger.error(f"Error generating activity with Qwen: {str(e)}")
            return random.choice(ANIMAL_ACTIVITIES)

def reiterate_expression(prompt, use_predefined=True):
    """Regenerate facial expression for the current prompt using QWEN, with expanded animal-specific expressions"""
    if not prompt or prompt.strip() == "":
        return "Please provide a prompt first"
    
    try:
        # Add a counter to track the number of iterations
        if not hasattr(reiterate_expression, 'iteration_counter'):
            reiterate_expression.iteration_counter = 0
        
        # First 5 predefined expressions as requested by the user
        first_five_expressions = [
            "sparkling eyes, gentle up-curved mouth, joyful expression, beaming smile",  # Happiness
            "drooping eyes, down-turned mouth, teary gaze, melancholic expression",     # Sadness
            "narrowed eyes, mouth pulled in snarl, furrowed brow, intense glare",       # Anger
            "huge round eyes, parted mouth, startled expression, trembling look",       # Fear
            "popped wide eyes, o-shaped mouth, shocked expression, astonished face"     # Surprise
        ]
        
        # Extract the subject from the prompt to ensure expressions are appropriate
        subject_match = re.search(r'a\s+([\w\s-]+)', prompt, re.IGNORECASE)
        subject = subject_match.group(1).strip() if subject_match else "character"
        
        # Check if the subject is an animal
        animal_terms = ["cat", "dog", "kitten", "puppy", "animal", "pet", "fox", "wolf", "rabbit", 
                        "bunny", "hamster", "guinea pig", "bird", "parrot", "owl", "ferret"]
        
        is_animal = any(animal in subject.lower() for animal in animal_terms)
        print(f"\n[QWEN EXPRESSION GENERATION - SUBJECT] '{subject}', Is animal: {is_animal}\n")
        
        # Extensive list of preset expressions that are appropriate for animals
        animal_expressions = [
            # Surprise expressions
            "wide eyes with perked ears, alert posture",
            "startled eyes with raised brows, slight gape",
            "wide-eyed surprise with twitching whiskers",
            "alert eyes with stiffened whiskers, attentive",
            "shocked expression with dilated pupils, frozen",
            
            # Sadness expressions
            "droopy eyes with downturned whiskers, forlorn",
            "melancholic gaze with slumped posture, dejected",
            "half-lidded eyes with limp whiskers, disheartened",
            "pleading puppy eyes with subtle frown",
            "distant stare with drooping ears, withdrawn",
            
            # Questioning expressions
            "questioning, head-tilted curiosity with focused eyes, attentive",
            "perplexed stare with twitching ear, confused",
            "inquisitive gaze with whisker twitch, thoughtful",
            "quizzical look with raised eyebrow, puzzled",
            "concentrated focus with head tilt, analytical",
            
            # Happiness expressions
            "bright eyes with relaxed whiskers, contented",
            "squinted eyes with upturned mouth, joyful",
            "sparkling gaze with perked ears, delighted",
            "playful eyes with eager posture, excited",
            "gentle eyes with soft smile, satisfied",
            
            # Curiosity expressions
            "wide-eyed wonder with forward ears, intrigued",
            "attentive gaze with whisker twitch, curious",
            "alert eyes with head tilt, investigating",
            "interested fascinated look stare with forward posture, engaged",
            "fascinated look with twitching nose, observant",
            
            # Sleepiness expressions
            "half-closed eyes with peaceful smile, drowsy",
            "heavy-lidded gaze with relaxed jaw, tired",
            "drooping eyelids with content expression, sleepy",
            "peaceful eyes with gentle breathing, resting",
            "dozing expression with occasional eye flutter",
            
            # Playfulness expressions
            "mischievous eyes with playful smirk, frisky",
            "bright-eyed excitement with eager posture, playful",
            "enthusiastic gaze with ready stance, energetic",
            "gleeful expression with bouncy movement, spirited",
            "impish look with twitching tail, mischievous",
            
            # Contentment expressions
            "slow-blinking eyes with relaxed whiskers, serene",
            "gentle gaze with soft features, peaceful",
            "satisfied look with casual posture, comfortable",
            "tranquil expression with steady breathing, content",
            "harmonious features with gentle look, balanced",
            
            # Focused expressions
            "narrowed eyes with fixed gaze, concentrated",
            "intent stare with frozen posture, fixated",
            "laser-focused eyes with slight head tilt, absorbed",
            "tracking gaze with alert whiskers, watchful",
            "calculated stare with minimal movement, stalking",
            
            # Cautious expressions
            "wary eyes with tentative posture, hesitant",
            "vigilant gaze with ready stance, guarded",
            "careful observation with slow movements, cautious",
            "uncertain look with retreating posture, apprehensive",
            "distrustful squint with tense features, suspicious"
        ]
        
        # Generate a detailed facial expression
        if use_predefined:
            # For the first 5 iterations, use only the 5 specified expressions
            if reiterate_expression.iteration_counter < 5:
                # Use the expression that corresponds to the current iteration counter
                facial_expression = first_five_expressions[reiterate_expression.iteration_counter]
                logger.info(f"Selected expression from first five (iteration {reiterate_expression.iteration_counter + 1}): {facial_expression}")
                print(f"[QWEN EXPRESSION GENERATION - FIRST FIVE] Iteration {reiterate_expression.iteration_counter + 1}: {facial_expression}")
                
                # Increment the counter for the next iteration
                reiterate_expression.iteration_counter += 1
                return facial_expression
            elif is_animal:  # After 5 iterations, use predefined expressions for animals
                # Select a preset animal expression
                facial_expression = random.choice(animal_expressions)
                logger.info(f"Selected preset animal expression: {facial_expression}")
                print(f"[QWEN EXPRESSION GENERATION - PRESET SELECTED] Expression: {facial_expression}")
                return facial_expression
        
        # For non-animals or if use_predefined is false, use the AI to generate a custom expression
        # Update the system prompt to allow for more diverse expressions
        system_prompt = """Generate a detailed facial expression that conveys emotion through eyes, mouth, and posture.
Choose from a wide range of emotions including but not limited to: happiness, sadness, anger, fear, surprise, 
curiosity, contentment, sleepiness, playfulness, focus, caution, questioning, or confusion.

Descriptions should be vivid and detailed, appropriate for the subject, and include:
1. How the eyes look (wide, narrowed, sparkling, teary, droopy, alert, etc.)
2. How the mouth is shaped (smiling, frowning, O-shaped, relaxed, etc.)
3. Additional features related to the subject (ear position, whisker movement, posture, etc. if relevant)
4. The overall emotional quality (joyful, melancholic, startled, curious, etc.)

For ANIMALS, be specific about unique facial features:
- For cats/dogs: Include whisker position, ear orientation, and head tilt
- For birds: Include beak position, feather fluffing, and head movements
- For other animals: Focus on species-specific expressions

Examples for animals:
- "wide eyes with perked ears, alert posture" (surprise)
- "droopy eyes with downturned whiskers, forlorn" (sadness)
- "head-tilted curiosity with focused eyes, attentive" (questioning)
- "bright eyes with relaxed whiskers, contented" (happiness)"""

        user_prompt = f"""The subject is: {subject}
Based on this prompt: "{prompt}"

Generate ONE detailed facial expression (5-8 words) that clearly conveys a specific emotion.
The expression should be appropriate for the subject and include descriptions of eyes AND other relevant features.
If the subject is an animal, include animal-specific details like ear position or whisker movement.
Respond with ONLY the facial expression, no other text."""

        print(f"\n[QWEN EXPRESSION GENERATION - PROMPT]\nSystem: {system_prompt[:200]}...\nUser: {user_prompt}\n")
        
        response = inference_with_api(
            image_path=None,
            prompt=user_prompt, 
            sys_prompt=system_prompt,
            model_id="qwen2.5-72b-instruct"
        )
        
        print(f"\n[QWEN EXPRESSION GENERATION - RAW RESPONSE]\n{response}\n")
        
        if response:
            # Remove any periods, quotes or extra whitespace
            facial_expression = response.strip().rstrip('.').strip('"\'')
            
            # If the response is too long, keep only essential parts
            words = facial_expression.split()
            if len(words) > 10:
                facial_expression = ' '.join(words[:10])
                print(f"[QWEN EXPRESSION GENERATION - TRUNCATED] Original length: {len(words)}, truncated to: {len(facial_expression.split())}")
            
            logger.info(f"Generated facial expression: {facial_expression}")
            print(f"[QWEN EXPRESSION GENERATION - FINAL] {facial_expression}")
            return facial_expression or "Could not generate expression"
        
        print("[QWEN EXPRESSION GENERATION - EMPTY RESPONSE] Could not generate expression")
        return "Could not generate expression"
    except Exception as e:
        logger.error(f"Error regenerating facial expression: {str(e)}")
        print(f"[QWEN EXPRESSION GENERATION - ERROR] {str(e)}")
        return "Error generating facial expression"
        
def reiterate_fur_color(prompt, use_predefined=True):
    """Regenerate fur color for the current prompt using QWEN, ensuring variation from previous generations"""
    if not prompt or prompt.strip() == "":
        return "Please provide a prompt first"
    
    try:
        # Get previously used fur colors from global tracking
        if not hasattr(reiterate_fur_color, 'previously_used_fur_colors'):
            reiterate_fur_color.previously_used_fur_colors = set()
        
        import random
        
        # Extensive list of creative fur colors and patterns
        preset_fur_colors = [
            # Solid colors with rich descriptions
            "iridescent silver with blue undertones",
            "deep midnight blue with silver tips",
            "warm honey gold with amber highlights",
            "rich chocolate brown with caramel accents",
            "soft dove gray with white undercoat",
            "velvety charcoal black with blue sheen",
            "creamy ivory with pale gold highlights",
            "deep emerald green with teal shimmers",
            "royal purple with silver flecks",
            "burgundy red with copper undertones",
            
            # Patterns and combinations
            "silver tabby with charcoal stripes",
            "tortoiseshell with amber patches",
            "calico with rust, cream, and ebony patches",
            "tuxedo with glossy black and pearl white",
            "blue-gray with silvery tiger stripes",
            "marbled swirls of copper and chocolate",
            "spotted pattern of tan and dark brown",
            "dappled golden spots on cream base",
            "brindle pattern with copper and mahogany",
            "salt and pepper with silver tips",
            
            # Fantasy colors
            "celestial blue with star-like silver speckles",
            "nebula purple with galactic swirls",
            "aurora green with shifting blue highlights",
            "sunset ombré from orange to pink",
            "crystalline white with rainbow shimmers",
            "dragon scale green with ruby undertones",
            "moonlight silver with pearlescent sheen",
            "ethereal teal with luminescent edges",
            "stardust sprinkled over deep indigo",
            "twilight gradient from navy to purple",
            
            # Seasonal inspirations
            "autumn russet with golden-leaf patterns",
            "winter frost white with silver-blue tips",
            "spring meadow green with wildflower speckles",
            "summer sunshine gold with amber waves",
            "harvest amber with cinnamon highlights",
            "cherry blossom pink with white patches",
            "forest moss green with bark-like markings",
            "desert sand with terracotta patterns",
            "ocean blue with seafoam white tips",
            "wildfire orange with smoky gray tips",
            
            # Gemstone and metallic
            "sapphire blue with faceted reflections",
            "emerald green with gold shimmer",
            "ruby red with crystal highlights",
            "amethyst purple with lavender undertones",
            "jade green with pearlescent finish",
            "amber gold with honey inclusions",
            "opal white with rainbow reflections",
            "onyx black with silver sparkles",
            "turquoise blue with copper matrix patterns",
            "rose quartz pink with crystalline structure",
            "brushed copper with patina highlights",
            "antique bronze with golden flecks",
            
            # Texture-focused descriptions
            "silky cream with satin finish",
            "fluffy cloud white with downy texture",
            "sleek obsidian black with glossy sheen",
            "plush cinnamon with velvety softness",
            "wispy smoke gray with feathered texture",
            "dense charcoal with cashmere feel",
            "rippled sandy beige with wave patterns",
            "thick russet with woolly undercoat",
            
            # Color combinations with detailed patterns
            "lavender gray with silver tabby markings",
            "champagne beige with chocolate points",
            "misty blue with cloud-like white patches",
            "caramel swirled with cream marble pattern",
            "slate gray with lightning-like white streaks",
            "dusty rose with silver tipped guard hairs",
            "coffee brown with golden dappled spots",
            "stormy gray with electric blue highlights",
            "alabaster white with subtle vanilla stripes",
            "mahogany red with black smoke overlay"
        ]
        
        # Use preset fur colors if use_predefined is True
        if use_predefined:
            # Try to get a color that hasn't been used recently
            for _ in range(5):  # Try up to 5 times to find an unused color
                fur_color = random.choice(preset_fur_colors)
                
                # Check if this color is not in the recently used set
                if fur_color.lower() not in reiterate_fur_color.previously_used_fur_colors:
                    # Add this color to the tracking set
                    reiterate_fur_color.previously_used_fur_colors.add(fur_color.lower())
                    
                    # Keep set size manageable
                    if len(reiterate_fur_color.previously_used_fur_colors) > 15:
                        color_list = list(reiterate_fur_color.previously_used_fur_colors)
                        reiterate_fur_color.previously_used_fur_colors = set(color_list[5:])
                    
                    logger.info(f"Selected preset fur color: {fur_color}")
                    print(f"\n[QWEN FUR COLOR GENERATION - PRESET SELECTED] Color: {fur_color}\n")
                    return fur_color
            
            # If all presets were recently used, still choose a random preset but allow reuse
            fur_color = random.choice(preset_fur_colors)
            logger.info(f"All preset colors recently used, reusing: {fur_color}")
            print(f"\n[QWEN FUR COLOR GENERATION - REUSING PRESET] Color: {fur_color} (all presets recently used)\n")
            return fur_color
                
        # Extract the subject from the prompt for appropriate fur colors
        subject_match = re.search(r'a\s+([\w\s-]+)', prompt, re.IGNORECASE)
        subject = subject_match.group(1).strip() if subject_match else "character"
        print(f"\n[QWEN FUR COLOR GENERATION - SUBJECT] '{subject}'\n")
        
        # Generate a unique and detailed fur color description
        system_prompt = """Generate a CREATIVE and DETAILED fur color description that follows these guidelines:
1. The fur color should be APPROPRIATE and REALISTIC for the subject
2. Be SPECIFIC and DETAILED, mentioning both primary color and any highlights, patterns, or textures
3. Use rich, vivid language that evokes a clear visual image
4. Keep descriptions concise (3-6 words) but descriptive
5. Include interesting undertones, highlights, or patterns where appropriate
6. Avoid generic colors like just "brown" or "white" - be more specific

Examples of great fur colors:
- silver tabby with charcoal stripes
- deep sapphire blue with starry speckles
- warm russet with golden undertones
- frosted gray with silver tips
- dappled honey gold with amber spots"""

        user_prompt = f"""The subject is: {subject}

Generate ONE concise fur color description (3-6 words) that would be appropriate for this subject.
Use rich, specific language and avoid generic color terms.
Include texture, pattern, or highlighting details when appropriate.
Respond with ONLY the fur color description, no other text."""

        print(f"\n[QWEN FUR COLOR GENERATION - PROMPT]\nSystem: {system_prompt}\nUser: {user_prompt}\n")
        
        # Call Qwen API to generate a fur color
        response = inference_with_api(
            image_path=None,
            prompt=user_prompt, 
            sys_prompt=system_prompt,
            model_id="qwen2.5-72b-instruct"
        )
        
        print(f"\n[QWEN FUR COLOR GENERATION - RAW RESPONSE]\n{response}\n")
        
        # Clean up the response
        if response:
            # Remove any periods, quotes or extra whitespace
            fur_color = response.strip().rstrip('.').strip('"\'')
            
            # If the response is too long, keep only essential parts
            words = fur_color.split()
            if len(words) > 8:
                fur_color = ' '.join(words[:8])
                print(f"[QWEN FUR COLOR GENERATION - TRUNCATED] Original length: {len(words)}, truncated to: {len(fur_color.split())}")
            
            # Add to tracking set to avoid repetition
            reiterate_fur_color.previously_used_fur_colors.add(fur_color.lower())
            
            # Keep set size manageable
            if len(reiterate_fur_color.previously_used_fur_colors) > 15:
                color_list = list(reiterate_fur_color.previously_used_fur_colors)
                reiterate_fur_color.previously_used_fur_colors = set(color_list[5:])
            
            logger.info(f"Generated new fur color: {fur_color}")
            print(f"[QWEN FUR COLOR GENERATION - FINAL] {fur_color}")
            return fur_color or "Could not generate fur color"
            
        # Fallback to original method if the direct approach fails
        print("[QWEN FUR COLOR GENERATION - EMPTY RESPONSE] Falling back to secondary generation method")
        fallback_fur_color = generate_fur_color_for_prompt(prompt)
        print(f"[QWEN FUR COLOR GENERATION - FALLBACK RESULT] {fallback_fur_color}")
        return fallback_fur_color or "Could not generate fur color"
    except Exception as e:
        logger.error(f"Error regenerating fur color: {str(e)}")
        print(f"[QWEN FUR COLOR GENERATION - ERROR] {str(e)}")
        return "Error generating fur color"

# New helper functions to update both the input field and modified prompt
def update_with_new_activity_and_prompt(prompt, current_activity=None, current_expression=None, current_fur_color=None, use_predefined_options=True):
    """Update activity and modified prompt when the re-iterate button is clicked"""
    # Always generate a new activity when re-iterate button is clicked
    logger.info(f"Regenerating activity for prompt: {prompt[:50]}..." if prompt and len(prompt) > 50 else f"Regenerating activity for prompt: {prompt}")
    logger.info(f"Current activity: {current_activity}, Using predefined options: {use_predefined_options}")
    print(f"[REGENERATING ACTIVITY] Current: '{current_activity}', Using predefined: {use_predefined_options}")
    
    new_activity = reiterate_activity(prompt, use_predefined=use_predefined_options)
    
    # Create a modified prompt with the activity
    modified_prompt = enhance_prompt_with_activity_expression(
        prompt,
        activity=new_activity,
        facial_expression=current_expression if current_expression else None,
        fur_color=current_fur_color if current_fur_color else None
    )
    
    logger.info(f"New activity generated: {new_activity}")
    print(f"[NEW ACTIVITY GENERATED] {new_activity}")
    logger.info(f"Modified prompt: {modified_prompt[:50]}..." if modified_prompt and len(modified_prompt) > 50 else f"Modified prompt: {modified_prompt}")
    
    return new_activity, modified_prompt

def update_with_new_expression_and_prompt(prompt, current_activity=None, current_expression=None, current_fur_color=None, use_predefined_options=True):
    """Update facial expression and modified prompt when the re-iterate button is clicked"""
    # Always generate a new expression when re-iterate button is clicked
    logger.info(f"Regenerating facial expression for prompt: {prompt[:50]}..." if prompt and len(prompt) > 50 else f"Regenerating facial expression for prompt: {prompt}")
    logger.info(f"Current expression: {current_expression}, Using predefined options: {use_predefined_options}")
    print(f"[REGENERATING FACIAL EXPRESSION] Current: '{current_expression}', Using predefined: {use_predefined_options}")
    
    new_expression = reiterate_expression(prompt, use_predefined=use_predefined_options)
    
    # Create a modified prompt with the new expression
    modified_prompt = enhance_prompt_with_activity_expression(
        prompt,
        activity=current_activity if current_activity else None,
        facial_expression=new_expression,
        fur_color=current_fur_color if current_fur_color else None
    )
    
    logger.info(f"New facial expression generated: {new_expression}")
    print(f"[NEW FACIAL EXPRESSION GENERATED] {new_expression}")
    logger.info(f"Modified prompt: {modified_prompt[:50]}..." if modified_prompt and len(modified_prompt) > 50 else f"Modified prompt: {modified_prompt}")
    
    return new_expression, modified_prompt

def update_with_new_fur_color_and_prompt(prompt, current_activity=None, current_expression=None, current_fur_color=None, use_predefined_options=True):
    """Update fur color and modified prompt when the re-iterate button is clicked"""
    # Always generate a new fur color when re-iterate button is clicked
    logger.info(f"Regenerating fur color for prompt: {prompt[:50]}..." if prompt and len(prompt) > 50 else f"Regenerating fur color for prompt: {prompt}")
    logger.info(f"Current fur color: {current_fur_color}, Using predefined options: {use_predefined_options}")
    print(f"[REGENERATING FUR COLOR] Current: '{current_fur_color}', Using predefined: {use_predefined_options}")
    
    new_fur_color = reiterate_fur_color(prompt, use_predefined=use_predefined_options)
    
    # Create a modified prompt with the new fur color
    modified_prompt = enhance_prompt_with_activity_expression(
        prompt,
        activity=current_activity if current_activity else None,
        facial_expression=current_expression if current_expression else None,
        fur_color=new_fur_color
    )
    
    logger.info(f"New fur color generated: {new_fur_color}")
    print(f"[NEW FUR COLOR GENERATED] {new_fur_color}")
    logger.info(f"Modified prompt: {modified_prompt[:50]}..." if modified_prompt and len(modified_prompt) > 50 else f"Modified prompt: {modified_prompt}")
    
    return new_fur_color, modified_prompt

def create_requirements_file():
    """Create a requirements.txt file if it doesn't exist with all necessary dependencies"""
    if os.path.exists('requirements.txt'):
        logger.info("requirements.txt already exists, skipping creation")
        return
    
    requirements = [
        "fastapi==0.104.1",
        "uvicorn==0.24.0.post1",
        "python-multipart==0.0.6",
        "aiohttp==3.9.0",
        "aiofiles==23.2.1",
        "Pillow==10.1.0",
        "requests-toolbelt==1.0.0",
        "gradio==4.26.0", # Updated Gradio version
        "beautifulsoup4==4.12.2",
        "opencv-python==4.8.1.78",
        "numpy==1.26.2",
        "scikit-image==0.22.0",
        "rembg==2.0.50", # Make sure this is a compatible version
        "scipy==1.11.4", # Added for gaussian_filter
        "werkzeug==3.0.1", # For secure filenames
        "boto3==1.34.8", # For S3 integration
        "botocore==1.34.8", # For S3 integration
        "tqdm==4.66.2",
        "requests==2.31.0",
        "python-dotenv==1.0.1",
        "pytz==2023.3",  # Added for GMT+7 timezone support
        "google-api-python-client==2.111.0", # Added for GDrive
        "google-auth-httplib2==0.1.1",   # Added for GDrive
        "google-auth-oauthlib==1.1.0"    # Added for GDrive
    ]
    
    try:
        with open('requirements.txt', 'w') as f:
            for req in requirements:
                f.write(f"{req}\n")
        logger.info("Created requirements.txt file")
    except Exception as e:
        logger.error(f"Error creating requirements.txt: {str(e)}")

# Add this wrapper function before the create_gradio_ui function
def sync_upload_and_generate_image(*args, **kwargs):
    """Synchronous wrapper for the async upload_and_generate_image function"""
    import asyncio
    
    # Create a new event loop
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    
    try:
        # Run the async function and get the result
        result = loop.run_until_complete(upload_and_generate_image(*args, **kwargs))
        return result
    finally:
        # Clean up the loop
        loop.close()

# Add generate_wrapper function to fix "truth value of array is ambiguous" error
def remove_background_with_white_background(image_path):
    try:
        img = Image.open(image_path).convert("RGBA")
        output = remove(img)
        
        # Create a new image with a white background
        white_bg = Image.new("RGBA", output.size, "WHITE")
        white_bg.paste(output, (0, 0), output)
        
        # Convert to RGB for saving as JPEG or to remove alpha for PNG
        final_image = white_bg.convert("RGB")
        
        # Save the new image
        nobg_path = image_path.replace(os.path.splitext(image_path)[1], "_nobg.png")
        final_image.save(nobg_path)
        logger.info(f"Successfully removed background and saved to {nobg_path}")
        return nobg_path
    except Exception as e:
        logger.error(f"Error in remove_background_with_white_background: {e}")
        return None

def apply_image_to_card_template(image_path, card_template_path, output_size=(1024, 1024)):
    try:
        base_image = Image.open(image_path).convert("RGBA")
        card_template = Image.open(card_template_path).convert("RGBA")

        # Resize card template to output size
        card_template = card_template.resize(output_size)

        # Resize base image to fit within the card template (e.g., 80% of width)
        scale_factor = 0.8
        new_width = int(output_size[0] * scale_factor)
        aspect_ratio = base_image.height / base_image.width
        new_height = int(new_width * aspect_ratio)
        
        resized_image = base_image.resize((new_width, new_height), Image.LANCZOS)

        # Paste the resized image onto the card template
        paste_x = (output_size[0] - new_width) // 2
        paste_y = (output_size[1] - new_height) // 2
        
        card_template.paste(resized_image, (paste_x, paste_y), resized_image)

        # Save the final image
        card_path = image_path.replace(os.path.splitext(image_path)[1], "_card.png")
        card_template.convert("RGB").save(card_path)
        logger.info(f"Successfully applied card template and saved to {card_path}")
        return card_path
    except Exception as e:
        logger.error(f"Error applying card template: {e}")
        return None

def process_dual_outputs(image_paths, card_template_path, theme, category):
    nobg_images = []
    card_images = []

    for img_path in image_paths:
        # 1. Remove background and add white background
        nobg_img = remove_background_with_white_background(img_path)
        if nobg_img:
            nobg_images.append(nobg_img)

        # 2. Apply card template if it exists
        if card_template_path and os.path.exists(card_template_path):
            card_img = apply_image_to_card_template(img_path, card_template_path)
            if card_img:
                card_images.append(card_img)

    return nobg_images, card_images

def create_comprehensive_zip(original_images, nobg_images, card_images, theme, category):
    if not original_images and not nobg_images and not card_images:
        return None

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        # Add original images with "Original_" prefix
        for img_path in original_images or []:
            if os.path.exists(img_path):
                zf.write(img_path, f"Original_{os.path.basename(img_path)}")
        
        # Add white background images with "WhiteBG_" prefix
        for img_path in nobg_images or []:
            if os.path.exists(img_path):
                zf.write(img_path, f"WhiteBG_{os.path.basename(img_path)}")
        
        # Add card template images with "CardTemplate_" prefix
        for img_path in card_images or []:
            if os.path.exists(img_path):
                zf.write(img_path, f"CardTemplate_{os.path.basename(img_path)}")
    
    zip_buffer.seek(0)
    
    # Save the zip file
    zip_filename = f"{theme}_{category}_all_outputs.zip"
    # Use the first available image path to determine the directory
    base_dir = None
    for img_list in [original_images, nobg_images, card_images]:
        if img_list:
            base_dir = os.path.dirname(img_list[0])
            break
    
    if base_dir:
        zip_path = os.path.join(base_dir, zip_filename)
        with open(zip_path, 'wb') as f:
            f.write(zip_buffer.getvalue())
        
        logger.info(f"Created comprehensive ZIP file with all three outputs at {zip_path}")
        return zip_path
    
    return None

def generate_wrapper(
    provider, ref_img, card_img, theme, category, subcategory, 
    # Leonardo params
    leo_model, guidance_scale, generated_prompt, neg_p,  # Updated param name
    # Legacy parameters removed - now using multi-reference system
    preset, leo_num_img,
    # Ideogram params
    ideogram_model, ideogram_style, ideogram_num_img,
    # Common params
    output_f,
    # Additional params for multiple prompts
    extracted_images, all_prompts, current_image_index,
    # Prompt modification
    modification_type=None, modification_details=None, modified_prompt=None,
    # Reference image filename
    ref_img_filename=None,
    # Filename convention
    filename_convention="Current Filename Setting",
    # S3 upload
    upload_to_s3_bucket=False,
    # Optional seed for reproducibility
    seed=None,
    # Activity and expression
    activity=None, facial_expression=None, fur_color=None, ethnicity="Auto",
    # Base64 encoding option
    encode_to_base64=False,
    # Stop flag to cancel generation
    stop_flag=False,
    # Google Drive upload
    upload_to_gdrive=False,
    # Background removal is now always active - both original and processed images are saved
    # Multi-reference image support for Leonardo
    reference_image_1=None, ref_type_1="None", ref_strength_1="Mid",
    reference_image_2=None, ref_type_2="None", ref_strength_2="Mid",
    reference_image_3=None, ref_type_3="None", ref_strength_3="Mid",
    # Ideogram style reference
    ideogram_disable_style_reference=False,
    # Ideogram rendering speed for V3 model
    ideogram_rendering_speed="DEFAULT",

    # Imagen-4 model
    imagen4_model="google/imagen-4",
    # Imagen-4 aspect ratio
    imagen4_aspect_ratio="1:1",
    # Imagen-4 safety filter
    imagen4_safety_filter="block_only_high",
    # Imagen-4 number of images
    imagen4_num_images=1,
    # Counter override parameter
    counter_override=None
):
    """
    Wrapper function for image generation that safely handles parameters 
    and properly converts numpy arrays to Python types
    """
    import shutil  # Import shutil locally to ensure it's available for Imagen-4 processing
    
    # List to keep track of temporary directories to clean up
    temp_dirs_to_cleanup = []
    
    try:
        # Log starting information
        logger.info(f"Starting generate_wrapper with provider: {provider}")
        logger.info(f"Theme: {theme}, Category: {category}, Subcategory: {subcategory}")
        logger.info(f"🔍 SUBCATEGORY DEBUG: subcategory value = '{subcategory}' (type: {type(subcategory)})")
        if subcategory is None:
            logger.warning("⚠️ SUBCATEGORY IS NONE - This may cause uploads to stop at category level")
        
        # DEBUGGING: Print out all input values for key parameters to track the order
        logger.info("=== PARAMETER DEBUGGING INFO ===")
        logger.info(f"seed (raw value, type: {type(seed).__name__}): {seed}")
        logger.info(f"activity (type: {type(activity).__name__}): {activity}")
        logger.info(f"facial_expression (type: {type(facial_expression).__name__}): {facial_expression}")
        logger.info(f"fur_color (type: {type(fur_color).__name__}): {fur_color}")
        logger.info(f"Provider: {provider}")
        logger.info(f"Model: {leo_model if provider == 'Leonardo' else ideogram_model}")
        logger.info(f"Generated Prompt: {generated_prompt[:50]}..." if generated_prompt and len(generated_prompt) > 50 else f"Generated Prompt: {generated_prompt}")
        if modified_prompt:
            logger.info(f"Modified Prompt: {modified_prompt[:50]}..." if len(modified_prompt) > 50 else f"Modified Prompt: {modified_prompt}")
        logger.info(f"Current Image Index: {current_image_index}")
        logger.info(f"S3 Upload: {upload_to_s3_bucket}")
        logger.info(f"Filename Convention: {filename_convention}")
        logger.info(f"Google Drive Upload: {upload_to_gdrive}")
        print(f"[GENERATION STARTING] Provider: {provider}, Theme: {theme}, Category: {category}")
        print(f"[GENERATION PROMPT] {generated_prompt[:100]}..." if generated_prompt and len(generated_prompt) > 100 else f"[GENERATION PROMPT] {generated_prompt}")
        if activity:
            print(f"[ACTIVITY] {activity}")
        if facial_expression:
            print(f"[FACIAL EXPRESSION] {facial_expression}")
        if fur_color:
            print(f"[FUR COLOR] {fur_color}")
        logger.info("=================================")
        
        # Critical fix for the seed parameter
        # Try to directly access and use the raw seed value
        validated_seed = None
        
        if seed is not None:
            # Handle different seed input formats
            try:
                if isinstance(seed, str):
                    if seed.strip():
                        # Check if it's a numeric string
                        if seed.strip().replace('-', '').isdigit():
                            validated_seed = int(seed.strip())
                            logger.info(f"Converted string seed '{seed}' to integer: {validated_seed}")
                        else:
                            logger.warning(f"Seed '{seed}' is not a valid numeric string, ignoring")
                elif isinstance(seed, (int, float)):
                    validated_seed = int(seed)
                    logger.info(f"Using numeric seed directly: {validated_seed}")
                elif isinstance(seed, np.number):
                    validated_seed = int(seed)
                    logger.info(f"Converted NumPy number {seed} to integer: {validated_seed}")
                elif hasattr(seed, 'value') and isinstance(seed.value, (int, float, str)):
                    # Handle Gradio Number component value
                    if isinstance(seed.value, str) and seed.value.strip().replace('-', '').isdigit():
                        validated_seed = int(seed.value.strip())
                    elif isinstance(seed.value, (int, float)):
                        validated_seed = int(seed.value)
                    logger.info(f"Extracted seed value from object with 'value' attribute: {validated_seed}")
                else:
                    logger.warning(f"Unsupported seed type: {type(seed).__name__}, value: {seed}")
            except (ValueError, TypeError) as e:
                logger.error(f"Error converting seed: {str(e)}")
                validated_seed = None
        
        # Use the act_param, exp_param, and fur_param variables for activity, expression, and fur color
        act_param = activity
        exp_param = facial_expression
        fur_param = fur_color

        # Convert numpy arrays to Python native types to prevent Boolean ambiguity errors
        # Handle activity parameter
        if isinstance(act_param, np.ndarray):
            activity_str = str(act_param[0]) if len(act_param) > 0 else ""
            logger.info(f"Converted activity from numpy array to string: {activity_str}")
            act_param = activity_str
        
        # Handle facial_expression parameter
        if isinstance(exp_param, np.ndarray):
            expression_str = str(exp_param[0]) if len(exp_param) > 0 else ""
            logger.info(f"Converted facial_expression from numpy array to string: {expression_str}")
            exp_param = expression_str
            
        # Handle fur_color parameter
        if isinstance(fur_param, np.ndarray):
            fur_color_str = str(fur_param[0]) if len(fur_param) > 0 else ""
            logger.info(f"Converted fur_color from numpy array to string: {fur_color_str}")
            fur_param = fur_color_str
            
        # Handle modification_details parameter
        if isinstance(modification_details, np.ndarray):
            mod_details_str = str(modification_details[0]) if len(modification_details) > 0 else ""
            logger.info(f"Converted modification_details from numpy array to string: {mod_details_str}")
            modification_details = mod_details_str
            
        # Handle modified_prompt parameter
        if isinstance(modified_prompt, np.ndarray):
            mod_prompt_str = str(modified_prompt[0]) if len(modified_prompt) > 0 else ""
            logger.info(f"Converted modified_prompt from numpy array to string: {mod_prompt_str}")
            modified_prompt = mod_prompt_str
        
        # Log the final validated seed value
        logger.info(f"Final validated seed value: {validated_seed}")
        
        # Safe boolean checks for primitives that should be boolean
        safe_upload_to_s3 = bool(upload_to_s3_bucket) if upload_to_s3_bucket is not None else False
        safe_upload_to_gdrive = bool(upload_to_gdrive) if upload_to_gdrive is not None else False # Added for GDrive
        
        # AUTOMATIC GOOGLE DRIVE UPLOAD FOR IMAGEN-4: Always enable Google Drive upload for Imagen-4
        if provider == "Imagen-4":
            safe_upload_to_gdrive = True
            logger.info("🚀 IMAGEN-4 AUTO GDRIVE: Automatically enabled Google Drive upload for Imagen-4")
        
        safe_stop_flag = bool(stop_flag) if stop_flag is not None else False
        safe_encode_to_base64 = bool(encode_to_base64) if encode_to_base64 is not None else False # Added for Base64 encoding
        
        # Log enabled features
        if safe_upload_to_gdrive:
            logger.info("✅ GDRIVE UPLOAD: Enabled - Will upload to Google Drive")
        if safe_encode_to_base64:
            logger.info("Base64 encoding enabled - will encode images to Base64")
        
        # Process image reference safely
        selected_category = category
        logger.info(f"Selected theme: {theme}, category: {selected_category}")
        
        # Fix filename_convention - remove any "Current Filename Setting" text
        if isinstance(filename_convention, str) and filename_convention == "Current Filename Setting":
            filename_convention = "numeric"  # Default to numeric
            logger.info("Changed 'Current Filename Setting' to 'numeric'")
        
        # Determine which prompt to use based on modifications
        # Priority: modified_prompt > modification_details > generated_prompt
        prompt_to_use = generated_prompt
        
        # First check if we have a modified_prompt (from activity/expression button)
        if modified_prompt and isinstance(modified_prompt, str) and len(modified_prompt.strip()) > 0:
            logger.info(f"Using modified prompt (activity/expression): {modified_prompt[:50]}...")
            prompt_to_use = modified_prompt
        else:
            # Safe string check for modification_details 
            has_mod_details = False
            if modification_details is not None:
                if isinstance(modification_details, str):
                    has_mod_details = len(modification_details.strip()) > 0
            
            if has_mod_details:
                # Use Qwen to intelligently mix the initial prompt with the added features
                logger.info(f"Using additional features: {modification_details[:50]}...")
                # Make sure modification_type is not None before calling modify_prompt
                if modification_type is not None:
                    prompt_to_use = modify_prompt(generated_prompt, modification_type, modification_details, current_ref_img)
                    logger.info(f"Modified prompt with features: {prompt_to_use[:50]}...")
                else:
                    # If modification_type is None, just append the details to the prompt
                    prompt_to_use = f"{generated_prompt}, {modification_details}"
                    logger.info(f"Appended features to prompt: {prompt_to_use[:50]}...")
            
        # Remove ", Current Filename Setting" if it somehow got appended to the prompt
        if isinstance(prompt_to_use, str) and prompt_to_use.strip().endswith(", Current Filename Setting"):
            prompt_to_use = prompt_to_use.strip()[:-len(", Current Filename Setting")]
            logger.info(f"Removed 'Current Filename Setting' from prompt: {prompt_to_use[:100]}...")
        
        # For multiple images, handle the ZIP file case
        current_ref_img = ref_img
        
        # Safe array checks for extracted_images
        has_extracted_images = False
        if extracted_images is not None:
            if isinstance(extracted_images, list):
                has_extracted_images = len(extracted_images) > 0
        
        # Safe integer check for current_image_index
        safe_current_index = 0
        if current_image_index is not None:
            try:
                safe_current_index = int(current_image_index)
            except (ValueError, TypeError):
                safe_current_index = 0
        
        # Initialize activity, expression, and fur_color flags before batch processing
        # Safe string checks for activity
        has_activity = False
        if activity is not None:
            if isinstance(activity, str):
                has_activity = len(activity.strip()) > 0
        
        # Safe string checks for facial_expression
        has_expression = False 
        if facial_expression is not None:
            if isinstance(facial_expression, str):
                has_expression = len(facial_expression.strip()) > 0
                
        # Safe string checks for fur_color
        has_fur_color = False
        if fur_color is not None:
            if isinstance(fur_color, str):
                has_fur_color = len(fur_color.strip()) > 0
        
        # Check if ethnicity should be applied
        has_ethnicity = ethnicity and ethnicity != "Auto"
        
        # Check if we have a ZIP file with multiple images for batch processing
        # BUT NOT if we have prompt variations (which use placeholder paths starting with "variation_")
        has_real_multiple_images = False
        if has_extracted_images and len(extracted_images) > 1:
            # Check if these are real image files or just prompt variations
            first_item = extracted_images[0] if extracted_images else ""
            is_variation_mode = isinstance(first_item, str) and first_item.startswith("variation_")
            has_real_multiple_images = not is_variation_mode
        
        # Check if we have prompt variations that need to be processed
        has_prompt_variations = False
        if has_extracted_images and len(extracted_images) > 1:
            first_item = extracted_images[0] if extracted_images else ""
            has_prompt_variations = isinstance(first_item, str) and first_item.startswith("variation_")
        
        if has_real_multiple_images:
            logger.info(f"🔄 COMPREHENSIVE BATCH PROCESSING: Found {len(extracted_images)} images in ZIP file")
            print(f"🔄 COMPREHENSIVE BATCH PROCESSING: Processing {len(extracted_images)} images from ZIP file with comprehensive outputs")
            
            # Use comprehensive output generation for multiple images
            return generate_comprehensive_outputs_for_zip_batch(
                extracted_images, provider, card_img, theme, selected_category,
                leo_model, guidance_scale, prompt_to_use, neg_p, preset, leo_num_img,
                ideogram_model, ideogram_style, ideogram_num_img, output_f,
                filename_convention, safe_upload_to_s3, validated_seed, act_param, exp_param,
                fur_param, ethnicity, safe_stop_flag, safe_upload_to_gdrive,
                safe_encode_to_base64, modification_type, has_activity, has_expression, has_fur_color,
                has_ethnicity, reference_image_1, ref_type_1, ref_strength_1, reference_image_2,
                ref_type_2, ref_strength_2, reference_image_3, ref_type_3, ref_strength_3,
                ideogram_disable_style_reference, ideogram_rendering_speed, modified_prompt,
                all_prompts
            )
        elif has_prompt_variations and all_prompts and len(all_prompts) > 1:
            logger.info(f"🎯 PROMPT VARIATION BATCH PROCESSING: Found {len(all_prompts)} prompt variations")
            print(f"🎯 PROMPT VARIATION BATCH PROCESSING: Processing {len(all_prompts)} prompt variations with single reference image")
            
            # For prompt variations, always use the original reference image (ref_img), not the placeholder paths
            variation_ref_img = ref_img
            
            # Use variation-specific batch processing for multiple prompts with single reference image
            return generate_comprehensive_outputs_for_prompt_variations(
                variation_ref_img, all_prompts, provider, card_img, theme, selected_category, subcategory,
                leo_model, guidance_scale, neg_p, preset, leo_num_img,
                ideogram_model, ideogram_style, ideogram_num_img, output_f,
                filename_convention, safe_upload_to_s3, validated_seed, act_param, exp_param,
                fur_param, ethnicity, safe_stop_flag, safe_upload_to_gdrive,
                safe_encode_to_base64, modification_type, has_activity, has_expression, has_fur_color,
                has_ethnicity, reference_image_1, ref_type_1, ref_strength_1, reference_image_2,
                ref_type_2, ref_strength_2, reference_image_3, ref_type_3, ref_strength_3,
                ideogram_disable_style_reference, ideogram_rendering_speed, modified_prompt,
                imagen4_model, imagen4_aspect_ratio, imagen4_safety_filter, imagen4_num_images,
                counter_override
            )

        
        # Single image processing (original logic)
        if has_extracted_images and safe_current_index < len(extracted_images):
            # Get the reference image for the current index
            current_ref_img = extracted_images[safe_current_index]
            
            # Safe array checks for all_prompts
            has_all_prompts = False
            if all_prompts is not None:
                if isinstance(all_prompts, list):
                    has_all_prompts = len(all_prompts) > 0
            
            if has_all_prompts and safe_current_index < len(all_prompts):
                prompt_to_use = all_prompts[safe_current_index]
                logger.info(f"Using prompt for image {safe_current_index+1}: {prompt_to_use[:50]}...")
        
        # Make sure we have an actual prompt
        if not prompt_to_use or (isinstance(prompt_to_use, str) and prompt_to_use.strip() == ""):
            return (
                [],  # No images 
                "Error: No prompt available for generation. Try uploading a different image.",  # Status
                None,  # No ZIP file
                None,  # No modified images
                None,  # No modified ZIP
                None   # Duplicate output for modified_zip_file_output
            )
        
        # Initialize parameters outside the conditional block to avoid UnboundLocalError
        act_param = None
        exp_param = None
        fur_param = None
        original_prompt = prompt_to_use  # Initialize original_prompt to avoid UnboundLocalError
        
        if has_activity or has_expression or has_fur_color or has_ethnicity:
            original_prompt = prompt_to_use
            # Convert all possible None values to empty strings for safer function calls
            act_param = activity if has_activity else None
            exp_param = facial_expression if has_expression else None
            fur_param = fur_color if has_fur_color else None
        
        # Check if we already have a modified prompt that should NOT be changed
        using_existing_modified_prompt = (modified_prompt and isinstance(modified_prompt, str) and len(modified_prompt.strip()) > 0)
        
        if not using_existing_modified_prompt:
            # Only enhance the prompt if we're not using an existing modified prompt
            # Detect if the image contains humans for fur color decision
            contains_human = False
            if current_ref_img is not None and isinstance(current_ref_img, str):
                contains_human = detect_human_in_image(current_ref_img)
            
            prompt_to_use = enhance_prompt_with_activity_expression(prompt_to_use, act_param, exp_param, fur_param, ethnicity, contains_human)
            logger.info(f"Enhanced prompt with activity/expression/fur color/ethnicity: {prompt_to_use[:100]}...")
        else:
            logger.info(f"🔒 PRESERVING EXISTING MODIFIED PROMPT - NOT applying activity/expression/fur color/ethnicity enhancements")
        
        if original_prompt != prompt_to_use:
            enhancements = []
            if has_activity:
                enhancements.append(f"activity '{activity}'")
            if has_expression:
                enhancements.append(f"facial expression '{facial_expression}'")
            if has_fur_color:
                enhancements.append(f"fur color '{fur_color}'")
            if has_ethnicity:
                enhancements.append(f"ethnicity '{ethnicity}'")
            
            logger.info(f"Prompt enhancement changes: Added {', '.join(enhancements)}")
        
        try:
            # Log the generation attempt
            logger.info(f"Starting image generation with {provider}")
            logger.info(f"Prompt: {prompt_to_use[:100]}...")
            logger.info(f"Theme: {theme}, Category: {selected_category}")
            
            # Check if we are dealing with a file upload that needs processing
            # Handle numpy arrays before boolean check
            if isinstance(current_ref_img, np.ndarray):
                logger.info(f"Reference image is a numpy array of shape {current_ref_img.shape if hasattr(current_ref_img, 'shape') else 'unknown'}")
                # Add debugging to print a sample of the array data
                sample_str = str(current_ref_img[:5, :5] if hasattr(current_ref_img, 'shape') and len(current_ref_img.shape) >= 2 else 'cannot display sample')
                logger.info(f"Reference image array data sample: {sample_str}") 
                
                # Check if it's an actual image array (2D or 3D array with image data)
                if hasattr(current_ref_img, 'ndim') and current_ref_img.ndim in [2, 3]:
                    try:
                        # It's actual image data, save it to a temporary file
                        import tempfile
                        from PIL import Image
                        
                        # Create a temporary directory if needed
                        temp_dir = tempfile.mkdtemp()
                        temp_image_path = os.path.join(temp_dir, "reference_image.png")
                        
                        # Add to cleanup list
                        temp_dirs_to_cleanup.append(temp_dir)
                        
                        # Convert numpy array to PIL Image and save
                        try:
                            # Check array dimensions safely
                            is_grayscale = (current_ref_img.ndim == 2) or (current_ref_img.ndim == 3 and current_ref_img.shape[2] == 1)
                            if is_grayscale:
                                # Grayscale image
                                Image.fromarray(current_ref_img.astype(np.uint8)).save(temp_image_path)
                            else:
                                # Color image (RGB or RGBA)
                                Image.fromarray(current_ref_img.astype(np.uint8)).save(temp_image_path)
                        except Exception as e:
                            logger.warning(f"Error handling numpy array dimensions: {str(e)}, using fallback conversion")
                            # Fallback: try direct conversion
                            Image.fromarray(current_ref_img.astype(np.uint8)).save(temp_image_path)
                        
                        logger.info(f"Saved numpy array image data to temporary file: {temp_image_path}")
                        current_ref_img = temp_image_path
                    except Exception as e:
                        logger.error(f"Failed to save numpy array image data: {str(e)}")
                        current_ref_img = None
                else:
                    # It's not valid image data, try to convert the first element if it's a string
                    if current_ref_img.size > 0:
                        try:
                            if isinstance(current_ref_img[0], (str, bytes)):
                                current_ref_img = str(current_ref_img[0])
                                logger.info(f"Converted numpy array to path string: {current_ref_img}")
                            else:
                                logger.warning(f"Numpy array first element is not a string: {type(current_ref_img[0])}")
                                current_ref_img = None
                        except (IndexError, TypeError) as e:
                            logger.error(f"Error accessing numpy array first element: {str(e)}")
                            current_ref_img = None
                    else:
                        logger.warning("Numpy array doesn't contain valid image data or a path string")
                        current_ref_img = None
            
            # Now safe to use in boolean context
            if current_ref_img and isinstance(current_ref_img, dict) and 'name' in current_ref_img:
                current_ref_img = current_ref_img['name']
            
            # Validate reference images exist if controlnet options are enabled (new multi-reference system)
            any_reference_type_selected = (
                (ref_type_1 and ref_type_1 != "None") or 
                (ref_type_2 and ref_type_2 != "None") or 
                (ref_type_3 and ref_type_3 != "None")
            )
            
            if any_reference_type_selected and provider == "Leonardo":
                # Check if we have at least one reference image when reference types are selected
                has_reference_images = (
                    reference_image_1 is not None or 
                    reference_image_2 is not None or 
                    reference_image_3 is not None or
                    current_ref_img is not None  # Legacy reference image support
                )
                
                if not has_reference_images:
                    logger.error("ControlNet reference types selected, but no reference images provided")
                    return (
                        [],  # No images
                        "Error: ControlNet reference types selected, but no reference images were provided. Please upload at least one reference image.",
                        None,  # No ZIP file
                        None,  # No modified images
                        None,  # No modified ZIP
                        None   # Duplicate output for modified_zip_file_output
                    )
                else:
                    logger.info(f"ControlNet reference types selected with available reference images")
            
            # Same for card template
            if isinstance(card_img, np.ndarray):
                logger.info(f"Converting numpy array card template to string")
                if card_img.size > 0:
                    card_img = str(card_img[0])
                else:
                    card_img = None
            
            # Handle tuple input for card_img
            if isinstance(card_img, tuple):
                logger.info(f"Card template is a tuple: {card_img}")
                if len(card_img) > 0 and isinstance(card_img[0], str):
                    card_img = card_img[0]
                    logger.info(f"Extracted file path from tuple: {card_img}")
                else:
                    logger.warning(f"Cannot extract valid file path from tuple: {card_img}")
                    card_img = None

            if card_img and isinstance(card_img, dict) and 'name' in card_img:
                card_img = card_img['name']
                
            # Verify that the card_img exists
            if card_img and isinstance(card_img, str):
                if not os.path.exists(card_img):
                    logger.warning(f"Card template path does not exist: {card_img}")
                else:
                    logger.info(f"Using card template: {card_img}")
            
            # Extract reference image filename for optional filename convention
            ref_filename = None
            
            # Handle possible numpy array in ref_img_filename
            if isinstance(ref_img_filename, np.ndarray):
                logger.info(f"Converting numpy array ref_img_filename to string")
                if ref_img_filename.size > 0:
                    ref_img_filename = str(ref_img_filename[0])
                else:
                    ref_img_filename = None
                    
            if ref_img_filename and isinstance(ref_img_filename, str) and ref_img_filename.strip():
                ref_filename = ref_img_filename
            elif current_ref_img is not None:
                # Handle possible numpy array in current_ref_img
                if isinstance(current_ref_img, np.ndarray):
                    if current_ref_img.size > 0:
                        # Convert numpy array to string for filename extraction
                        current_ref_img_str = str(current_ref_img[0])
                        ref_filename = os.path.basename(current_ref_img_str)
                    else:
                        ref_filename = None
                elif isinstance(current_ref_img, str):
                    ref_filename = os.path.basename(current_ref_img)
                    if '.' in ref_filename:
                        ref_filename = ref_filename.rsplit('.', 1)[0]  # Remove extension
            
            # Initialize payload dictionary for parameters
            payload = {}
            
            # Validate seed parameter to ensure it's a valid integer
            validated_seed = None
            if seed is not None:
                try:
                    # We already did the detailed validation at the start of the function,
                    # so we can just convert to int directly if possible
                    validated_seed = int(seed) if seed is not None else None
                    if validated_seed is not None:
                        payload["seed"] = validated_seed
                        logger.info(f"Using seed value: {validated_seed} for generation")
                except (ValueError, TypeError):
                    logger.warning(f"Invalid seed value during final validation: {seed}, ignoring")
                    validated_seed = None
            
            # Check if Imagen-4 is selected as provider
            if provider == "Imagen-4":
                logger.info("🚀 Using Imagen-4 Image Generation")
                
                # Validate reference image for Imagen-4 (fix array ambiguity)
                has_ref_img = False
                if current_ref_img is not None:
                    if isinstance(current_ref_img, str):
                        has_ref_img = len(current_ref_img.strip()) > 0
                    elif isinstance(current_ref_img, np.ndarray):
                        has_ref_img = current_ref_img.size > 0
                    else:
                        has_ref_img = True
                
                if not has_ref_img:
                    return (
                        [],  # No images
                        "❌ Imagen-4 requires a reference image. Please upload Reference Image 1.",  # Status
                        None,  # No ZIP file
                        [],  # No images for inpainting
                        None,  # No modified ZIP
                        None   # Duplicate output for modified_zip_file_output
                    )
                
                # Check if we have prompts for Imagen-4 (fix array ambiguity)
                has_prompt = False
                if prompt_to_use is not None:
                    if isinstance(prompt_to_use, str):
                        has_prompt = len(prompt_to_use.strip()) > 0
                    elif isinstance(prompt_to_use, np.ndarray):
                        has_prompt = prompt_to_use.size > 0
                    else:
                        has_prompt = True
                
                if not has_prompt:
                    return (
                        [],  # No images
                        "❌ Imagen-4 requires a prompt. Please generate or provide a prompt.",  # Status
                        None,  # No ZIP file
                        [],  # No images for inpainting
                        None,  # No modified ZIP
                        None   # Duplicate output for modified_zip_file_output
                    )
                
                try:
                    # Create output directories
                    out_folder = "imagen4_processed_images"
                    if os.path.exists(out_folder):
                        shutil.rmtree(out_folder)
                    os.makedirs(out_folder)
                    
                    # Extract reference filename for Imagen-4
                    ref_filename = None
                    if current_ref_img and isinstance(current_ref_img, str):
                        ref_filename = os.path.basename(current_ref_img)
                        if '.' in ref_filename:
                            ref_filename = ref_filename.rsplit('.', 1)[0]  # Remove extension
                    
                    # Process the reference image with Imagen-4
                    ref_bytes = safe_file_handler_imagen4(current_ref_img)
                    if not ref_bytes:
                        return (
                            [],  # No images
                            "❌ Failed to process reference image for Imagen-4",  # Status
                            None,  # No ZIP file
                            [],  # No images for inpainting
                            None,  # No modified ZIP
                            None   # Duplicate output for modified_zip_file_output
                        )
                    
                    logger.info(f"🎨 Imagen-4 FULL PROMPT: {prompt_to_use}")
                    logger.info(f"🎨 Imagen-4 MODEL: {imagen4_model}")
                    logger.info(f"Imagen-4 Settings - Aspect ratio: {imagen4_aspect_ratio}, Safety filter: {imagen4_safety_filter}")
                    
                    # Ensure Imagen-4 uses the correct prompt (prioritizing modified_prompt for activity/expression)
                    if modified_prompt and isinstance(modified_prompt, str) and len(modified_prompt.strip()) > 0:
                        logger.info(f"🎨 Imagen-4: Using modified prompt from activity/expression: {modified_prompt[:100]}...")
                        imagen4_prompt = modified_prompt
                    else:
                        logger.info(f"🎨 Imagen-4: Using standard prompt: {prompt_to_use[:100]}...")
                        imagen4_prompt = prompt_to_use
                    
                    # Generate images with Imagen-4 using the reference image and actual UI parameters
                    logger.info(f"🎨 Imagen-4: Generating {imagen4_num_images} image(s) with settings: aspect_ratio={imagen4_aspect_ratio}, safety_filter={imagen4_safety_filter}")
                    result_b64_list = generate_single_image_imagen4(ref_bytes, [], imagen4_prompt, imagen4_aspect_ratio, imagen4_num_images, imagen4_safety_filter, imagen4_model)
                    
                    if result_b64_list:
                        # Process multiple images
                        image_paths = []
                        
                        for i, result_b64 in enumerate(result_b64_list):
                            # Convert to PIL Image
                            from PIL import Image
                            image_data = base64.b64decode(result_b64)
                            image = Image.open(io.BytesIO(image_data))
                            
                            # Ensure the image is exactly 1024x1024 (no portrait/landscape)
                            if image.size != (1024, 1024):
                                logger.info(f"🔄 Imagen-4: Resizing image from {image.size} to 1024x1024")
                                image = image.resize((1024, 1024), Image.Resampling.LANCZOS)
                                logger.info(f"✅ Imagen-4: Successfully resized to 1024x1024")
                            
                            # Use sequential numbering for multiple images (same as Leonardo)
                            if i > 0:
                                file_num = next_file_number + i
                                # Always use the numeric convention format: ThemeCodeCategoryCodeImageNumber
                                base_filename = f"{theme_code}{category_code}{file_num:05d}"
                            else:
                                # For the first image, use the base filename pattern directly
                                base_filename = base_filename_pattern
                            
                            # Create output filename using TTCCCNNNNN format
                            output_filename = f"{base_filename}.png"
                            output_path = os.path.join(out_folder, output_filename)
                            
                            # Save the initial generated image
                            image.save(output_path, "PNG", optimize=True)
                            logger.info(f"✅ Successfully generated Imagen-4 image {i+1}/{len(result_b64_list)}: {output_filename}")
                            
                            # IMPORTANT: Save the original image before background removal (same as Leonardo)
                            original_filename = f"{base_filename}_original.png"
                            original_path = os.path.join(out_folder, original_filename)
                            image.save(original_path, "PNG", optimize=True)
                            logger.info(f"✅ Saved original Imagen-4 image before background removal: {original_filename}")
                            
                            # Apply background removal process (same as other providers)
                            logger.info(f"🎨 Imagen-4: Applying automatic background removal with birefnet_hr to {output_path}")
                            processed_img = remove_background_birefnet_hr(output_path)
                            
                            if processed_img is not None:
                                # Apply improved alpha edge processing for better appearance
                                logger.info(f"🎨 Imagen-4: Improving alpha channel edges after background removal")
                                processed_img = improve_alpha_edges(processed_img, threshold=180, edge_feather=4, use_gaussian_blur=True, feather_intensity=0.6)
                                
                                # Save the transparent background version for card template use
                                transparent_img = processed_img.copy()
                                
                                # Apply white background for the main output image
                                canvas = Image.new("RGBA", processed_img.size, "WHITE")
                                try:
                                    img_bands = processed_img.split()
                                    if processed_img.mode == 'RGBA' and len(img_bands) == 4:
                                        canvas.paste(processed_img, mask=img_bands[3])
                                    else:
                                        canvas.paste(processed_img)
                                except Exception as e:
                                    logger.warning(f"Error handling image channels: {str(e)}, using fallback paste")
                                    canvas.paste(processed_img)
                                processed_img = canvas
                                
                                # Save the processed image (overwrites the original)
                                processed_img.save(output_path, "PNG", optimize=True)
                                logger.info(f"✅ Imagen-4: Applied background removal to {output_filename}")
                                
                                # Apply to card template if provided
                                if card_img and os.path.exists(card_img):
                                    try:
                                        logger.info(f"🎨 Imagen-4: Applying to card template: {card_img}")
                                        card_template = Image.open(card_img).convert("RGBA")
                                        
                                        # Ensure the transparent image is in RGBA mode
                                        if transparent_img.mode != 'RGBA':
                                            transparent_img = transparent_img.convert('RGBA')
                                        
                                        # Apply the transparent image to the card template
                                        card_with_image = place_image_on_card(card_template.copy(), transparent_img, preserve_original_alpha=True)
                                        
                                        # Create card filename using TTCCCNNNNN format
                                        card_filename = f"{base_filename}_card.png"
                                        card_path = os.path.join(out_folder, card_filename)
                                        card_with_image.save(card_path, format='PNG')
                                        image_paths.append(card_path)  # Add card image to paths
                                        logger.info(f"✅ Imagen-4: Applied to card template and saved to: {card_path}")
                                    except Exception as card_error:
                                        logger.error(f"❌ Imagen-4: Error applying to card template: {str(card_error)}")
                                else:
                                    logger.info(f"ℹ️ Imagen-4: No card template provided, skipping card application")
                            else:
                                logger.warning(f"⚠️ Imagen-4: Background removal failed for {output_path}, using original image")
                            
                            # Add both original and processed images to the list (same as Leonardo)
                            image_paths.append(original_path)  # Add original first
                            image_paths.append(output_path)   # Then processed
                        
                        # Create Excel file for Imagen-4 generation (same as Leonardo)
                        from openpyxl import Workbook
                        from openpyxl.drawing.image import Image as XLImage
                        
                        # Function to add image to cell while maintaining aspect ratio
                        def add_image_to_cell_with_aspect_ratio(worksheet, img_path, cell_reference, max_width=200, max_height=150):
                            """Add an image to a specific cell while maintaining aspect ratio"""
                            try:
                                # Open and process the image
                                with Image.open(img_path) as img:
                                    # Get original dimensions
                                    orig_width, orig_height = img.size
                                    aspect_ratio = orig_width / orig_height
                                    
                                    # Calculate new dimensions while maintaining aspect ratio
                                    if aspect_ratio > 1:  # Landscape
                                        new_width = min(max_width, orig_width)
                                        new_height = int(new_width / aspect_ratio)
                                        if new_height > max_height:
                                            new_height = max_height
                                            new_width = int(new_height * aspect_ratio)
                                    else:  # Portrait or square
                                        new_height = min(max_height, orig_height)
                                        new_width = int(new_height * aspect_ratio)
                                        if new_width > max_width:
                                            new_width = max_width
                                            new_height = int(new_width / aspect_ratio)
                                    
                                    # Resize image while maintaining aspect ratio
                                    img_resized = img.resize((new_width, new_height), Image.LANCZOS)
                                    
                                    # Create an in-memory file-like object for the image
                                    img_buffer = io.BytesIO()
                                    img_resized.save(img_buffer, format='PNG')
                                    img_buffer.seek(0)
                                    
                                    # Create an openpyxl image object
                                    xl_img = XLImage(img_buffer)
                                    
                                    # Get the cell to position the image properly
                                    cell = worksheet[cell_reference]
                                    
                                    # Position the image in the cell
                                    xl_img.anchor = cell_reference
                                    worksheet.add_image(xl_img)
                                    
                                    return True
                                    
                            except Exception as e:
                                logger.error(f"Error adding image to cell {cell_reference}: {str(e)}")
                                return False

                        # Generate Excel file with image details - use TTCCCNNNNN format
                        excel_filename = f"{base_filename}_imagen4.xlsx"
                        excel_path = os.path.join(out_folder, excel_filename)
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Generated Images"
                        
                        # Add comprehensive headers including all metadata (same as Leonardo)
                        headers = [
                            "Generated Prompt",          # A
                            "Modified Prompt",           # B  
                            "Output Filename",           # C
                            "Reference Image",           # D
                            "Generated Image Original",  # E - NEW: Original image before background removal
                            "Generated Image",           # F - After background removal
                            "Card Image",               # G - After background removal + card template
                            "Activity",                 # H
                            "Facial Expression",        # I
                            "Fur Color",                # J
                            "Theme",                    # K
                            "Category",                 # L
                            "Provider",                 # M
                            "Model",                    # N
                            "Timestamp"                 # O
                        ]
                        ws.append(headers)
                        
                        # Set column widths for better visibility
                        column_widths = {
                            'A': 60,  # Generated Prompt
                            'B': 60,  # Modified Prompt
                            'C': 35,  # Output Filename
                            'D': 25,  # Reference Image
                            'E': 25,  # Generated Image Original
                            'F': 25,  # Generated Image
                            'G': 25,  # Card Image
                            'H': 20,  # Activity
                            'I': 20,  # Facial Expression
                            'J': 20,  # Fur Color
                            'K': 15,  # Theme
                            'L': 15,  # Category
                            'M': 15,  # Provider
                            'N': 20,  # Model
                            'O': 20   # Timestamp
                        }
                        
                        for col_letter, width in column_widths.items():
                            ws.column_dimensions[col_letter].width = width
                        
                        # Get current timestamp for metadata
                        current_timestamp = get_gmt7_timestamp()
                        
                        # Organize image paths by type
                        original_images = []
                        processed_images = []
                        card_images = []
                        
                        for file_path in image_paths:
                            filename = os.path.basename(file_path)
                            if '_original.png' in filename:
                                original_images.append(file_path)
                            elif '_card' in filename:
                                card_images.append(file_path)
                            else:
                                processed_images.append(file_path)
                        
                        # Process each generated image and add to the Excel file
                        # Create one row per original image (base image)
                        for i in range(len(original_images)):
                            original_path = original_images[i] if i < len(original_images) else None
                            processed_path = processed_images[i] if i < len(processed_images) else None
                            card_path = card_images[i] if i < len(card_images) else None
                            
                            # Add a new row
                            row_num = i + 2  # Start from row 2
                            ws.row_dimensions[row_num].height = 150
                            
                            # Add generated prompt to column A
                            ws.cell(row=row_num, column=1, value=generated_prompt if generated_prompt else prompt_to_use)
                            
                            # Add modified prompt to column B
                            ws.cell(row=row_num, column=2, value=prompt_to_use)
                            
                            # Add filename to column C (use original filename without _original suffix)
                            if original_path:
                                original_filename = os.path.basename(original_path).replace('_original.png', '.png')
                                ws.cell(row=row_num, column=3, value=original_filename)
                            
                            # Add reference image to column D if available
                            if current_ref_img and os.path.exists(current_ref_img):
                                try:
                                    add_image_to_cell_with_aspect_ratio(ws, current_ref_img, f'D{row_num}')
                                except Exception as e:
                                    logger.error(f"Error adding reference image to Excel: {str(e)}")
                            
                            # Add original image to column E (before background removal)
                            if original_path and os.path.exists(original_path):
                                try:
                                    add_image_to_cell_with_aspect_ratio(ws, original_path, f'E{row_num}')
                                except Exception as e:
                                    logger.error(f"Error adding original image to Excel: {str(e)}")
                            
                            # Add processed image to column F (after background removal)
                            if processed_path and os.path.exists(processed_path):
                                try:
                                    add_image_to_cell_with_aspect_ratio(ws, processed_path, f'F{row_num}')
                                except Exception as e:
                                    logger.error(f"Error adding processed image to Excel: {str(e)}")
                            
                            # Add card image to column G (after background removal + card template)
                            if card_path and os.path.exists(card_path):
                                try:
                                    add_image_to_cell_with_aspect_ratio(ws, card_path, f'G{row_num}')
                                except Exception as e:
                                    logger.error(f"Error adding card image to Excel: {str(e)}")
                            
                            # Add metadata columns (adjusted for new column structure)
                            ws.cell(row=row_num, column=8, value=activity if activity else '')  # Activity (H)
                            ws.cell(row=row_num, column=9, value=facial_expression if facial_expression else '')  # Facial Expression (I)
                            ws.cell(row=row_num, column=10, value=fur_color if fur_color else '')  # Fur Color (J)
                            ws.cell(row=row_num, column=11, value=theme if theme else '')  # Theme (K)
                            ws.cell(row=row_num, column=12, value=selected_category if selected_category else '')  # Category (L)
                            ws.cell(row=row_num, column=13, value='Imagen-4')  # Provider (M)
                            ws.cell(row=row_num, column=14, value='google/imagen-4')  # Model (N)
                            ws.cell(row=row_num, column=15, value=current_timestamp)  # Timestamp (O)
                        
                        # Save the Excel file
                        try:
                            wb.save(excel_path)
                            logger.info(f"✅ Imagen-4: Excel file generated at {excel_path}")
                        except Exception as excel_save_error:
                            logger.error(f"❌ Imagen-4: Error saving Excel file: {excel_save_error}")
                            excel_path = None
                        
                        # Create ZIP file with S3/GDrive integration for Imagen-4
                        try:
                            import zipfile
                            import tempfile
                            
                            # Create ZIP file with all generated images, Excel report, and base64 data
                            zip_filename = f"{base_filename}_imagen4.zip"
                            zip_path = os.path.join(out_folder, zip_filename)
                            
                            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                                # Add all generated images to ZIP
                                for img_path in image_paths:
                                    if os.path.exists(img_path):
                                        arcname = os.path.basename(img_path)
                                        zipf.write(img_path, arcname)
                                        logger.info(f"Added {arcname} to ZIP file")
                                
                                # Add Excel report to ZIP
                                if excel_path and os.path.exists(excel_path):
                                    excel_arcname = os.path.basename(excel_path)
                                    zipf.write(excel_path, excel_arcname)
                                    logger.info(f"Added Excel report {excel_arcname} to ZIP file")
                                
                                # Add base64 data as text files for each image
                                for i, result_b64 in enumerate(result_b64_list):
                                    b64_filename = f"imagen4_image_{i+1}_base64.txt"
                                    zipf.writestr(b64_filename, result_b64)
                                    logger.info(f"Added base64 data as {b64_filename} to ZIP file")
                                
                                # Add metadata file
                                metadata = {
                                    "provider": "Imagen-4",
                                    "model": "google/imagen-4",
                                    "prompt": prompt_to_use,
                                    "aspect_ratio": imagen4_aspect_ratio,
                                    "safety_filter": imagen4_safety_filter,
                                    "num_images": len(result_b64_list),
                                    "timestamp": timestamp,
                                    "theme": theme,
                                    "category": selected_category,
                                    "reference_image": ref_filename if ref_filename else None,
                                    "excel_report": excel_arcname if excel_path and os.path.exists(excel_path) else None
                                }
                                import json
                                zipf.writestr("metadata.json", json.dumps(metadata, indent=2))
                            
                            logger.info(f"✅ Created ZIP file: {zip_path}")
                            
                            # Handle S3 upload if enabled
                            s3_url = None
                            if safe_upload_to_s3:
                                try:
                                    s3_url = upload_to_s3(zip_path, theme, selected_category)
                                    if s3_url:
                                        logger.info(f"✅ Imagen-4 ZIP uploaded to S3: {s3_url}")
                                    else:
                                        logger.warning("❌ Failed to upload Imagen-4 ZIP to S3")
                                except Exception as s3_error:
                                    logger.error(f"❌ S3 upload error for Imagen-4: {s3_error}")
                            
                            # Handle Google Drive upload if enabled
                            gdrive_url = None
                            if safe_upload_to_gdrive:
                                try:
                                    gdrive_result = upload_to_google_drive(
                                        zip_path, 
                                        theme=theme, 
                                        category=selected_category,
                                        subcategory=subcategory  # FIXED: Add missing subcategory parameter
                                    )
                                    if gdrive_result:
                                        gdrive_url = gdrive_result
                                        logger.info(f"✅ Imagen-4 ZIP uploaded to Google Drive")
                                    else:
                                        logger.warning("❌ Failed to upload Imagen-4 ZIP to Google Drive")
                                except Exception as gdrive_error:
                                    logger.error(f"❌ Google Drive upload error for Imagen-4: {gdrive_error}")
                            
                            # Upload individual images and Excel file to S3/GDrive if enabled
                            files_to_upload = image_paths.copy()
                            if excel_path and os.path.exists(excel_path):
                                files_to_upload.append(excel_path)
                            
                            for file_path in files_to_upload:
                                if safe_upload_to_s3:
                                    try:
                                        # Create bucket folder path from theme and category
                                        bucket_folder = "imagen4_outputs"
                                        if theme:
                                            bucket_folder = f"{bucket_folder}/{theme}"
                                            if selected_category:
                                                bucket_folder = f"{bucket_folder}/{selected_category}"
                                        
                                        file_s3_url = upload_to_s3(file_path, bucket_folder=bucket_folder)
                                        if file_s3_url:
                                            logger.info(f"✅ Uploaded {os.path.basename(file_path)} to S3")
                                    except Exception as file_s3_error:
                                        logger.error(f"❌ Failed to upload {os.path.basename(file_path)} to S3: {file_s3_error}")
                                
                                if safe_upload_to_gdrive:
                                    try:
                                        file_gdrive_result = upload_to_google_drive(
                                            file_path, 
                                            theme=theme, 
                                            category=selected_category,
                                            subcategory=subcategory  # Pass subcategory
                                        )
                                        if file_gdrive_result:
                                            logger.info(f"✅ Uploaded {os.path.basename(file_path)} to Google Drive")
                                    except Exception as file_gdrive_error:
                                        logger.error(f"❌ Failed to upload {os.path.basename(file_path)} to Google Drive: {file_gdrive_error}")
                            
                            # Create display images with metadata
                            display_images, ref_image_path = create_display_images_with_metadata(
                                image_paths, 
                                [current_ref_img] if current_ref_img else [], 
                                [],
                                reference_filename=ref_filename
                            )
                            
                            # Create status message with upload info (consistent with other providers)
                            original_count = len(original_images)
                            processed_count = len(processed_images)
                            card_count = len(card_images)
                            
                            status_parts = [f"Generation complete with Imagen-4! Generated {original_count} image(s) (aspect ratio: {imagen4_aspect_ratio}, safety filter: {imagen4_safety_filter}) with automatic background removal"]
                            if current_ref_img and os.path.exists(current_ref_img):
                                status_parts.append("Using reference image with style reference")
                            if card_count > 0:
                                status_parts[0] = f"Generation complete with Imagen-4! Generated {original_count} image(s) and {card_count} card image(s) (aspect ratio: {imagen4_aspect_ratio}, safety filter: {imagen4_safety_filter})"
                            if excel_path and os.path.exists(excel_path):
                                status_parts.append("Excel report generated")
                            if s3_url:
                                status_parts.append("Uploaded to S3")
                            if gdrive_url:
                                status_parts.append("Uploaded to Google Drive")
                            status_message = ". ".join(status_parts)
                            
                            return (
                                display_images,  # Images for gallery display
                                status_message,  # Status with upload info
                                zip_path,        # ZIP file path for download
                                image_paths,     # Store original image paths for inpainting
                                None,           # No modified ZIP
                                None            # Duplicate output for modified_zip_file_output
                            )
                            
                        except Exception as zip_error:
                            logger.error(f"❌ Error creating ZIP file for Imagen-4: {zip_error}")
                            # Fallback to basic processing
                            # Create fallback status message
                            fallback_status = f"Generation complete with Imagen-4! Generated {len(result_b64_list)} image(s) with automatic background removal"
                            if current_ref_img and os.path.exists(current_ref_img):
                                fallback_status += " (Using reference image with style reference)"
                            if excel_path and os.path.exists(excel_path):
                                fallback_status += " (Excel report generated)"
                            
                            gallery_results, status_text, zip_file, processed_paths, variation_nums = process_generated_images(
                                image_paths,                    # images
                                fallback_status,               # status
                                None,                          # zip_file
                                current_ref_img,               # ref_image
                                prompt_to_use,                 # prompt
                                0,                             # counter
                                []                             # variation_nums
                            )
                            
                            # Create final fallback status message
                            final_fallback_status = f"Generation complete with Imagen-4! Generated {len(result_b64_list)} image(s) with automatic background removal (ZIP creation failed)"
                            if current_ref_img and os.path.exists(current_ref_img):
                                final_fallback_status = f"Generation complete with Imagen-4! Generated {len(result_b64_list)} image(s) with automatic background removal (Using reference image with style reference, ZIP creation failed)"
                            if excel_path and os.path.exists(excel_path):
                                if current_ref_img and os.path.exists(current_ref_img):
                                    final_fallback_status = f"Generation complete with Imagen-4! Generated {len(result_b64_list)} image(s) with automatic background removal (Using reference image with style reference, Excel report generated, ZIP creation failed)"
                                else:
                                    final_fallback_status = f"Generation complete with Imagen-4! Generated {len(result_b64_list)} image(s) with automatic background removal (Excel report generated, ZIP creation failed)"
                            
                            return (
                                gallery_results,  # Images for gallery display
                                final_fallback_status,  # Status
                                zip_file,    # ZIP file
                                image_paths,  # Store original image paths for inpainting
                                None,        # No modified ZIP
                                None         # Duplicate output for modified_zip_file_output
                            )
                    else:
                        return (
                            [],  # No images
                            "❌ Imagen-4 failed to generate image. Please check your Replicate API key and try again.",  # Status
                            None,  # No ZIP file
                            [],  # No images for inpainting
                            None,  # No modified ZIP
                            None   # Duplicate output for modified_zip_file_output
                        )
                        
                except Exception as imagen4_error:
                    logger.error(f"❌ Imagen-4 Error: {str(imagen4_error)}")
                    return (
                        [],  # No images
                        f"❌ Imagen-4 Error: {str(imagen4_error)}",  # Status
                        None,  # No ZIP file
                        [],  # No images for inpainting
                        None,  # No modified ZIP
                        None   # Duplicate output for modified_zip_file_output
                    )
            
            # Generate images using the correct provider and settings
            # Fixed: Use sync version to avoid asyncio nested loop issues
            results = sync_upload_and_generate_image(
                provider=provider,
                reference_images=current_ref_img,
                card_template=None,  # DISABLED: Let dual output system handle card templates
                theme=theme,
                category=selected_category,
                subcategory=subcategory,
                # Leonardo specific parameters
                model_name=leo_model,
                width=1024,
                height=1024,
                guidance_scale=guidance_scale,  # Use guidance_scale instead of magic_strength
                generated_prompt=prompt_to_use,
                negative_prompt=neg_p,
                # Legacy parameters removed - now using multi-reference system
                preset_style=preset,
                num_images=leo_num_img,
                # Ideogram parameters
                ideogram_model=ideogram_model,
                ideogram_style=ideogram_style,
                ideogram_num_images=ideogram_num_img,
                # Common parameters
                output_format=output_f,
                # Filename convention
                filename_convention=filename_convention,
                # S3 upload settings
                upload_to_s3_bucket=safe_upload_to_s3,
                # Seed for reproducibility - use validated_seed here
                seed=validated_seed,
                # Activity and expression
                activity=act_param,
                facial_expression=exp_param,
                fur_color=fur_param,
                ethnicity=ethnicity,
                # Stop flag
                stop_flag=safe_stop_flag,
                # Google Drive upload
                upload_to_gdrive=safe_upload_to_gdrive,
                # Post-QC folder selection
                # Base64 encoding option
                encode_to_base64=safe_encode_to_base64,
                # Generation type - determine based on whether we have activity/expression modifications
                generation_type="activity" if (modification_type == "activity" or (has_activity or has_expression or has_fur_color or has_ethnicity)) else "standard",
                # Multi-reference image support for Leonardo
                reference_image_1=reference_image_1, ref_type_1=ref_type_1, ref_strength_1=ref_strength_1,
                reference_image_2=reference_image_2, ref_type_2=ref_type_2, ref_strength_2=ref_strength_2,
                reference_image_3=reference_image_3, ref_type_3=ref_type_3, ref_strength_3=ref_strength_3,
                # Ideogram style reference control
                ideogram_disable_style_reference=ideogram_disable_style_reference,
                # Ideogram rendering speed for V3 model
                ideogram_rendering_speed=ideogram_rendering_speed,
                # Modified prompt for Excel generation
                modified_prompt=modified_prompt,
                # Imagen-4 parameters
                imagen4_aspect_ratio=imagen4_aspect_ratio,
                imagen4_safety_filter=imagen4_safety_filter,
                imagen4_num_images=imagen4_num_images,
                # Counter override parameter
                counter_override=counter_override
            )
            
            # Check generation results
            if not results:
                return (
                    [],  # No images
                    "Error: Failed to generate images. Check logs for details.",  # Status
                    None,  # No ZIP file
                    None,  # No modified images
                    None,  # No modified ZIP
                    None   # Duplicate output for modified_zip_file_output
                )
            
            # Unpack results
            if len(results) == 3:
                # Handle case where upload_and_generate_image returns 3 values
                images, status, download_url = results
                variation_numbers = []  # Default empty list for variation numbers
            else:
                # Handle case where upload_and_generate_image returns 4+ values
                images, status, download_url = results[:3]
                variation_numbers = results[3] if len(results) > 3 else []
            
            # Check if we have images
            if not images or len(images) == 0:
                return (
                    [],  # No images
                    f"Error: {status}",  # Status with error details
                    None,  # No ZIP file
                    [],   # No images for inpainting
                    None,  # No modified ZIP
                    None   # Duplicate output for modified_zip_file_output
                )
            
            # Update the counter for batch display
            counter_text = f"Generated {len(images)} image(s)"
            
            # Create display images with metadata (original images first)
            display_images, ref_image_path = create_display_images_with_metadata(
                images, 
                [current_ref_img] if current_ref_img else [], 
                variation_numbers,
                reference_filename=ref_filename
            )
            
            # Only show original images during generation (no dual output processing)
            # Dual output processing should only happen in "Remove Background & Apply to Card Template" function
            
            # Create display images with only original images
            display_images = []
            
            # Add original images to gallery
            if images:
                # Handle both file paths and base64 strings
                processed_images = []
                for i, img in enumerate(images):
                    if isinstance(img, str) and not img.startswith('data:') and os.path.exists(img):
                        # It's a file path
                        processed_images.append((img, f"{os.path.basename(img)} (Original)"))
                    else:
                        # It's a base64 string or image object
                        processed_images.append((img, f"Generated Image {i+1} (Original)"))
                display_images.extend(processed_images)
                logger.info(f"Added {len(images)} original images to gallery")

            # Create ZIP file with only original images
            zip_path = None
            if images:
                zip_path = create_zip_file(images, theme, selected_category)

            # Update status message to reflect only original images
            status_message = f"Generated {len(images)} original images."
            if zip_path:
                status_message += " ZIP file created."

            # Return results to the UI (only original images)
            return (
                display_images, 
                status_message,
                zip_path, # Original images ZIP
                images, # Pass original generated images to inpainting state
                None, # No nobg zip during generation
                None, # No card zip during generation
            )
            
            # Store the original image paths for inpainting (not display objects)
            generated_images_for_inpainting = images if images else []
            
            # No dual output processing during image generation
            # This should only happen in "Remove Background & Apply to Card Template" function
            
            # Simple status message for original images only
            enhanced_status = f"Generated {len(images)} images with {provider}. {status}"
            
            # Return only original images (no dual output processing)
            return (
                display_images,  # Only original images
                enhanced_status,  # Simple status
                zip_path,    # Original images ZIP
                generated_images_for_inpainting,  # Store original image paths for inpainting
                None,   # No white background images ZIP during generation
                None    # No card template images ZIP during generation
            )
        
        except Exception as e:
            logger.error(f"Error in generate_wrapper: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return (
                [],  # No images
                f"Error during generation: {str(e)}",  # Status with error details
                None,  # No ZIP file
                [],  # Return empty list instead of None for generated_images_state
                None,  # No modified ZIP
                None   # Duplicate output for modified_zip_file_output
            )
    
    except Exception as outer_e:
        logger.error(f"Critical error in generate_wrapper outer block: {str(outer_e)}")
        import traceback
        logger.error(traceback.format_exc())
        return (
            [],  # No images
            f"Critical error: {str(outer_e)}",  # Status with error details
            None,  # No ZIP file
            [],  # Return empty list instead of None for generated_images_state
            None,  # No modified ZIP
            None   # Duplicate output for modified_zip_file_output
        )
    finally:
        # Clean up any temporary directories created
        for temp_dir in temp_dirs_to_cleanup:
            try:
                import shutil
                if os.path.exists(temp_dir):
                    shutil.rmtree(temp_dir)
                    logger.info(f"Cleaned up temporary directory: {temp_dir}")
            except Exception as cleanup_error:
                logger.warning(f"Failed to clean up temporary directory {temp_dir}: {str(cleanup_error)}")

def remove_background_with_white_background(image_path):
    """
    Remove background from image and replace with white background (not transparent).
    Ensures output has exact same dimensions as input.
    """
    try:
        from rembg import remove
        from PIL import Image
        
        logger.info(f"Processing white background for: {image_path}")
        
        # Load the input image
        with open(image_path, 'rb') as input_file:
            input_data = input_file.read()
        
        # Remove background using rembg
        output_data = remove(input_data)
        
        # Convert to PIL Image
        img_no_bg = Image.open(io.BytesIO(output_data)).convert('RGBA')
        
        # Create white background with same dimensions
        white_bg = Image.new('RGB', img_no_bg.size, (255, 255, 255))
        
        # Paste the image with removed background onto white background
        white_bg.paste(img_no_bg, mask=img_no_bg.split()[-1])
        
        logger.info(f"✅ White background created successfully for {image_path}")
        return white_bg
        
    except Exception as e:
        logger.error(f"❌ Error creating white background for {image_path}: {str(e)}")
        # Fallback: return original image if background removal fails
        try:
            return Image.open(image_path).convert('RGB')
        except:
            return None

def apply_image_to_card_template(image, card_template_path, output_size=(512, 512)):
    """
    Apply a processed image onto a card template.
    """
    try:
        logger.info(f"Applying image to card template: {card_template_path}")
        
        # Load card template
        card_template = Image.open(card_template_path).convert('RGBA')
        
        # Resize card template to output size
        card_template = card_template.resize(output_size, Image.LANCZOS)
        
        # Convert image to RGBA if needed
        if image.mode != 'RGBA':
            image = image.convert('RGBA')
        
        # Calculate sizing - make image fit within 70% of card template
        max_size = (int(output_size[0] * 0.7), int(output_size[1] * 0.7))
        image_resized = image.copy()
        image_resized.thumbnail(max_size, Image.LANCZOS)
        
        # Calculate position to center the image on the template
        x_offset = (output_size[0] - image_resized.size[0]) // 2
        y_offset = (output_size[1] - image_resized.size[1]) // 2
        
        # Create result image
        result = Image.new('RGBA', output_size, (255, 255, 255, 0))
        result.paste(card_template, (0, 0))
        result.paste(image_resized, (x_offset, y_offset), image_resized)
        
        # Convert to RGB for final output
        final_result = Image.new('RGB', output_size, (255, 255, 255))
        final_result.paste(result, mask=result.split()[-1] if result.mode == 'RGBA' else None)
        
        logger.info(f"✅ Card template application successful")
        return final_result
        
    except Exception as e:
        logger.error(f"❌ Error applying image to card template: {str(e)}")
        return None

def process_dual_outputs(image_paths, card_template_path, theme, category):
    """
    Process images to create both white background and card template outputs.
    """
    nobg_images = []
    card_images = []
    
    try:
        logger.info(f"🔄 Processing dual outputs for {len(image_paths)} images")
        
        for i, image_path in enumerate(image_paths):
            try:
                # Generate white background version (ALWAYS)
                white_bg_image = remove_background_with_white_background(image_path)
                if white_bg_image:
                    # Save white background image
                    nobg_filename = f"{theme}{category}{str(i+1).zfill(3)}_nobg.png"
                    nobg_path = os.path.join(tempfile.gettempdir(), nobg_filename)
                    white_bg_image.save(nobg_path, 'PNG')
                    nobg_images.append(nobg_path)
                    logger.info(f"✅ White background saved: {nobg_path}")
                else:
                    logger.error(f"❌ Failed to create white background for {image_path}")
                
                # Generate card template version (ONLY if template provided)
                if card_template_path and os.path.exists(card_template_path):
                    if white_bg_image:  # Use the white background image for card template
                        card_image = apply_image_to_card_template(white_bg_image, card_template_path)
                        if card_image:
                            # Save card template image
                            card_filename = f"{theme}{category}{str(i+1).zfill(3)}_card.png"
                            card_path = os.path.join(tempfile.gettempdir(), card_filename)
                            card_image.save(card_path, 'PNG')
                            card_images.append(card_path)
                            logger.info(f"✅ Card template saved: {card_path}")
                        else:
                            logger.error(f"❌ Failed to apply card template for {image_path}")
                else:
                    logger.info(f"No card template provided, skipping card template processing")
                    
            except Exception as img_error:
                logger.error(f"❌ Error processing image {image_path}: {str(img_error)}")
                continue
        
        logger.info(f"✅ Dual output processing complete: {len(nobg_images)} white background, {len(card_images)} card template images")
        return nobg_images, card_images
        
    except Exception as e:
        logger.error(f"❌ Error in process_dual_outputs: {str(e)}")
        return [], []

def create_individual_download_files(nobg_images, card_images, theme, category):
    """
    Create individual ZIP files for white background and card template downloads.
    """
    nobg_zip_path = None
    card_zip_path = None
    
    try:
        # Create white background ZIP
        if nobg_images:
            nobg_zip_filename = f"{theme}{category}_white_backgrounds.zip"
            nobg_zip_path = os.path.join(tempfile.gettempdir(), nobg_zip_filename)
            
            with zipfile.ZipFile(nobg_zip_path, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for img_path in nobg_images:
                    if os.path.exists(img_path):
                        zip_file.write(img_path, os.path.basename(img_path))
            
            logger.info(f"✅ White background ZIP created: {nobg_zip_path}")
        
        # Create card template ZIP
        if card_images:
            card_zip_filename = f"{theme}{category}_card_templates.zip"
            card_zip_path = os.path.join(tempfile.gettempdir(), card_zip_filename)
            
            with zipfile.ZipFile(card_zip_path, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for img_path in card_images:
                    if os.path.exists(img_path):
                        zip_file.write(img_path, os.path.basename(img_path))
            
            logger.info(f"✅ Card template ZIP created: {card_zip_path}")
        
        return nobg_zip_path, card_zip_path
        
    except Exception as e:
        logger.error(f"❌ Error creating individual download files: {str(e)}")
        return None, None

def remove_watermark(image, is_photoroom=False):
    """
    Remove watermarks from the bottom right corner of the image without affecting the card content.
    
    Args:
        image (PIL.Image): The image to process
        is_photoroom (bool): Flag to indicate if this is from PhotoRoom API, for more targeted removal
        
    Returns:
        PIL.Image: Image with watermark removed
    """
    try:
        logger.info("Checking for watermarks to remove")
        
        # Create a copy of the image to work with
        result = image.copy()
        width, height = result.size
        
        # Define the watermark region (specifically narrow band at bottom right for watermarks)
        # PhotoRoom watermarks are typically in a very specific location in the bottom right
        watermark_region_width = int(width * 0.15)  # Adjust based on known watermark size
        watermark_region_height = int(height * 0.05)  # Make this smaller to avoid affecting card content
            
        watermark_x = width - watermark_region_width
        watermark_y = height - watermark_region_height
        
        # Check if there's a likely watermark by looking for text-like patterns
        # or semi-transparent overlays in the region
        watermark_found = False
        region = result.crop((watermark_x, watermark_y, width, height))
        
        # Convert to RGBA if not already
        if region.mode != 'RGBA':
            region = region.convert('RGBA')
        
        # For PhotoRoom, we know the watermark is likely present
        if is_photoroom:
            # PhotoRoom adds "remove.bg" watermark
            watermark_found = True
            logger.info("PhotoRoom output - looking for watermark")
            
            # Scan the region for text-like patterns with specific color profile common in PhotoRoom watermarks
            pixels = region.load()
            watermark_pixels = []
            
            # PhotoRoom text watermarks often have specific RGB values and alpha patterns
            for y in range(region.height):
                for x in range(region.width):
                    r, g, b, a = pixels[x, y]
                    # Look for watermark text pixels (often white or black text)
                    if ((r > 200 and g > 200 and b > 200) or 
                                         (r < 50 and g < 50 and b < 50)) and a > 180:
                        watermark_pixels.append((x, y))
            
            if len(watermark_pixels) > 20:  # Threshold for confirming watermark presence
                watermark_found = True
                logger.info(f"Found {len(watermark_pixels)} potential watermark pixels")
            else:
                # Alternative detection: check for "remove" text pattern
                region_data = np.array(region)
                # Simple edge detection to find text-like features
                if np.std(region_data[:, :, :3]) > 20:  # High variance indicates potential text
                    watermark_found = True
                    logger.info("Detected potential watermark text pattern")
        else:
            # Generic watermark detection
            # Look for text-like patterns or semi-transparent overlays in the bottom right
            region_data = list(region.getdata())
            semi_transparent_pixels = sum(1 for r, g, b, a in region_data if 0 < a < 255 and a > 100)
            high_contrast_pixels = sum(1 for r, g, b, a in region_data if 
                                      (max(r, g, b) - min(r, g, b) > 100) and a > 200)
            
            # If there are enough semi-transparent or high contrast pixels, it's likely a watermark
            if semi_transparent_pixels > 20 or high_contrast_pixels > 40:
                watermark_found = True
                logger.info(f"Detected possible watermark with {semi_transparent_pixels} semi-transparent pixels and {high_contrast_pixels} high-contrast pixels")
        
        if watermark_found:
            logger.info("Removing watermark from image")
            
            # Create a mask that only affects the watermark area
            mask = Image.new('RGBA', result.size, (0, 0, 0, 0))  # Fully transparent mask
            draw = ImageDraw.Draw(mask)
            
            # Precisely locate "remove" text in the bottom right if possible
            # This is a more targeted approach than blanking out the entire region
            
            # Option 1: Just make the watermark area transparent (no filling)
            draw.rectangle((watermark_x, watermark_y, width, height), fill=(0, 0, 0, 0))
            
            # Create a version of the image without the watermark
            result_without_watermark = Image.new('RGBA', result.size, (0, 0, 0, 0))
            result_without_watermark.paste(result, (0, 0))
            
            # Apply the mask to keep everything except the watermark region
            for x in range(watermark_x, width):
                for y in range(watermark_y, height):
                    # Get current pixel
                    r, g, b, a = result.getpixel((x, y))
                    
                    # Check if this pixel looks like part of a watermark
                    # (high brightness or darkness with moderate to high alpha)
                    is_watermark_pixel = ((r > 200 and g > 200 and b > 200) or 
                                         (r < 50 and g < 50 and b < 50)) and a > 150
                    
                    if is_watermark_pixel:
                        # Make this pixel transparent
                        result_without_watermark.putpixel((x, y), (0, 0, 0, 0))
            
            logger.info("Watermark removal completed")
            print("[Success] Watermark detected and removed from image")
            return result_without_watermark
        else:
            logger.info("No obvious watermark detected")
            return result
    except Exception as e:
        logger.error(f"Error removing watermark: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        # Return the original image if watermark removal fails
        return image

# Constants and configuration
logger = logging.getLogger(__name__)

# === DUAL OUTPUT PROCESSING FUNCTIONS ===

def remove_background_with_white_background(image_path):
    """
    Remove background from image and replace with white background (not transparent)
    
    Args:
        image_path (str): Path to the input image
        
    Returns:
        PIL.Image: Image with white background or None if failed
    """
    try:
        logger.info(f"Processing background removal for: {image_path}")
        
        # Open the input image
        with open(image_path, 'rb') as f:
            input_data = f.read()
        
        # Remove background using rembg
        output_data = remove(input_data)
        
        # Convert to PIL Image
        img_no_bg = Image.open(io.BytesIO(output_data)).convert('RGBA')
        
        # Create white background with same dimensions
        white_bg = Image.new('RGB', img_no_bg.size, (255, 255, 255))
        
        # Paste the no-bg image onto white background
        white_bg.paste(img_no_bg, mask=img_no_bg.split()[-1])  # Use alpha channel as mask
        
        logger.info(f"Successfully removed background and applied white background for: {image_path}")
        return white_bg
        
    except Exception as e:
        logger.error(f"Error removing background from {image_path}: {str(e)}")
        return None

def apply_image_to_card_template(image, card_template_path, output_size=(512, 512)):
    """
    Apply processed image to card template
    
    Args:
        image (PIL.Image): The processed image to apply
        card_template_path (str): Path to the card template
        output_size (tuple): Output size for the final card
        
    Returns:
        PIL.Image: Card with applied image or None if failed
    """
    try:
        if not card_template_path or not os.path.exists(card_template_path):
            logger.warning("No card template provided or template not found")
            return None
            
        logger.info(f"Applying image to card template: {card_template_path}")
        
        # Open card template
        card_template = Image.open(card_template_path).convert('RGBA')
        
        # Resize image to fit template (you may want to adjust this logic)
        # For now, we'll resize the image to 80% of template size and center it
        template_width, template_height = card_template.size
        
        # Calculate target size for the image (80% of template)
        target_width = int(template_width * 0.8)
        target_height = int(template_height * 0.8)
        
        # Resize image while maintaining aspect ratio
        image_resized = image.copy()
        image_resized.thumbnail((target_width, target_height), Image.Resampling.LANCZOS)
        
        # Convert image to RGBA for proper compositing
        if image_resized.mode != 'RGBA':
            image_resized = image_resized.convert('RGBA')
        
        # Create a new image with template as base
        result = card_template.copy()
        
        # Calculate position to center the image on template
        x_offset = (template_width - image_resized.width) // 2
        y_offset = (template_height - image_resized.height) // 2
        
        # Paste the image onto the template
        result.paste(image_resized, (x_offset, y_offset), image_resized)
        
        # Convert to RGB and resize to output size
        result = result.convert('RGB')
        result = result.resize(output_size, Image.Resampling.LANCZOS)
        
        logger.info("Successfully applied image to card template")
        return result
        
    except Exception as e:
        logger.error(f"Error applying image to card template: {str(e)}")
        return None

def process_dual_outputs(image_paths, card_template_path=None, theme=None, category=None):
    """
    Process images to create dual outputs: white background and card template applied
    
    Args:
        image_paths (list): List of image file paths to process
        card_template_path (str): Path to card template (optional)
        theme (str): Theme for naming convention
        category (str): Category for naming convention
        
    Returns:
        tuple: (nobg_image_paths, card_image_paths) - paths to processed images
    """
    import tempfile
    
    nobg_image_paths = []
    card_image_paths = []
    
    try:
        logger.info(f"Starting dual output processing for {len(image_paths)} images")
        
        # Additional safety check: Convert any PIL Image objects to file paths
        safe_image_paths = []
        for img in image_paths:
            if isinstance(img, str):
                safe_image_paths.append(img)
            elif hasattr(img, 'save'):  # PIL Image object
                logger.info(f"Converting PIL Image to temporary file: {type(img)}")
                temp_dir = tempfile.mkdtemp()
                temp_path = os.path.join(temp_dir, f"temp_image_{len(safe_image_paths)}.png")
                img.save(temp_path, 'PNG')
                safe_image_paths.append(temp_path)
                logger.info(f"Saved PIL Image to: {temp_path}")
            else:
                logger.warning(f"Unsupported image type in process_dual_outputs: {type(img)}")
                continue
        
        logger.info(f"Converted {len(image_paths)} images to {len(safe_image_paths)} safe paths")
        
        for i, image_path in enumerate(safe_image_paths):
            if not image_path or not os.path.exists(image_path):
                logger.warning(f"Image path not found: {image_path}")
                continue
                
            logger.info(f"Processing image {i+1}/{len(safe_image_paths)}: {image_path}")
            
            # Create temporary directory for outputs
            temp_dir = tempfile.mkdtemp()
            
            # Generate filenames with proper naming convention
            base_filename = os.path.splitext(os.path.basename(image_path))[0]
            
            if theme and category:
                theme_code = THEME_MAPPING.get(theme, "00")
                category_code = CATEGORY_MAPPING.get(category, "000")
                nobg_filename = f"{theme_code}{category_code}{i+1:05d}_nobg.png"
                card_filename = f"{theme_code}{category_code}{i+1:05d}_card.png"
            else:
                nobg_filename = f"{base_filename}_nobg.png"
                card_filename = f"{base_filename}_card.png"
                
            # 1. Create white background image
            white_bg_image = remove_background_with_white_background(image_path)
            if white_bg_image:
                nobg_path = os.path.join(temp_dir, nobg_filename)
                white_bg_image.save(nobg_path, 'PNG')
                nobg_image_paths.append(nobg_path)
                logger.info(f"Saved white background image: {nobg_path}")
            else:
                logger.error(f"Failed to create white background image for: {image_path}")
                
            # 2. Create card template applied image (if template provided)
            if card_template_path and white_bg_image:
                card_image = apply_image_to_card_template(white_bg_image, card_template_path)
                if card_image:
                    card_path = os.path.join(temp_dir, card_filename)
                    card_image.save(card_path, 'PNG')
                    card_image_paths.append(card_path)
                    logger.info(f"Saved card template image: {card_path}")
                else:
                    logger.error(f"Failed to create card template image for: {image_path}")
            elif card_template_path:
                logger.warning(f"Cannot create card template image - white background processing failed for: {image_path}")
                
        logger.info(f"Dual output processing complete. Created {len(nobg_image_paths)} white background images and {len(card_image_paths)} card images")
        return nobg_image_paths, card_image_paths
        
    except Exception as e:
        logger.error(f"Error in dual output processing: {str(e)}")
        return [], []

def create_individual_download_files(nobg_images, card_images, theme=None, category=None):
    """
    Create individual downloadable files for both output types
    
    Args:
        nobg_images (list): List of white background image paths
        card_images (list): List of card template image paths
        theme (str): Theme for naming
        category (str): Category for naming
        
    Returns:
        tuple: (nobg_zip_path, card_zip_path) - paths to individual ZIP files
    """
    import tempfile
    import zipfile
    
    timestamp = get_gmt7_filename_timestamp()
    
    # Create individual ZIP files
    nobg_zip_path = None
    card_zip_path = None
    
    try:
        # Create ZIP for white background images
        if nobg_images:
            temp_dir = tempfile.mkdtemp()
            
            if theme and category:
                theme_code = THEME_MAPPING.get(theme, "00")
                category_code = CATEGORY_MAPPING.get(category, "000")
                nobg_zip_name = f"{theme_code}{category_code}_white_background_{timestamp}.zip"
            else:
                nobg_zip_name = f"white_background_images_{timestamp}.zip"
                
            nobg_zip_path = os.path.join(temp_dir, nobg_zip_name)
            
            with zipfile.ZipFile(nobg_zip_path, 'w', compression=zipfile.ZIP_DEFLATED) as zipf:
                for i, img_path in enumerate(nobg_images):
                    if os.path.exists(img_path):
                        zipf.write(img_path, os.path.basename(img_path))
                        
            logger.info(f"Created white background ZIP: {nobg_zip_path}")
            
        # Create ZIP for card template images
        if card_images:
            temp_dir = tempfile.mkdtemp()
            
            if theme and category:
                theme_code = THEME_MAPPING.get(theme, "00")
                category_code = CATEGORY_MAPPING.get(category, "000")
                card_zip_name = f"{theme_code}{category_code}_card_template_{timestamp}.zip"
            else:
                card_zip_name = f"card_template_images_{timestamp}.zip"
                
            card_zip_path = os.path.join(temp_dir, card_zip_name)
            
            with zipfile.ZipFile(card_zip_path, 'w', compression=zipfile.ZIP_DEFLATED) as zipf:
                for i, img_path in enumerate(card_images):
                    if os.path.exists(img_path):
                        zipf.write(img_path, os.path.basename(img_path))
                        
            logger.info(f"Created card template ZIP: {card_zip_path}")
            
        return nobg_zip_path, card_zip_path
        
    except Exception as e:
        logger.error(f"Error creating individual download files: {str(e)}")
        return None, None

# Add constant for animal activities
ANIMAL_ACTIVITIES = [
    # Pose varieties
    "standing with feet firmly planted",
    "striking an elegant pose",
    "in a graceful ballet stance",
    "sitting with perfect posture",
    "in a playful yoga pose",
    "stretching in a warrior pose",
    "with paws tucked underneath",
    "in a meditative sitting pose",
    "with head tilted inquisitively",
    "in a candid mid-movement pose",
    "leaping mid air",
    
    # Simple positions
    "sitting on a colorful cushion",
    "perched on a wooden stool",
    "resting on a small pedestal",
    "lounging on a velvet pillow",
    "sitting proudly on a rock",
    "positioned on a decorative tile",
    "resting on a smooth surface",
    "seated on a small ottoman",
    "perched on a tiny footstool",
    "sitting on a plush carpet",
    
    # Indoor scenes without problematic objects
    "exploring a cozy living room",
    "investigating a small wooden box",
    "balancing on a stack of colorful blocks",
    "stretching to reach upward",
    "nestled in a knitted basket",
    "sliding across a polished floor",
    "hiding under a blanket",
    "curled up in a small basket",
    "sitting in an empty flowerpot",
    "exploring a fabric tunnel",
    
    # Actions with appropriate objects
    "playing with a small rubber ball",
    "investigating a wooden puzzle",
    "batting at a soft fabric toy",
    "pouncing on a small plush cube",
    "pushing a small wooden block",
    "climbing a short fabric ramp",
    "examining a colorful toy ring",
    "balancing on a small cushion",
    "with a tiny basket of berries",
    "touching a smooth stone",
    
    # Emotive poses
    "looking surprised with wide eyes",
    "with an inquisitive expression",
    "showing a playful demeanor",
    "with a majestic expression",
    "looking alert and attentive",
    "with a relaxed, content posture",
    "showing a curious expression",
    "with an adorable sleepy look",
    "with whiskers forward in interest",
    "with tall, alert ears",
    
    # Simple activities
    "mid-pounce position",
    "in a stretching position",
    "performing an acrobatic pose",
    "crouched ready to leap",
    "in a playful pouncing stance",
    "showcasing perfect balance",
    "in an elegant sitting pose",
    "mid-jump in perfect form",
    "in a graceful landing pose",
    "demonstrating perfect stillness",
    
    # More pose varieties
    "in a proud standing position",
    "showing off perfect whiskers",
    "with tail curved gracefully",
    "with paws neatly positioned",
    "in a symmetrical seated pose",
    "showing confident posture",
    "in a professional portrait pose",
    "with an aristocratic bearing",
    "in a famous sculpture pose",
    "showing theatrical expression",
    
    # With small decorative elements
    "next to a tiny potted succulent",
    "beside a small decorative vase",
    "near a miniature sculpture",
    "with a tiny decorative lantern",
    "next to small ceramic figurines",
    "with a decorative fabric swatch",
    "beside a small wooden artifact",
    "near a small decorative clock",
    "with a tiny treasure chest",
    "next to a decorative jewelry box",
    
    # Active but contained poses
    "mid-spin in perfect form",
    "in a graceful twirling pose",
    "balancing on one paw",
    "in a perfect hunting crouch",
    "showing off climbing ability",
    "demonstrating agility",
    "in a perfect jumping form",
    "displaying athletic prowess",
    "in a dynamic action pose",
    "demonstrating perfect coordination"
]

# Function to extract the subject from a prompt
def extract_subject_from_prompt(prompt):
    """Extract the main subject from a prompt"""
    # Try various patterns to find the subject
    patterns = [
        r'a\s+([\w\s-]+?)(?:\s+with|\s+in|\s+on|\s+,|\s+is|\s+that|\s+\.|$)',  # 'a cat with...'
        r'the\s+([\w\s-]+?)(?:\s+with|\s+in|\s+on|\s+,|\s+is|\s+that|\s+\.|$)',  # 'the rabbit is...'
        r'(?:of|showing)\s+(?:a|an)\s+([\w\s-]+?)(?:\s+with|\s+in|\s+on|\s+,|\s+is|\s+that|\s+\.|$)',  # 'showing a fox with...'
    ]
    
    for pattern in patterns:
        match = re.search(pattern, prompt, re.IGNORECASE)
        if match:
            subject = match.group(1).strip()
            # If subject is too long, take just the last word or two
            if len(subject.split()) > 2:
                subject_parts = subject.split()
                subject = ' '.join(subject_parts[-2:])
            return subject
    
    # Fallback if no pattern matches
    return "animal"

from PIL import Image, ImageFilter
import logging

logger = logging.getLogger(__name__)

# --- New Google Drive Utility Functions (provided by user, with corrections) ---
# --- Google Drive Utility Functions (Original Placeholders) ---
# It is recommended to move these to a separate utility file if they grow.

def check_google_drive_dependencies():
    """Placeholder for checking Google Drive dependencies."""
    # For now, assume dependencies are met or handled by direct imports.
    logger.warning("check_google_drive_dependencies is a placeholder. Implement actual dependency check.")
    print("check_google_drive_dependencies is a placeholder. Implement actual dependency check.")
    return True

# --- New Google Drive Utility Functions (provided by user, with corrections) ---
# The create_google_drive_service function follows, which you have already modified.
def create_google_drive_service():
    """Create and return a Google Drive service object."""
    # Check if Google Drive dependencies are installed
    if not check_google_drive_dependencies():
        logger.error("Google Drive dependencies not installed")
        print("Google Drive functionality is disabled. To enable, install required packages:")
        print("pip install google-api-python-client google-auth-httplib2 google-auth-oauthlib")
        return None

    # Import dependencies
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        from google.auth.transport.requests import Request # ADDED for credential refresh
        # MediaFileUpload is imported lower down, closer to its use if needed.
    except ImportError as e:
        logger.error(f"Failed to import Google Drive dependencies: {str(e)}")
        print(f"Google Drive error: {str(e)}")
        return None

    SERVICE_ACCOUNT_FILE = ''
    API_NAME = 'drive'
    API_VERSION = 'v3'
    SCOPES = ['https://www.googleapis.com/auth/drive']
    SHARED_DRIVE_ID =''
    SUBJECT = ''
    
    # Check if service account file exists
    if not os.path.exists(SERVICE_ACCOUNT_FILE):
        logger.error(f"Service account key file not found: {SERVICE_ACCOUNT_FILE}")
        print(f"Error: Google Drive service account key file not found: {SERVICE_ACCOUNT_FILE}")
        print("Please ensure the service account key file is in the correct location.")
        return None

    cred = None
    try:
        cred = service_account.Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SCOPES, subject=SUBJECT)
        logger.info("Successfully loaded credentials from service account file.")
        print("[Google Drive Debug] Loaded credentials from service account file.")

        if cred and not cred.valid: # If loaded but not initially valid
            logger.warning("[DEBUG] Service account credentials loaded but are not immediately valid. Attempting refresh.")
            print("[Google Drive Debug] Service account credentials loaded but are not immediately valid. Attempting refresh.")
            try:
                cred.refresh(Request())
                logger.info("[DEBUG] Service account credentials refresh attempt complete.")
                print("[Google Drive Debug] Service account credentials refresh attempt complete.")
                if not cred.valid:
                    logger.warning("[DEBUG] Credentials still not valid after refresh attempt.")
                    print("[Google Drive Debug] Credentials still not valid after refresh attempt.")
            except Exception as refresh_err:
                logger.error(f"[DEBUG] Failed to refresh service account credentials: {str(refresh_err)}")
                print(f"[Google Drive Debug] Failed to refresh service account credentials: {str(refresh_err)}")
                # Optionally, you might want to return None here if refresh failure is critical
    
    except Exception as e:
        logger.error(f"Failed to load service account credentials: {str(e)}")
        print(f"[Google Drive Error] Failed to load service account credentials: {str(e)}")
        return None
    
    # Final check on credentials before building service
    if not cred or not cred.valid: 
        logger.error("[DEBUG] Credentials are not valid after attempting to load (and potentially refresh) from service account file.")
        print("[Google Drive Debug] Failed to obtain valid credentials from service account file (even after refresh attempt). Service creation aborted.")
        return None
        
    logger.info("[DEBUG] Proceeding to build Google Drive service with obtained credentials.")
    print("[Google Drive Debug] Credentials appear valid. Building service.")
    
    try:
        # Build and return the Drive service
        service = build(API_NAME, API_VERSION, credentials=cred)
        logger.info("Google Drive service created successfully")
        print("Google Drive service initialized successfully")
        
        try:
            file_metadata = service.files().get(
                fileId=SHARED_DRIVE_ID,
                supportsAllDrives=True,
                fields='id, name, mimeType, driveId'
            ).execute()
            print(f"[Google Drive Debug] SHARED_DRIVE_ID metadata: {file_metadata}")
            if 'driveId' in file_metadata:
                print(f"[Google Drive Debug] This is in a shared drive with driveId: {file_metadata['driveId']}")
            else:
                print("[Google Drive Debug] This is not in a shared drive.")
        except Exception as e:
            print(f"[Google Drive Debug] Error checking SHARED_DRIVE_ID: {str(e)}")

        return service
    except Exception as e:
        logger.error(f"Failed to create Google Drive service: {str(e)}")
        print(f"Error connecting to Google Drive: {str(e)}")
        return None

def get_or_create_folder(service, parent_folder_id, folder_name):
    """Get the ID of a folder by name, create it if it doesn't exist."""
    # Escape single quotes in folder_name for the query
    safe_folder_name = folder_name.replace("'", "\\'")
    query = f"mimeType='application/vnd.google-apps.folder' and name='{safe_folder_name}' and '{parent_folder_id}' in parents and trashed=false"
    
    try:
        response = service.files().list(q=query, spaces='drive', fields='files(id, name)', supportsAllDrives=True, includeItemsFromAllDrives=True).execute()
        if response.get('files'):
            # Folder exists
            return response['files'][0]['id']
        else:
            # Folder does not exist, create it
            file_metadata = {
                'name': folder_name,
                'mimeType': 'application/vnd.google-apps.folder',
                'parents': [parent_folder_id]
            }
            folder = service.files().create(body=file_metadata, fields='id', supportsAllDrives=True).execute()
            logger.info(f"Created folder '{folder_name}' with ID: {folder.get('id')}")
            return folder.get('id')
    except Exception as e:
        logger.error(f"Error finding or creating folder '{folder_name}': {str(e)}")
        print(f"Google Drive error: Could not find/create folder '{folder_name}': {str(e)}")
        return None

def get_or_create_preqc_folder(service, root_folder_id='1rGP_5RIbYkqrJ1KEgXH8SHN8bFaBtkpm'):
    """Get or create the 'Pre-QC' parent folder for all uploads."""
    return get_or_create_folder(service, root_folder_id, 'Pre-QC')

def get_or_create_postqc_folder(service, root_folder_id='1rGP_5RIbYkqrJ1KEgXH8SHN8bFaBtkpm'):
    """Get or create the 'Post-QC' parent folder for all uploads."""
    return get_or_create_folder(service, root_folder_id, 'Post-QC')

def upload_multiple_files_to_google_drive_hierarchical(file_paths, theme, category, subcategory, root_folder_id='1rGP_5RIbYkqrJ1KEgXH8SHN8bFaBtkpm'):
    """
    Uploads multiple files to Google Drive using the specified hierarchical folder structure.
    This function contains the full logic and does not depend on the old upload function.
    """
    logger.info(f"🔍 UPLOAD DEBUG: upload_multiple_files_to_google_drive_hierarchical called with theme='{theme}', category='{category}', subcategory='{subcategory}'")
    service = create_google_drive_service()
    if not service:
        logger.error("Failed to create Google Drive service for multiple file upload.")
        return []
    
    # Use the provided root folder ID directly
    if not root_folder_id:
        logger.error("No root folder ID provided.")
        return []

    # 1. Resolve Theme Folder
    theme_folder_name = THEME_FOLDER_MAPPING.get(theme)
    if not theme_folder_name:
        logger.error(f"Invalid theme '{theme}'. No mapping found.")
        return []
    theme_folder_id = get_or_create_folder(service, root_folder_id, theme_folder_name)
    if not theme_folder_id:
        logger.error(f"Failed to get or create theme folder: {theme_folder_name}")
        return []

    # 2. Resolve Category Folder
    category_folder_name = CATEGORY_FOLDER_MAPPING.get(theme, {}).get(category)
    if not category_folder_name:
        logger.error(f"Invalid category '{category}' for theme '{theme}'.")
        return []
    category_folder_id = get_or_create_folder(service, theme_folder_id, category_folder_name)
    if not category_folder_id:
        logger.error(f"Failed to get or create category folder: {category_folder_name}")
        return []

    # 3. Resolve Subcategory Folder
    subcategory_folder_id = category_folder_id
    logger.info(f"🔍 SUBCATEGORY RESOLUTION: subcategory='{subcategory}', category='{category}'")
    if subcategory:
        logger.info(f"🔍 SUBCATEGORY PROVIDED: Looking up folder mapping for category='{category}', subcategory='{subcategory}'")
        subcategory_folder_name = SUBCATEGORY_FOLDER_MAPPING.get(category, {}).get(subcategory)
        logger.info(f"🔍 SUBCATEGORY MAPPING RESULT: subcategory_folder_name='{subcategory_folder_name}'")
        if subcategory_folder_name:
            logger.info(f"🔍 CREATING SUBCATEGORY FOLDER: '{subcategory_folder_name}' under category folder ID: {category_folder_id}")
            resolved_id = get_or_create_folder(service, category_folder_id, subcategory_folder_name)
            if resolved_id:
                subcategory_folder_id = resolved_id
                logger.info(f"✅ SUBCATEGORY FOLDER CREATED/FOUND: ID = {subcategory_folder_id}")
            else:
                logger.error(f"Failed to get or create subcategory folder: {subcategory_folder_name}")
                return []
        else:
            logger.warning(f"⚠️ SUBCATEGORY MAPPING NOT FOUND: No mapping found for category='{category}', subcategory='{subcategory}'. Using raw subcategory name.")
            # Fallback: use the raw subcategory name if no mapping found
            resolved_id = get_or_create_folder(service, category_folder_id, subcategory)
            if resolved_id:
                subcategory_folder_id = resolved_id
                logger.info(f"📁 SUBCATEGORY FOLDER (FALLBACK): Successfully created/retrieved folder '{subcategory}' with ID: {resolved_id}")
            else:
                logger.error(f"❌ SUBCATEGORY FOLDER FALLBACK FAILED: Failed to get or create subcategory folder: {subcategory}")
                return []
    else:
        logger.info(f"🔍 NO SUBCATEGORY PROVIDED: Using category folder ID: {category_folder_id}")

    uploaded_file_urls = []
    for file_path in file_paths:
        try:
            if not os.path.exists(file_path):
                logger.warning(f"File not found, skipping upload: {file_path}")
                continue

            file_name = os.path.basename(file_path)
            parent_folder_id = subcategory_folder_id

            # 4. Handle 'Card' Folder for specific files
            if "_card" in file_name or file_name.endswith('.zip'):
                card_folder_id = get_or_create_folder(service, subcategory_folder_id, "Card")
                if card_folder_id:
                    parent_folder_id = card_folder_id
                else:
                    logger.error("Failed to get or create 'Card' subfolder. Uploading to subcategory folder instead.")

            # 5. Upload the file
            mime_type, _ = mimetypes.guess_type(file_path)
            if not mime_type:
                mime_type = 'application/octet-stream'
            
            file_metadata = {'name': file_name, 'parents': [parent_folder_id]}
            media = MediaFileUpload(file_path, mimetype=mime_type)
            
            file = service.files().create(body=file_metadata, media_body=media, fields='id, webViewLink', supportsAllDrives=True).execute()
            file_id = file.get('id')

            if file_id:
                logger.info(f"Successfully uploaded '{file_name}' with ID: {file_id}")
                uploaded_file_urls.append(file.get('webViewLink'))
            else:
                logger.error(f"Upload succeeded but no file ID returned for '{file_name}'")

        except Exception as e:
            logger.error(f"Error uploading file '{file_path}' to Google Drive: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            
    return uploaded_file_urls

def upload_multiple_base64_to_google_drive_hierarchical(file_paths, theme, category, subcategory, root_folder_id='1rGP_5RIbYkqrJ1KEgXH8SHN8bFaBtkpm'):
    """
    Uploads multiple base64 text files to a dedicated 'Base64' subfolder within the hierarchical structure.
    """
    service = create_google_drive_service()
    if not service:
        logger.error("Failed to create Google Drive service for base64 file upload.")
        return []
    
    # Use the provided root folder ID directly
    if not root_folder_id:
        logger.error("No root folder ID provided.")
        return []

    # 1. Resolve Theme Folder
    theme_folder_name = THEME_FOLDER_MAPPING.get(theme)
    if not theme_folder_name: return []
    theme_folder_id = get_or_create_folder(service, root_folder_id, theme_folder_name)
    if not theme_folder_id: return []

    # 2. Resolve Category Folder
    category_folder_name = CATEGORY_FOLDER_MAPPING.get(theme, {}).get(category)
    if not category_folder_name: return []
    category_folder_id = get_or_create_folder(service, theme_folder_id, category_folder_name)
    if not category_folder_id: return []

    # 3. Resolve Subcategory Folder
    subcategory_folder_id = category_folder_id
    if subcategory:
        subcategory_folder_name = SUBCATEGORY_FOLDER_MAPPING.get(category, {}).get(subcategory)
        if subcategory_folder_name:
            resolved_id = get_or_create_folder(service, category_folder_id, subcategory_folder_name)
            if resolved_id:
                subcategory_folder_id = resolved_id
            else:
                return []

    # 4. Create a dedicated 'Base64' folder
    base64_parent_folder_id = get_or_create_folder(service, subcategory_folder_id, "Base64")
    if not base64_parent_folder_id:
        logger.error("Failed to get or create 'Base64' subfolder.")
        return []
        
    # 5. Upload all base64 files into the 'Base64' folder
    uploaded_file_urls = []
    for file_path in file_paths:
        try:
            if not os.path.exists(file_path):
                logger.warning(f"Base64 file not found, skipping: {file_path}")
                continue
            
            file_name = os.path.basename(file_path)
            mime_type = 'text/plain'
            
            file_metadata = {'name': file_name, 'parents': [base64_parent_folder_id]}
            media = MediaFileUpload(file_path, mimetype=mime_type)
            
            file = service.files().create(body=file_metadata, media_body=media, fields='id, webViewLink', supportsAllDrives=True).execute()
            file_id = file.get('id')

            if file_id:
                logger.info(f"Successfully uploaded base64 file '{file_name}' with ID: {file_id}")
                uploaded_file_urls.append(file.get('webViewLink'))
        except Exception as e:
            logger.error(f"Error uploading base64 file '{file_path}': {str(e)}")

    return uploaded_file_urls

def get_or_create_base64_folder(service, root_folder_id='1rGP_5RIbYkqrJ1KEgXH8SHN8bFaBtkpm'):
    """Get or create the 'Base64' parent folder for base64 encoded files."""
    return get_or_create_folder(service, root_folder_id, 'Base64')

def upload_base64_to_google_drive(file_path, theme=None, category=None, use_postqc=False):
    """Upload a base64 file specifically to the Base64 folder structure on Google Drive."""
    try:
        # Check if file exists first
        if not os.path.exists(file_path):
            logger.error(f"Base64 file does not exist: {file_path}")
            print(f"Google Drive base64 upload failed: File not found: {file_path}")
            return None
        
        # Create Google Drive service
        service = create_google_drive_service()
        if not service:
            logger.error("Failed to create Google Drive service for base64 upload")
            print("Google Drive base64 upload failed: Could not initialize Drive service")
            return None
        
        # Get or create the Base64 parent folder
        root_folder_id = '1rGP_5RIbYkqrJ1KEgXH8SHN8bFaBtkpm'
        base64_folder_id = get_or_create_base64_folder(service, root_folder_id)
        
        if not base64_folder_id:
            logger.error("Failed to get or create Base64 parent folder")
            print("Google Drive base64 upload failed: Could not access Base64 folder")
            return None
        
        parent_folder_id = base64_folder_id
        logger.info(f"Using Base64 folder as parent: {parent_folder_id}")
        print(f"[Google Drive Base64] Using Base64 folder as parent: {parent_folder_id}")
        
        # Create folder structure based on theme and category within Base64 folder
        if theme:
            theme_folder_id = get_or_create_folder(service, parent_folder_id, theme)
            if not theme_folder_id:
                logger.error(f"Failed to get or create theme folder '{theme}' in Base64")
                return None
            if category:
                category_folder_id = get_or_create_folder(service, theme_folder_id, category)
                if not category_folder_id:
                    logger.error(f"Failed to get or create category folder '{category}' in Base64")
                    return None
                parent_folder_id = category_folder_id
            else:
                parent_folder_id = theme_folder_id
        
        # Get file name from path
        file_name = os.path.basename(file_path)
        
        # Check if file already exists to avoid duplicates
        safe_file_name = file_name.replace("'", "\\'")
        query = f"name='{safe_file_name}' and '{parent_folder_id}' in parents and trashed=false"
        response = service.files().list(q=query, spaces='drive', fields='files(id, name)').execute()
        if response.get('files'):
            existing_file_id = response['files'][0]['id']
            logger.info(f"Base64 file '{file_name}' already exists with ID: {existing_file_id}")
            return existing_file_id
        
        # Determine MIME type
        mime_type, _ = mimetypes.guess_type(file_path)
        if not mime_type:
            mime_type = 'text/plain'  # Base64 files are typically text files
        
        # Upload file to Base64 folder
        file_metadata = {
            'name': file_name,
            'parents': [parent_folder_id]
        }
        
        try:
            media = MediaFileUpload(file_path, mimetype=mime_type)
            file = service.files().create(
                body=file_metadata,
                media_body=media,
                fields='id',
                supportsAllDrives=True).execute()
            file_id = file.get('id')
            
            if file_id:
                logger.info(f"Base64 file '{file_name}' uploaded to Google Drive Base64 folder with ID: {file_id}")
                print(f"Successfully uploaded base64 file '{file_name}' to Google Drive Base64 folder")
                return file_id
            else:
                logger.error(f"Base64 upload succeeded but no file ID returned for '{file_name}'")
                print(f"Base64 upload issue: No file ID returned for '{file_name}'")
                return None
                
        except Exception as upload_error:
            logger.error(f"Error during base64 file upload operation for '{file_name}': {str(upload_error)}")
            print(f"Google Drive base64 upload error for '{file_name}': {str(upload_error)}")
            return None
            
    except Exception as e:
        logger.error(f"Error in upload_base64_to_google_drive for '{file_path}': {str(e)}")
        print(f"Failed to upload base64 file to Google Drive ('{file_path}'): {str(e)}")
        return None

def upload_multiple_base64_to_google_drive(file_paths, theme=None, category=None, use_postqc=False):
    """Upload multiple base64 files specifically to the Base64 folder on Google Drive."""
    if not file_paths:
        logger.warning("No base64 files provided for Google Drive upload")
        print("No base64 files provided for Google Drive upload.")
        return []
    
    file_ids = []
    total_files = len(file_paths)
    successful = 0
    failed = 0
    
    logger.info(f"Starting upload of {total_files} base64 files to Google Drive Base64 folder")
    print(f"Starting upload of {total_files} base64 files to Google Drive Base64 folder...")
    
    for index, file_path in enumerate(file_paths):
        try:
            print(f"Uploading base64 file {index+1}/{total_files}: {os.path.basename(file_path)}")
            file_id = upload_base64_to_google_drive(file_path, theme, category, use_postqc)
            
            if file_id:
                file_ids.append(file_id)
                successful += 1
                logger.info(f"Base64 file {index+1}/{total_files} uploaded successfully: {file_id}")
            else:
                failed += 1
                logger.warning(f"Base64 file {index+1}/{total_files} upload failed: {file_path}")
        except Exception as e:
            failed += 1
            logger.error(f"Error uploading base64 file {index+1}/{total_files} ('{file_path}'): {str(e)}")
            print(f"Error during upload of base64 file {os.path.basename(file_path)}: {str(e)}")

    if total_files > 0:
        summary_message = f"Google Drive Base64 upload complete: {successful} successful, {failed} failed out of {total_files} total files."
        if failed > 0:
            logger.warning(summary_message)
        else:
            logger.info(summary_message)
        print(summary_message)
    
    if successful == 0 and total_files > 0:
        logger.error("All Google Drive Base64 uploads failed")

    return file_ids


def upload_multiple_files_to_google_drive(file_paths, parent_folder_id=None, theme=None, category=None, subcategory=None, use_postqc=False):    
    """Upload multiple files to Google Drive."""    
    if not file_paths:        
        logger.warning("No files provided for Google Drive upload")
        print("No files provided for Google Drive upload.") # Added print statement
        return []
    
    file_ids = []    
    total_files = len(file_paths)    
    successful = 0    
    failed = 0        
    logger.info(f"Starting upload of {total_files} files to Google Drive")    
    print(f"Starting upload of {total_files} files to Google Drive...")        
    for index, file_path in enumerate(file_paths):        
        try:            
            print(f"Uploading file {index+1}/{total_files}: {os.path.basename(file_path)}")
            # Use hierarchical upload if theme, category, and subcategory are provided
            if theme and category:
                urls = upload_multiple_files_to_google_drive_hierarchical(
                    [file_path], theme, category, subcategory
                )
                file_id = urls[0] if urls else None
            else:
                file_id = upload_to_google_drive(file_path, parent_folder_id, theme, category, subcategory, use_postqc)                        
            if file_id:                
                file_ids.append(file_id)                
                successful += 1                
                logger.info(f"File {index+1}/{total_files} uploaded successfully: {file_id}")            
            else:                
                failed += 1                
                logger.warning(f"File {index+1}/{total_files} upload failed: {file_path}")        
        except Exception as e:            
            failed += 1            
            logger.error(f"Error uploading file {index+1}/{total_files} ('{file_path}'): {str(e)}")
            print(f"Error during upload of {os.path.basename(file_path)}: {str(e)}")

    if total_files > 0: # Print summary only if there were files to process
        summary_message = f"Google Drive upload complete: {successful} successful, {failed} failed out of {total_files} total files."
        if failed > 0:
            logger.warning(summary_message)
        else:
            logger.info(summary_message)
        print(summary_message)
    
    if successful == 0 and total_files > 0:
        logger.error("All Google Drive uploads failed")

    return file_ids

def create_comprehensive_zip_with_all_files(
    original_images, nobg_images, card_images, excel_path=None, 
    theme=None, category=None, encode_to_base64=False
):
    """
    Create a comprehensive ZIP file containing all types of files:
    - Original generated images
    - Background-removed images
    - Card-applied images
    - Excel file with metadata
    
    Args:
        original_images (list): List of original generated image paths
        nobg_images (list): List of background-removed image paths
        card_images (list): List of card-applied image paths
        excel_path (str): Path to Excel file to include
        theme (str): Theme name for proper naming convention
        category (str): Category name for proper naming convention
        encode_to_base64 (bool): Whether to include base64 encoded files
        
    Returns:
        str: Path to the created ZIP file or None if failed
    """
    import tempfile
    import zipfile
    
    if not original_images and not nobg_images and not card_images:
        logger.warning("No images provided to create comprehensive ZIP file")
        return None
        
    # Create timestamp for unique filename
    timestamp = get_gmt7_filename_timestamp()
    
    # Create proper ZIP filename using theme/category codes
    if theme and category:
        theme_code = THEME_MAPPING.get(theme, "00")
        category_code = CATEGORY_MAPPING.get(category, "000")
        zip_name = f"{theme_code}{category_code}_comprehensive_{timestamp}.zip"
    else:
        # Fallback to generic name if theme/category not provided
        zip_name = f"comprehensive_images_{timestamp}.zip"
    
    # Create ZIP in temp directory
    temp_dir = tempfile.mkdtemp()
    zip_filepath = os.path.join(temp_dir, zip_name)
    
    try:
        # Create ZIP file
        file_counter = 1  # Counter for sequential numbering
        processed_image_paths = []  # Keep track of actual file paths for base64 processing
        
        with zipfile.ZipFile(zip_filepath, 'w', compression=zipfile.ZIP_DEFLATED) as zipf:
            # Add original images with proper naming
            for img_path in original_images:
                actual_path = None
                
                # Handle different types of image paths/objects
                if isinstance(img_path, str):
                    actual_path = img_path
                elif hasattr(img_path, 'save'):
                    # It's a PIL Image, save it to temporary file first
                    temp_dir_img = tempfile.mkdtemp()
                    temp_filename = f"temp_image_{len(os.listdir(temp_dir_img))}.png"
                    actual_path = os.path.join(temp_dir_img, temp_filename)
                    
                    try:
                        img_path.save(actual_path)
                        logger.info(f"Saved PIL Image to temporary file: {actual_path}")
                    except Exception as save_error:
                        logger.error(f"Error saving PIL Image to file: {str(save_error)}")
                        continue
                elif isinstance(img_path, (tuple, list)) and len(img_path) > 0:
                    actual_path = img_path[0] if isinstance(img_path[0], str) else None
                elif isinstance(img_path, dict) and 'path' in img_path:
                    actual_path = img_path['path']
                else:
                    logger.warning(f"Unsupported image path format in ZIP creation: {type(img_path)}")
                    continue
                
                if actual_path and os.path.exists(actual_path):
                    # Get file extension
                    file_ext = os.path.splitext(actual_path)[1].lower()
                    if not file_ext:
                        file_ext = '.png'  # Default to PNG if no extension
                    
                    # Create proper filename using TTCCCNNNNN convention
                    if theme and category:
                        theme_code = THEME_MAPPING.get(theme, "00")
                        category_code = CATEGORY_MAPPING.get(category, "000")
                        # Format: TTCCCNNNNN_original.ext (e.g., 0100200001_original.png for Pets/Cats)
                        proper_filename = f"{theme_code}{category_code}{file_counter:05d}_original{file_ext}"
                    else:
                        # Fallback to generic naming if theme/category not provided
                        proper_filename = f"image_{file_counter:03d}_original{file_ext}"
                    
                    # Add file to ZIP with proper filename
                    zipf.write(actual_path, proper_filename)
                    processed_image_paths.append(actual_path)  # Track for base64 processing
                    logger.info(f"Added original image to ZIP with proper naming: {proper_filename}")
                    file_counter += 1
                else:
                    logger.warning(f"Original image file not found for ZIP: {actual_path}")
            
            # Add background-removed images with proper naming
            for img_path in nobg_images:
                if isinstance(img_path, str) and os.path.exists(img_path):
                    # Get file extension
                    file_ext = os.path.splitext(img_path)[1].lower()
                    if not file_ext:
                        file_ext = '.png'  # Default to PNG if no extension
                    
                    # Create proper filename using TTCCCNNNNN convention
                    if theme and category:
                        theme_code = THEME_MAPPING.get(theme, "00")
                        category_code = CATEGORY_MAPPING.get(category, "000")
                        # Format: TTCCCNNNNN_nobg.ext (e.g., 0100200001_nobg.png for Pets/Cats)
                        proper_filename = f"{theme_code}{category_code}{file_counter:05d}_nobg{file_ext}"
                    else:
                        # Fallback to generic naming if theme/category not provided
                        proper_filename = f"image_{file_counter:03d}_nobg{file_ext}"
                    
                    # Add file to ZIP with proper filename
                    zipf.write(img_path, proper_filename)
                    processed_image_paths.append(img_path)  # Track for base64 processing
                    logger.info(f"Added no-bg image to ZIP with proper naming: {proper_filename}")
                    file_counter += 1
                else:
                    logger.warning(f"No-bg image file not found for ZIP: {img_path}")
            
            # Add card-applied images with proper naming
            for img_path in card_images:
                if isinstance(img_path, str) and os.path.exists(img_path):
                    # Get file extension
                    file_ext = os.path.splitext(img_path)[1].lower()
                    if not file_ext:
                        file_ext = '.png'  # Default to PNG if no extension
                    
                    # Create proper filename using TTCCCNNNNN convention
                    if theme and category:
                        theme_code = THEME_MAPPING.get(theme, "00")
                        category_code = CATEGORY_MAPPING.get(category, "000")
                        # Format: TTCCCNNNNN_card.ext (e.g., 0100200001_card.png for Pets/Cats)
                        proper_filename = f"{theme_code}{category_code}{file_counter:05d}_card{file_ext}"
                    else:
                        # Fallback to generic naming if theme/category not provided
                        proper_filename = f"image_{file_counter:03d}_card{file_ext}"
                    
                    # Add file to ZIP with proper filename
                    zipf.write(img_path, proper_filename)
                    processed_image_paths.append(img_path)  # Track for base64 processing
                    logger.info(f"Added card image to ZIP with proper naming: {proper_filename}")
                    file_counter += 1
                else:
                    logger.warning(f"Card image file not found for ZIP: {img_path}")
            
            # Add base64 encoded files if requested
            if encode_to_base64 and processed_image_paths:
                logger.info("Base64 encoding enabled - adding base64 files to ZIP")
                try:
                    base64_files = batch_encode_images_to_base64(processed_image_paths)
                    for base64_file in base64_files:
                        if os.path.exists(base64_file):
                            # Add base64 file to ZIP using just the filename
                            zipf.write(base64_file, os.path.basename(base64_file))
                            logger.info(f"Added base64 file to ZIP: {os.path.basename(base64_file)}")
                        else:
                            logger.warning(f"Base64 file not found for ZIP: {base64_file}")
                    logger.info(f"Successfully added {len(base64_files)} base64 files to ZIP")
                except Exception as base64_error:
                    logger.error(f"Error adding base64 files to ZIP: {str(base64_error)}")
                    # Continue without base64 files - don't fail the entire ZIP creation
            
            # Add Excel file if provided
            if excel_path and os.path.exists(excel_path):
                zipf.write(excel_path, os.path.basename(excel_path))
                logger.info(f"Added Excel file to ZIP: {os.path.basename(excel_path)}")
        
        logger.info(f"Created comprehensive ZIP file with all file types: {zip_filepath}")
        return zip_filepath
        
    except Exception as e:
        logger.error(f"Error creating comprehensive ZIP file: {str(e)}")
        return None

if __name__ == "__main__":
    # Create requirements.txt file if it doesn't exist
    create_requirements_file()
    
    # Ensure logging is properly set up
    print("="*60)
    print("Starting  Image Generator")
    print("Logging level:", logging.getLevelName(logger.level))
    print("Logger handlers:", logger.handlers)
    print("="*60)
    logger.info("Application starting")
    
    # Run standalone Gradio app with sharing enabled
    gradio_app = create_gradio_ui()
    gradio_app.launch(server_name="0.0.0.0", share=True)
